from datetime import datetime
from decimal import Decimal
from zoneinfo import ZoneInfo
from django.conf import settings
from django.db.models import Q
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet
from rest_framework.views import APIView
from rest_framework.generics import GenericAPIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import ValidationError, MethodNotAllowed
from django.db import IntegrityError, models
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.urls import reverse
import json

from user.utils import send_excel_link_email, send_pdf_link_email
from .models import (
    Form, Question, QuestionType, StageAssignment, FormType,
    Folder, StageAccessType, FormAssignment, FormAssignType,
    Stage, CustomUser, StageAccess, Answer,
    Groups, FormSubmision, StageSubmissionHistory,
    GroupAssignment, AuditGroup, AuditInfo, AuditFormSubmissionHistory, FormResponseShare, FormPayloadFiles,LogicFollowUp,
    FormAutoShareConfig,
    TaskCloseQuestion, Logic, Option,
    FollowUpTask, FollowUpTaskCloseQuestionResponse
)
from task.models import Task, TaskAssignee, TaskAuditLog
from reportlab.lib.styles import ParagraphStyle
from io import BytesIO
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import A4 ,landscape
# from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle, HRFlowable, Image, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import HRFlowable
from django.http import HttpResponse, HttpResponseRedirect
from django.core.mail import EmailMessage
import requests
from PIL import Image as PILImage
import pandas as pd
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
from .serializers import (
    FormSerializer, FolderSerializer, FormAssignmentSerializer,
    AnswerSerializer, CompactFormSerializer, FormSubmissionSerializer,
    StageSubmissionHistorySerializer, FormListSerializer,StageSerializer,
    FormSubmissionShareSerializer, FollowUpTaskSerializer,
    FollowUpTaskResponseSerializer, FormAnswerEditSerializer, FormResponseSummarySerializer, FormToggleSerializer, FormPayloadFilesSerializer,
    AuditFormSubmissionHistorySerializer, QuestionSerializer
    )
from vibro.permissions import IsAdmin, IsAdminOrSuperAdmin, IsEndUser, IsEndUserOrAdmin,IsLocationLeader
from vibro.views import userContextAPIView
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import action
from django.shortcuts import get_object_or_404
from user.models import OrganizationAdmin, Groups
from django.db import transaction
from rest_framework.permissions import IsAuthenticated
import uuid
import logging
import copy
from django.db.models import Max, Q
logger = logging.getLogger(__name__)
from rest_framework.permissions import AllowAny
from reportlab.platypus import PageBreak
from reportlab.platypus.doctemplate import PageTemplate, Frame
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from .optimized_form_serializer import OptimizedFormSerializer
import re
from urllib.parse import urlparse
from vibro.utils import UtilsFunctions
import json
import io
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Q, Case, When, IntegerField, Sum, OuterRef, Subquery
import threading
PAYLOAD_THRESHOLD_BYTES = 1024 * 1024  # 256 KB
from django.core.cache import cache
from django.db import connection


def _is_local_host(host):
    if not host:
        return False
    return host.split(":")[0].lower() in {"localhost", "127.0.0.1", "0.0.0.0"}


def _media_url(path):
    """Build a public media URL using BACKEND_BASE_URL from settings."""
    base = getattr(settings, 'BACKEND_BASE_URL', 'http://localhost:8000').rstrip('/')
    return f"{base}/media/{path}"


def _ensure_production_https(url):
    """
    Use HTTP for local dev (runserver) and HTTPS for deployed hosts.
    """
    if not url:
        return url
    parsed = urlparse(url)
    host = parsed.hostname or (parsed.netloc.split(":")[0] if parsed.netloc else "")
    if _is_local_host(host):
        if parsed.scheme == "https":
            return "http://" + url[len("https://") :]
        return url
    if parsed.scheme == "http":
        return "https://" + url[len("http://") :]
    return url


def _build_public_download_url(request, relative_path):
    base = getattr(settings, 'BACKEND_BASE_URL', None)
    if base:
        return f"{base.rstrip('/')}{relative_path}"
    return _ensure_production_https(request.build_absolute_uri(relative_path))


def _build_public_base_url(request):
    base = getattr(settings, 'BACKEND_BASE_URL', None)
    if base:
        return base.rstrip('/')
    return _ensure_production_https(request.build_absolute_uri("/")).rstrip("/")


def _auto_share_submission_response(submission, actor, base_url=None):
    """
    Create automatic response shares from the form's stored auto-share configuration.

    This intentionally does not modify manual sharing behavior. It only creates
    missing share records after a submission is completed.
    """
    form = submission.form
    logger.info(
        "Auto-share check for submission %s: form_id=%s auto_share_response=%s",
        submission.id,
        form.id,
        form.auto_share_response,
    )

    if not form.auto_share_response:
        return

    config = (
        FormAutoShareConfig.objects.filter(form=form, organization=submission.organization)
        .prefetch_related('users', 'groups__members', 'location_leaders')
        .first()
    )
    if not config:
        logger.info(
            "Auto-share enabled but no config found for form_id=%s submission_id=%s",
            form.id,
            submission.id,
        )
        return

    selected_users = list(
        config.users.filter(organization=submission.organization).values_list('id', flat=True)
    )
    selected_groups = list(
        config.groups.filter(organization=submission.organization).values_list('id', flat=True)
    )
    selected_leaders = list(
        config.location_leaders.filter(organization=submission.organization).values_list('id', flat=True)
    )

    resolved_user_ids = set(selected_users)
    for group in config.groups.filter(organization=submission.organization).prefetch_related('members'):
        resolved_user_ids.update(
            group.members.filter(organization=submission.organization).values_list('id', flat=True)
        )
    resolved_user_ids.update(selected_leaders)

    logger.info(
        "Auto-share config for submission %s: selected_user_ids=%s selected_group_ids=%s selected_location_leader_ids=%s resolved_shared_user_ids=%s",
        submission.id,
        selected_users,
        selected_groups,
        selected_leaders,
        sorted(resolved_user_ids),
    )

    created_share_records = []

    for user_id in selected_users:
        share, created = FormResponseShare.objects.get_or_create(
            form_submission=submission,
            shared_to_user_id=user_id,
            share_type='user',
            defaults={
                'shared_by': actor,
                'organization': submission.organization,
            },
        )
        if created:
            created_share_records.append({'type': 'user', 'recipient_id': user_id, 'share_id': share.id})

    for group_id in selected_groups:
        share, created = FormResponseShare.objects.get_or_create(
            form_submission=submission,
            shared_to_group_id=group_id,
            share_type='group',
            defaults={
                'shared_by': actor,
                'organization': submission.organization,
            },
        )
        if created:
            created_share_records.append({'type': 'group', 'recipient_id': group_id, 'share_id': share.id})

    for leader_id in selected_leaders:
        share, created = FormResponseShare.objects.get_or_create(
            form_submission=submission,
            shared_to_leader_id=leader_id,
            share_type='location_leader',
            defaults={
                'shared_by': actor,
                'organization': submission.organization,
            },
        )
        if created:
            created_share_records.append({'type': 'location_leader', 'recipient_id': leader_id, 'share_id': share.id})

    logger.info(
        "Auto-share created records for submission %s: created_share_records=%s",
        submission.id,
        created_share_records,
    )
    logger.info(
        "Auto-share leader share records for submission %s: share_ids=%s leader_user_ids=%s",
        submission.id,
        list(
            FormResponseShare.objects.filter(
                form_submission=submission,
                share_type='location_leader'
            ).values_list('id', flat=True)
        ),
        list(
            FormResponseShare.objects.filter(
                form_submission=submission,
                share_type='location_leader'
            ).values_list('shared_to_leader_id', flat=True)
        ),
    )

    _email_auto_shared_response_pdf(submission, actor, config, resolved_user_ids, base_url)


def _email_auto_shared_response_pdf(submission, actor, config, resolved_user_ids, base_url=None):
    """
    Email the completed response PDF to all resolved auto-share recipients.
    Runs in a background thread to avoid affecting submission latency.
    """
    recipient_qs = CustomUser.objects.filter(
        id__in=resolved_user_ids,
        organization=submission.organization
    ).exclude(email__isnull=True).exclude(email__exact="")
    recipient_emails = list(recipient_qs.values_list('email', flat=True).distinct())

    logger.info(
        "Auto-share email recipients for submission %s: recipient_user_ids=%s recipient_emails=%s",
        submission.id,
        sorted(resolved_user_ids),
        recipient_emails,
    )

    if not recipient_emails:
        logger.info("Auto-share email skipped for submission %s: no recipient emails found", submission.id)
        return

    def background_auto_share_email():
        try:
            if not base_url:
                logger.warning(
                    "Auto-share PDF link email skipped for submission %s: base_url unavailable",
                    submission.id,
                )
                return

            form = (
                Form.objects
                .filter(id=submission.form_id, organization=submission.organization, is_deleted=False)
                .select_related('folder', 'form_admin', 'deletedBy', 'archivedBy')
                .first()
            )
            if not form:
                logger.warning("Auto-share email skipped: form %s not found", submission.form_id)
                return

            stage_access_qs = StageAccess.objects.select_related('allow_user', 'allow_group', 'form', 'stage')
            if form.form_type == FormType.AUDIT:
                qs = Form.objects.filter(pk=form.id).select_related(
                    'folder', 'form_admin', 'deletedBy', 'archivedBy'
                ).prefetch_related(
                    'assignee__user', 'assignee__group', 'assignee__leader',
                    'audit_info__questions',
                    'audit_info__questions__options',
                    'audit_info__questions__child_questions',
                    'audit_info__questions__child_questions__options',
                    'audit_info__questions__logic_parent_question__logic_questions__options',
                    'audit_info__questions__logic_parent_question__follow_ups',
                    'audit_info__questions__logic_parent_question__follow_ups__task_close_questions__options',
                    'audit_group__questions',
                    'audit_group__questions__options',
                    'audit_group__questions__child_questions',
                    'audit_group__questions__child_questions__options',
                    'audit_group__questions__logic_parent_question__logic_questions__options',
                    'audit_group__questions__logic_parent_question__follow_ups',
                    'audit_group__questions__logic_parent_question__follow_ups__task_close_questions__options',
                )
            else:
                question_qs = (
                    Question.objects
                    .select_related('form', 'stage', 'parent_question')
                    .prefetch_related(
                        'options',
                        'child_questions',
                        'child_questions__options',
                        'child_questions__child_questions',
                        'child_questions__child_questions__options',
                        'child_questions__logic_parent_question__logic_questions__options',
                        'child_questions__logic_parent_question__follow_ups',
                        'child_questions__logic_parent_question__follow_ups__task_close_questions__options',
                        'logic_parent_question__logic_questions__options',
                        'logic_parent_question__follow_ups',
                        'logic_parent_question__follow_ups__task_close_questions__options',
                    )
                )
                stage_qs = (
                    Stage.objects
                    .select_related('form')
                    .prefetch_related(
                        models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                        models.Prefetch('questions', queryset=question_qs),
                    )
                )
                qs = Form.objects.filter(pk=form.id).select_related(
                    'folder', 'form_admin', 'deletedBy', 'archivedBy'
                ).prefetch_related(
                    'assignee__user', 'assignee__group', 'assignee__leader',
                    models.Prefetch('stages', queryset=stage_qs),
                )

            optimized_form = qs.get(pk=form.id)
            form_schema = FormSerializer(optimized_form, many=False).data

            submission_instance = FormSubmision.objects.filter(
                id=submission.id,
                form_id=form.id,
                organization=submission.organization
            ).select_related(
                'submission_initiated_by',
                'completed_by'
            ).prefetch_related(
                models.Prefetch(
                    'answers',
                    queryset=Answer.objects.select_related('question').prefetch_related('question__options'),
                    to_attr='prefetched_answers'
                ),
                'stage_submissions_history',
                'group_submissions_history'
            ).first()

            if not submission_instance:
                logger.warning("Auto-share email skipped: submission %s not found", submission.id)
                return

            response_data = {
                'id': submission_instance.id,
                'submission_initiated_on': submission_instance.submission_initiated_on.isoformat() if submission_instance.submission_initiated_on else None,
                'submission_initiated_by': getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                'initiator_designation': str(getattr(submission_instance.submission_initiated_by, 'designation', 'N/A')) if submission_instance.submission_initiated_by else 'N/A',
                'initiator_department': str(getattr(submission_instance.submission_initiated_by, 'department', 'N/A')) if submission_instance.submission_initiated_by else 'N/A',
                'initiator_location': str(getattr(submission_instance.submission_initiated_by, 'location', 'N/A')) if submission_instance.submission_initiated_by else 'N/A',
                'current_owner': getattr(submission_instance.completed_by, 'first_name', '') + ' ' + getattr(submission_instance.completed_by, 'last_name', '') if submission_instance.completed_by else 'N/A',
                'is_completed': submission_instance.is_completed,
                'stages': []
            }
            answer_lookup = {a.question_id: a for a in getattr(submission_instance, 'prefetched_answers', [])}

            def attach_answer_data(answer_obj):
                return AnswerSerializer(answer_obj).data if answer_obj else {}

            def replace_option_id_with_text(answers_data, options):
                if answers_data and 'answer' in answers_data and options:
                    answer_value = answers_data['answer']
                    raw_answer_value = answers_data.get('answer_id', answer_value)
                    other_text = answers_data.get('other_text')

                    def get_export_option_text(option_label):
                        if option_label is None:
                            return option_label
                        if other_text and str(option_label).strip().lower() == 'other':
                            return other_text
                        return option_label

                    def find_option_by_id(option_id):
                        option_id = str(option_id).strip()
                        for opt in options:
                            if opt and str(opt.get('id', '')).strip() == option_id:
                                return opt
                        return None

                    def replace_text_if_other(text_value):
                        if other_text and str(text_value).strip().lower() == 'other':
                            return other_text
                        return text_value

                    if answer_value:
                        answer_str = str(answer_value).strip()
                        raw_answer_str = str(raw_answer_value).strip()

                        if other_text and answer_str.lower() == 'other':
                            answers_data['answer'] = other_text
                        elif ',' in answer_str:
                            ids = [id.strip() for id in answer_str.split(',') if id.strip()]
                            replaced = []
                            for id_val in ids:
                                if id_val.isdigit():
                                    opt = find_option_by_id(id_val)
                                    if opt:
                                        replaced.append(get_export_option_text(opt.get('option')))
                                    else:
                                        replaced.append(id_val)
                                else:
                                    replaced.append(replace_text_if_other(id_val))
                            answers_data['answer'] = ', '.join(replaced)
                        elif answer_str.isdigit():
                            opt = find_option_by_id(answer_str)
                            if opt:
                                answers_data['answer'] = get_export_option_text(opt.get('option'))
                        else:
                            raw_option_ids = re.findall(r'\d+', raw_answer_str)
                            if raw_option_ids:
                                matched_options = []
                                for option_id in raw_option_ids:
                                    opt = find_option_by_id(option_id)
                                    if opt:
                                        matched_options.append(get_export_option_text(opt.get('option')))
                                if matched_options:
                                    answers_data['answer'] = ', '.join(matched_options)
                                    return

                            answers_data['answer'] = replace_text_if_other(answer_value)

            if form_schema.get("form_type") == FormType.AUDIT:
                stages = []
                audit_info = form_schema.get("audit_info")
                if audit_info:
                    audit_info_copy = copy.deepcopy(audit_info)
                    audit_info_stage = {'id': audit_info_copy.get('id', 'audit_info'), 'name': audit_info_copy.get('name', 'Audit Info'), 'is_audit_info': True, 'order': 0, 'questions': audit_info_copy.get("questions", [])}
                    for question in audit_info_stage['questions']:
                        question['answers'] = attach_answer_data(answer_lookup.get(question['id']))
                        replace_option_id_with_text(question['answers'], question.get('options'))
                        for sub_question in question.get('sub_questions', []):
                            sub_question['answers'] = attach_answer_data(answer_lookup.get(sub_question['id']))
                            replace_option_id_with_text(sub_question['answers'], sub_question.get('options'))
                        for logic in question.get("logics", []):
                            for logic_question in logic.get("logic_questions", []):
                                logic_question['answers'] = attach_answer_data(answer_lookup.get(logic_question['id']))
                                replace_option_id_with_text(logic_question['answers'], logic_question.get('options'))
                    stages.append(audit_info_stage)
                for audit_group in form_schema.get("audit_group", []):
                    audit_group_copy = copy.deepcopy(audit_group)
                    for audit_group_question in audit_group_copy.get("questions", []):
                        audit_group_question['answers'] = attach_answer_data(answer_lookup.get(audit_group_question['id']))
                        replace_option_id_with_text(audit_group_question['answers'], audit_group_question.get('options'))
                        for logic in audit_group_question.get("logics", []):
                            for logic_question in logic.get("logic_questions", []):
                                logic_question['answers'] = attach_answer_data(answer_lookup.get(logic_question['id']))
                                replace_option_id_with_text(logic_question['answers'], logic_question.get('options'))
                        for sub_question in audit_group_question.get('sub_questions', []):
                            sub_question['answers'] = attach_answer_data(answer_lookup.get(sub_question['id']))
                            replace_option_id_with_text(sub_question['answers'], sub_question.get('options'))
                    stages.append(audit_group_copy)
                response_data['stages'] = stages
            else:
                stage_history_by_id = {h.stage_id: h for h in submission_instance.stage_submissions_history.all()}
                for stage in form_schema.get('stages', []):
                    stage_copy = copy.deepcopy(stage)
                    stage_history = stage_history_by_id.get(stage_copy['id'])
                    history_data = StageSubmissionHistorySerializer(stage_history, many=False).data if stage_history else {}
                    stage_data = {
                        'id': stage_copy['id'],
                        'name': stage_copy.get('name', ''),
                        'order': stage_copy.get('order', 0),
                        'is_completed': stage_history is not None,
                        'completed_by': history_data.get('completed_by'),
                        'completed_on': history_data.get('completed_on'),
                        'questions': []
                    }
                    for question in stage_copy.get('questions', []):
                        question_data = {
                            'id': question['id'],
                            'question': question.get('question', ''),
                            'question_type': question.get('question_type', ''),
                            'order': question.get('order', 0),
                            'answers': attach_answer_data(answer_lookup.get(question['id']))
                        }
                        replace_option_id_with_text(question_data['answers'], question.get('options'))
                        sub_questions = []
                        for sub_question in question.get('sub_questions', []):
                            sub_question_data = {
                                'id': sub_question['id'],
                                'question': sub_question.get('question', ''),
                                'question_type': sub_question.get('question_type', ''),
                                'answers': attach_answer_data(answer_lookup.get(sub_question['id']))
                            }
                            replace_option_id_with_text(sub_question_data['answers'], sub_question.get('options'))
                            sub_questions.append(sub_question_data)
                        if sub_questions:
                            question_data['sub_questions'] = sub_questions
                        logic_blocks = []
                        for logic in question.get("logics", []):
                            logic_questions = []
                            for logic_question in logic.get("logic_questions", []):
                                logic_question_data = {
                                    'id': logic_question['id'],
                                    'question': logic_question.get('question', ''),
                                    'question_type': logic_question.get('question_type', ''),
                                    'answers': attach_answer_data(answer_lookup.get(logic_question['id']))
                                }
                                replace_option_id_with_text(logic_question_data['answers'], logic_question.get('options'))
                                logic_questions.append(logic_question_data)
                            if logic_questions:
                                logic_blocks.append({'logic_questions': logic_questions})
                        if logic_blocks:
                            question_data['logics'] = logic_blocks
                        stage_data['questions'].append(question_data)
                    response_data['stages'].append(stage_data)

            form_info = {
                'title': form.title or f'Form {form.id}',
                'form_type': form.get_form_type_display() or 'standard',
                'created_by': getattr(form.form_admin, 'first_name', '') + ' ' + getattr(form.form_admin, 'last_name', '') if form.form_admin else 'N/A',
                'created_at': form.created_at.isoformat() if form.created_at else None,
                'prefix': getattr(form, 'prefix', None) or ''
            }
            pdf = generate_audit_pdf([response_data], form_info, form.id, form.organization, is_audit_form=(form.form_type == FormType.AUDIT))
            timestamp = timezone.now().strftime("%d_%m_%y_%H_%M")
            filename = f"{timestamp}_form_{form.id}_submission_{submission.id}.pdf"
            tracking_id = uuid.uuid4()
            cache_key = f"pdf_job:{tracking_id}"
            upload_timestamp = timezone.now().strftime("%d%m%y_%H%M%S")
            local_file_path = f"reports/pdfs/{submission.organization.id}/{upload_timestamp}_{filename}"
            bucket_name = getattr(settings, "S3_BUCKET_NAME_ANNOUNCEMENT", None) or settings.S3_BUCKET_NAME
            pdf_buffer = io.BytesIO(pdf if isinstance(pdf, (bytes, bytearray)) else pdf.getvalue())
            uf = UtilsFunctions()
            
            # Save file locally for development
            import os
            local_dir = os.path.join(settings.BASE_DIR, "media", "reports", "pdfs", str(submission.organization.id))
            os.makedirs(local_dir, exist_ok=True)
            local_file_full_path = os.path.join(local_dir, f"{upload_timestamp}_{filename}")
            with open(local_file_full_path, 'wb') as f:
                f.write(pdf_buffer.getvalue())
            
            # Use local file URL instead of S3
            s3_url = _media_url(local_file_path)
            
            expires_in_seconds = 86400
            short_download_url = s3_url
            now_iso = timezone.now().isoformat()
            cache.set(cache_key, {
                "status": "SUCCESS",
                "message": "Auto-share PDF link ready",
                "form_id": form.id,
                "submission_ids": [submission.id],
                "email": recipient_emails,
                "filename": filename,
                "bucket_name": bucket_name,
                "s3_key": local_file_path,
                "expires_in_seconds": expires_in_seconds,
                "created_at": now_iso,
                "started_at": now_iso,
                "completed_at": now_iso,
                "error": None,
            }, timeout=86400)

            sent_emails = []
            for email in recipient_emails:
                try:
                    send_pdf_link_email(email, short_download_url, filename, expires_in_seconds)
                    sent_emails.append(email)
                except Exception as email_error:
                    logger.error(
                        "Auto-share PDF email failed for submission %s to %s: %s",
                        submission.id,
                        email,
                        str(email_error),
                    )

            logger.info(
                "Auto-share PDF email completed for submission %s: sent_emails=%s",
                submission.id,
                sent_emails,
            )
        except Exception as e:
            logger.error("Auto-share PDF email generation failed for submission %s: %s", submission.id, str(e))

    threading.Thread(target=background_auto_share_email, daemon=False).start()


# userContextAPIView - this view take the current user snapshot and store it in a global stage.
# Where-ever logged user information rrequired we can call get_current_user to get current user data

_FORM_EMAIL_TOGGLE_COLUMN = None
_FORM_EMAIL_TOGGLE_COLUMN_CHECKED = False

def _is_form_email_notification_enabled(form):
    """
    Returns True only when the form's 'Trigger Email Notifications' toggle is enabled.
    Safe across environments where the underlying column name differs or is not mapped in the Form model.
    """
    # 1) Prefer model fields/properties if present (fast, no SQL).
    candidate_attrs = (
        "trigger_email_notifications",
        "trigger_email_notification",
        "email_notifications",
        "email_notification",
        "email_notifications_enabled",
        "notification",  # fallback if DB column is named like this for forms
    )
    for attr in candidate_attrs:
        if hasattr(form, attr):
            value = getattr(form, attr, None)
            if value is not None:
                return bool(value)

    # 2) Fallback: look for a matching DB column on form_master and read it via SQL.
    # This avoids adding model fields (and potential DB mismatch) while still honoring the toggle.
    global _FORM_EMAIL_TOGGLE_COLUMN, _FORM_EMAIL_TOGGLE_COLUMN_CHECKED
    try:
        if not _FORM_EMAIL_TOGGLE_COLUMN_CHECKED:
            table_name = str(getattr(form._meta, "db_table", "") or "")
            if not table_name:
                _FORM_EMAIL_TOGGLE_COLUMN_CHECKED = True
                return False

            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = current_schema()
                      AND table_name = %s
                    """,
                    [table_name],
                )
                columns = {row[0] for row in cursor.fetchall()}

            for col in candidate_attrs:
                if col in columns:
                    _FORM_EMAIL_TOGGLE_COLUMN = col
                    break

            _FORM_EMAIL_TOGGLE_COLUMN_CHECKED = True

        if not _FORM_EMAIL_TOGGLE_COLUMN:
            return False

        table_sql = connection.ops.quote_name(form._meta.db_table)
        col_sql = connection.ops.quote_name(_FORM_EMAIL_TOGGLE_COLUMN)

        with connection.cursor() as cursor:
            cursor.execute(
                f"SELECT {col_sql} FROM {table_sql} WHERE id = %s",
                [form.id],
            )
            row = cursor.fetchone()
        return bool(row[0]) if row else False
    except Exception:
        # Never fail submissions due to notification toggle lookups.
        return False

def _notify_form_admin_on_submission_commit(request, form, submission, stage=None):
    """
    Best-effort: notify form admin when a submission is made.
    Runs via transaction.on_commit so email failures don't roll back form submission.
    """
    try:
        if not _is_form_email_notification_enabled(form):
            return

        admin_user = getattr(form, "form_admin", None)
        admin_email = (getattr(admin_user, "email", None) or "").strip()
        if not admin_email:
            return

        submitted_by = getattr(request, "user", None)
        submitted_by_name = ""
        if submitted_by:
            submitted_by_name = (
                f"{getattr(submitted_by, 'first_name', '')} {getattr(submitted_by, 'last_name', '')}"
            ).strip() or getattr(submitted_by, "username", "") or getattr(submitted_by, "email", "") or ""

        response_path = f"/api/forms/{form.id}/submissions/pdf/download?submission_id={submission.id}"
        response_url = _build_public_download_url(request, response_path) if request else response_path

        stage_line = ""
        if stage is not None:
            stage_name = getattr(stage, "name", "") or ""
            stage_order = getattr(stage, "order", None)
            stage_line = f"Stage: {stage_name or 'Stage'}{f' (Order {stage_order})' if stage_order is not None else ''}\n"

        submitted_on = getattr(submission, "completed_on", None) or getattr(submission, "submission_initiated_on", None)
        submitted_on_str = submitted_on.strftime("%d-%b-%Y %I:%M %p") if submitted_on else ""

        from_email = getattr(settings, "EMAIL_FROM_ADDRESS", None) or getattr(settings, "DEFAULT_FROM_EMAIL", None) or "no-reply@vibro.com"
        submission_status = "Completed" if getattr(submission, "is_completed", False) else "In progress"

        email = EmailMessage(
            subject=f"New form submission: {getattr(form, 'title', 'Untitled')}",
            body=(
                "A new form submission has been received.\n\n"
                f"Form: {getattr(form, 'title', 'Untitled')}\n"
                f"{stage_line}"
                f"Submission ID: {getattr(submission, 'id', '')}\n"
                f"Status: {submission_status}\n"
                f"Submitted by: {submitted_by_name}\n"
                f"Submitted on: {submitted_on_str}\n\n"
                f"View response: {response_url}\n"
            ),
            from_email=from_email,
            to=[admin_email],
        )
        email.send(fail_silently=False)
    except Exception as e:
        logger.error(f"Failed to notify form admin for form {getattr(form, 'id', None)} submission {getattr(submission, 'id', None)}: {str(e)}")

class FormCountsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        organization = request.user.organization

        total_forms = Form.objects.filter(organization=organization, is_deleted=False, is_archived=False).count()
        standard_forms = Form.objects.filter(organization=organization, is_deleted=False, is_archived=False, form_type=FormType.STANDARD).count()
        location_forms = Form.objects.filter(organization=organization, is_deleted=False, is_archived=False, form_type=FormType.LOCATION).count()
        audit_forms = Form.objects.filter(organization=organization, is_deleted=False, is_archived=False, form_type=FormType.AUDIT).count()

        return Response({
            'total_forms': total_forms,
            'standard_forms': standard_forms,
            'location_forms': location_forms,
            'audit_forms': audit_forms
        })


class QuestionTypeView(APIView):
    permission_classes = [IsAdmin]
    def get(self, request):
        
        questionTypeList = []
        for choice in QuestionType:
            if choice.value == QuestionType.AUDIT and request.query_params.get('form_type', None) != FormType.AUDIT:
                continue
            questionTypeList.append({ "key": choice.name, "value": choice.value, "label": choice.label })
        return Response(questionTypeList)


class FormTypeView(APIView):
    permission_classes = [IsAdmin]
    def get(self, request):
        return Response([{ "key": choice.name, "value": choice.value, "label": choice.label } for choice in FormType])
    
    
class FormAssignmentTypeView(APIView):
    permission_classes = [IsAdmin]
    def get(self, request):
        return Response([{ "key": choice.name, "value": choice.value, "label": choice.label } for choice in FormAssignType])
    

class StageAccessTypeView(APIView):
    permission_classes = [IsAdmin]
    def get(self, request):
        return Response([{ "key": choice.name, "value": choice.value, "label": choice.label } for choice in StageAccessType])



class FolderViewSet(userContextAPIView, ModelViewSet):
    permission_classes = [IsEndUserOrAdmin]
    serializer_class = FolderSerializer

    def get_queryset(self):
        queryset = Folder.objects.filter(organization=self.request.user.organization)
        if self.action == 'list':
            parent_id = self.request.query_params.get('parent')
            if parent_id:
                queryset = queryset.filter(parent_id=parent_id)
            else:
                queryset = queryset.filter(parent__isnull=True)
        return queryset

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except IntegrityError as e:
            return Response(
                {"error": "A folder with this name already exists in your organization."},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except IntegrityError as e:
            return Response(
                {"error": "A folder with this name already exists in your organization."},
                status=status.HTTP_400_BAD_REQUEST
            )


class FormViewSet(userContextAPIView, ModelViewSet):
    # permission_classes = [IsEndUserOrAdmin]
    permission_classes = [AllowAny]
    queryset = Form.objects.filter(is_deleted=False)
    serializer_class = FormSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            print(f"=== FORM CREATE VALIDATION ERROR ===")
            print(f"Errors: {serializer.errors}")
            import json as _json
            try:
                print(_json.dumps(serializer.errors, indent=2, default=str)[:5000])
            except Exception:
                print(str(serializer.errors)[:5000])
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def get_queryset(self):
        base_qs = (
            Form.objects
            .filter(is_deleted=False)
            .select_related('folder', 'form_admin', 'deletedBy', 'archivedBy')
        )

        # Stage subtree prefetch: access, questions, options, children, logics, followups
        stage_access_qs = StageAccess.objects.select_related('allow_user', 'allow_group', 'form', 'stage')

        question_qs = (
            Question.objects
            .select_related('form', 'stage', 'parent_question')
            .prefetch_related(
                'options',
                'child_questions',
                'child_questions__options',
                'child_questions__logic_parent_question__logic_questions__options',
                'logic_parent_question__logic_questions__options',
                'logic_parent_question__follow_ups',
                'logic_parent_question__follow_ups__task_close_questions__options',
                'logic_parent_question__follow_ups__assign_form',
            )
        )

        stage_qs = (
            Stage.objects
            .select_related('form')
            .prefetch_related(
                models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                models.Prefetch('questions', queryset=question_qs),
            )
        )

        return (
            base_qs
            .prefetch_related(
                # assignments on form
                'assignee__user',
                'assignee__group',
                'assignee__leader',
                # audit structures
                'audit_info__questions',
                'audit_info__questions__options',
                'audit_info__questions__child_questions',
                'audit_info__questions__child_questions__options',
                'audit_info__questions__logic_parent_question__logic_questions__options',
                'audit_info__questions__logic_parent_question__follow_ups',
                'audit_info__questions__logic_parent_question__follow_ups__task_close_questions__options',
                'audit_group__questions',
                'audit_group__questions__options',
                'audit_group__questions__child_questions',
                'audit_group__questions__child_questions__options',
                'audit_group__questions__logic_parent_question__logic_questions__options',
                'audit_group__questions__logic_parent_question__follow_ups',
                'audit_group__questions__logic_parent_question__follow_ups__task_close_questions__options',
                # stages tree
                models.Prefetch('stages', queryset=stage_qs),
            )
        )

    @method_decorator(cache_page(300))
    def retrieve(self, request, *args, **kwargs):
        # Tailored, single-object optimized load to avoid over-prefetching
        pk = kwargs.get(self.lookup_field or 'pk')

        # Progressive loading: ?groups=1,2 returns only those stage/group orders with full data
        # ?group_id=5 filters by AuditGroup.id (for collaborative mode)
        groups_param = request.query_params.get('groups', None)
        group_id_param = request.query_params.get('group_id', None)
        requested_orders = None
        if groups_param:
            try:
                requested_orders = [int(g.strip()) for g in groups_param.split(',') if g.strip()]
            except (ValueError, TypeError):
                requested_orders = None
        elif group_id_param:
            try:
                requested_group_id = int(group_id_param.strip())
                ag = AuditGroup.objects.filter(id=requested_group_id).first()
                if ag:
                    requested_orders = [ag.order]
            except (ValueError, TypeError):
                pass

        base = (
            Form.objects
            .filter(is_deleted=False, pk=pk)
            .select_related('folder', 'form_admin', 'deletedBy', 'archivedBy')
            .prefetch_related(
                'auto_share_config',
                'auto_share_config__users',
                'auto_share_config__groups',
                'auto_share_config__location_leaders',
            )
        )

        # Determine form type with a lightweight fetch
        form_type = (
            Form.objects.filter(pk=pk).values_list('form_type', flat=True).first()
        )

        stage_access_qs = StageAccess.objects.select_related('allow_user', 'allow_group', 'form', 'stage', 'stage__organization')
        logic_qs = (
            Question.objects.none()  # placeholder for type hints
        )
        question_qs = (
            Question.objects
            .select_related('form', 'stage', 'parent_question')
            .prefetch_related(
                'options',
                'child_questions',
                'child_questions__options',
                'child_questions__child_questions',
                'child_questions__child_questions__options',
                'child_questions__logic_parent_question__logic_questions__options',
                'child_questions__logic_parent_question__follow_ups',
                'child_questions__logic_parent_question__follow_ups__task_close_questions__options',
                'logic_parent_question__logic_questions__options',
                'logic_parent_question__follow_ups',
                'logic_parent_question__follow_ups__task_close_questions__options',
            )
        )

        if form_type == FormType.AUDIT:
            if requested_orders:
                audit_group_qs = AuditGroup.objects.filter(order__in=requested_orders)
                audit_info_qs = AuditInfo.objects.none()  # skip audit_info for partial loads
            else:
                audit_group_qs = AuditGroup.objects.all()
                audit_info_qs = AuditInfo.objects.all()

            qs = base.prefetch_related(
                'assignee__user', 'assignee__group', 'assignee__leader',
                models.Prefetch('audit_info', queryset=audit_info_qs) if not requested_orders else
                'audit_info',
                'audit_info__questions',
                'audit_info__questions__options',
                'audit_info__questions__child_questions',
                'audit_info__questions__child_questions__options',
                'audit_info__questions__logic_parent_question__logic_questions__options',
                'audit_info__questions__logic_parent_question__follow_ups',
                'audit_info__questions__logic_parent_question__follow_ups__task_close_questions__options',
                models.Prefetch('audit_group', queryset=audit_group_qs),
                'audit_group__questions',
                'audit_group__questions__options',
                'audit_group__questions__child_questions',
                'audit_group__questions__child_questions__options',
                'audit_group__questions__logic_parent_question__logic_questions__options',
                'audit_group__questions__logic_parent_question__follow_ups',
                'audit_group__questions__logic_parent_question__follow_ups__task_close_questions__options',
            )
        else:
            if requested_orders:
                stage_qs = (
                    Stage.objects
                    .filter(order__in=requested_orders)
                    .select_related('form', 'organization')
                    .prefetch_related(
                        models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                        models.Prefetch('questions', queryset=question_qs),
                    )
                )
            else:
                stage_qs = (
                    Stage.objects
                    .select_related('form', 'organization')
                    .prefetch_related(
                        models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                        models.Prefetch('questions', queryset=question_qs),
                    )
                )
            qs = base.prefetch_related(
                'assignee__user', 'assignee__group', 'assignee__leader',
                models.Prefetch('stages', queryset=stage_qs),
            )

        instance = get_object_or_404(qs)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

class FormBatchMetadataView(APIView):
    """Lightweight endpoint returning only id, title, form_type, prefix for multiple forms.
    Accepts ?ids=1,2,3 query param. Used by mobile app to avoid slow full form fetches."""
    permission_classes = [AllowAny]

    def get(self, request):
        ids_param = request.query_params.get('ids', '')
        try:
            form_ids = [int(x.strip()) for x in ids_param.split(',') if x.strip()]
        except (ValueError, TypeError):
            return Response({'error': 'Invalid ids parameter'}, status=400)

        if not form_ids:
            return Response([], status=200)

        forms = Form.objects.filter(
            id__in=form_ids, is_deleted=False
        ).values('id', 'title', 'form_type', 'prefix')

        result = {
            str(f['id']): {
                'id': f['id'],
                'title': f['title'] or '',
                'form_type': f['form_type'] or '',
                'prefix': f['prefix'] or '',
            }
            for f in forms
        }
        return Response(result, status=200)

class FormStageMetadataView(APIView):
    """Lightweight endpoint that returns only stage/group metadata (no questions).
    Used by mobile app for progressive loading — know how many groups exist before fetching them."""
    permission_classes = [AllowAny]

    def get(self, request, form_id):
        form = get_object_or_404(Form, id=form_id, is_deleted=False)
        form_type = form.form_type

        result = {
            'form_id': form.id,
            'form_title': form.title or '',
            'form_type': form_type,
            'stages': [],
            'audit_groups': [],
        }

        if form_type == 'audit':
            groups = AuditGroup.objects.filter(form=form).order_by('order')
            for g in groups:
                result['audit_groups'].append({
                    'id': g.id,
                    'name': g.name,
                    'order': g.order,
                    'group_uuid': str(g.group_uuid) if hasattr(g, 'group_uuid') else None,
                })
        else:
            stages = Stage.objects.filter(form=form).order_by('order')
            for s in stages:
                result['stages'].append({
                    'id': s.id,
                    'name': s.name,
                    'order': s.order,
                    'stage_uuid': str(s.stage_uuid) if hasattr(s, 'stage_uuid') else None,
                    'is_completed': False,
                })

        return Response(result)


class FormFastView(APIView):
    """Ultra-fast form endpoint that bypasses DRF serializers entirely.
    Builds JSON from bulk .values() queries — 5-10x faster than serializer-based approach.
    Supports ?groups=1,2 for progressive loading."""
    permission_classes = [AllowAny]

    def get(self, request, form_id):
        form = get_object_or_404(Form, id=form_id, is_deleted=False)

        # Progressive loading: ?groups=1,2 (by order) or ?group_id=5 (by AuditGroup.id for collaborative mode)
        groups_param = request.query_params.get('groups', None)
        group_id_param = request.query_params.get('group_id', None)
        requested_orders = None
        if groups_param:
            try:
                requested_orders = [int(g.strip()) for g in groups_param.split(',') if g.strip()]
            except (ValueError, TypeError):
                requested_orders = None
        elif group_id_param:
            try:
                requested_group_id = int(group_id_param.strip())
                ag = AuditGroup.objects.filter(id=requested_group_id, form=form).first()
                if ag:
                    requested_orders = [ag.order]
            except (ValueError, TypeError):
                pass

        form_type = form.form_type

        # Base form fields
        result = {
            'id': form.id,
            'form_type': form_type,
            'title': form.title,
            'prefix': form.prefix,
            'GPS': form.GPS,
            'trigger_email_notifications': form.trigger_email_notifications,
            'share_response': form.share_response,
            'allow_editing': form.allow_editing,
            'can_edit_previous_state': form.can_edit_previous_state,
            'auto_share_response': form.auto_share_response,
            'pass_percentage': form.pass_percentage,
            'max_score': form.max_score,
            'is_archived': form.is_archived if hasattr(form, 'is_archived') else False,
            'is_deleted': form.is_deleted,
            'is_disabled': form.is_disabled,
            'folder': form.folder_id,
            'folder_name': form.folder.name if form.folder else None,
            'form_admin': form.form_admin_id,
            'form_admin_display': self._get_user_display(form.form_admin),
            'created_at': form.created_at.isoformat() if form.created_at else None,
            'updated_at': form.updated_at.isoformat() if form.updated_at else None,
            'stages': [],
            'audit_group': [],
            'audit_info': None,
            'assignments': [],
            'auto_share_config': None,
        }

        # Assignments
        assignments = FormAssignment.objects.filter(form=form).values(
            'id', 'assign_type', 'user_id', 'group_id', 'leader_id'
        )
        result['assignments'] = list(assignments)

        # Auto share config
        try:
            config = FormAutoShareConfig.objects.get(form=form)
            result['auto_share_config'] = {
                'users': list(config.users.values_list('id', flat=True)),
                'groups': list(config.groups.values_list('id', flat=True)),
                'location_leaders': list(config.location_leaders.values_list('id', flat=True)),
            }
        except FormAutoShareConfig.DoesNotExist:
            pass

        if form_type == 'audit':
            result.update(self._build_audit_fast(form, requested_orders))
        else:
            result.update(self._build_stages_fast(form, requested_orders))

        return Response(result)

    def _get_user_display(self, user):
        if not user:
            return "Unknown"
        first = user.first_name or ""
        last = user.last_name or ""
        return f"{first.strip()} {last.strip()}".strip() or user.username or user.email

    @staticmethod
    def _compact(d):
        """Remove None and empty string values from a dict.
        Keeps False, 0, and [] which the frontend expects."""
        return {k: v for k, v in d.items() if v is not None and v != ''}

    def _build_stages_fast(self, form, requested_orders):
        """Build stages with questions, options, logics, follow-ups using bulk queries."""
        stage_qs = Stage.objects.filter(form=form).order_by('order')
        if requested_orders:
            stage_qs = stage_qs.filter(order__in=requested_orders)

        stages = list(stage_qs.values('id', 'name', 'stage_uuid', 'order', 'is_completed'))
        if not stages:
            return {'stages': []}

        stage_ids = [s['id'] for s in stages]

        # Stage access
        stage_access = list(StageAccess.objects.filter(stage_id__in=stage_ids).values(
            'id', 'access_type', 'allow_user_id', 'allow_group_id', 'allow_stage',
            'stage_id', 'stage_approvals'
        ))
        access_by_stage = {}
        for sa in stage_access:
            sid = sa.pop('stage_id')
            access_by_stage.setdefault(sid, []).append(sa)

        # All questions for these stages (top-level only: no parent, not logic/task-close children)
        questions = list(
            Question.objects.filter(
                stage_id__in=stage_ids, parent_question__isnull=True,
                is_logic_question=False, is_task_close_question=False
            )
            .order_by('order')
            .values(
                'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                'question_type', 'question_sub_type', 'question_hint', 'order',
                'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                'form_id', 'stage_id', 'parent_question_id'
            )
        )
        all_question_ids = [q['id'] for q in questions]

        # Child questions (sub_questions) — 2 levels deep
        child_qs = Question.objects.filter(parent_question_id__in=all_question_ids).order_by('order').values(
            'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
            'question_type', 'question_sub_type', 'question_hint', 'order',
            'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
            'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
            'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
            'form_id', 'stage_id', 'parent_question_id'
        )
        children_l1 = list(child_qs)
        child_ids_l1 = [c['id'] for c in children_l1]

        # Grandchildren
        children_l2 = []
        if child_ids_l1:
            children_l2 = list(
                Question.objects.filter(parent_question_id__in=child_ids_l1).order_by('order').values(
                    'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                    'question_type', 'question_sub_type', 'question_hint', 'order',
                    'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                    'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                    'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                    'form_id', 'stage_id', 'parent_question_id'
                )
            )

        all_q_ids = all_question_ids + child_ids_l1 + [c['id'] for c in children_l2]

        # Options — single query
        options = list(
            Option.objects.filter(question_id__in=all_q_ids).order_by('order').values(
                'id', 'option', 'score', 'failed', 'order', 'question_id'
            )
        )
        options_by_q = {}
        for opt in options:
            qid = opt.pop('question_id')
            options_by_q.setdefault(qid, []).append(opt)

        # Logics — single query
        logics = list(
            Logic.objects.filter(question_id__in=all_q_ids).order_by('order').values(
                'id', 'logic_type', 'comparison', 'logic_value', 'notification',
                'email', 'order', 'question_id', 'user_id', 'group_id'
            )
        )
        logic_ids = [l['id'] for l in logics]
        logics_by_q = {}
        for lg in logics:
            qid = lg.pop('question_id')
            logics_by_q.setdefault(qid, []).append(lg)

        # Logic questions (M2M) — use ORM through table
        logic_question_links = []
        if logic_ids:
            logic_question_links = list(
                Logic.objects.filter(id__in=logic_ids).values_list('id', 'logic_questions__id')
            )
            logic_question_links = [(l_id, q_id) for l_id, q_id in logic_question_links if q_id is not None]

        # Fetch logic child question data
        logic_child_q_ids = set(lq[1] for lq in logic_question_links)
        logic_child_questions = {}
        if logic_child_q_ids:
            lq_data = list(
                Question.objects.filter(id__in=logic_child_q_ids).values(
                    'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                    'question_type', 'question_sub_type', 'question_hint', 'order',
                    'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                    'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                    'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                    'form_id', 'stage_id', 'parent_question_id'
                )
            )
            logic_child_questions = {q['id']: q for q in lq_data}

            # Options for logic child questions
            lq_options = list(
                Option.objects.filter(question_id__in=logic_child_q_ids).order_by('order').values(
                    'id', 'option', 'score', 'failed', 'order', 'question_id'
                )
            )
            for opt in lq_options:
                qid = opt.pop('question_id')
                if qid in logic_child_questions:
                    logic_child_questions[qid].setdefault('options', []).append(opt)

        # Group logic questions by logic_id
        logic_qs_by_logic = {}
        for logic_id, q_id in logic_question_links:
            if q_id in logic_child_questions:
                logic_qs_by_logic.setdefault(logic_id, []).append(logic_child_questions[q_id])

        # Follow-ups — single query
        follow_ups = list(
            LogicFollowUp.objects.filter(question_id__in=all_q_ids).values(
                'id', 'logic_id', 'title', 'description', 'deadline', 'assign_form_id',
                'assign_to', 'user_id', 'group_id', 'leader_id',
                'assign_user_ids', 'assign_group_ids', 'assign_leader_ids',
                'followup_toggle', 'form_id', 'stage_id', 'question_id'
            )
        )
        follow_ups_by_q = {}
        for fu in follow_ups:
            qid = fu.pop('question_id')
            fu['assigned_form_title'] = None
            if fu.get('assign_form_id'):
                try:
                    fu['assigned_form_title'] = Form.objects.get(id=fu['assign_form_id']).title
                except Form.DoesNotExist:
                    pass
            # Resolve assigned users/groups names
            fu['assigned_users'] = self._resolve_user_names(fu.get('assign_user_ids', []))
            fu['assigned_groups'] = self._resolve_group_names(fu.get('assign_group_ids', []))
            fu['assigned_leaders'] = self._resolve_user_names(fu.get('assign_leader_ids', []))
            follow_ups_by_q.setdefault(qid, []).append(fu)

        # Task close questions — M2M through follow_up
        follow_up_ids = [fu['id'] for fqs in follow_ups_by_q.values() for fu in fqs]
        task_close_links = []
        if follow_up_ids:
            task_close_links = list(
                LogicFollowUp.objects.filter(id__in=follow_up_ids).values_list('id', 'task_close_questions__id')
            )
            task_close_links = [(fu_id, q_id) for fu_id, q_id in task_close_links if q_id is not None]

        task_close_q_ids = set(tc[1] for tc in task_close_links)
        task_close_questions = {}
        if task_close_q_ids:
            tc_data = list(
                Question.objects.filter(id__in=task_close_q_ids).values(
                    'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                    'question_type', 'question_sub_type', 'question_hint', 'order',
                    'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                    'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                    'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                    'form_id', 'stage_id', 'parent_question_id'
                )
            )
            task_close_questions = {q['id']: q for q in tc_data}
            tc_options = list(
                Option.objects.filter(question_id__in=task_close_q_ids).order_by('order').values(
                    'id', 'option', 'score', 'failed', 'order', 'question_id'
                )
            )
            for opt in tc_options:
                qid = opt.pop('question_id')
                if qid in task_close_questions:
                    task_close_questions[qid].setdefault('options', []).append(opt)

        task_close_by_fu = {}
        for fu_id, q_id in task_close_links:
            if q_id in task_close_questions:
                task_close_by_fu.setdefault(fu_id, []).append(task_close_questions[q_id])

        # Attach task_close_questions to follow_ups
        for fqs in follow_ups_by_q.values():
            for fu in fqs:
                fu['task_close_questions'] = task_close_by_fu.get(fu['id'], [])

        # Build question tree
        def build_question(qdata):
            qid = qdata['id']
            qdata['options'] = options_by_q.get(qid, [])
            qdata['sub_questions'] = []
            qdata['logics'] = []

            # Attach logics
            for lg in logics_by_q.get(qid, []):
                lg['logic_questions'] = logic_qs_by_logic.get(lg['id'], [])
                lg['follow_up'] = None
                # Find follow-up for this logic
                for fu in follow_ups_by_q.get(qid, []):
                    if fu.get('logic_id') == lg['id']:
                        lg['follow_up'] = fu
                        break
                lg['followup_toggle'] = lg['follow_up']['followup_toggle'] if lg['follow_up'] else False
                qdata['logics'].append(lg)

            # Attach follow_ups directly (for questions without logics match)
            if not qdata['logics']:
                for fu in follow_ups_by_q.get(qid, []):
                    qdata.setdefault('_follow_ups', []).append(fu)

            return qdata

        # Build children maps
        children_by_parent = {}
        for c in children_l1 + children_l2:
            children_by_parent.setdefault(c['parent_question_id'], []).append(c)

        def build_with_children(qdata):
            qdata = build_question(qdata)
            qdata['sub_questions'] = []
            for child in children_by_parent.get(qdata['id'], []):
                child = build_with_children(dict(child))
                qdata['sub_questions'].append(child)
            return qdata

        # Group questions by stage
        questions_by_stage = {}
        for q in questions:
            sid = q.pop('stage_id')
            questions_by_stage.setdefault(sid, []).append(q)

        # Assemble stages
        for stage in stages:
            stage['questions'] = []
            stage['stage_access'] = []
            stage['assignments'] = access_by_stage.get(stage['id'], [])
            for sa in stage['assignments']:
                sa['stage_order'] = stage['order']
                sa['organization'] = form.organization_id
            for q in questions_by_stage.get(stage['id'], []):
                stage['questions'].append(build_with_children(dict(q)))

        return {'stages': stages}

    def _build_audit_fast(self, form, requested_orders):
        """Build audit groups with questions using bulk queries."""
        group_qs = AuditGroup.objects.filter(form=form).order_by('order')
        if requested_orders:
            group_qs = group_qs.filter(order__in=requested_orders)

        groups = list(group_qs.values('id', 'name', 'group_uuid', 'order', 'form_id'))
        if not groups:
            return {'audit_group': [], 'audit_info': None}

        group_ids = [g['id'] for g in groups]

        # Audit info (always include — it's lightweight)
        audit_info_data = None
        ai = AuditInfo.objects.filter(form=form).first()
        if ai:
            ai_questions = list(
                Question.objects.filter(audit_info_id=ai.id, parent_question__isnull=True)
                .order_by('order').values(
                    'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                    'question_type', 'question_sub_type', 'question_hint', 'order',
                    'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                    'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                    'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                    'form_id', 'stage_id', 'parent_question_id',
                    'audit_info_id', 'audit_group_id'
                )
            )
            ai_q_ids = [q['id'] for q in ai_questions]
            ai_options = list(Option.objects.filter(question_id__in=ai_q_ids).order_by('order').values(
                'id', 'option', 'score', 'failed', 'order', 'question_id'
            ))
            ai_opts_by_q = {}
            for opt in ai_options:
                qid = opt.pop('question_id')
                ai_opts_by_q.setdefault(qid, []).append(opt)

            for q in ai_questions:
                q['options'] = ai_opts_by_q.get(q['id'], [])
                q['sub_questions'] = []
                q['logics'] = []

            audit_info_data = {
                'id': ai.id,
                'name': ai.name,
                'group_uuid': ai.group_uuid,
                'questions': ai_questions,
            }

        # All questions for these groups (top-level only: no parent, not logic/task-close children)
        questions = list(
            Question.objects.filter(
                audit_group_id__in=group_ids, parent_question__isnull=True,
                is_logic_question=False, is_task_close_question=False
            )
            .order_by('order')
            .values(
                'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                'question_type', 'question_sub_type', 'question_hint', 'order',
                'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                'form_id', 'stage_id', 'parent_question_id',
                'audit_info_id', 'audit_group_id'
            )
        )
        all_q_ids = [q['id'] for q in questions]

        # Child questions
        children_l1 = list(
            Question.objects.filter(parent_question_id__in=all_q_ids).order_by('order').values(
                'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                'question_type', 'question_sub_type', 'question_hint', 'order',
                'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                'form_id', 'stage_id', 'parent_question_id',
                'audit_info_id', 'audit_group_id'
            )
        )
        child_ids_l1 = [c['id'] for c in children_l1]
        children_l2 = []
        if child_ids_l1:
            children_l2 = list(
                Question.objects.filter(parent_question_id__in=child_ids_l1).order_by('order').values(
                    'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                    'question_type', 'question_sub_type', 'question_hint', 'order',
                    'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                    'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                    'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                    'form_id', 'stage_id', 'parent_question_id',
                    'audit_info_id', 'audit_group_id'
                )
            )

        all_q_ids_total = all_q_ids + child_ids_l1 + [c['id'] for c in children_l2]

        # Options
        options = list(
            Option.objects.filter(question_id__in=all_q_ids_total).order_by('order').values(
                'id', 'option', 'score', 'failed', 'order', 'question_id'
            )
        )
        opts_by_q = {}
        for opt in options:
            qid = opt.pop('question_id')
            opts_by_q.setdefault(qid, []).append(opt)

        # Logics
        logics = list(
            Logic.objects.filter(question_id__in=all_q_ids_total).order_by('order').values(
                'id', 'logic_type', 'comparison', 'logic_value', 'notification',
                'email', 'order', 'question_id', 'user_id', 'group_id'
            )
        )
        logic_ids = [l['id'] for l in logics]
        logics_by_q = {}
        for lg in logics:
            qid = lg.pop('question_id')
            logics_by_q.setdefault(qid, []).append(lg)

        # Logic questions M2M — use ORM
        logic_question_links = []
        if logic_ids:
            logic_question_links = list(
                Logic.objects.filter(id__in=logic_ids).values_list('id', 'logic_questions__id')
            )
            logic_question_links = [(l_id, q_id) for l_id, q_id in logic_question_links if q_id is not None]

        logic_child_q_ids = set(lq[1] for lq in logic_question_links)
        logic_child_questions = {}
        if logic_child_q_ids:
            lq_data = list(
                Question.objects.filter(id__in=logic_child_q_ids).values(
                    'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                    'question_type', 'question_sub_type', 'question_hint', 'order',
                    'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                    'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                    'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                    'form_id', 'stage_id', 'parent_question_id',
                    'audit_info_id', 'audit_group_id'
                )
            )
            logic_child_questions = {q['id']: q for q in lq_data}
            lq_options = list(
                Option.objects.filter(question_id__in=logic_child_q_ids).order_by('order').values(
                    'id', 'option', 'score', 'failed', 'order', 'question_id'
                )
            )
            for opt in lq_options:
                qid = opt.pop('question_id')
                if qid in logic_child_questions:
                    logic_child_questions[qid].setdefault('options', []).append(opt)

        logic_qs_by_logic = {}
        for logic_id, q_id in logic_question_links:
            if q_id in logic_child_questions:
                logic_qs_by_logic.setdefault(logic_id, []).append(logic_child_questions[q_id])

        # Follow-ups
        follow_ups = list(
            LogicFollowUp.objects.filter(question_id__in=all_q_ids_total).values(
                'id', 'logic_id', 'title', 'description', 'deadline', 'assign_form_id',
                'assign_to', 'user_id', 'group_id', 'leader_id',
                'assign_user_ids', 'assign_group_ids', 'assign_leader_ids',
                'followup_toggle', 'form_id', 'stage_id', 'question_id'
            )
        )
        follow_ups_by_q = {}
        for fu in follow_ups:
            qid = fu.pop('question_id')
            fu['assigned_form_title'] = None
            if fu.get('assign_form_id'):
                try:
                    fu['assigned_form_title'] = Form.objects.get(id=fu['assign_form_id']).title
                except Form.DoesNotExist:
                    pass
            fu['assigned_users'] = self._resolve_user_names(fu.get('assign_user_ids', []))
            fu['assigned_groups'] = self._resolve_group_names(fu.get('assign_group_ids', []))
            fu['assigned_leaders'] = self._resolve_user_names(fu.get('assign_leader_ids', []))
            follow_ups_by_q.setdefault(qid, []).append(fu)

        # Task close questions
        follow_up_ids = [fu['id'] for fqs in follow_ups_by_q.values() for fu in fqs]
        task_close_links = []
        if follow_up_ids:
            task_close_links = list(
                LogicFollowUp.objects.filter(id__in=follow_up_ids).values_list('id', 'task_close_questions__id')
            )
            task_close_links = [(fu_id, q_id) for fu_id, q_id in task_close_links if q_id is not None]

        task_close_q_ids = set(tc[1] for tc in task_close_links)
        task_close_questions = {}
        if task_close_q_ids:
            tc_data = list(
                Question.objects.filter(id__in=task_close_q_ids).values(
                    'id', 'question_uuid', 'question', 'description', 'critical', 'formula',
                    'question_type', 'question_sub_type', 'question_hint', 'order',
                    'is_required', 'require_live', 'number_of_file_allowed', 'min_value',
                    'max_value', 'max_score', 'is_logic_question', 'is_task_close_question',
                    'is_audit_info_question', 'is_other', 'reference_images', 'reference_videos',
                    'form_id', 'stage_id', 'parent_question_id',
                    'audit_info_id', 'audit_group_id'
                )
            )
            task_close_questions = {q['id']: q for q in tc_data}
            tc_options = list(
                Option.objects.filter(question_id__in=task_close_q_ids).order_by('order').values(
                    'id', 'option', 'score', 'failed', 'order', 'question_id'
                )
            )
            for opt in tc_options:
                qid = opt.pop('question_id')
                if qid in task_close_questions:
                    task_close_questions[qid].setdefault('options', []).append(opt)

        task_close_by_fu = {}
        for fu_id, q_id in task_close_links:
            if q_id in task_close_questions:
                task_close_by_fu.setdefault(fu_id, []).append(task_close_questions[q_id])

        for fqs in follow_ups_by_q.values():
            for fu in fqs:
                fu['task_close_questions'] = task_close_by_fu.get(fu['id'], [])

        # Build children map
        children_by_parent = {}
        for c in children_l1 + children_l2:
            children_by_parent.setdefault(c['parent_question_id'], []).append(c)

        def build_question(qdata):
            qid = qdata['id']
            qdata['options'] = opts_by_q.get(qid, [])
            qdata['sub_questions'] = []
            qdata['logics'] = []
            for lg in logics_by_q.get(qid, []):
                lg['logic_questions'] = logic_qs_by_logic.get(lg['id'], [])
                lg['follow_up'] = None
                for fu in follow_ups_by_q.get(qid, []):
                    if fu.get('logic_id') == lg['id']:
                        lg['follow_up'] = fu
                        break
                lg['followup_toggle'] = lg['follow_up']['followup_toggle'] if lg['follow_up'] else False
                qdata['logics'].append(lg)
            return qdata

        def build_with_children(qdata):
            qdata = build_question(dict(qdata))
            qdata['sub_questions'] = []
            for child in children_by_parent.get(qdata['id'], []):
                qdata['sub_questions'].append(build_with_children(dict(child)))
            return qdata

        # Group questions by audit_group
        questions_by_group = {}
        for q in questions:
            gid = q.pop('audit_group_id')
            questions_by_group.setdefault(gid, []).append(q)

        for group in groups:
            group['questions'] = []
            for q in questions_by_group.get(group['id'], []):
                group['questions'].append(build_with_children(dict(q)))

        return {'audit_group': groups, 'audit_info': audit_info_data}

    def _resolve_user_names(self, user_ids):
        if not user_ids:
            return []
        users = CustomUser.objects.filter(id__in=user_ids).values('id', 'first_name', 'last_name', 'username')
        result = []
        for u in users:
            name = f"{u.get('first_name', '')} {u.get('last_name', '')}".strip() or u.get('username', '')
            result.append({'id': u['id'], 'name': name})
        return result

    def _resolve_group_names(self, group_ids):
        if not group_ids:
            return []
        groups = Groups.objects.filter(id__in=group_ids).values('id', 'name')
        return [{'id': g['id'], 'name': g['name']} for g in groups]

    def create(self, request, *args, **kwargs):
        UF = UtilsFunctions()
        form_name = request.data.get("title", "unknown_form")
        json_data = json.dumps(request.data, indent=2)
        file_obj = io.BytesIO(json_data.encode("utf-8"))
        timestamp = datetime.now().strftime("%d%m%y_%H%M%S")
        file_path = f"{form_name}_upload_create_{timestamp}.json"
        
        # Temporarily bypass S3 upload for development
        file_url = UF.upload_file_to_s3(file_path, file_obj, content_type="application/json", bucket=settings.S3_BUCKET_NAME)
        
        # Skip S3 upload requirement for development
        # if not file_url:
        #     return Response({"error": "Upload to S3 failed"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        print(">>> DATA BEFORE SERIALIZER:", file_url)

        # Only create FormPayloadFiles if S3 upload succeeded
        file = None
        if file_url:
            file = FormPayloadFiles.objects.create(
                title=request.data.get('title', None),
                form_type=request.data.get('form_type', None),
                organization=request.user.organization,
                form_admin=request.user,
                file_path=file_url,
                status=FormPayloadFiles.Status.INPROGRESS,
                method=FormPayloadFiles.Method.POST
            )

        serializer = self.get_serializer(data=request.data, context={**self.get_serializer_context(), 'skip_unique_checks': True})
        if not serializer.is_valid():
            clean = UF.extract_error_message(serializer.errors)
            print("clean ::", clean)
            logger.error("FormSerializer validation failed: %s", serializer.errors)
            if file:
                file.status = FormPayloadFiles.Status.FAILED
                file.error_message = str(clean)
                file.save()
            print("error Message ::", serializer.errors)
            raise ValidationError(serializer.errors)
        # Pass skip_unique_checks for bulk/nested creation to avoid O(N) lookups
        instance = serializer.save()

        if file:
            file.form = instance
            file.status = FormPayloadFiles.Status.SUCCESS
            file.save()

        return Response({'id': instance.id, 'title': instance.title, 'form_type': instance.form_type}, status=status.HTTP_201_CREATED)
    
    @transaction.atomic
    def update(self, request, *args, **kwargs):
        # Get the existing form
        form = self.get_object()

        UF = UtilsFunctions()
        form_name = request.data.get("title", "unknown_form")
        json_data = json.dumps(request.data, indent=2)
        file_obj = io.BytesIO(json_data.encode("utf-8"))
        timestamp = datetime.now().strftime("%d%m%y_%H%M%S")
        file_path = f"{form_name}_upload_update_{timestamp}.json"
        file_url = UF.upload_file_to_s3(file_path, file_obj, content_type="application/json")
        if not file_url:
            return Response({"error": "Upload to S3 failed"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        print(">>> DATA BEFORE SERIALIZER:", file_url)

        file = FormPayloadFiles.objects.create(
            title=request.data.get('title', None),
            organization=request.user.organization,
            form_admin=request.user,
            file_path=file_url,
            status=FormPayloadFiles.Status.INPROGRESS,
            method=FormPayloadFiles.Method.PUT
        )



        # Validate and update the form with new data
        serializer = self.get_serializer(form, data=request.data, partial=True)
        if not serializer.is_valid():
            clean = UF.extract_error_message(serializer.errors)
            file.status = FormPayloadFiles.Status.FAILED
            file.error_message = str(clean)
            file.save()
            raise ValidationError(serializer.errors)
        updated_form = serializer.save()

        file.form = updated_form
        file.status = FormPayloadFiles.Status.SUCCESS
        file.save()


        return Response({'id': updated_form.id, 'title': updated_form.title, 'form_type': updated_form.form_type}, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        raise MethodNotAllowed("DELETE")
    
   

class FormCloneViewSet(userContextAPIView, GenericAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Form.objects.all()
    serializer_class = FormSerializer

    @transaction.atomic
    def put(self, request, form_id):
        disableForm = get_object_or_404(
            Form,
            id= form_id
        )
        
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            logger.error("Form clone validation failed for source form %s: %s", form_id, serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        newFormId = serializer.data.get('id')

        return Response(serializer.data, status=status.HTTP_201_CREATED)
     
    
class FormAssignmentViewSet(ModelViewSet):
    permission_classes = [IsAdmin]
    queryset = FormAssignment.objects.all()
    serializer_class = FormAssignmentSerializer

    def create(self, request, *args, **kwargs):
        data = request.data
        assign_type = data.get('assign_type')
        form_id = data.get('form')

        try:
            # Validate and extract list based on assign_type
            formInstance = get_object_or_404(
                Form,
                id=form_id,
                is_deleted=False,
                is_archived=False,
                organization=request.user.organization
            )
            
            if not Form.objects.filter(id=form_id, is_deleted=False, is_archived=False).exists():
                return Response(
                    {"detail": "Form not found or is deleted/archived."},
                    status=status.HTTP_404_NOT_FOUND
                )
                
            ids = data.get(assign_type)
            if not isinstance(ids, list):
                return Response(
                    {"detail": f"{assign_type} must be a list of IDs."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            # Get the form to check its type
            form = Form.objects.get(id=form_id, is_deleted=False, is_archived=False)

            created = []
            for obj_id in ids:
                # Check for duplicates before creating
                existing_assignment = None
                if assign_type == FormAssignType.USER:
                    existing_assignment = FormAssignment.objects.filter(
                        form_id=form_id,
                        assign_type=assign_type,
                        user_id=obj_id
                    ).first()
                elif assign_type == FormAssignType.GROUP:
                    existing_assignment = FormAssignment.objects.filter(
                        form_id=form_id,
                        assign_type=assign_type,
                        group_id=obj_id
                    ).first()
                elif assign_type == FormAssignType.LOCATION_LEADER:
                    existing_assignment = FormAssignment.objects.filter(
                        form_id=form_id,
                        assign_type=assign_type,
                        leader_id=obj_id
                    ).first()

                if existing_assignment:
                    # Skip duplicate assignment but don't fail the entire operation
                    continue
                else:
                    kwargs = {
                        'assign_type': assign_type,
                        'form_id': form_id,
                        f'{assign_type}_id': obj_id
                    }
                    created.append(FormAssignment.objects.create(**kwargs))

            serializer = self.get_serializer(created, many=True)
            
            
            # 1st Stage assignment
            if assign_type == FormAssignType.USER or assign_type == FormAssignType.LOCATION_LEADER:
                for user_id in ids:
                    if form.form_type == FormType.LOCATION or form.form_type == FormType.STANDARD:
                        StageAssignment.objects.create(
                            form_id=form_id,
                            user_id=user_id,
                            stage_order=1,
                            stage=Stage.objects.filter(form_id=form_id, order=1).first(),
                            assignment_uuid=uuid.uuid4(),
                            organization=request.user.organization
                        )
                    elif form.form_type == FormType.AUDIT:
                        GroupAssignment.objects.create(
                            form_id=form_id,
                            user_id=user_id,
                            group_order=1,
                            group=AuditGroup.objects.filter(form_id=form_id, organization=request.user.organization).first(),
                            assignment_uuid=uuid.uuid4(),
                            organization=request.user.organization
                        )  
            elif assign_type == FormAssignType.GROUP:
                allUsers = []
                for group_id in ids:
                    group = get_object_or_404(Groups, id=group_id, organization=request.user.organization)
                    allUsers.extend(group.members.all().distinct())
                
                allUsers = list(set(allUsers))
                for user in allUsers:
                    if form.form_type == FormType.LOCATION or form.form_type == FormType.STANDARD:
                        StageAssignment.objects.create(
                            form_id=form_id,
                            user_id=user.id,
                            stage_order=1,
                            stage=Stage.objects.filter(form_id=form_id, order=1).first(),
                            assignment_uuid=uuid.uuid4(),
                            organization=request.user.organization
                        )
                    elif form.form_type == FormType.AUDIT:
                        GroupAssignment.objects.create(
                            form_id=form_id,
                            user_id=user.id,
                            group_order=1,
                            group=AuditGroup.objects.filter(form_id=form_id, organization=request.user.organization).first(),
                            assignment_uuid=uuid.uuid4(),
                            organization=request.user.organization
                        )    
                        
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except:
            return Response({"message": "Unable to assign form"}, status=status.HTTP_400_BAD_REQUEST)
            
    def update(self, request, *args, **kwargs):
        raise MethodNotAllowed("UPDATE")

    def destroy(self, request, *args, **kwargs):
        raise MethodNotAllowed("DELETE")
    

class FormUnAssignmentViewSet(ModelViewSet):
    permission_classes = [IsAdmin]
    queryset = FormAssignment.objects.all()
    serializer_class = FormAssignmentSerializer

    def create(self, request, *args, **kwargs):
        data = request.data
        form_id = data.get('form')
        assign_type = data.get('assign_type')
        ids = data.get(assign_type, [])
         
        formInstance = get_object_or_404(
            Form,
            id=form_id,
            is_deleted=False,
            is_archived=False,
            organization=request.user.organization
        )
        
        if assign_type == FormAssignType.USER or assign_type == FormAssignType.LOCATION_LEADER:
            FormAssignment.objects.filter(
                form=formInstance,
                user__in=ids,
            ).delete()
            
        elif assign_type == FormAssignType.GROUP:
            for group_id in ids:
                group = get_object_or_404(Groups, id=group_id, organization=request.user.organization)
                FormAssignment.objects.filter(
                    form=formInstance,
                    group__in=ids,
                ).delete()
        
        userIds = []
        if assign_type == FormAssignType.USER or assign_type == FormAssignType.LOCATION_LEADER:
            userIds = ids
            
        elif assign_type == FormAssignType.GROUP:
            for group_id in ids:
                group = get_object_or_404(Groups, id=group_id, organization=request.user.organization)
                userIds.extend(group.members.all().distinct().values_list('id', flat=True))
        
        userIds = list(set(userIds))
        
        if formInstance.form_type == FormType.LOCATION or formInstance.form_type == FormType.STANDARD:
            for user_id in userIds:
                StageAssignment.objects.filter(
                    form_id=form_id,
                    user_id=user_id,
                    organization=request.user.organization,
                ).delete()
        elif formInstance.form_type == FormType.AUDIT:
            for user_id in userIds:
                GroupAssignment.objects.filter(
                    form_id=form_id,
                    user_id=user_id,
                    organization=request.user.organization,
                ).delete()
        
        return Response({"message": "Formassignments deleted successfully."}, status=status.HTTP_201_CREATED)
   
    def update(self, request, *args, **kwargs):
        raise MethodNotAllowed("UPDATE")

    def destroy(self, request, *args, **kwargs):
        raise MethodNotAllowed("DELETE")
        

class StageAssignmentViewSet(ModelViewSet):
    permission_classes = [IsAdmin]
    queryset = FormAssignment.objects.all()
    serializer_class = FormAssignmentSerializer

    def create(self, request, *args, **kwargs):
        data = request.data
        form_id = data.get('form')
        stage_id = data.get('stage')
        assign_type = data.get('assign_type')
        ids = data.get('ids', [])
        assignment_uuid = uuid.uuid4()
        
        stage = get_object_or_404(
            Stage,
            id=stage_id,
            form__id=form_id,
            form__organization=request.user.organization
        )
        
        formSubmissionInstance = get_object_or_404(
            FormSubmision,
            id=data.get('form_submission_id'),
            organization=request.user.organization
        )
        
        userIds = []
        if assign_type == StageAccessType.USER:
            userIds = ids
        elif assign_type == StageAccessType.GROUP:
            for group_id in ids:
                group = get_object_or_404(Groups, id=group_id, organization=request.user.organization)
                userIds.extend(group.members.all().distinct().values_list('id', flat=True))
        
        userIds = list(set(userIds))
        for user_id in userIds:
            StageAssignment.objects.create(
                form_id=form_id,
                stage=stage,
                stage_order=stage.order,
                user_id=user_id,
                assignment_uuid=assignment_uuid,
                form_submission=formSubmissionInstance,
                organization=request.user.organization,
                assigned_by=request.user,
                assigned_on=timezone.now(),
            )
        
        return Response({"message": "Stage assignments created successfully.", 'assignment_uuid': assignment_uuid}, status=status.HTTP_201_CREATED)
   
   
class UserSentFormsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, user_id):
        try:
            user = get_object_or_404(CustomUser,id=user_id, organization=request.user.organization)
            if not user:
                return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

            stageAssignments = StageAssignment.objects.filter(assigned_by=user, organization=request.user.organization)
            
            assignedForms = [{
                'form': CompactFormSerializer(assignment.form, many=False, context={'request': request}).data,
                'stage_id': assignment.stage.id,
                'stage_order': assignment.stage.order,
                'stage_assignment_id': assignment.id,
                'assignment_uuid': assignment.assignment_uuid,
                'form_submission_id': assignment.form_submission.id if assignment.form_submission else None,
                'is_stage_submission_pending': not assignment.is_assignment_fullfilled,
                'is_form_submission_pending': not assignment.form_submission.is_completed if assignment.form_submission else True
            } for assignment in stageAssignments]
        
            return Response(assignedForms, status=status.HTTP_200_OK)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)   
        
        
class UserReceivedFormsView(userContextAPIView, APIView):
    # permission_classes = [IsEndUserOrAdmin,IsLocationLeader]
    permission_classes = [AllowAny]

    def get(self, request, user_id=None):
        try:
            if not request.user.is_authenticated:
                return Response({"error": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
            
            if not user_id:
                raise ValidationError("User Id Required.")
                            
            user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)
            stageAssignments = StageAssignment.objects.filter(
                user=user,
                form__is_deleted=False,
                form__is_archived=False,
                form__organization=user.organization
            ).select_related('form', 'stage', 'form_submission').order_by(
                '-form_submission__submission_initiated_on',
                '-form_submission__id',
                '-id',
            ).distinct()
            
            assignedForms = [{
                'form': CompactFormSerializer(assignment.form, many=False, context={'request': request}).data,
                'stage_id': assignment.stage.id,
                'stage_name': assignment.stage.name,
                'stage_order': assignment.stage.order,
                'stage_assignment_id': assignment.id,
                'assignment_uuid': assignment.assignment_uuid,
                'form_submission_id': assignment.form_submission.id if assignment.form_submission else None,
                'submission_initiated_on': (
                    assignment.form_submission.submission_initiated_on.isoformat()
                    if assignment.form_submission and assignment.form_submission.submission_initiated_on
                    else None
                ),
                'is_stage_submission_pending': not assignment.is_assignment_fullfilled,
                'is_form_submission_pending': not assignment.form_submission.is_completed if assignment.form_submission else True
            } for assignment in stageAssignments]
            
            return Response(assignedForms, status=status.HTTP_200_OK)
        except Exception as e:
            user_id_for_log = user_id if user_id else (getattr(request.user, 'id', None) if request.user.is_authenticated else None)
            logger.error(f"Error retrieving assigned forms for user {user_id_for_log}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    
class FormArchiveView(APIView):
    permission_classes = [IsAdmin]
    def post(self, request, pk):
        try:
            re_active_param = request.query_params.get('re-activate', 'false')
            form = Form.objects.get(pk=pk)
            form.is_archived = False if re_active_param == 'true' else True
            form.archivedBy = request.user if re_active_param == 'false' else None
            form.save()
            
            message = 'Form re-activated successfully.' if form.is_archived is False else 'Form archived successfully.'
            logger.info(message, "form ID:", pk, "Re-active param:", re_active_param)
            return Response({'message': message}, status=status.HTTP_200_OK)
        
        except Form.DoesNotExist:
            return Response({'error': 'Form not found.'}, status=status.HTTP_404_NOT_FOUND)
    
    
class FormArchiveListView(APIView):
    permission_classes = [IsAdmin]
    def get(self, request):
        try:
            form = Form.objects.filter(is_archived=True)
            serializer = FormSerializer(form, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Form.DoesNotExist:
            return Response({'error': 'Form not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        
class FormDeleteView(APIView):
    permissoin_classses =[IsAdmin]
    def post(self, request, pk):
        try:
            re_active_param = request.query_params.get('re-activate', 'false')
            form = Form.objects.get(pk=pk)
            form.is_deleted = False if re_active_param == 'true' else True
            form.deletedBy = request.user if re_active_param == 'false' else None
            form.save()
            
            message = 'Form re-activated successfully.' if form.is_deleted is False else 'Form maked as deleted successfully.'
            logger.info(message, "Form ID:", pk, "Re-active param:", re_active_param)
            return Response({'message': message}, status=status.HTTP_200_OK)
        
        except Form.DoesNotExist:
            return Response({'error': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)
        
    def delete(self, request, pk):
        try:
            form = get_object_or_404(Form, pk=pk)
            commit_param = request.query_params.get('commit', 'false')
            print("commit_param ", commit_param)
            if commit_param == 'true':
                form.delete()
                logger.info("Form permanently deleted: %s", pk)
                return Response({'message': 'From permanently deleted.'}, status=status.HTTP_204_NO_CONTENT)
            else:
                # Soft delete
                form.is_deleted = True
                form.deletedBy = request.user
                form.save()
                logger.info("Form soft deleted (is_deleted=True): %s", pk)
                return Response({'message': 'Form marked as deleted.'}, status=status.HTTP_200_OK)
        except Form.DoesNotExist:
            return Response({'error': 'Form not Found.'}, status=status.HTTP_404_NOT_FOUND)


class FormDeletedListView(APIView):
    permission_classes = [IsAdmin]
    def get(self, request):
        try:
            form = Form.objects.filter(is_deleted=True)
            serializer = FormSerializer(form, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Form.DoesNotExist:
            return Response({'error': 'Form not found.'}, status=status.HTTP_404_NOT_FOUND)            
        

class FormBulkDeleteView(APIView):
    permissoin_classses =[IsAdmin]
    def post(self, request):
        ids = request.data.get('ids', [])
        commit_param = request.data.get('commit', False)
        
        print("commit_param ", commit_param)
        
        if not ids:
            return Response({"error": "No IDs provided."}, status=status.HTTP_400_BAD_REQUEST)

        if commit_param:
            Form.objects.filter(id__in=ids).delete()
            logger.info("Bulk permanently delete Forms: %s", ids)
            return Response({"message": "Forms permanently deleted."}, status=status.HTTP_204_NO_CONTENT)
        else:
            form = Form.objects.filter(id__in=ids)
            form.update(is_deleted=True, deletedBy=request.user, last_deleted_date = timezone.now())
            logger.info("Bulk delete Forms: %s", ids)
            return Response({"message": "Forms marked as deleted."}, status=status.HTTP_200_OK)
            
            
class FormBulkArchiveView(APIView):
    permissoin_classses =[IsAdmin]
    def post(self, request):
        ids = request.data.get('ids', [])
        
        if not ids:
            return Response({"error": "No IDs provided."}, status=status.HTTP_400_BAD_REQUEST)

        form = Form.objects.filter(id__in=ids)
        form.update(is_archived=True, archivedBy=request.user, last_archived_date = timezone.now())
        logger.info("Bulk Archived Forms: %s", ids)
        return Response({"message": "Forms marked as Archived."}, status=status.HTTP_200_OK)
            

class FormAllDeleteView(APIView):
    permissoin_classses =[IsAdmin]
    def post(self, request):
        Form.objects.all().delete()
        return Response({"message": "Forms permanently deleted."}, status=status.HTTP_200_OK)
  
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data
        
        # Replace form_admin ID with admin's name or email
        for item in data:
            try:
                user = CustomUser.objects.get(id=item['form_admin'])
                item['form_admin'] = user.get_full_name() or user.email
            except CustomUser.DoesNotExist:
                item['form_admin'] = "Unknown"
                
        return Response(data)
    
    def destroy(self, request, *args, **kwargs):
        """
        Handle deletion of a single form.
        Maps DELETE /api/form/{pk}/ to this method (overriding default destroy).
        """
        pk = self.kwargs.get('pk')  # Get pk from URL kwargs
        if pk is None:
            return Response({"error": "Form ID is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        instance = get_object_or_404(Form, id=pk, organization=request.user.organization)
        user = request.user
        
        # Check if user is an organization admin for this form's organization
        if not OrganizationAdmin.objects.filter(organization=instance.organization, admin_user=user).exists():
            raise PermissionDenied("Only organization admins can delete forms.")

        # Perform deletion
        instance.delete()
        return Response({"message": "Form deleted successfully"}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """
        Handle bulk deletion of forms.
        Expects a POST request with a list of form IDs in the 'ids' field.
        """
        form_ids = request.data.get('ids', [])
        if not form_ids:
            return Response({"error": "No form IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = request.user
            forms = Form.objects.filter(id__in=form_ids, organization=user.organization)
            if not forms.exists():
                return Response({"error": "No forms found for the provided IDs"}, status=status.HTTP_404_NOT_FOUND)
            
            # Check permission for each form
            for form in forms:
                if not OrganizationAdmin.objects.filter(organization=form.organization, admin_user=user).exists():
                    raise PermissionDenied(f"Only organization admins can delete form with ID {form.id}.")

            # Perform bulk deletion
            forms.delete()
            return Response({"message": f"Selected form IDs {', '.join(map(str, form_ids))} deleted successfully"}, status=status.HTTP_200_OK)
        except PermissionDenied as e:
            return Response({"error": str(e)}, status=status.HTTP_403_FORBIDDEN)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class FormDetailsViewSet(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        # Fetch all forms for the user's organization
        queryset = Form.objects.filter(organization=request.user.organization)

        # Transform data to match frontend requirements
        forms_data = []
        for form in queryset:
            form_data = {
                'id': form.id,
                'title': form.title,
                'author': form.form_admin.first_name and form.form_admin.last_name,
                'created_date': form.created_at.strftime('%Y-%m-%d'),
                'latest_response': '—',  # Placeholder until FormResponse model is confirmed
                'form_type': form.get_form_type_display(),
                'repeat_schedule': 'None',  # Placeholder until repeat_schedule field is confirmed
                'responses': 0,  # Placeholder until FormResponse model is confirmed
            }
            forms_data.append(form_data)

        return Response(forms_data, status=status.HTTP_200_OK)
        

class SubmitStageAnswerView(APIView):
    permission_classes = [IsAuthenticated]
    
    @transaction.atomic
    def post(self, request):
        try:
            responses = []

            form_id = request.data.get('form')
            stage_id = request.data.get('stage')
            assignment_uuid = request.data.get('stage_assignment_uuid')
            form_submission_id = request.data.get('form_submission_id')

            if not form_id or not stage_id:
                return Response({"error": "Form ID and Stage ID are required."}, status=status.HTTP_400_BAD_REQUEST)

            # Get form — allow archived forms to be submitted (planners may reference archived forms)
            form = get_object_or_404(
                Form,
                id=form_id,
                is_deleted=False,
                organization=request.user.organization
            )

            # Get stage
            stage = get_object_or_404(Stage, id=stage_id, form=form)

            # Check if this is a todo workflow submission
            # For todo workflows, stages can be submitted independently regardless of order
            # Check if there are any tasks assigned to this user for this form
            from task.models import TaskAssignee
            is_todo_workflow = TaskAssignee.objects.filter(
                task__form_id=form_id,
                task__organization=request.user.organization,
                assigned_user=request.user
            ).exists() or TaskAssignee.objects.filter(
                task__form_id=form_id,
                task__organization=request.user.organization,
                assigned_group__members=request.user
            ).exists()

            # Check if this is a todo workflow submission
            # For todo workflows, stages can be submitted independently regardless of order
            # Check if there are any tasks assigned to this user for this form
            from task.models import TaskAssignee
            is_todo_workflow = TaskAssignee.objects.filter(
                task__form_id=form_id,
                task__organization=request.user.organization,
                assigned_user=request.user
            ).exists() or TaskAssignee.objects.filter(
                task__form_id=form_id,
                task__organization=request.user.organization,
                assigned_group__members=request.user
            ).exists()

            # Handle optional assignment_uuid - allow collaborative submissions
            if assignment_uuid is None or assignment_uuid == '':

                # For collaborative workflows (received todos), allow submission without assignment UUID
                # Create a temp assignment for the current user

                assignment_uuid = uuid.uuid4()
                StageAssignment.objects.create(
                    form=form,
                    stage=stage,
                    stage_order=stage.order,
                    user=request.user,
                    assignment_uuid=assignment_uuid,
                    organization=request.user.organization,
                )
                logger.info(f"Created temp StageAssignment {assignment_uuid} for collaborative stage {stage.id}")

            # For collaborative workflows (received todos), allow submission
            # if assignment exists, regardless of which user it was originally assigned to

            # Skip stage completion check for todo workflows since they allow multiple submissions
            # For regular workflows, check stage not already submitted
            if not is_todo_workflow:
                stageCompletionHistory = StageSubmissionHistory.objects.filter(
                    stage=stage,
                    stage_order=stage.order,
                    stage_assignment_uuid=assignment_uuid,
                    organization=request.user.organization
                )
                if stageCompletionHistory.exists():
                    raise PermissionDenied("Stage Already submitted")
                
                
            
            
            # For todo workflows, skip the assignment existence check since assignments are created during submission
            # For regular workflows, check assignment exists
            if not is_todo_workflow:
                stageAssignment = StageAssignment.objects.filter(
                    stage=stage,
                    form=form,
                    assignment_uuid=assignment_uuid
                )
                if not stageAssignment.exists():
                    raise PermissionDenied("Assignment does not exist")

            # Check previous stage completion (skip for todo workflows)
            if not is_todo_workflow and stage.order > 1:
                previousStage = StageSubmissionHistory.objects.filter(
                    stage_order=stage.order - 1,
                    form_submission_id = form_submission_id,
                )

                if not previousStage.exists():
                    return Response(
                        {"error": "Previous stage must be submitted before submitting this stage."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Create or fetch FormSubmission
            if stage.order == 1 or is_todo_workflow:
                # For todo workflows and stage 1, always create a new FormSubmission
                # since each form fill should be a separate submission
                formSubmissionInstance = FormSubmision.objects.create(
                    form=form,
                    submission_initiated_by=request.user,
                    organization=request.user.organization,
                    submission_initiated_stage=stage
                )
                # Mark associated planner as completed when form is submitted
                from planner.models import PlannerAssignment, PlannerSubmission
                from django.utils import timezone
                planner_assignment_id = request.data.get('planner_assignment_id')
                if planner_assignment_id:
                    # Mark only the specific planner that was started
                    planner_to_update = PlannerAssignment.objects.filter(
                        id=planner_assignment_id,
                        organization=request.user.organization,
                        is_completed=False
                    ).first()
                    if planner_to_update:
                        if not planner_to_update.started_by:
                            planner_to_update.started_by = request.user
                            planner_to_update.started_on = timezone.now()
                        planner_to_update.is_completed = True
                        planner_to_update.completed_on = timezone.now()
                        planner_to_update.completed_by = request.user
                        planner_to_update.save()
                        PlannerSubmission.objects.get_or_create(
                            planner_assignment=planner_to_update,
                            form_submission=formSubmissionInstance,
                            defaults={'submitted_by': request.user}
                        )
                # If no planner_assignment_id provided, do NOT complete any planner
            else:
                # For regular forms, require existing form submission ID for stage.order > 1
                formSubmissionInstance_qs = FormSubmision.objects.filter(id=form_submission_id)
                if not formSubmissionInstance_qs.exists():
                    raise ValidationError("Valid Form submission ID required when stage order is greater than 1.")
                formSubmissionInstance = formSubmissionInstance_qs.first()

            # Save each answer
            for answer_data in request.data.get('answers', []):
                answer = {
                    'question': answer_data.get('question'),
                    'question_type': answer_data.get('question_type'),
                    'answer': answer_data.get('answer'),
                    'Form': form_id,
                    'stage': stage_id,
                    'division': answer_data.get('division'),
                    'sub_division': answer_data.get('sub_division'),
                    'location': answer_data.get('location'),
                    'user': answer_data.get('user'),
                    'submitted_by': request.user.id,
                    'submission': formSubmissionInstance.id,
                    'organization': request.user.organization.id,
                    'other_text': answer_data.get('other_text'),
                    'remarks': answer_data.get('remarks'),
                    'approved_stages': answer_data.get('approved_stages'),
                    'signature': answer_data.get('signature'),
                }

                serializer = AnswerSerializer(data=answer)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                responses.append(serializer.data)

            
            #update form_submission to stage assignment for current user only
            StageSubmissionHistory.objects.create(
                stage = stage,
                stage_order = stage.order,
                form_submission = formSubmissionInstance,
                stage_assignment_uuid = assignment_uuid,
                completed_by = request.user,
                organization = request.user.organization
            )

            # Notify form admin on every stage submission
            transaction.on_commit(
                lambda: _notify_form_admin_on_submission_commit(request, form, formSubmissionInstance, stage=stage)
            )

            # Mark the corresponding StageAssignment as fulfilled for todo workflows
            try:
                stage_assignment = StageAssignment.objects.filter(
                    stage=stage,
                    user=request.user,
                    assignment_uuid=assignment_uuid,
                    organization=request.user.organization
                ).first()

                if stage_assignment:
                    stage_assignment.is_assignment_fullfilled = True
                    stage_assignment.save()
                    print(f"Marked StageAssignment {stage_assignment.id} as fulfilled for stage {stage.id}")
            except Exception as e:
                print(f"Error marking StageAssignment as fulfilled: {str(e)}")
                # Continue without failing the submission
            
            
            # Mark delete previous stage assignments
            StageAssignment.objects.filter( assignment_uuid=assignment_uuid ).delete()
            
            
            # Assign same stage to current user if stage = 1
            if stage.order == 1:
                StageAssignment.objects.create(
                    form_id=form_id,
                    stage=stage,
                    stage_order=stage.order,
                    user=request.user,
                    assignment_uuid=uuid.uuid4(),
                    organization=request.user.organization,
                )
            
            approved_stage = any(
                ans.get("approved_stages") is True
                for ans in request.data.get("answers", [])
            )
            if approved_stage:
                formSubmissionInstance.completed_by = request.user
                formSubmissionInstance.completed_on = timezone.now()
                formSubmissionInstance.is_completed = True
                formSubmissionInstance.save()
                _auto_share_submission_response(formSubmissionInstance, request.user, _build_public_base_url(request))

                StageAssignment.objects.filter(
                    form=form,
                    form_submission=formSubmissionInstance,
                    organization=request.user.organization,
                    is_assignment_fullfilled=False
                ).update(is_assignment_fullfilled=True)

                return Response(
                    {
                        'message': 'Answers submitted successfully',
                        'answers': responses,
                        'next_stage_assigning_required': False,
                        'next_stage_id': None,
                        'form_submission_id': formSubmissionInstance.id,
                        'assignment_uuid': assignment_uuid,
                        'is_form_completed': True
                    },
                    status=status.HTTP_201_CREATED
                )
            # Prepare next stage if needed
            nextStageInstance = Stage.objects.filter(
                form=form,
                order=stage.order + 1,
                form__is_deleted=False,
                form__is_archived=False,
                form__organization=request.user.organization
            )

            next_stage_assigning_required=False
            nextStageId = None
            is_form_completed = False

            # If more stages exist, check next stage access
            if nextStageInstance.exists():
                nextStage = nextStageInstance.first()
                nextStageId = nextStage.id
                nextStageOrder = nextStage.order
                nextAssignmentUUID = uuid.uuid4()

                print("Next stage found for form:", form_id, "Stage order:", nextStageOrder)
                nextStageAccessInstances = StageAccess.objects.filter(stage=nextStage, form=form)

                # Check if this form is part of a todo workflow
                from task.models import TaskAssignee
                todo_workflow_assignees = TaskAssignee.objects.filter(
                    task__form=form,
                    task__organization=request.user.organization
                ).select_related('task', 'assigned_user', 'assigned_group')

                is_todo_workflow = todo_workflow_assignees.exists()

                # For todo workflows, ALWAYS create stage assignments for all task assignees
                # regardless of stage access rules
                if is_todo_workflow:
                    print("Form is part of todo workflow - creating dynamic stage assignments for all task users")

                    # Collect all unique users assigned to tasks for this form
                    todo_users = set()
                    task_to_users = {}  # Track which task each user belongs to

                    for assignee in todo_workflow_assignees:
                        if assignee.assigned_user:
                            todo_users.add(assignee.assigned_user.id)
                            task_to_users[assignee.assigned_user.id] = assignee.task
                        elif assignee.assigned_group:
                            for member in assignee.assigned_group.members.filter(organization=request.user.organization):
                                todo_users.add(member.id)
                                task_to_users[member.id] = assignee.task

                    todo_users_list = list(todo_users)
                    print(f"Creating stage assignments for {len(todo_users_list)} todo workflow users: {todo_users_list}")

                    for user_id in todo_users_list:
                        task_for_user = task_to_users.get(user_id)
                        StageAssignment.objects.create(
                            form_id=form_id,
                            stage=nextStage,
                            stage_order=nextStageOrder,
                            user_id=user_id,
                            assignment_uuid=nextAssignmentUUID,
                            form_submission=formSubmissionInstance,
                            organization=request.user.organization,
                            task=task_for_user  # Link to the originating task
                        )

                    print(f"Successfully created {len(todo_users_list)} stage assignments for next stage in todo workflow")
                elif nextStageAccessInstances.exists():
                    nextStageAccessType = nextStageAccessInstances.first().access_type
                    print("Next stage access found for form: ", form_id, " and Next Stage order: ", nextStageOrder, "and Next stage access type: ", nextStageAccessType)

                    # If next stage has access as ORGANIZATION, we need to assign next stage to users/group manually from frontend
                    next_stage_assigning_required = nextStageAccessType == StageAccessType.ORGANIZATION

                    # Assign next stage to corresponding users based on access type
                    if nextStageAccessType == StageAccessType.USER:
                        for user_id in nextStageAccessInstances.values_list('allow_user', flat=True).distinct():
                            StageAssignment.objects.create(
                                form_id=form_id,
                                stage=nextStage,
                                stage_order=nextStageOrder,
                                user_id=user_id,
                                assignment_uuid=nextAssignmentUUID,
                                form_submission=formSubmissionInstance,
                                organization=request.user.organization
                            )

                    # Assign next stage to corresponding groups users based on access type
                    elif nextStageAccessType == StageAccessType.GROUP:
                        allUsers = []

                        for group_id in nextStageAccessInstances.values_list('allow_group', flat=True).distinct():
                            groupInstance = Groups.objects.filter(id=group_id, organization=request.user.organization)
                            if groupInstance.exists():
                                group = groupInstance.first()
                                print("Group found: ", group.name, " for form: ", form_id, " and Next Stage order: ", nextStageOrder)
                                allUsers.extend(group.members.all().distinct())

                        allUsers = list(set(allUsers))
                        for user in allUsers:
                            StageAssignment.objects.create(
                                form_id=form_id,
                                stage=nextStage,
                                stage_order=nextStageOrder,
                                user_id=user.id,
                                assignment_uuid=nextAssignmentUUID,
                                form_submission=formSubmissionInstance,
                                organization=request.user.organization
                            )

                    # Assign next stage to previous stage users based on access type
                    elif nextStageAccessType == StageAccessType.PREVIOUS_STAGE:
                        sourceStgeUUID = nextStageAccessInstances.first().allow_stage
                        print("Previous stage access type detected for form:", form_id, "Stage order:", nextStageOrder)
                        print("Source stage UUID for next stage assignment: ", sourceStgeUUID)
                        preStages = Stage.objects.filter( form=form, stage_uuid=sourceStgeUUID)
                        if preStages.exists():
                            preStage = preStages.first()
                            stageHistroty = StageSubmissionHistory.objects.filter(
                                stage=preStage,
                                stage_order=preStage.order,
                                form_submission=formSubmissionInstance,
                                organization=request.user.organization
                            ).order_by('-completed_on').first()
                            if stageHistroty:
                                print("Last time the source stage ", preStage, "was completed by: ", stageHistroty.completed_by, "Assigning the next stage to him/her.")
                                StageAssignment.objects.create(
                                    form_id=form_id,
                                    stage=nextStage,
                                    stage_order=nextStageOrder,
                                    user_id=stageHistroty.completed_by.id,
                                    assignment_uuid=nextAssignmentUUID,
                                    form_submission=formSubmissionInstance,
                                    organization=request.user.organization
                                )
                else:
                    # No StageAccess records exist and not a todo workflow - no assignments created
                    print("No StageAccess records found and form is not part of todo workflow - no dynamic assignments created")
                    # For non-todo forms, if no StageAccess records exist, we cannot assign the next stage
                    # This maintains the existing behavior for regular forms

            else:
                print("No next stage found for form:", form_id, "Stage order:", stage.order)
                # Make the form submission as completed
                formSubmissionInstance.completed_by = request.user
                formSubmissionInstance.completed_on = timezone.now()
                formSubmissionInstance.is_completed = True
                formSubmissionInstance.save()
                _auto_share_submission_response(formSubmissionInstance, request.user, _build_public_base_url(request))

                is_form_completed = True

            logger.info(f"Submitted {len(responses)} answers by user {request.user.id} for form {form_id}")
            return Response(
                {
                    'message': 'Answers submitted successfully',
                    'answers': responses,
                    'next_stage_assigning_required': next_stage_assigning_required,
                    'next_stage_id': nextStageId,
                    'form_submission_id': formSubmissionInstance.id,
                    'assignment_uuid': assignment_uuid,
                    'is_form_completed': is_form_completed
                },
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            logger.error(f"Error submitting answers: {str(e)}")
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


def _logic_condition_met(logic, parent_answer):
    """Return True if the parent answer satisfies the logic condition.

    Mirrors the frontend `matchLogicCondition` helper so the backend
    can evaluate which logic questions are visible after a submission.
    """
    if parent_answer is None or parent_answer == "":
        return False

    parts = [p.strip() for p in str(parent_answer).split("|") if p.strip()]
    target = str(logic.logic_value).strip()

    def _normalize(part):
        # If the answer is an option ID, convert it to the option text so
        # we can compare against logic_value (which is stored as option text).
        if part.isdigit():
            option = Option.objects.filter(
                id=int(part), question_id=logic.question_id
            ).first()
            if option:
                return str(option.option).strip()
        return part

    if logic.comparison:
        try:
            target_num = float(target)
        except (ValueError, TypeError):
            target_num = 0

        def _compare(part):
            try:
                part_num = float(_normalize(part))
            except (ValueError, TypeError):
                return False
            comparison = logic.comparison
            if comparison == "equals":
                return part_num == target_num
            if comparison == "greater_than":
                return part_num > target_num
            if comparison == "less_than":
                return part_num < target_num
            if comparison == "greaterthan_or_equalto":
                return part_num >= target_num
            if comparison == "lessthan_or_equalto":
                return part_num <= target_num
            return False

        return any(_compare(p) for p in parts)

    normalized = [_normalize(p) for p in parts]
    if logic.logic_type == "is":
        return any(p == target for p in normalized)
    if logic.logic_type == "is_not":
        return any(p != target for p in normalized)
    return False


def _cleanup_stale_logic_answers(form, submission, submitted_answers_data, organization):
    """Delete answers for logic questions that are hidden after the latest edit.

    When a controlling option changes, the frontend only sends the currently
    visible answers. This cleanup removes the old answers for hidden logic
    children from the database so they are not re-hydrated on the next edit.
    """
    def _coerce_id(value):
        if value is None:
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return value

    submitted_question_ids = {
        _coerce_id(a.get("question"))
        for a in submitted_answers_data
        if a.get("question") is not None
    }

    current_answers = Answer.objects.filter(
        submission=submission,
        organization=organization,
    ).select_related("question")

    audit_group_ids = set()
    for answer in current_answers:
        if answer.question.audit_group_id:
            audit_group_ids.add(answer.question.audit_group_id)

    logic_qs = Logic.objects.filter(form=form)
    if audit_group_ids:
        logic_qs = logic_qs.filter(
            models.Q(audit_group_id__in=audit_group_ids)
            | models.Q(audit_group__isnull=True)
        )
    elif submission.submission_initiated_stage_id:
        logic_qs = logic_qs.filter(
            models.Q(stage_id=submission.submission_initiated_stage_id)
            | models.Q(stage__isnull=True)
        )
    logic_qs = logic_qs.prefetch_related("logic_questions")

    child_question_ids = set()
    for logic in logic_qs:
        for child in logic.logic_questions.all():
            child_question_ids.add(child.id)

    if not child_question_ids:
        return

    # Iteratively determine visible logic children. Hidden logic children are
    # treated as unanswered so nested logic is evaluated correctly.
    answer_dict = {a.question_id: a.answer for a in current_answers}
    visible_children = set()

    while True:
        newly_visible = set()
        for logic in logic_qs:
            parent_answer = answer_dict.get(logic.question_id)
            if parent_answer is None or parent_answer == "":
                continue
            if _logic_condition_met(logic, parent_answer):
                for child in logic.logic_questions.all():
                    if child.id not in visible_children:
                        newly_visible.add(child.id)
        if not newly_visible:
            break
        visible_children.update(newly_visible)
        for answer in current_answers:
            if (
                answer.question_id in child_question_ids
                and answer.question_id not in visible_children
                and answer_dict.get(answer.question_id) not in (None, "")
            ):
                answer_dict[answer.question_id] = None

    stale_ids = [
        answer.id
        for answer in current_answers
        if answer.question_id in child_question_ids
        and answer.question_id not in visible_children
        and answer.question_id not in submitted_question_ids
    ]
    if stale_ids:
        Answer.objects.filter(id__in=stale_ids).delete()


def _sync_planner_location_from_submission(form, submission, user, organization):
    """Sync PlannerAssignment.location when a form's location answer is edited.

    After a form submission is edited, find the latest location-type answer
    and update all PlannerAssignment records linked to this form + user so
    the planner reflects the new location.
    """
    try:
        from planner.models import PlannerAssignment
        from user.models import Locations

        # Only proceed if the form has a LOCATION question
        if not Question.objects.filter(form=form, question_type=QuestionType.LOCATION).exists():
            logger.debug(f"[sync_planner_loc] Form {form.id} has no LOCATION question, skipping")
            return

        location_answer = Answer.objects.filter(
            submission=submission,
            question_type=QuestionType.LOCATION,
            organization=organization,
        ).select_related('location').order_by('-submitted_on').first()

        if not location_answer:
            logger.debug(f"[sync_planner_loc] No location answer found for submission {submission.id}")
            return

        logger.info(
            f"[sync_planner_loc] Found location answer: location_id={location_answer.location_id}, "
            f"answer='{location_answer.answer}', other_text='{location_answer.other_text}'"
        )

        new_location = None
        if location_answer.location_id:
            new_location = location_answer.location
        elif location_answer.answer:
            loc_name = location_answer.answer.strip()
            new_location = Locations.objects.filter(
                name__iexact=loc_name,
                organization=organization,
            ).first()
            if not new_location:
                # Try partial match (contains)
                new_location = Locations.objects.filter(
                    name__icontains=loc_name,
                    organization=organization,
                ).first()
            if not new_location and location_answer.other_text:
                loc_name = location_answer.other_text.strip()
                new_location = Locations.objects.filter(
                    name__iexact=loc_name,
                    organization=organization,
                ).first()
                if not new_location:
                    new_location = Locations.objects.filter(
                        name__icontains=loc_name,
                        organization=organization,
                    ).first()
        elif location_answer.other_text:
            loc_name = location_answer.other_text.strip()
            new_location = Locations.objects.filter(
                name__iexact=loc_name,
                organization=organization,
            ).first()
            if not new_location:
                new_location = Locations.objects.filter(
                    name__icontains=loc_name,
                    organization=organization,
                ).first()

        logger.info(
            f"[sync_planner_loc] Resolved new_location: "
            f"{new_location.name if new_location else 'None'} (id={new_location.id if new_location else 'None'})"
        )

        # Find PlannerAssignment records for this form + user (including completed ones)
        user_groups = user.user_groups.all() if hasattr(user, 'user_groups') else []
        planner_qs = PlannerAssignment.objects.filter(
            form=form,
            organization=organization,
        ).filter(
            Q(user=user) | Q(group__in=user_groups) | Q(leader=user)
        )

        # Log what we found
        planner_count = planner_qs.count()
        logger.info(
            f"[sync_planner_loc] Found {planner_count} planner assignment(s) "
            f"for form={form.id}, user={user.id}, groups={[g.id for g in user_groups]}"
        )
        for pa in planner_qs:
            logger.info(
                f"[sync_planner_loc]   PA id={pa.id}, order_id={pa.order_id}, "
                f"current_location={pa.location.name if pa.location else 'None'}, "
                f"is_completed={pa.is_completed}"
            )

        updated_count = planner_qs.update(location=new_location)
        if updated_count:
            logger.info(
                f"[sync_planner_loc] Synced planner location to "
                f"'{new_location.name if new_location else 'None'}' for "
                f"{updated_count} assignment(s) after form {form.id} edit"
            )
        else:
            logger.warning(
                f"[sync_planner_loc] No planner assignments updated for form {form.id}"
            )
    except Exception as e:
        logger.warning(f"[sync_planner_loc] Failed to sync planner location from submission: {e}", exc_info=True)


class SubmitGroupAnswerView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        try:
            responses = []
            form_id = request.data.get('form')
            assignment_uuid = request.data.get('group_assignment_uuid')
            
            if not form_id:
                return Response({"error": "Form ID is required."}, status=status.HTTP_400_BAD_REQUEST)

            # Get form — allow archived forms to be submitted (planners may reference archived forms)
            form = get_object_or_404(
                Form,
                id=form_id,
                is_deleted=False,
                organization=request.user.organization
            )
            
            is_audit = form.form_type == 'audit'
            
            # For audit forms, generate UUID if not provided
            if is_audit and not assignment_uuid:
                assignment_uuid = str(uuid.uuid4())
            
            # For non-audit forms, require assignment_uuid
            if not is_audit and not assignment_uuid:
                return Response(
                    {"error": "group_assignment_uuid is required."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check GroupAssignment only for non-audit forms
            if not is_audit:
                groupAssignment = GroupAssignment.objects.filter(
                    form=form,
                    user=request.user,
                    assignment_uuid=assignment_uuid
                )
                if not groupAssignment.exists():
                    raise PermissionDenied("You are not assigned to this form")

            # If a submission ID is provided, update the existing group/audit submission
            # instead of creating a duplicate. This is used when editing a sent form.
            form_submission_id = request.data.get('form_submission_id')
            if form_submission_id:
                logger.info(f"[SubmitGroupAnswerView] EDIT PATH - form_submission_id={form_submission_id}, form_id={form_id}")
                # Log location-related answers in the payload
                for ans in request.data.get('answers', []):
                    if ans.get('question_type') == 'location':
                        logger.info(f"[SubmitGroupAnswerView] Location answer in payload: question={ans.get('question')}, answer={ans.get('answer')}, location={ans.get('location')}, other_text={ans.get('other_text')}")
                from django.utils import timezone
                try:
                    formSubmissionInstance = FormSubmision.objects.get(
                        id=form_submission_id,
                        form=form,
                        organization=request.user.organization
                    )
                except FormSubmision.DoesNotExist:
                    return Response(
                        {"error": "Form submission not found."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                if not form.allow_editing:
                    raise PermissionDenied("Editing is not allowed for this form.")

                if formSubmissionInstance.completed_by != request.user:
                    raise PermissionDenied("Only the user who completed this form can edit it.")

                # Update or create answers for the existing submission
                def coerce_fk(value):
                    if value is None:
                        return None
                    try:
                        return int(value)
                    except (ValueError, TypeError):
                        return value

                # Map payload field names to model _id attribute names for ForeignKey fields
                FK_FIELD_MAP = {
                    'question': 'question_id',
                    'Form': 'Form_id',
                    'division': 'division_id',
                    'sub_division': 'sub_division_id',
                    'location': 'location_id',
                    'user': 'user_id',
                    'submitted_by': 'submitted_by_id',
                    'submission': 'submission_id',
                    'organization': 'organization_id',
                }

                for auditAns in request.data.get('answers', []):
                    question_id = coerce_fk(auditAns.get('question'))
                    answer_qs = Answer.objects.filter(
                        question_id=question_id,
                        submission=formSubmissionInstance,
                        organization=request.user.organization
                    )
                    answer_payload = {
                        'question': question_id,
                        'question_type': auditAns.get('question_type'),
                        'answer': auditAns.get('answer'),
                        'Form': coerce_fk(form_id),
                        'division': coerce_fk(auditAns.get('division')),
                        'sub_division': coerce_fk(auditAns.get('sub_division')),
                        'location': coerce_fk(auditAns.get('location')),
                        'user': coerce_fk(auditAns.get('user')),
                        'submitted_by': coerce_fk(request.user.id),
                        'submission': coerce_fk(formSubmissionInstance.id),
                        'organization': coerce_fk(request.user.organization.id),
                        'other_text': auditAns.get('other_text'),
                    }
                    if answer_qs.exists():
                        answer_instance = answer_qs.first()
                        for attr, value in answer_payload.items():
                            if value is not None:
                                model_attr = FK_FIELD_MAP.get(attr, attr)
                                setattr(answer_instance, model_attr, value)
                        answer_instance.save()
                        if auditAns.get('question_type') == 'location':
                            logger.info(f"[SubmitGroupAnswerView] Saved location answer: id={answer_instance.id}, location_id={answer_instance.location_id}, answer={answer_instance.answer}")
                        responses.append(AnswerSerializer(answer_instance).data)
                    else:
                        serializer = AnswerSerializer(data=answer_payload)
                        serializer.is_valid(raise_exception=True)
                        serializer.save()
                        responses.append(serializer.data)

                # Refresh audit history summary when provided
                if is_audit:
                    form_overall_status = request.data.get('form_overall_status')
                    form_overall_score = request.data.get('form_overall_score')
                    form_critical_failed = request.data.get('form_critical_failed', 0)
                    groups_status_list = request.data.get('groups_status', [])

                    if form_overall_score:
                        try:
                            form_overall_score = float(form_overall_score)
                        except (ValueError, TypeError):
                            form_overall_score = None

                    try:
                        form_critical_failed = int(form_critical_failed)
                    except (ValueError, TypeError):
                        form_critical_failed = 0

                    if groups_status_list:
                        AuditFormSubmissionHistory.objects.filter(
                            form_submission=formSubmissionInstance
                        ).delete()
                        for group_data in groups_status_list:
                            group_id_value = group_data.get('group_id')
                            group_instance = None
                            if group_id_value:
                                try:
                                    group_instance = AuditGroup.objects.get(
                                        id=group_id_value,
                                        form=form,
                                        organization=request.user.organization
                                    )
                                except AuditGroup.DoesNotExist:
                                    continue

                            group_score = group_data.get('group_score')
                            if group_score:
                                try:
                                    group_score = float(group_score)
                                except (ValueError, TypeError):
                                    group_score = None

                            AuditFormSubmissionHistory.objects.create(
                                form_submission=formSubmissionInstance,
                                group_assignment_uuid=assignment_uuid,
                                completed_by=request.user,
                                organization=request.user.organization,
                                form_overall_status=form_overall_status,
                                form_overall_score=form_overall_score,
                                form_critical_failed=form_critical_failed,
                                groups_status=group_data.get('status'),
                                group_score=group_score,
                                group_percentage=str(group_data.get('group_percentage', '')),
                                group_critical_failed=1 if group_data.get('critical', False) else 0,
                                form_id=form,
                                group_id=group_instance
                            )
                    else:
                        history = AuditFormSubmissionHistory.objects.filter(
                            form_submission=formSubmissionInstance
                        ).first()
                        if history:
                            if form_overall_status is not None:
                                history.form_overall_status = form_overall_status
                            if form_overall_score is not None:
                                history.form_overall_score = form_overall_score
                            if form_critical_failed is not None:
                                history.form_critical_failed = form_critical_failed
                            history.save()

                # Remove answers for logic questions that are no longer visible
                # after this edit so stale data does not reappear on later re-edits.
                _cleanup_stale_logic_answers(
                    form, formSubmissionInstance, request.data.get('answers', []), request.user.organization
                )

                # Sync planner location to reflect edited location answer
                _sync_planner_location_from_submission(
                    form, formSubmissionInstance, request.user, request.user.organization
                )

                return Response(
                    {
                        'message': 'Answers updated successfully',
                        'answers': responses,
                        'form_submission_id': formSubmissionInstance.id,
                        'assignment_uuid': assignment_uuid,
                        'is_form_completed': True
                    },
                    status=status.HTTP_200_OK
                )

            formSubmissionInstance = FormSubmision.objects.create(
                form=form,
                submission_initiated_by=request.user,
                organization=request.user.organization
            )
            # Mark associated planner as completed when form is submitted
            from planner.models import PlannerAssignment, PlannerSubmission
            from django.utils import timezone
            planner_assignment_id = request.data.get('planner_assignment_id')
            if planner_assignment_id:
                planner_to_update = PlannerAssignment.objects.filter(
                    id=planner_assignment_id,
                    organization=request.user.organization
                ).first()
                if planner_to_update:
                    planner_to_update.is_completed = True
                    planner_to_update.completed_on = timezone.now()
                    planner_to_update.completed_by = request.user
                    planner_to_update.save(update_fields=['is_completed', 'completed_on', 'completed_by'])
                    PlannerSubmission.objects.get_or_create(
                        planner_assignment=planner_to_update,
                        form_submission=formSubmissionInstance,
                        defaults={'submitted_by': request.user}
                    )
            # If no planner_assignment_id provided, do NOT complete any planner
            
            groupInstance = AuditGroup.objects.filter(form_id=form_id).order_by('order').first()
            
            # Get group_id from request data if provided (for audit forms)
            group_id_from_request = request.data.get('group_id')
            if group_id_from_request and is_audit:
                try:
                    groupInstance = AuditGroup.objects.get(
                        id=group_id_from_request,
                        form_id=form_id,
                        organization=request.user.organization
                    )
                except AuditGroup.DoesNotExist:
                    pass  # Fall back to first group if not found
            
            # Save audit info answers
            for auditAns in request.data.get('answers', []):
                answer = {
                    'question': auditAns.get('question'),
                    'question_type': auditAns.get('question_type'),
                    'answer': auditAns.get('answer'),
                    'Form': form_id,
                    'division': auditAns.get('division'),
                    'sub_division': auditAns.get('sub_division'),
                    'location': auditAns.get('location'),
                    'user': auditAns.get('user'),
                    'submitted_by': request.user.id,
                    'submission': formSubmissionInstance.id,
                    'organization': request.user.organization.id,
                    'other_text': auditAns.get('other_text'),
                }

                serializer = AnswerSerializer(data=answer)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                responses.append(serializer.data)
                
            # Prepare audit score calculation data (only for audit forms)
            if is_audit:
                # Extract score calculation data from request payload
                # These fields will be sent from mobile app
                form_overall_status = request.data.get('form_overall_status')
                form_overall_score = request.data.get('form_overall_score')
                form_critical_failed = request.data.get('form_critical_failed', 0)
                groups_status_list = request.data.get('groups_status', [])

                # Convert string percentages to Decimal if provided as strings
                if form_overall_score:
                    try:
                        form_overall_score = float(form_overall_score)
                    except (ValueError, TypeError):
                        form_overall_score = None

                # Convert critical_failed to int
                try:
                    form_critical_failed = int(form_critical_failed)
                except (ValueError, TypeError):
                    form_critical_failed = 0

                # Create individual records for each group in groups_status
                for group_data in groups_status_list:
                    group_id_value = group_data.get('group_id')
                    group_instance = None
                    if group_id_value:
                        try:
                            group_instance = AuditGroup.objects.get(
                                id=group_id_value,
                                form=form,
                                organization=request.user.organization
                            )
                        except AuditGroup.DoesNotExist:
                            # Skip if group not found
                            continue

                    # Convert group score if needed
                    group_score = group_data.get('group_score')
                    if group_score:
                        try:
                            group_score = float(group_score)
                        except (ValueError, TypeError):
                            group_score = None

                    AuditFormSubmissionHistory.objects.create(
                        form_submission=formSubmissionInstance,
                        group_assignment_uuid=assignment_uuid,
                        completed_by=request.user,
                        organization=request.user.organization,
                        form_overall_status=form_overall_status,
                        form_overall_score=form_overall_score,
                        form_critical_failed=form_critical_failed,
                        groups_status=group_data.get('status'),  # Store individual status string
                        group_score=group_score,
                        group_percentage=str(group_data.get('group_percentage', '')),
                        group_critical_failed=1 if group_data.get('critical', False) else 0,
                        form_id=form,
                        group_id=group_instance
                    )

                # If no per-group summary was provided, create a fallback audit history
                # so the submission appears in submission-history for mobile clients.
                if not groups_status_list:
                    AuditFormSubmissionHistory.objects.create(
                        form_submission=formSubmissionInstance,
                        group_assignment_uuid=assignment_uuid,
                        completed_by=request.user,
                        organization=request.user.organization,
                        form_overall_status=form_overall_status or None,
                        form_overall_score=form_overall_score if form_overall_score is not None else None,
                        form_critical_failed=form_critical_failed or 0,
                        groups_status=None,
                        group_score=None,
                        group_percentage=None,
                        group_critical_failed=0,
                        form_id=form,
                        group_id=None
                    )
            
            # Only delete GroupAssignment for non-audit forms
            if not is_audit:
                GroupAssignment.objects.filter(assignment_uuid=assignment_uuid).delete()
            
            # Create new assignment for next stage (for all forms)
            GroupAssignment.objects.create(
                form_id=form_id,
                group_id=groupInstance.id,
                group_order=1,
                user=request.user,
                assignment_uuid=uuid.uuid4(),
                organization=request.user.organization,
            )
            
            formSubmissionInstance.completed_by = request.user
            formSubmissionInstance.completed_on = timezone.now()
            formSubmissionInstance.is_completed = True
            formSubmissionInstance.save()
            _auto_share_submission_response(formSubmissionInstance, request.user, _build_public_base_url(request))
            is_form_completed = True

            transaction.on_commit(
                lambda: _notify_form_admin_on_submission_commit(request, form, formSubmissionInstance)
            )

            logger.info(f"Submitted {len(responses)} answers by user {request.user.id} for form {form_id}")
            return Response(
                {
                    'message': 'Answers submitted successfully',
                    'answers': responses,
                    'form_submission_id': formSubmissionInstance.id,
                    'assignment_uuid': assignment_uuid,
                    'is_form_completed': is_form_completed
                },
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            logger.error(f"Error submitting answers: {str(e)}")
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )
 

class UserAssignedFormsView(userContextAPIView, APIView):
    # permission_classes = [IsEndUserOrAdmin]
    permission_classes = [AllowAny]

    def get(self, request, user_id=None):
        try:
            if not request.user.is_authenticated:
                return Response({"error": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
            
            if not user_id:
                raise ValidationError("User Id Required.")
                            
            user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)
            stageAssignments = StageAssignment.objects.filter(
                user=user,
                form__is_deleted=False,
                form__is_archived=False,
                form__organization=user.organization
            ).select_related('form').distinct()
            
            assignedForms = [{
                'form': CompactFormSerializer(assignment.form, many=False, context={'request': request}).data,
                'stage_id': assignment.stage.id,
                'stage_order': assignment.stage.order,
                'stage_assignment_id': assignment.id,
                'assignment_uuid': assignment.assignment_uuid,
                'form_submission_id': assignment.form_submission.id if assignment.form_submission else None,
                'is_stage_submission_pending': not assignment.is_assignment_fullfilled,
                'is_form_submission_pending': (not assignment.form_submission.is_completed) if assignment.form_submission else False
            } for assignment in stageAssignments]
            
            return Response(assignedForms, status=status.HTTP_200_OK)
        except Exception as e:
            user_id_for_log = user_id if user_id else (getattr(request.user, 'id', None) if request.user.is_authenticated else None)
            logger.error(f"Error retrieving assigned forms for user {user_id_for_log}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
class UserGroupAssignedFormsView(userContextAPIView, APIView):
    def get(self, request):
        try:
            if not request.user.is_authenticated:
                return Response({"error": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
            
            print("request user :", request.user)
            print("request user id :", request.user.id)
            print("request  :", request)
            user = get_object_or_404(CustomUser, id=request.user.id, organization=request.user.organization)
            print('user list :', user)
            groupAssignment = GroupAssignment.objects.filter(
                user=user,
                form__is_deleted=False,
                form__is_archived=False,
                form__organization=user.organization
            ).select_related('form').distinct()

            assingedGroups = [{
                'form': CompactFormSerializer(assignment.form, many=False, context={'request': request}).data,
                'group_id': assignment.group.id,
                'stage_order': assignment.group.order,
                'group_assignment_id': assignment.id,
                'assignment_uuid': assignment.assignment_uuid,
            } for assignment in groupAssignment]
            
            return Response(assingedGroups,status=status.HTTP_200_OK)
        except Exception as e:
            user_id = getattr(request.user, 'id', None) if request.user.is_authenticated else None
            user_name = getattr(request.user, 'name', None) if request.user.is_authenticated else None
            logger.error(f"Error retrieving assigned forms for user {user_name or user_id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class AssignedFormsInFolderView(userContextAPIView, APIView):
    def get(self, request, folder_id=None):
        try:
            user = request.user

            # Get forms assigned to the current user
            assigned_forms = FormAssignment.objects.filter(
                user=user,
                assign_type=FormAssignType.USER
            ).values_list('form_id', flat=True)

            # Base query for assigned forms within the folder
            forms = Form.objects.filter(
                id__in=assigned_forms,
                folder_id=folder_id,
                organization=user.organization,
                is_deleted=False,
                is_archived=False,
                form_admin=user  # Ensure the form's admin matches the logged-in user
            )

            # Serialize forms
            serializer = CompactFormSerializer(forms, many=True, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error retrieving assigned forms for user {user.id}, folder_id: {folder_id or 'all'}: {str(e)}")
            return Response({"error": f"Failed to retrieve assigned forms: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)      

        
class UsedFoldersByUserView(userContextAPIView, APIView):
    def get(self, request):
        try:
            user = request.user

            # Get folders containing assigned forms
            query = StageAssignment.objects.filter(
                user=user,
                form__is_deleted=False,
                form__is_archived=False,
                form__organization=user.organization
            ).select_related('form').distinct().select_related('form')

            folder_ids = query.values_list('form__folder__id', flat=True).distinct()
            serializer = FolderSerializer(Folder.objects.filter(id__in=folder_ids), many=True, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error retrieving assigned folders for user {user.id}: {str(e)}")
            return Response({"error": f"Failed to retrieve assigned folders: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


class FormSubmissionList(APIView):
    permission_classes = [IsEndUserOrAdmin]

    def post(self, request):
        try:
            data = request.data
            form_id = data.get('forms')
            # Filter to main form submissions only; exclude followup task-close submissions.
            # Standard/location main submissions always have submission_initiated_stage set.
            # Audit main submissions keep submission_initiated_stage NULL, but they always
            # have AuditFormSubmissionHistory rows (related_name='group_submissions_history').
            # Offline submissions can exist without `submission_initiated_stage` (stage link not yet created/assigned).
            # Keep them in the list if they have any related data (answers or stage history).
            # BUT exclude submissions that ONLY have task-close question answers (these are
            # generated when completing a followup task and should not appear as form responses).
            task_close_only_ids = FormSubmision.objects.filter(
                submission_initiated_stage__isnull=True,
                stage_submissions_history__isnull=True,
                answers__question__is_task_close_question=True,
            ).exclude(
                answers__question__is_task_close_question=False
            ).values_list('id', flat=True)

            form_submissions = FormSubmision.objects.filter(
                form__id__in=form_id,
                organization=request.user.organization,
            ).filter(
                Q(submission_initiated_stage__isnull=False)
                | Q(form__form_type=FormType.AUDIT, group_submissions_history__isnull=False)
                | Q(answers__isnull=False)
                | Q(stage_submissions_history__isnull=False)
            ).exclude(
                id__in=task_close_only_ids
            ).distinct().select_related(
                'form',
                'submission_initiated_by__designation',
                'submission_initiated_by__department',
                'submission_initiated_by__location',
                'completed_by',
                'submission_initiated_stage',
            ).prefetch_related(
                'stage_submissions_history__completed_by',
            )

            # Batch-prefetch location answers for all submissions (avoids N+1 in serializer)
            submission_ids = list(form_submissions.values_list('id', flat=True))
            location_answers = Answer.objects.filter(
                submission_id__in=submission_ids,
                question_type=QuestionType.LOCATION,
                location__isnull=False,
            ).select_related('location').order_by('id')

            # Group by submission_id and attach to each submission object
            loc_by_sub = {}
            for ans in location_answers:
                sid = ans.submission_id
                if sid not in loc_by_sub:
                    loc_by_sub[sid] = []
                loc_by_sub[sid].append(ans)

            # Batch-prefetch stage submission history
            stage_histories = StageSubmissionHistory.objects.filter(
                form_submission_id__in=submission_ids,
            ).select_related('completed_by')
            hist_by_sub = {}
            for sh in stage_histories:
                sid = sh.form_submission_id
                if sid not in hist_by_sub:
                    hist_by_sub[sid] = []
                hist_by_sub[sid].append(sh)

            # Batch-prefetch stages for all forms (single query per unique form)
            form_ids = list(set(form_submissions.values_list('form_id', flat=True)))
            all_stages = Stage.objects.filter(
                form_id__in=form_ids,
                form__is_deleted=False,
                form__is_archived=False,
            ).order_by('order')
            stages_by_form = {}
            for stage in all_stages:
                fid = stage.form_id
                if fid not in stages_by_form:
                    stages_by_form[fid] = []
                stages_by_form[fid].append(stage)

            for sub in form_submissions:
                sub._prefetched_location_answers = loc_by_sub.get(sub.id, [])
                sub._prefetched_stage_submissions_history = hist_by_sub.get(sub.id, [])
                sub._cached_stages = stages_by_form.get(sub.form_id, [])
            serializer = FormSubmissionSerializer(form_submissions, many=True, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error retrieving form submissions: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
#     permission_classes = [IsEndUserOrAdmin]

#     def get(self, request, form_id, submission_id):
#         try:
#             form = get_object_or_404(Form, id=form_id, organization=request.user.organization)
#             formSchema = FormSerializer(form, many=False, context={'request': request}).data
            
#             submissions = FormSubmision.objects.filter(id=submission_id, organization=request.user.organization)
#             if submissions.exists():
#                 submissionsData = FormSubmissionSerializer(submissions.first(), context={'request': request}).data
#                 formSchema['submissionsDetail'] = submissionsData
            
#             if formSchema.get("form_type") == FormType.AUDIT:
#                 for audit_group in formSchema.get("audit_group", []):
#                     for audit_group_question in audit_group.get("questions", []):
#                         # Get answers for the main audit group question
#                         answers = Answer.objects.filter(
#                             question=audit_group_question['id'],
#                             submission=submission_id,
#                             organization=request.user.organization
#                         )
#                         audit_group_question['answers'] = AnswerSerializer(answers.first()).data if answers.exists() else {}
                    
#                         # ✅ CORRECT: Process logic questions for THIS audit group question
#                         for logics in audit_group_question.get("logics", []):
#                             for logicQuestion in logics.get("logic_questions", []):
#                                 logic_answer = Answer.objects.filter(
#                                     question=logicQuestion['id'],
#                                     submission=submission_id,
#                                     organization=request.user.organization
#                                 )
#                                 logicQuestion['answers'] = AnswerSerializer(logic_answer.first()).data if logic_answer.exists() else {}   
                                
#                         # ✅ CORRECT: Process sub-questions for THIS audit group question
#                         for subQuestion in audit_group_question.get('sub_questions', []):
#                             sub_answers = Answer.objects.filter(
#                                 question=subQuestion['id'],
#                                 submission=submission_id,
#                                 organization=request.user.organization
#                             )
#                             subQuestion['answers'] = AnswerSerializer(sub_answers.first()).data if sub_answers.exists() else {}
#                             print("==========================5")
#             else:
#                 for stage in formSchema.get('stages', []):
#                     stage_histories = StageSubmissionHistory.objects.filter(
#                         stage__id=stage['id'],
#                         form_submission__id=submission_id,
#                         organization=request.user.organization
#                     )
                    
#                     historyData={}
#                     if stage_histories.exists():
#                         stageHistory = stage_histories.first()
#                         historyData = StageSubmissionHistorySerializer(stageHistory, many=False, context={'request': request}).data
                    
#                     stage['is_completed']=stage_histories.exists()
#                     stage['completed_by']= historyData.get('completed_by', None) if historyData else None
#                     stage['completed_on']= historyData.get('completed_on') if historyData else None
                    
#                     #Answers for the stage
#                     for question in stage.get('questions', []):
#                         answers = Answer.objects.filter(
#                             question=question['id'],
#                             stage=stage['id'],
#                             submission=submission_id,
#                             organization=request.user.organization
#                         )
#                         question['answers'] = AnswerSerializer(answers.first()).data if answers.exists() else {}
                        
#                         for subQuestion in question.get('sub_questions', []):
#                             sub_answers = Answer.objects.filter(
#                                 question=subQuestion['id'],
#                                 stage=stage['id'],
#                                 submission=submission_id,
#                                 organization=request.user.organization
#                             )
#                             subQuestion['answers'] = AnswerSerializer(sub_answers.first()).data if sub_answers.exists() else {}
                        
#                         for logics in question.get("logics", []):
#                             for logicQuestion in logics.get("logic_questions", []):
#                                 logic_answer = Answer.objects.filter(
#                                     question=logicQuestion['id'],
#                                     stage=stage['id'],
#                                     submission=submission_id,
#                                     organization=request.user.organization
#                                 )
#                                 logicQuestion['answers'] = AnswerSerializer(logic_answer.first()).data if logic_answer.exists() else {}   
            
            
                        
#             return Response(formSchema, status=status.HTTP_200_OK)
#         except Exception as e:
#             logger.error(f"Error retrieving form responses for form {form_id}: {str(e)}")
#             return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class FormResponsePDFView(APIView):
    """Legacy PDF endpoint: POST only. For 'send PDF by email' use FormResponsePDFDownloadView with GET .../pdf/download?email=...&submission_ids=..."""
    permission_classes = [IsAuthenticated]

    def post(self, request, form_id):
        import sys
        from datetime import datetime as _dt
        print(f"[FormResponsePDFView POST] form_id={form_id} (this is the /submissions/pdf endpoint, NOT /pdf/download) | Time: {_dt.now()}")
        sys.stdout.flush()
        try:
            # --- Fetch form ---
            form = get_object_or_404(Form, id=form_id, organization=request.user.organization)
        
            formSchema = FormSerializer(form, many=False, context={'request': request}).data
            
            # --- Fetch submission ---
            submissions = FormSubmision.objects.filter(form_id=form_id, organization=request.user.organization)
            submission = {}
            submission_instance = None
            if submissions.exists():
                submission_instance = submissions.first()
                submissionsData = FormSubmissionSerializer(submission_instance, context={'request': request}).data
                formSchema['submissionsDetail'] = submissionsData
                submission = submissionsData

            # --- Fetch related answers ---
            answers = Answer.objects.filter(
                Form_id=form_id,
                organization=request.user.organization
            )
            for answ in answers:
                print("answers on",answ.submitted_on)
            answers_data = AnswerSerializer(answers, many=True).data
            formSchema['answers'] = answers_data

            # --- PDF buffer ---
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=A4,
                rightMargin=50, leftMargin=50,
                topMargin=50, bottomMargin=50
            )

            # --- Styles ---
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='Heading1_custom', fontSize=24, leading=28,
                                      fontName='Helvetica-Bold', alignment=1, spaceAfter=12))
            styles.add(ParagraphStyle(name='Heading2_custom', fontSize=18, leading=22,
                                      fontName='Helvetica-Bold', spaceAfter=10,
                                      textColor=colors.HexColor('#005A9C')))
            styles.add(ParagraphStyle(name='Heading3_custom', fontSize=11, leading=15,
                                      fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=6,
                                      textColor=colors.HexColor('#333333'), leftIndent=10))
            styles.add(ParagraphStyle(name='Normal_custom', fontSize=10, leading=12,
                                      fontName='Helvetica', spaceAfter=4))
            styles.add(ParagraphStyle(name='Answer_custom', fontSize=10, leading=12,
                                      fontName='Helvetica-Bold', textColor=colors.HexColor('#4CAF50')))

            story = []

            # --- Title ---
            submission_date_str = submission.get('submitted_at', 'N/A')
            if submission_instance and submission_instance.submission_initiated_on:
                submission_date_str = submission_instance.submission_initiated_on.strftime("%d-%b-%Y %I:%M %p")

            story.append(Paragraph(formSchema.get('title', 'Untitled Form'), styles['Heading1_custom']))
            story.append(Spacer(1, 10))
            story.append(Paragraph(f"<b>Submission Date:</b> {submission_date_str}", styles['Normal_custom']))
            story.append(Spacer(1, 20))

            # --- Metadata ---
            metadata_data = [
                ["Sent by:", getattr(request.user, "username", "")],
                ["Designation, Department:", f"{getattr(request.user, 'designation', '')}, {getattr(request.user, 'department', '')}"],
                ["Identifier:", getattr(request.user, "identifier", "")],
                ["Sent At:", submission_date_str],
            ]

            submitted_on = answers.first().submitted_on if answers.exists() else ""
            if submitted_on:
                metadata_data.append(["Submitted On:", submitted_on])

            metadata_table = Table(metadata_data, colWidths=[150, 350], hAlign="LEFT")
            metadata_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(metadata_table)
            story.append(Spacer(1, 20))

            # --- Action history ---
            history = submission.get("history", [])
            if history:
                story.append(Paragraph("<b>Submission History</b>", styles['Heading2_custom']))
                history_data = [["Date", "Name", "Action", "Remarks", "Stage"]]
                for h in history:
                    history_data.append([
                        h.get("date", ""),
                        h.get("name", ""),
                        h.get("action", ""),
                        h.get("remarks", ""),
                        h.get("stage", "")
                    ])
                history_table = Table(history_data, colWidths=[100, 100, 80, 120, 80])
                history_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#005A9C')),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor('#005A9C')),
                ]))
                story.append(history_table)
                story.append(Spacer(1, 20))

            # --- Location (GPS only) ---
            story.append(Paragraph("<b>Location</b>", styles["Heading3_custom"]))

            lat, lon = submission.get("location_lat"), submission.get("location_long")
            if submission.get("location_address"):
                story.append(Paragraph(submission["location_address"], styles["Normal_custom"]))
            if lat and lon:
                map_url = f"http://maps.google.com/maps?q={lat},{lon}"
                story.append(Paragraph(f'<a href="{map_url}">View on Google Maps</a>', styles["Normal_custom"]))
            if not (lat and lon):
                story.append(Paragraph("Location unavailable", styles["Normal_custom"]))

            story.append(Spacer(1, 20))

            # --- Recursive question rendering ---
            def render_question(q, level=0):
                indent = " " * (level * 2)
                question_text = q.get("question", "")
                question_type = q.get("question_type", "")
                question_id = q.get("id") or q.get("question_uuid")

                # Debug logging for table questions
                if question_type == "table":
                    print(f"[PDF DEBUG] Table Question: {question_text}")
                    print(f"[PDF DEBUG] Question ID: {question_id}")
                    print(f"[PDF DEBUG] Question data keys: {q.keys()}")
                    # Look for table answers in the answers array
                    table_answers = [a for a in formSchema.get('answers', []) if str(a.get('question')) == str(question_id)]
                    print(f"[PDF DEBUG] Found {len(table_answers)} table answers for this question")
                    for ta in table_answers:
                        print(f"[PDF DEBUG] Table answer: {ta.get('answer')}")
                    sys.stdout.flush()

                # Special handling for table type questions
                if question_type == "table":
                    story.append(Paragraph(f"<b>{indent}{question_text}</b>", styles['Normal_custom']))
                    
                    # Look for table answers in the answers array instead of question's answer field
                    table_answers = [a for a in formSchema.get('answers', []) if str(a.get('question')) == str(question_id)]
                    
                    table_rows = []
                    if table_answers:
                        # Try to get table rows from the first table answer
                        first_answer = table_answers[0].get('answer', '')
                        if isinstance(first_answer, str):
                            try:
                                table_rows = json.loads(first_answer)
                                print(f"[PDF DEBUG] Parsed table_rows from answer: {table_rows}")
                                sys.stdout.flush()
                            except Exception as e:
                                print(f"[PDF DEBUG] Failed to parse JSON: {e}")
                                sys.stdout.flush()
                        elif isinstance(first_answer, list):
                            table_rows = first_answer
                            print(f"[PDF DEBUG] Table rows from list: {table_rows}")
                            sys.stdout.flush()
                    
                    # If no rows from answer field, try sub-questions answers
                    if not table_rows and table_answers:
                        # Build rows from individual sub-question answers
                        sub_questions = q.get("sub_questions", [])
                        if sub_questions:
                            # Create one row per table answer
                            for table_answer in table_answers:
                                row = {}
                                for sub_q in sub_questions:
                                    sub_q_id = sub_q.get("id") or sub_q.get("question_uuid")
                                    # Look for sub-question answer
                                    sub_answers = [a for a in formSchema.get('answers', []) 
                                                  if str(a.get('question')) == str(sub_q_id)]
                                    if sub_answers:
                                        row[sub_q_id] = sub_answers[0].get('answer', '')
                                if row:
                                    table_rows.append(row)
                    
                    if table_rows and len(table_rows) > 0:
                        sub_questions = q.get("sub_questions", [])
                        print(f"[PDF DEBUG] Number of table rows to render: {len(table_rows)}")
                        print(f"[PDF DEBUG] Number of sub_questions: {len(sub_questions)}")
                        sys.stdout.flush()
                        
                        # Render each row
                        for row_index, row in enumerate(table_rows):
                            story.append(Paragraph(f"{indent}  <b>Row {row_index + 1}</b>", styles['Normal_custom']))
                            
                            # Render sub-question answers for this row
                            for sub_q in sub_questions:
                                sub_q_uuid = sub_q.get("question_uuid") or sub_q.get("id")
                                sub_q_text = sub_q.get("question", "")
                                row_answer = row.get(sub_q_uuid, "") if isinstance(row, dict) else ""
                                
                                story.append(Paragraph(f"{indent}    - {sub_q_text}: {row_answer}", styles['Normal_custom']))
                    else:
                        story.append(Paragraph(f"{indent}  No data", styles['Normal_custom']))
                    
                    story.append(Spacer(1, 6))
                    return

                # Regular question rendering for non-table types
                answer = q.get("answers", {}).get("answer", "")
                story.append(Paragraph(f"<b>{indent}{question_text}</b>", styles['Normal_custom']))
                story.append(Paragraph(f"{indent}<b>Answer:</b> {answer}", styles['Answer_custom']))

                for opt in q.get("options", []):
                    story.append(Paragraph(f"{indent}  - Option: {opt.get('option', '')}", styles['Normal_custom']))

                # --- Dynamic reference images ---
                for img in q.get("reference_images", []):
                    try:
                        img_response = requests.get(img, timeout=10)
                        if img_response.status_code == 200:
                            img_data = BytesIO(img_response.content)
                            pil_image = PILImage.open(img_data)
                            pil_image = pil_image.convert("RGB")
                            buf = BytesIO()
                            pil_image.save(buf, format="JPEG")
                            buf.seek(0)
                            story.append(Image(buf, width=300, height=200))
                    except Exception:
                        story.append(Paragraph(f"{indent}  - [Image could not be processed: {img}]", styles['Normal_custom']))

                for vid in q.get("reference_videos", []):
                    story.append(Paragraph(f"{indent}  - Video: {vid}", styles['Normal_custom']))

                for sub_q in q.get("sub_questions", []):
                    render_question(sub_q, level + 1)

                for logic in q.get("logics", []):
                    for logic_q in logic.get("logic_questions", []):
                        render_question(logic_q, level + 1)

                story.append(Spacer(1, 6))

            # --- Stages with Questions ---
            for stage in formSchema.get("stages", []):
                stage_name = stage.get("name") or stage.get("title") or f"Stage {stage.get('order', '')}"
                story.append(Paragraph(f"<b>{stage_name}</b>", styles['Heading3_custom']))
                
                # Separate table questions from regular questions
                table_questions = []
                regular_questions = []
                
                for q in stage.get("questions", []):
                    q_type = q.get("question_type", "")
                    q_text = q.get("question", "")
                    print(f"[PDF DEBUG] Question: {q_text}, Type: {q_type}")
                    sys.stdout.flush()
                    
                    if q_type == "table":
                        table_questions.append(q)
                    else:
                        regular_questions.append(q)
                
                print(f"[PDF DEBUG] Found {len(table_questions)} table questions and {len(regular_questions)} regular questions")
                sys.stdout.flush()
                
                # Render regular questions first
                for q in regular_questions:
                    render_question(q)
                
                # Then render all table questions with their rows grouped together
                for idx, q in enumerate(table_questions):
                    print(f"[PDF DEBUG] Rendering table question {idx + 1}: {q.get('question', '')}")
                    sys.stdout.flush()
                    render_question(q)
                
                story.append(Spacer(1, 12))

            # --- Answers Section ---
            if formSchema.get("answers"):
                story.append(Paragraph("<b>All Answers</b>", styles['Heading2_custom']))
                answers_table_data = [["Question", "Answer"]]

                for ans in formSchema['answers']:
                    q_text = Question.objects.filter(id=ans['question']).first()
                    q_label = q_text.question if q_text else f"Q-{ans['question']}"
                    answers_table_data.append([Paragraph(q_label, styles['Normal_custom']),
                                               Paragraph(ans['answer'], styles['Normal_custom'])])

                answers_table = Table(answers_table_data, colWidths=[250, 250])
                answers_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#005A9C')),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor('#005A9C')),
                ]))
                story.append(answers_table)
                story.append(Spacer(1, 12))

            # --- Build PDF ---
            doc.build(story)
            pdf = buffer.getvalue()
            buffer.close()

            emails = request.data.get("emails", [])
            # --- Send Email ---
            email = EmailMessage(
                subject=f"Form {formSchema.get('title', 'Untitled Form')}",
                body=f"""New response for {formSchema.get('title', 'Untitled Form')} Form @ {datetime.now().strftime('%I:%M %p')} from {getattr(request.user, 'name', '')}.
                        Click HERE to view the response.
                        Form Name : {formSchema.get('title', 'Untitled Form')} @ {datetime.now().strftime('%I:%M %p')}
                        Submitted by : {answ.submitted_by or ''}
                        Dept & Designation : {getattr(request.user, 'designation', '')}, {getattr(request.user, 'department', '')}
                        Submitted on : {answ.submitted_on or ''}""",
                from_email="no-reply@example.com",
                to=emails
            )
            email.attach(f"form_{form_id}_submission_{form_id}.pdf", pdf, "application/pdf")
            email.send(fail_silently=False)

            # --- Return PDF as response ---
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'inline; filename=form_{form_id}_submission_{form_id}.pdf'
            response.write(pdf)
            return response

        except Exception as e:
            logger.error(f"Error generating PDF for form {form_id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


def generate_pdf(formSchema, submission):
    """
    Generates PDF bytes from form schema and submission data.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=50, leftMargin=50,
        topMargin=50, bottomMargin=50
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Heading1_custom', fontSize=24, leading=28,
                              fontName='Helvetica-Bold', alignment=1, spaceAfter=12))
    styles.add(ParagraphStyle(name='Normal_custom', fontSize=12, leading=14, fontName='Helvetica', spaceAfter=6))
    styles.add(ParagraphStyle(name='Answer_custom', fontSize=12, leading=14, fontName='Helvetica-Bold', textColor=colors.HexColor('#4CAF50')))

    story = []

    # --- Title & submission date ---
    submission_date_str = submission.get('submitted_at', 'N/A')
    if submission.get('submission_initiated_on'):
        initiated_on = submission['submission_initiated_on']
        if isinstance(initiated_on, str):
            submission_date_str = initiated_on
        else:
            submission_date_str = initiated_on.strftime("%d-%b-%Y %I:%M %p")

    story.append(Paragraph(formSchema.get('title', 'Untitled Form'), styles['Heading1_custom']))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"<b>Submission Date:</b> {submission_date_str}", styles['Normal_custom']))
    story.append(Spacer(1, 20))

    # --- Questions & Answers ---
    def render_question(q, level=0):
        indent = " " * (level * 2)
        question_text = q.get("question", "")
        answer = q.get("answers", {}).get("answer", "")

        story.append(Paragraph(f"<b>{indent}{question_text}</b>", styles['Normal_custom']))

        # Check if answer contains image URLs and process each one
        image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp']
        processed_images = False

        if isinstance(answer, str) and 'http' in answer:
            # Split the answer by "|" first to handle pipe-separated URLs
            if "|" in answer:
                potential_urls = answer.split("|")
            else:
                potential_urls = answer.split()

            # Filter and process only image URLs
            for url in potential_urls:
                url = url.strip()
                if url and any(url.lower().endswith(ext) for ext in image_extensions):
                    try:
                        img_response = requests.get(url, timeout=10)
                        img_response.raise_for_status()
                        img_data = BytesIO(img_response.content)
                        pil_image = PILImage.open(img_data).convert("RGB")

                        # Maintain aspect ratio
                        max_width = 400
                        max_height = 300
                        img_width, img_height = pil_image.size
                        ratio = min(max_width / img_width, max_height / img_height)
                        img_width = int(img_width * ratio)
                        img_height = int(img_height * ratio)

                        buf = BytesIO()
                        pil_image.save(buf, format="JPEG")
                        buf.seek(0)
                        story.append(Image(buf, width=img_width, height=img_height))
                        story.append(Spacer(1, 6))
                        processed_images = True
                    except Exception as e:
                        story.append(Paragraph(f"{indent}<b>Answer:</b> [Image could not be loaded: {url}]", styles['Answer_custom']))
                        processed_images = True

        # If no images were processed, display the original answer
        if isinstance(answer, str) and not processed_images:
            story.append(Paragraph(f"{indent}<b>Answer:</b> {answer}", styles['Answer_custom']))

        # Options
        for opt in q.get("options", []):
            story.append(Paragraph(f"{indent}  - Option: {opt.get('option', '')}", styles['Normal_custom']))

        # Reference Images
        for img_url in q.get("reference_images", []):
            try:
                img_response = requests.get(img_url, timeout=10)
                img_response.raise_for_status()
                img_data = BytesIO(img_response.content)
                pil_image = PILImage.open(img_data).convert("RGB")

                # Maintain aspect ratio
                max_width = 400
                max_height = 300
                img_width, img_height = pil_image.size
                ratio = min(max_width / img_width, max_height / img_height)
                img_width = int(img_width * ratio)
                img_height = int(img_height * ratio)

                buf = BytesIO()
                pil_image.save(buf, format="JPEG")
                buf.seek(0)
                story.append(Image(buf, width=img_width, height=img_height))
                story.append(Spacer(1, 6))
            except Exception as e:
                story.append(Paragraph(f"{indent}  - [Image could not be processed: {img_url}]", styles['Normal_custom']))

        # Sub-questions
        for sub_q in q.get("sub_questions", []):
            render_question(sub_q, level + 1)

        story.append(Spacer(1, 6))

    # Prepare stages to process based on form type
    if formSchema.get("form_type") == "audit":
        stages_to_process = []
        if formSchema.get("audit_info"):
            audit_info_stage = {
                'id': formSchema["audit_info"].get('id', 'audit_info'),
                'name': formSchema["audit_info"].get('name', 'Audit Info'),
                'is_audit_info': True,
                'order': 0,
                'questions': formSchema["audit_info"].get("questions", [])
            }
            stages_to_process.append(audit_info_stage)
        stages_to_process.extend(formSchema.get("audit_group", []))
    else:
        stages_to_process = formSchema.get("stages", [])

    # Iterate stages
    for stage in stages_to_process:
        stage_name = stage.get("name") or stage.get("title") or f"Stage {stage.get('order', '')}"
        story.append(Paragraph(f"<b>{stage_name}</b>", styles['Normal_custom']))
        for q in stage.get("questions", []):
            render_question(q)
        story.append(Spacer(1, 12))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


# def generate_enhanced_pdf(responses, form_info, form_id):
#     """
#     Enhanced PDF generation function matching frontend structure.
#     """
#     buffer = BytesIO()
#     doc = SimpleDocTemplate(
#         buffer, pagesize=A4,
#         rightMargin=50, leftMargin=50,
#         topMargin=50, bottomMargin=50
#     )
#     dim_black = colors.Color(0.15, 0.15, 0.15)
#     light_black = colors.Color(0.2, 0.2, 0.2)

#     styles = getSampleStyleSheet()
#     styles.add(ParagraphStyle(name='Heading1_custom', fontSize=24, leading=28,
#                               fontName='Helvetica-Bold', alignment=0, spaceAfter=12))
#     styles.add(ParagraphStyle(name='Heading2_custom', fontSize=18, leading=22,
#                               fontName='Helvetica-Bold', spaceAfter=10,
#                               textColor=colors.HexColor('#005A9C')))
#     styles.add(ParagraphStyle(name='Heading3_custom', fontSize=14, leading=18,
#                               fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=6,
#                               textColor=colors.HexColor('#333333')))
#     styles.add(ParagraphStyle(name='Normal_custom', fontSize=10, leading=12,
#                               fontName='Helvetica', spaceAfter=4))
#     styles.add(ParagraphStyle(name='Answer_custom', fontSize=10, leading=12,
#                               fontName='Helvetica-Bold', textColor=light_black))

#     story = []

#     # Header
#     title = form_info.get('title', f'Form {form_id}')
#     story.append(Paragraph(title, styles['Heading1_custom']))
#     story.append(Spacer(1, 10))

#     # Form info
#     story.append(Paragraph(f"<b>Form Type:</b> {form_info.get('form_type', 'N/A')}", styles['Normal_custom']))
#     story.append(Paragraph(f"<b>Created by:</b> {form_info.get('created_by', 'N/A')}", styles['Normal_custom']))
#     if form_info.get('created_at'):
#         try:
#             created_date = datetime.fromisoformat(form_info['created_at'].replace('Z', '+00:00')).strftime('%d-%b-%Y')
#             story.append(Paragraph(f"<b>Created on:</b> {created_date}", styles['Normal_custom']))
#         except:
#             story.append(Paragraph(f"<b>Created on:</b> {form_info.get('created_at', 'N/A')}", styles['Normal_custom']))
#     story.append(Spacer(1, 20))

#     # Process each response (should be only one for download view)
#     # for response in responses:
#     for idx, response in enumerate(responses):
#         if idx > 0:
#             story.append(PageBreak()) 
        
#         # Response header
#         story.append(Paragraph(f"<b>Response ID:</b> {response.get('id', 'N/A')}", styles['Heading2_custom']))
#         story.append(Spacer(1, 10))

#         # Response details - create a table
#         details_data = [
#             ["Submission Date:", response.get('submission_initiated_on', 'N/A')],
#             ["Initiated By:", response.get('submission_initiated_by', 'N/A')],
#             ["Designation:", response.get('initiator_designation', 'N/A')],
#             ["Department:", response.get('initiator_department', 'N/A')],
#             ["Location:", response.get('initiator_location', 'N/A')],
#             ["Current Owner:", response.get('current_owner', 'N/A')],
#             ["Status:", 'Completed' if response.get('is_completed', False) else 'Pending']
#         ]

#         details_table = Table(details_data, colWidths=[150, 350])
#         details_table.setStyle(TableStyle([
#             ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
#             ("FONTSIZE", (0, 0), (-1, -1), 10),
#             ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
#             ("TOPPADDING", (0, 0), (-1, -1), 4),
#         ]))
#         story.append(details_table)
#         story.append(Spacer(1, 15))

#         # Form details with stages and questions
#         if response.get('stages'):
#             story.append(Paragraph("<b>Form Details:</b>", styles['Heading3_custom']))

#             for stage in response['stages']:
#                 # Add top spacing before stage
#                 story.append(Spacer(1, 10))

#                 # Stage header with background color
#                 stage_status = '✓' if stage.get('is_completed', False) else '(Pending)'
#                 stage_name = stage.get('name', f'Stage {stage.get("order", "")}')

#                 # Create a table for stage header with background color
#                 stage_header_data = [[f"{stage_status} {stage_name}"]]
#                 stage_header_table = Table(stage_header_data, colWidths=[480])  # Reduced to allow 10pt left margin
#                 stage_header_table.setStyle(TableStyle([
#                     ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor('#E3F2FD')),  # Light blue background
#                     ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor('#1565C0')),   # Dark blue text
#                     ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
#                     ("FONTSIZE", (0, 0), (-1, -1), 14),
#                     ("ALIGN", (0, 0), (-1, -1), "LEFT"),
#                     ("VALIGN", (0, 0), (-1, -1), "TOP"),
#                     ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
#                     ("TOPPADDING", (0, 0), (-1, -1), 8),
#                     ("LEFTPADDING", (0, 0), (-1, -1), 0),  # Added 15pt left padding to align with questions
#                 ]))
#                 story.append(stage_header_table)
#                 story.append(Spacer(1, 10))  # Add bottom spacing after stage header

#                 # Questions and answers
#                 if stage.get('questions'):
#                     for question in stage['questions']:
#                         # Question
#                         question_text = question.get('question', '')
#                         story.append(Paragraph(f"<b>{question_text}</b>", styles['Normal_custom']))
                        
#                         story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
                        
#                         story.append(Spacer(1, 4))
                        
#                         # Answer
#                         if question.get('answers') and question['answers'].get('answer'):
#                             answer = question['answers']['answer']

#                             # Check if answer contains image URLs and process each one
#                             image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
#                             processed_images = False

#                             if isinstance(answer, str):
#                                 # Split the answer by "|" first to handle pipe-separated URLs or text
#                                 potential_urls = answer.split("|") if "|" in answer else [answer]

#                                 urls_to_display = []
#                                 for url in potential_urls:
#                                     url = url.strip()
#                                     if url:
#                                         if any(url.lower().endswith(ext) for ext in image_extensions):
#                                             try:
#                                                 # Add image loading with timeout and size limit
#                                                 img_response = requests.get(url, timeout=10)
#                                                 img_response.raise_for_status()

#                                                 # Check content size (limit to 5MB)
#                                                 if len(img_response.content) > 5 * 1024 * 1024:
#                                                     story.append(Paragraph(f"<i>Image (too large):</i> {url}", styles['Answer_custom']))
#                                                 else:
#                                                     # Convert to PIL Image and resize if needed
#                                                     pil_img = PILImage.open(BytesIO(img_response.content)).convert("RGB")

#                                                     # Resize if too large (max width 400, height 300)
#                                                     max_width, max_height = 400, 300
#                                                     ratio = min(max_width / pil_img.width, max_height / pil_img.height)
#                                                     if ratio < 1:
#                                                         new_width = int(pil_img.width * ratio)
#                                                         new_height = int(pil_img.height * ratio)
#                                                         pil_img = pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)

#                                                     # Convert back to bytes
#                                                     buf = BytesIO()
#                                                     pil_img.save(buf, format="JPEG", quality=85)
#                                                     buf.seek(0)

#                                                     # Add to PDF - left aligned
#                                                     img_width = min(pil_img.width, max_width)
#                                                     img_height = min(pil_img.height, max_height)

# # Create image flowable directly with left alignment
#                                                     img = Image(buf, width=img_width, height=img_height)
#                                                     img.hAlign = 'LEFT'  # ✅ Forces image to align left

#                                                     # story.append(img)
#                                                     # story.append(Spacer(1, 5))

#                                                     # Remove the default space below
#                                                     img._restrictSize(img_width, img_height)  # Forces ReportLab to use exact size
#                                                     story.append(img)

#                                                 processed_images = True
#                                             except Exception as e:
#                                                 print(f"Failed to load image {url}: {e}")
#                                                 urls_to_display.append(f"<i>Image (failed to load):</i> {url}")
#                                                 processed_images = True
#                                         else:
#                                             # This is text content, not an image
#                                             if 'http' in url:
#                                                 urls_to_display.append(url)
#                                             else:
#                                                 # Regular text answer
#                                                 answer_text = url[:500] + '...' if len(url) > 500 else url
#                                                 story.append(Paragraph(f"{answer_text}", styles['Answer_custom']))
#                                                 processed_images = True

#                                 # Display all URLs that weren't images
#                                 if urls_to_display:
#                                     answer_text = ', '.join(urls_to_display)
#                                     answer_text = answer_text[:500] + '...' if len(answer_text) > 500 else answer_text
#                                     story.append(Paragraph(f"<i>Answer:</i> {answer_text}", styles['Answer_custom']))
#                                     processed_images = True

#                             # Fallback for non-string answers
#                             if not processed_images:
#                                 if isinstance(answer, str):
#                                     answer_text = answer[:500] + '...' if len(answer) > 500 else answer
#                                     story.append(Paragraph(f"<i>Answer:</i> {answer_text}", styles['Answer_custom']))
#                                 else:
#                                     answer_text = str(answer)[:500] + '...' if len(str(answer)) > 500 else str(answer)
#                                     story.append(Paragraph(f"<i>Answer:</i> {answer_text}", styles['Answer_custom']))

#                         # Handle sub-questions
#                         if question.get('sub_questions'):
#                             for sub_q in question['sub_questions']:
#                                 story.append(Paragraph(f"  • Sub: {sub_q.get('question', '')}", styles['Normal_custom']))
#                                 if sub_q.get('answers') and sub_q['answers'].get('answer'):
#                                     story.append(Paragraph(f"    <i>Answer:</i> {sub_q['answers']['answer']}", styles['Answer_custom']))
#                         story.append(Spacer(1, 6))
#                         # Add divider line between questions
#                         story.append(Spacer(1, 4))

#                         story.append(Spacer(1, 8))

#                 story.append(Spacer(1, 10))

#     # Footer
#     story.append(Spacer(1, 20))
#     story.append(Paragraph(f"Generated on {datetime.now().strftime('%d-%b-%Y %I:%M %p')} | Form ID: {form_id}",
#                           styles['Normal_custom']))

#     doc.build(story)
#     buffer.seek(0)
#     return buffer.getvalue()

def generate_enhanced_pdf(responses, form_info, form_id):
    """
    Enhanced PDF generation function with a premium, professional template.
    """
    buffer = BytesIO()

    # --- Setup Styles and Colors ---
    # Premium Color Palette
    primary_blue = colors.HexColor('#007bff')  # A professional, modern blue
    accent_gray = colors.HexColor('#f8f9fa')  # Light gray for backgrounds
    text_dark = colors.HexColor('#343a40')     # Dark gray for body text
    text_light = colors.HexColor('#6c757d')    # Lighter gray for secondary text

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Heading1_premium', fontSize=28, leading=34,
                             fontName='Helvetica-Bold', alignment=0, spaceAfter=20,
                             textColor=primary_blue))
    styles.add(ParagraphStyle(name='Heading2_premium', fontSize=18, leading=22,
                             fontName='Helvetica-Bold', spaceBefore=15, spaceAfter=8,
                             textColor=text_dark))
    styles.add(ParagraphStyle(name='Heading3_premium', fontSize=14, leading=18,
                             fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=6,
                             textColor=primary_blue))
    styles.add(ParagraphStyle(name='Question_premium', fontSize=12, leading=16,
                             fontName='Helvetica-Bold', textColor=text_dark, spaceAfter=4))
    styles.add(ParagraphStyle(name='Answer_premium', fontSize=11, leading=14,
                             fontName='Helvetica', textColor=text_light, spaceAfter=12))
    styles.add(ParagraphStyle(name='Footer_premium', fontSize=9, leading=10,
                             fontName='Helvetica', textColor=text_light, alignment=TA_RIGHT, spaceAfter=0))
    styles.add(ParagraphStyle(name='Details_label', fontSize=10, leading=12,
                             fontName='Helvetica-Bold', textColor=text_dark))
    styles.add(ParagraphStyle(name='Details_value', fontSize=10, leading=12,
                             fontName='Helvetica', textColor=text_light))

    # --- Build the Document Structure ---
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=50, leftMargin=50, topMargin=70, bottomMargin=50
    )
    story = []

    # --- Header and Footer Functions ---
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.drawString(doc.rightMargin, doc.bottomMargin - 15, text)
        canvas.restoreState()

    def add_custom_header(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica-Bold', 12)
        canvas.setFillColor(text_dark)
        canvas.drawString(doc.leftMargin, A4[1] - 40, "Vibro")

        canvas.setFont('Helvetica', 10)
        canvas.setFillColor(text_light)
        canvas.drawRightString(A4[0] - doc.rightMargin, A4[1] - 40, f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}")

        canvas.setStrokeColor(primary_blue)
        canvas.line(doc.leftMargin, A4[1] - 50, A4[0] - doc.rightMargin, A4[1] - 50)
        canvas.restoreState()

    doc.build(story, onFirstPage=add_custom_header, onLaterPages=add_custom_header)

    # --- Content Generation ---
    for idx, response in enumerate(responses):
        if idx > 0:
            story.append(PageBreak()) 
        
        # Form Title
        story.append(Paragraph(form_info.get('title', f'Form {form_id}'), styles['Heading1_premium']))
        
        # Form Info as a concise table
        form_info_table_data = [
            [
                Paragraph(f"<b>Form Type:</b>", styles['Details_label']),
                Paragraph(form_info.get('form_type', 'N/A'), styles['Details_value']),
                Paragraph(f"<b>Created By:</b>", styles['Details_label']),
                Paragraph(form_info.get('created_by', 'N/A'), styles['Details_value'])
            ],
            [
                Paragraph(f"<b>Created On:</b>", styles['Details_label']),
                Paragraph(datetime.fromisoformat(form_info['created_at'].replace('Z', '+00:00')).strftime('%d-%b-%Y') if form_info.get('created_at') else 'N/A', styles['Details_value']),
                Paragraph(f"<b>Form ID:</b>", styles['Details_label']),
                Paragraph(str(form_id), styles['Details_value'])
            ]
        ]
        info_table = Table(form_info_table_data, colWidths=[65, 160, 65, 160])
        info_table.setStyle(TableStyle([
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('VALIGN', (0,0), (-1,-1), 'TOP')
        ]))
        story.append(info_table)
        story.append(Spacer(1, 15))

        # Horizontal separator
        story.append(HRFlowable(width="100%", thickness=1, color=primary_blue, spaceBefore=0, spaceAfter=20))

        # Response Details
        story.append(Paragraph(f"Response ID: {response.get('id', 'N/A')}", styles['Heading2_premium']))
        
        details_data = [
            [
                Paragraph("<b>Submission Date:</b>", styles['Details_label']),
                Paragraph(response.get('submission_initiated_on', 'N/A'), styles['Details_value']),
                Paragraph("<b>Initiated By:</b>", styles['Details_label']),
                Paragraph(response.get('submission_initiated_by', 'N/A'), styles['Details_value'])
            ],
            [
                Paragraph("<b>Designation:</b>", styles['Details_label']),
                Paragraph(response.get('initiator_designation', 'N/A'), styles['Details_value']),
                Paragraph("<b>Department:</b>", styles['Details_label']),
                Paragraph(response.get('initiator_department', 'N/A'), styles['Details_value'])
            ],
            [
                Paragraph("<b>Location:</b>", styles['Details_label']),
                Paragraph(_resolve_pdf_response_location(response), styles['Details_value']),
                Paragraph("<b>Status:</b>", styles['Details_label']),
                Paragraph('Completed' if response.get('is_completed', False) else 'Pending', styles['Details_value'])
            ]
        ]
        details_table = Table(details_data, colWidths=[80, 160, 80, 160])
        details_table.setStyle(TableStyle([
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('VALIGN', (0,0), (-1,-1), 'TOP')
        ]))
        story.append(details_table)
        story.append(Spacer(1, 25))

        # Form Details
        story.append(Paragraph("Form Details:", styles['Heading2_premium']))

        # Stages
        if response.get('stages'):
            for stage in response['stages']:
                story.append(Spacer(1, 10))
                stage_status = '✓' if stage.get('is_completed', False) else '○'
                stage_name = stage.get('name', f'Stage {stage.get("order", "")}')
                
                # A stage header with background color for visual separation
                stage_header_data = [[f"{stage_status} {stage_name}"]]
                stage_header_table = Table(stage_header_data, colWidths=[A4[0] - doc.leftMargin - doc.rightMargin])
                stage_header_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), accent_gray),
                    ("TEXTCOLOR", (0, 0), (-1, -1), primary_blue),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 14),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10)
                ]))
                story.append(stage_header_table)
                story.append(Spacer(1, 10))

                # Questions and answers
                if stage.get('questions'):
                    for question in stage['questions']:
                        question_text = question.get('question', '')
                        story.append(Paragraph(question_text, styles['Question_premium']))
                        
                        # Answer handling
                        if question.get('answers') and question['answers'].get('answer'):
                            answer = question['answers']['answer']
                            
                            # Handle different types of answers (text, image, signature)
                            image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
                            
                            if isinstance(answer, str) and any(ext in answer.lower() for ext in image_extensions) and ('http://' in answer or 'https://' in answer):
                                try:
                                    url = answer  # Define url here
                                    img_response = requests.get(url, timeout=10)
                                    img_response.raise_for_status()
                                    img_data = BytesIO(img_response.content)

                                    pil_img = PILImage.open(img_data).convert("RGB")
                                    max_width, max_height = 450, 300
                                    ratio = min(max_width / pil_img.width, max_height / pil_img.height)
                                    new_width = int(pil_img.width * ratio)
                                    new_height = int(pil_img.height * ratio)
                                    img_flowable = Image(img_data, width=new_width, height=new_height)
                                    img_flowable.hAlign = 'LEFT'
                                    story.append(img_flowable)
                                    story.append(Spacer(1, 8))
                                except Exception as e:
                                    story.append(Paragraph(f"<i>Image (failed to load):</i> {url}", styles['Answer_premium']))
                            elif isinstance(answer, str) and 'base64,' in answer:
                                try:
                                    import base64
                                    base64_part = answer.split(',')[1]
                                    img_data = BytesIO(base64.b64decode(base64_part))

                                    pil_img = PILImage.open(img_data).convert("RGB")
                                    max_width, max_height = 450, 300
                                    ratio = min(max_width / pil_img.width, max_height / pil_img.height)
                                    new_width = int(pil_img.width * ratio)
                                    new_height = int(pil_img.height * ratio)
                                    img_flowable = Image(img_data, width=new_width, height=new_height)
                                    img_flowable.hAlign = 'LEFT'
                                    story.append(img_flowable)
                                    story.append(Spacer(1, 8))
                                except Exception as e:
                                    story.append(Paragraph(f"<i>Signature (failed to decode):</i>", styles['Answer_premium']))
                            else:
                                # Regular text answer
                                story.append(Paragraph(str(answer), styles['Answer_premium']))
                            
                        # Handle sub-questions and logics
                        for sub_questions in [question.get('sub_questions', []), question.get('logics', [])]:
                            for sub_q_group in sub_questions:
                                sub_q_list = sub_q_group.get('logic_questions', []) if 'logic_questions' in sub_q_group else [sub_q_group]
                                for sub_q in sub_q_list:
                                    sub_q_text = sub_q.get('question', '')
                                    story.append(Paragraph(f"  • {sub_q_text}", styles['Question_premium']))
                                    if sub_q.get('answers') and sub_q['answers'].get('answer'):
                                        sub_answer = sub_q['answers']['answer']
                                        story.append(Paragraph(f"    - {sub_answer}", styles['Answer_premium']))
                        
                        # Add a separator between questions
                        story.append(Spacer(1, 5))
                        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceBefore=0, spaceAfter=8))
                        story.append(Spacer(1, 5))

    doc.build(story, onFirstPage=add_custom_header, onLaterPages=add_custom_header)
    buffer.seek(0)
    return buffer.getvalue()

def generate_audit_pdf(responses, form_info, form_id, organization=None, is_audit_form=True):
    """
    Generate PDF using the audit-style landscape structure.
    
    - Header section: Title, form info, submission details
    - Audit Summary section: Overall and group scores (audit only)
    - Stages section: Landscape orientation with table-based layout
    """
    buffer = BytesIO()

    HEADER_TOP_OFFSET = 16
    HEADER_BOX_HEIGHT = 26
    HEADER_CONTENT_GAP = 18
    HEADER_RESERVED_HEIGHT = HEADER_TOP_OFFSET + HEADER_BOX_HEIGHT + HEADER_CONTENT_GAP

    # --- Landscape orientation ---
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=40,
        leftMargin=40,
        topMargin=HEADER_RESERVED_HEIGHT,
        bottomMargin=50
    )
    
    # --- Setup Styles ---
    styles = getSampleStyleSheet()
    
    # Header styles (from enhanced PDF)
    title_style = ParagraphStyle(
        name='AuditTitle',
        fontSize=20,
        fontName='Helvetica-Bold',
        alignment=0,
        textColor=colors.HexColor('#3D3938'),
        spaceAfter=24
    )
    
    info_style = ParagraphStyle(
        name='AuditInfo',
        fontSize=9,
        fontName='Helvetica',
        textColor=colors.HexColor('#666666'),
        spaceAfter=4
    )
    
    details_label = ParagraphStyle(
        name='Details_label',
        fontSize=10,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#343a40')
    )
    
    details_value = ParagraphStyle(
        name='Details_value',
        fontSize=10,
        fontName='Helvetica',
        textColor=colors.HexColor('#6c757d')
    )
    
    # Table styles
    header_style = ParagraphStyle(
        name='AuditHeader',
        fontSize=9,
        fontName='Helvetica-Bold',
        alignment=0,
        textColor=colors.black,
        valign='TOP'
    )
    
    group_style = ParagraphStyle(
        name='GroupHeader',
        fontSize=10,
        fontName='Helvetica-Bold',
        alignment=0,
        textColor=colors.black,
        valign='MIDDLE'
    )
    
    question_style = ParagraphStyle(
        name='AuditQuestion',
        fontSize=9,
        fontName='Helvetica',
        alignment=0,
        textColor=colors.black,
        valign='TOP',
        wordWrap='CJK'
    )
    
    answer_style = ParagraphStyle(
        name='AuditAnswer',
        fontSize=9,
        fontName='Helvetica',
        alignment=0,
        textColor=colors.HexColor('#2E5090'),
        valign='TOP',
        wordWrap='CJK'
    )
    
    story = []

    def _compute_submission_task_metrics(response_data):
        """
        Compute task metrics for a single main-form submission.
        Returns overall counts and group-wise counts keyed by group/stage id.
        """
        metrics = {
            'overall': {'total': 0, 'completed': 0, 'overdue': 0, 'reopened': 0},
            'group': {}
        }

        if not organization or not response_data.get('id'):
            return metrics

        submission = FormSubmision.objects.filter(
            id=response_data.get('id'),
            organization=organization
        ).select_related('submission_initiated_by').first()

        if not submission:
            return metrics

        submission_anchor = submission.completed_on or submission.submission_initiated_on
        if not submission_anchor:
            return metrics

        # Match by form_submission_id first (preferred), then fall back to start_date
        base_tasks = Task.objects.filter(
            organization=organization,
            followup_task_form_id=form_id,
            form_submission_id=submission.id
        )

        if not base_tasks.exists() and submission_anchor:
            base_tasks = Task.objects.filter(
                organization=organization,
                followup_task_form_id=form_id,
                start_date=submission_anchor
            )

        # Fallback for edge-cases where timestamps differ slightly.
        if not base_tasks.exists():
            fallback_filter = Q(start_date__gte=submission_anchor - timezone.timedelta(minutes=2)) & Q(
                start_date__lte=submission_anchor + timezone.timedelta(minutes=2)
            )
            if submission.submission_initiated_by_id:
                fallback_filter &= Q(created_by_id=submission.submission_initiated_by_id)
            base_tasks = Task.objects.filter(
                organization=organization,
                followup_task_form_id=form_id
            ).filter(fallback_filter)

        now_ts = timezone.now()

        reopened_filter = (
            (Q(reopened_remarks__isnull=False) & ~Q(reopened_remarks='')) |
            Q(audit_logs__task_action__in=['Reopened', 'Followup_Reopened'])
        )

        metrics['overall'] = {
            'total': base_tasks.count(),
            'completed': base_tasks.filter(status='completed').count(),
            'overdue': base_tasks.filter(end_date__lt=now_ts).count(),
            'reopened': base_tasks.filter(reopened_filter).distinct().count(),
        }

        # Build stage/group question id map from response payload.
        stage_question_map = {}

        def _collect_question_ids(question_item):
            ids = set()
            qid = question_item.get('id')
            if qid:
                ids.add(qid)
            for sub_q in question_item.get('sub_questions', []):
                ids.update(_collect_question_ids(sub_q))
            for logic_item in question_item.get('logics', []):
                for logic_q in logic_item.get('logic_questions', []):
                    ids.update(_collect_question_ids(logic_q))
            return ids

        for stage in response_data.get('stages', []):
            stage_id = stage.get('id')
            if not stage_id or stage_id == 'audit_info':
                continue
            question_ids = set()
            for q in stage.get('questions', []):
                question_ids.update(_collect_question_ids(q))
            if question_ids:
                stage_question_map[str(stage_id)] = question_ids

        for stage_id, qids in stage_question_map.items():
            group_tasks = base_tasks.filter(follow_task_sub_question_id__in=qids)
            metrics['group'][stage_id] = {
                'total': group_tasks.count(),
                'completed': group_tasks.filter(status='completed').count(),
                'overdue': group_tasks.filter(end_date__lt=now_ts).count(),
                'reopened': group_tasks.filter(reopened_filter).distinct().count(),
            }

        return metrics
    
    # Helper to check if a question text should be excluded globally (for all groups)
    def should_exclude_question(question_text):
        """Check if a question should be completely excluded from PDF"""
        if not is_audit_form:
            return False
        question_text = str(question_text or '')
        # Exclude Observations and Photo questions from audit groups
        return 'Observations' in question_text or 'Photo' in question_text
    
    # --- Helper function to extract subquestions ---
    def get_all_subquestions(question, exclude_audit_info_fields=False):
        """Extract all unique subquestions from a question (includes sub_questions and logic questions)
        
        If exclude_audit_info_fields is True, exclude Observations and Photo subquestions
        """
        subquestions = []
        
        # Fields to exclude from audit_info questions (audit forms only)
        exclude_fields = ['Observations', 'Photo'] if (is_audit_form and exclude_audit_info_fields) else []
        
        for sub_q in question.get('sub_questions', []):
            sub_q_text = sub_q.get('question', '')
            # Exclude both audit_info specific fields and globally excluded questions
            if sub_q_text and sub_q_text not in exclude_fields and not should_exclude_question(sub_q_text):
                subquestions.append(sub_q)
        
        for logic in question.get('logics', []):
            for logic_q in logic.get('logic_questions', []):
                logic_q_text = logic_q.get('question', '')
                # Exclude both audit_info specific fields and globally excluded questions
                if logic_q_text and logic_q_text not in exclude_fields and not should_exclude_question(logic_q_text):
                    subquestions.append(logic_q)
        
        return subquestions
    
    # --- Helper: answer value -> flowable (Paragraph or Image for URLs/base64) ---
    # Smaller default image size and lower JPEG quality to reduce PDF size and email send time
    def _answer_to_flowable(answer_value, answer_style, max_img_width=100, max_img_height=60, max_chars=200, jpeg_quality=72):
        """Return a Paragraph or Image flowable for table cells. Renders image URLs and base64 as images."""
        if answer_value is None or (isinstance(answer_value, str) and not answer_value.strip()):
            return Paragraph('', answer_style)
        text = str(answer_value).strip()
        # Image URL (http/https with image extension or common image path)
        image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')
        if text.startswith(('http://', 'https://')) and any(text.lower().endswith(ext) for ext in image_extensions):
            try:
                resp = requests.get(text, timeout=10)
                resp.raise_for_status()
                if len(resp.content) > 10 * 1024 * 1024:
                    return Paragraph(text, answer_style)
                img_data = BytesIO(resp.content)
                pil_img = PILImage.open(img_data).convert("RGB")
                w, h = pil_img.size
                ratio = min(max_img_width / w, max_img_height / h, 1.0)
                nw, nh = int(w * ratio), int(h * ratio)
                buf = BytesIO()
                pil_img.save(buf, format="JPEG", quality=jpeg_quality)
                buf.seek(0)
                return Image(buf, width=nw, height=nh)
            except Exception:
                return Paragraph(text, answer_style)
        # Base64 image
        if 'base64,' in text:
            try:
                import base64 as b64
                part = text.split('base64,', 1)[-1].strip()
                raw = b64.b64decode(part)
                img_data = BytesIO(raw)
                pil_img = PILImage.open(img_data).convert("RGB")
                w, h = pil_img.size
                ratio = min(max_img_width / w, max_img_height / h, 1.0)
                nw, nh = int(w * ratio), int(h * ratio)
                buf = BytesIO()
                pil_img.save(buf, format="JPEG", quality=jpeg_quality)
                buf.seek(0)
                return Image(buf, width=nw, height=nh)
            except Exception:
                return Paragraph(text, answer_style)
        # Multiple URLs (e.g. pipe-separated): try first image
        if '|' in text:
            first = text.split('|')[0].strip()
            if first.startswith(('http://', 'https://')) and any(first.lower().endswith(ext) for ext in image_extensions):
                return _answer_to_flowable(first, answer_style, max_img_width, max_img_height, max_chars, jpeg_quality)
        return Paragraph(text, answer_style)

    # Max rows per Form Details table chunk to avoid ReportLab LayoutError (table too large for page)
    _MAX_AUDIT_TABLE_ROWS = 28

    # --- Helper function to build tables per group (avoiding large combined table issues) ---
    def build_audit_tables_per_group(groups, styles_dict, all_subquestions_across_groups=None):
        """Build separate tables for each group to avoid LayoutError with large tables."""
        if all_subquestions_across_groups is None:
            all_subquestions_across_groups = {}

        # Helper to safely truncate text for PDF cells
        def truncate_for_pdf(text, max_chars=140):
            if not text:
                return ''
            return str(text)

        # Use all subquestions across all groups for consistent column structure
        active_subquestions = list(all_subquestions_across_groups.keys())
        
        # Limit subquestions to prevent extremely narrow columns and keep PDF compact
        max_subquestions = 5
        if len(active_subquestions) > max_subquestions:
            logger.warning(f"Too many subquestions ({len(active_subquestions)}), limiting to {max_subquestions}")
            active_subquestions = active_subquestions[:max_subquestions]
        
        # Truncate subquestion texts for headers
        truncated_subquestions = [truncate_for_pdf(q, 55) for q in active_subquestions]
        header_row = ['Item', 'Response'] + truncated_subquestions
        num_cols = len(header_row)
        # Column header as flowables for continuation chunks (repeatRows)
        col_header_style = ParagraphStyle(name='ColHeader', fontSize=9, fontName='Helvetica-Bold')
        column_header_row = [Paragraph('Item', col_header_style), Paragraph('Response', col_header_style)] + [Paragraph(t, col_header_style) for t in truncated_subquestions]

        tables = []

        # Calculate column widths
        available_width = landscape(A4)[0] - 80  # Subtract margins (40+40)
        item_width = available_width * 0.25
        response_width = available_width * 0.25
        col_widths = [item_width, response_width]

        if len(truncated_subquestions) > 0:
            remaining_width = available_width * 0.5
            subq_width = remaining_width / len(truncated_subquestions)
            col_widths.extend([subq_width] * len(truncated_subquestions))

        for group_idx, group in enumerate(groups):
            is_audit_info_group = group.get('is_audit_info') or group.get('name') == 'Audit Info' or group.get('id') == 'audit_info'
            group_name = group.get('name') or group.get('title', 'Group')
            truncated_group_name = truncate_for_pdf(group_name, 140)

            table_data = []
            style_commands = []

            # Group header row
            group_cells = [Paragraph(truncated_group_name, styles_dict['group'])] + [''] * (num_cols - 1)
            table_data.append(group_cells)
            style_commands.extend([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCCCCC')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
            ])

            # Question rows for this group
            question_count = 0
            for question in group.get('questions', []):
                question_text = question.get('question', '')
                question_type = question.get('question_type', '')

                # Skip Observations and Photo questions from all groups globally
                if should_exclude_question(question_text):
                    continue

                # Special handling for table type questions
                if question_type == 'table':
                    # Get table rows from answer
                    answer_value = None
                    if question.get('answers') and question['answers'].get('answer'):
                        answer_value = question['answers']['answer']
                    
                    table_rows = []
                    if isinstance(answer_value, str):
                        try:
                            table_rows = json.loads(answer_value)
                        except:
                            table_rows = []
                    elif isinstance(answer_value, list):
                        table_rows = answer_value
                    
                    # Get subquestions for the table
                    subqs = get_all_subquestions(question, exclude_audit_info_fields=is_audit_info_group)
                    
                    # Add each table row as a separate entry
                    if table_rows and len(table_rows) > 0:
                        for row_index, row in enumerate(table_rows):
                            if isinstance(row, dict):
                                # Truncate question text with row number
                                truncated_question = truncate_for_pdf(f"{question_text} - Row {row_index + 1}", 140)
                                row_data = [
                                    Paragraph(truncated_question, styles_dict['question']),
                                    Paragraph('', styles_dict['answer'])
                                ]
                                
                                # Add subquestion answers - match with active_subquestions by text
                                subquestion_answers = {}
                                for subq in subqs:
                                    subq_id = subq.get('id') or subq.get('question_uuid')
                                    subq_text = subq.get('question', '')
                                    subq_answer = row.get(str(subq_id), '') if isinstance(row, dict) else row.get(subq_id, '')
                                    subquestion_answers[subq_text] = subq_answer
                                
                                # Add answers in the same order as active_subquestions
                                for subq_text in active_subquestions:
                                    subq_answer = subquestion_answers.get(subq_text, '')
                                    row_data.append(_answer_to_flowable(subq_answer, styles_dict['answer']))
                                
                                table_data.append(row_data)
                                question_count += 1
                    else:
                        # No table rows, add empty entry
                        truncated_question = truncate_for_pdf(question_text, 140)
                        row_data = [
                            Paragraph(truncated_question, styles_dict['question']),
                            Paragraph('No data', styles_dict['answer'])
                        ]
                        
                        for subq_text in active_subquestions:
                            row_data.append(Paragraph('', styles_dict['answer']))
                        
                        table_data.append(row_data)
                        question_count += 1
                else:
                    # Regular question handling
                    answer_value = None
                    if question.get('answers') and question['answers'].get('answer'):
                        answer_value = question['answers']['answer']

                    # Get subquestion answers
                    subquestion_answers = {}
                    subqs = get_all_subquestions(question, exclude_audit_info_fields=is_audit_info_group)
                    for subq in subqs:
                        subq_question_text = subq.get('question', '')
                        if subq.get('answers') and subq['answers'].get('answer'):
                            subquestion_answers[subq_question_text] = subq['answers']['answer']

                    # Truncate question text
                    truncated_question = truncate_for_pdf(question_text, 140)
                    row = [
                        Paragraph(truncated_question, styles_dict['question']),
                        _answer_to_flowable(answer_value, styles_dict['answer'])
                    ]

                    # Add subquestion answers
                    for subq_text in active_subquestions:
                        subq_answer = subquestion_answers.get(subq_text, '')
                        row.append(_answer_to_flowable(subq_answer, styles_dict['answer']))

                    table_data.append(row)
                    question_count += 1

            if len(table_data) > 1:  # Only create table if there are questions
                # Apply base styling (will be applied to each chunk)
                style_commands.extend([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
                ])

                # Chunk large tables to avoid ReportLab LayoutError (table too large for page)
                # Every chunk has column header row (Item, Response, subquestions) so headers are visible like before.
                max_rows = _MAX_AUDIT_TABLE_ROWS
                data_rows = table_data[1:]  # skip group header row
                chunk_start = 0
                while chunk_start < len(data_rows):
                    if chunk_start == 0:
                        # First chunk: group header + column header (Item, Response, subquestions) + data rows
                        n_data = min(max_rows - 2, len(data_rows))  # -2 for group row and column header row
                        chunk_data = [table_data[0], column_header_row] + data_rows[:n_data]
                        repeat_rows = 2  # repeat group + column header when table spans pages
                    else:
                        # Continuation chunk: column header + data rows
                        n_data = min(max_rows - 1, len(data_rows) - chunk_start)  # -1 for column header row
                        chunk_data = [column_header_row] + data_rows[chunk_start:chunk_start + n_data]
                        repeat_rows = 1  # repeat column header when table spans pages
                    nrows = len(chunk_data)
                    chunk_style = list(style_commands)
                    # Style header row(s): row 0 always; row 1 is column header on first chunk
                    chunk_style.extend([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCCCCC')),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                        ('TOPPADDING', (0, 0), (-1, 0), 6),
                    ])
                    if nrows >= 2:
                        # Column header row (row 1 on first chunk, row 0 on continuation) - same style
                        col_hdr_row = 1 if chunk_start == 0 else 0
                        chunk_style.extend([
                            ('BACKGROUND', (0, col_hdr_row), (-1, col_hdr_row), colors.HexColor('#E0E0E0')),
                            ('FONTNAME', (0, col_hdr_row), (-1, col_hdr_row), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, col_hdr_row), (-1, col_hdr_row), 9),
                            ('TOPPADDING', (0, col_hdr_row), (-1, col_hdr_row), 4),
                            ('BOTTOMPADDING', (0, col_hdr_row), (-1, col_hdr_row), 4),
                        ])
                    if nrows > (2 if chunk_start == 0 else 1):
                        data_start = 2 if chunk_start == 0 else 1
                        chunk_style.append(('ROWBACKGROUNDS', (0, data_start), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]))
                    table = Table(
                        chunk_data,
                        colWidths=col_widths,
                        repeatRows=repeat_rows,
                        splitByRow=1,
                    )
                    table.setStyle(TableStyle(chunk_style))
                    tables.append(table)
                    chunk_start += n_data
                    if chunk_start < len(data_rows):
                        tables.append(Spacer(1, 8))
                tables.append(Spacer(1, 15))  # Add space between groups

        return tables
    
    # --- Header on each page: generated datetime and form title (smaller text) ---
    _form_title_header = form_info.get('title', f'Form {form_id}')
    _generated_dt_header = datetime.now().strftime('%d-%b-%Y %I:%M %p')
    _app_name_header = 'VIBRO'
    _organization_header = (
        str(getattr(organization, 'organization_name', '') or '').strip()
        if organization else ''
    ) or 'N/A'

    def add_audit_page_number_only(canvas, doc):
        canvas.saveState()
        page_w = doc.pagesize[0]
        canvas.setFillColor(colors.HexColor('#343a40'))
        canvas.setFont('Helvetica', 9)
        canvas.drawRightString(page_w - doc.rightMargin, 18, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    def add_audit_page_header(canvas, doc):
        canvas.saveState()
        page_w = doc.pagesize[0]
        page_h = doc.pagesize[1]
        y_top = page_h - HEADER_TOP_OFFSET
        # Green boxed header area for app/org/datetime

        box_left = doc.leftMargin
        box_right = page_w - doc.rightMargin
        box_height = HEADER_BOX_HEIGHT
        box_bottom = y_top - box_height
        canvas.setFillColor(colors.HexColor('#d4edda'))
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(0.6)
        canvas.rect(box_left, box_bottom, box_right - box_left, box_height, stroke=1, fill=1)

        # Vertically center the header text inside the green box.
        text_y = box_bottom + (box_height / 2) - 4
        canvas.setFillColor(colors.HexColor('#343a40'))
        canvas.setFont('Helvetica-Bold', 11)
        canvas.drawString(doc.leftMargin + 8, text_y, _app_name_header)
        canvas.drawCentredString(page_w / 2, text_y, _organization_header)
        canvas.setFont('Helvetica', 9)
        canvas.drawRightString(page_w - doc.rightMargin - 8, text_y, _generated_dt_header)
        canvas.setStrokeColor(colors.HexColor('#6c757d'))
        canvas.setLineWidth(0.5)
        canvas.line(box_left, box_bottom - 4, box_right, box_bottom - 4)
        canvas.drawRightString(page_w - doc.rightMargin, 18, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    # --- Generate PDF content ---
    for idx, response in enumerate(responses):
        if idx > 0:
            story.append(PageBreak())
        
        # --- Form Title ---
        story.append(Paragraph(form_info.get('title', f'Form {form_id}'), title_style))
        base_subtitle = str(form_info.get('prefix') or form_id)
        subtitle_response_id = response.get('id')
        form_subtitle = f"{base_subtitle} -{subtitle_response_id}" if subtitle_response_id is not None else base_subtitle
        story.append(Paragraph(str(form_subtitle), ParagraphStyle(name='FormIdSubtitle', fontSize=10, fontName='Helvetica', textColor=colors.HexColor('#1a1a1a'), spaceAfter=10)))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1a1a1a'), spaceBefore=0, spaceAfter=12))
        
        # --- Resolve status and score for green banner (from audit history when available) ---
        banner_status = 'Completed' if response.get('is_completed', False) else 'Pending'
        banner_score = 'N/A'
        if is_audit_form and organization:
            try:
                _audit_first = AuditFormSubmissionHistory.objects.filter(
                    form_submission_id=response.get('id'),
                    organization=organization
                ).first()
                if _audit_first:
                    banner_status = _audit_first.form_overall_status or banner_status
                    banner_score = f"{_audit_first.form_overall_score:.1f}%" if _audit_first.form_overall_score is not None else 'N/A'
            except Exception:
                pass
        
        # --- Green metrics banner (3 equal columns: label above value) ---
        banner_label_style = ParagraphStyle(name='BannerLabel', fontSize=9, fontName='Helvetica-Bold', textColor=colors.HexColor('#2d2d2d'), spaceAfter=0)
        banner_value_style = ParagraphStyle(name='BannerValue', fontSize=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#1a1a1a'), spaceAfter=0)
        available_width_banner = landscape(A4)[0] - 80
        col_w = available_width_banner / 3
        banner_data = [
            [
                Paragraph("Location:", banner_label_style),
                Paragraph("Status:", banner_label_style),
                Paragraph("Score:", banner_label_style)
            ],
            [
                Paragraph(_resolve_pdf_response_location(response), banner_value_style),
                Paragraph(banner_status, banner_value_style),
                Paragraph(banner_score, banner_value_style)
            ]
        ]
        banner_table = Table(banner_data, colWidths=[col_w, col_w, col_w])
        banner_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#d4edda')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            # Vertical separators between columns
            ('LINEAFTER', (0, 0), (0, -1), 0.5, colors.HexColor('#6c757d')),
            ('LINEAFTER', (1, 0), (1, -1), 0.5, colors.HexColor('#6c757d')),
        ]))
        story.append(banner_table)
        story.append(Spacer(1, 16))
        
        # --- Two-column audit details (label : value pairs) ---
        detail_label_style = ParagraphStyle(name='DetailLabel', fontSize=9, fontName='Helvetica', textColor=colors.HexColor('#2d2d2d'))
        detail_value_style = ParagraphStyle(name='DetailValue', fontSize=9, fontName='Helvetica-Bold', textColor=colors.HexColor('#1a1a1a'))
        submission_date_raw = response.get('submission_initiated_on')
        if submission_date_raw:
            try:
                submission_dt = datetime.fromisoformat(str(submission_date_raw).replace('Z', '+00:00'))
                submission_date = submission_dt.astimezone(ZoneInfo('Asia/Kolkata')).strftime('%d-%b-%Y %I:%M %p')
            except Exception:
                submission_date = str(submission_date_raw)
        else:
            submission_date = 'N/A'
        # created_at_str = datetime.fromisoformat(form_info.get('created_at', '').replace('Z', '+00:00')).strftime('%d-%b-%Y') if form_info.get('created_at') else 'N/A'
        raw_prefix = str(form_info.get('prefix') or 'N/A')
        response_id_value = response.get('id')
        if raw_prefix != 'N/A' and response_id_value is not None:
            display_prefix = f"{raw_prefix} -{response_id_value}"
        else:
            display_prefix = raw_prefix
        
        left_col = [
            [Paragraph("Initiated By :", detail_label_style), Paragraph(response.get('submission_initiated_by', 'N/A'), detail_value_style)],
            [Paragraph("Submission Date :", detail_label_style), Paragraph(submission_date, detail_value_style)],
            # [Paragraph("Form Created On :", detail_label_style), Paragraph(created_at_str, detail_value_style)],
        ]
        right_col = [
            [Paragraph("Prefix :", detail_label_style), Paragraph(display_prefix, detail_value_style)],
            [Paragraph("Designation, Dept :", detail_label_style), Paragraph(f"{response.get('initiator_designation', 'N/A')}, {response.get('initiator_department', 'N/A')}", detail_value_style)],
            # [Paragraph("Response ID :", detail_label_style), Paragraph(str(response.get('id', 'N/A')), detail_value_style)],
        ]
        detail_col_width = (landscape(A4)[0] - 80) / 2
        left_table = Table(left_col, colWidths=[100, detail_col_width - 100])
        left_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (0, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        right_table = Table(right_col, colWidths=[110, detail_col_width - 110])
        right_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (0, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        details_two_col = Table([[left_table, right_table]], colWidths=[detail_col_width, detail_col_width])
        details_two_col.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (0, -1), 0),
            ('LEFTPADDING', (1, 0), (1, -1), 20),
        ]))
        story.append(details_two_col)
        story.append(Spacer(1, 25))
        
        # --- Audit Summary Section (only for audit forms) ---
        if is_audit_form and organization:
            try:
                task_metrics = _compute_submission_task_metrics(response)

                # Fetch audit submission history for this response
                audit_history = AuditFormSubmissionHistory.objects.filter(
                    form_submission_id=response.get('id'),
                    organization=organization
                ).select_related('group_id').order_by('group_id__order', 'group_id__id')
                
                if audit_history.exists():
                    # Get overall form scores from the first record (all records have the same overall scores)
                    first_history = audit_history.first()
                    overall_status = first_history.form_overall_status or 'N/A'
                    overall_score = f"{first_history.form_overall_score:.1f}%" if first_history.form_overall_score is not None else 'N/A'
                    overall_critical_fail = f"{first_history.form_critical_failed}" if first_history.form_critical_failed is not None else '0'
                    
                    # Calculate total questions for critical fail display (e.g., "53/58")
                    # We'll use form_critical_failed and estimate total, or use a simpler format
                    total_questions = first_history.form_critical_failed if first_history.form_critical_failed else 0
                    # For now, show just the critical failed count, or we can enhance this later
                    critical_fail_display = overall_critical_fail
                    
                    # Audit Summary Heading
                    story.append(Paragraph("Audit Summary", ParagraphStyle(
                        name='AuditSummaryHeading',
                        fontSize=14,
                        fontName='Helvetica-Bold',
                        textColor=colors.HexColor('#343a40'),
                        spaceAfter=10
                    )))
                    
                    # Build audit summary table
                    summary_table_data = []
                    
                    # Header row
                    header_row = [
                        Paragraph("<b>Audit Summary</b>", ParagraphStyle(name='SummaryHeader', fontSize=9, fontName='Helvetica-Bold', alignment=1)),
                        Paragraph("<b>Score (%)</b>", ParagraphStyle(name='SummaryHeader', fontSize=9, fontName='Helvetica-Bold', alignment=1)),
                        Paragraph("<b>Score Critical Fail</b>", ParagraphStyle(name='SummaryHeader', fontSize=9, fontName='Helvetica-Bold', alignment=1)),
                        Paragraph("<b>Task Completion (%)</b>", ParagraphStyle(name='SummaryHeader', fontSize=9, fontName='Helvetica-Bold', alignment=1)),
                        Paragraph("<b>Overdue (%)</b>", ParagraphStyle(name='SummaryHeader', fontSize=9, fontName='Helvetica-Bold', alignment=1)),
                        Paragraph("<b>Reopened (%)</b>", ParagraphStyle(name='SummaryHeader', fontSize=9, fontName='Helvetica-Bold', alignment=1))
                    ]
                    summary_table_data.append(header_row)
                    
                    # Overall summary row (with green background)
                    _ov = task_metrics['overall']
                    _ov_total = _ov['total'] or 0
                    _ov_completion = f"{round(_ov['completed'] / _ov_total * 100)}%" if _ov_total > 0 else '0%'
                    _ov_overdue = f"{round(_ov['overdue'] / _ov_total * 100)}%" if _ov_total > 0 else '0%'
                    _ov_reopened = f"{round(_ov['reopened'] / _ov_total * 100)}%" if _ov_total > 0 else '0%'
                    overall_row = [
                        Paragraph(f"Status: {overall_status}", ParagraphStyle(name='SummaryOverall', fontSize=9, fontName='Helvetica-Bold', alignment=0)),
                        Paragraph(overall_score, ParagraphStyle(name='SummaryOverall', fontSize=9, fontName='Helvetica', alignment=1)),
                        Paragraph(critical_fail_display, ParagraphStyle(name='SummaryOverall', fontSize=9, fontName='Helvetica', alignment=1)),
                        Paragraph(_ov_completion, ParagraphStyle(name='SummaryOverall', fontSize=9, fontName='Helvetica', alignment=1)),
                        Paragraph(_ov_overdue, ParagraphStyle(name='SummaryOverall', fontSize=9, fontName='Helvetica', alignment=1)),
                        Paragraph(_ov_reopened, ParagraphStyle(name='SummaryOverall', fontSize=9, fontName='Helvetica', alignment=1))
                    ]
                    summary_table_data.append(overall_row)
                    
                    # Group rows
                    for history in audit_history:
                        if history.group_id:
                            group_name = history.group_id.name or 'N/A'
                            group_score = f"{history.group_percentage}" if history.group_percentage else (f"{history.group_score:.1f}%" if history.group_score is not None else '0.0%')
                            group_critical_fail = f"{history.group_critical_failed}" if history.group_critical_failed is not None else '0'
                            group_metrics = task_metrics['group'].get(str(history.group_id.id), {'total': 0, 'completed': 0, 'overdue': 0, 'reopened': 0})
                            _g_total = group_metrics.get('total', 0) or 0
                            _g_completion = f"{round(group_metrics['completed'] / _g_total * 100)}%" if _g_total > 0 else '0%'
                            _g_overdue = f"{round(group_metrics['overdue'] / _g_total * 100)}%" if _g_total > 0 else '0%'
                            _g_reopened = f"{round(group_metrics['reopened'] / _g_total * 100)}%" if _g_total > 0 else '0%'
                            
                            group_row = [
                                Paragraph(group_name, ParagraphStyle(name='SummaryGroup', fontSize=9, fontName='Helvetica', alignment=0)),
                                Paragraph(f"{group_score}%", ParagraphStyle(name='SummaryGroup', fontSize=9, fontName='Helvetica', alignment=1)),
                                Paragraph(group_critical_fail, ParagraphStyle(name='SummaryGroup', fontSize=9, fontName='Helvetica', alignment=1)),
                                Paragraph(_g_completion, ParagraphStyle(name='SummaryGroup', fontSize=9, fontName='Helvetica', alignment=1)),
                                Paragraph(_g_overdue, ParagraphStyle(name='SummaryGroup', fontSize=9, fontName='Helvetica', alignment=1)),
                                Paragraph(_g_reopened, ParagraphStyle(name='SummaryGroup', fontSize=9, fontName='Helvetica', alignment=1))
                            ]
                            summary_table_data.append(group_row)
                    
                    # Create and style the summary table
                    available_width = landscape(A4)[0] - 80
                    col_widths = [
                        available_width * 0.35,  # Audit Summary column
                        available_width * 0.15,  # Score (%)
                        available_width * 0.15,  # Score Critical Fail
                        available_width * 0.12,  # Task Completion (%)
                        available_width * 0.12,  # Overdue (%)
                        available_width * 0.11   # Reopened (%)
                    ]
                    
                    summary_table = Table(summary_table_data, colWidths=col_widths)
                    summary_table.setStyle(TableStyle([
                        # Header row styling
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E0E0E0')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                        ('TOPPADDING', (0, 0), (-1, 0), 6),
                        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.black),
                        
                        # Overall row styling (green background)
                        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#90EE90')),  # Light green
                        ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),
                        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 1), (-1, 1), 9),
                        ('ALIGN', (0, 1), (0, 1), 'LEFT'),  # First column left-aligned
                        ('ALIGN', (1, 1), (-1, 1), 'CENTER'),  # Other columns centered
                        ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 1), (-1, 1), 6),
                        ('TOPPADDING', (0, 1), (-1, 1), 6),
                        
                        # Group rows styling
                        ('FONTNAME', (0, 2), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 2), (-1, -1), 9),
                        ('ALIGN', (0, 2), (0, -1), 'LEFT'),  # First column left-aligned
                        ('ALIGN', (1, 2), (-1, -1), 'CENTER'),  # Other columns centered
                        ('VALIGN', (0, 2), (-1, -1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 2), (-1, -1), 4),
                        ('TOPPADDING', (0, 2), (-1, -1), 4),
                        
                        # Grid lines
                        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('LEFTPADDING', (0, 0), (-1, -1), 4),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    
                    story.append(summary_table)
                    story.append(Spacer(1, 20))
            except Exception as e:
                # If audit summary fails, log but don't break PDF generation
                logger.warning(f"Failed to generate audit summary for response {response.get('id')}: {str(e)}")
        
        # --- Form Details Heading ---
        story.append(Paragraph("Form Details:", ParagraphStyle(
            name='FormDetailsHeading',
            fontSize=14,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#343a40'),
            spaceAfter=15
        )))
        
        # --- First pass: collect ALL subquestions across ALL groups ---
        all_subquestions_across_groups = {}
        if response.get('stages'):
            for group in response['stages']:
                is_audit_info_group = group.get('is_audit_info') or group.get('name') == 'Audit Info' or group.get('id') == 'audit_info'
                for question in group.get('questions', []):
                    # Skip Observations and Photo questions globally (from all groups)
                    if should_exclude_question(question.get('question', '')):
                        continue
                    
                    subqs = get_all_subquestions(question, exclude_audit_info_fields=is_audit_info_group)
                    for subq in subqs:
                        subq_text = subq.get('question', '')
                        if subq_text and subq_text not in all_subquestions_across_groups:
                            all_subquestions_across_groups[subq_text] = subq
        
        # --- Process audit groups ---
        if response.get('stages'):
            styles_dict = {
                'header': header_style,
                'group': group_style,
                'question': question_style,
                'answer': answer_style
            }
            
            # Build separate tables per group to avoid LayoutError
            audit_tables = build_audit_tables_per_group(
                response['stages'],
                styles_dict,
                all_subquestions_across_groups
            )
            for table in audit_tables:
                story.append(table)
    
    # --- Footer ---
    story.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        name='AuditFooter',
        fontSize=8,
        fontName='Helvetica',
        alignment=1,
        textColor=colors.HexColor('#999999'),
        spaceAfter=0
    )
    story.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d-%b-%Y %I:%M %p')} | Form ID: {form_id}",
        footer_style
    ))
    
    # --- Build PDF ---
    print(
        f"[PDF DEBUG] Using generate_audit_pdf | form_id={form_id} | responses={len(responses)} | "
        f"is_audit_form={is_audit_form}"
    )
    print("[PDF DEBUG] onFirstPage=add_audit_page_header | onLaterPages=add_audit_page_number_only")
    doc.build(
        story,
        onFirstPage=add_audit_page_header,
        onLaterPages=add_audit_page_number_only,
    )

    buffer.seek(0)
    return buffer.getvalue()


def _resolve_pdf_response_location(response):
    """
    Prefer the location chosen in the submitted form answers.
    If no location answer exists, show a hyphen so we do not reuse profile data.
    """

    def _walk(node):
        if isinstance(node, dict):
            question_type = str(node.get('question_type', '')).strip().lower()
            answers = node.get('answers') or {}
            if question_type == str(QuestionType.LOCATION).strip().lower():
                if isinstance(answers, dict):
                    answer_value = answers.get('answer')
                    if answer_value not in (None, ''):
                        return str(answer_value)

            for key in ('questions', 'sub_questions'):
                for child in node.get(key, []) or []:
                    resolved = _walk(child)
                    if resolved:
                        return resolved

            for logic in node.get('logics', []) or []:
                for child in logic.get('logic_questions', []) or []:
                    resolved = _walk(child)
                    if resolved:
                        return resolved

        elif isinstance(node, list):
            for item in node:
                resolved = _walk(item)
                if resolved:
                    return resolved

        return None

    for stage in response.get('stages', []) or []:
        resolved_location = _walk(stage)
        if resolved_location:
            return resolved_location

    return '-'


def generate_excel_data(responses, form_info, form_id, organization):
    """
    Generate Excel data from form responses in a single sheet format.
    Each row represents one complete response with all its details.
    For audit forms, inserts per-group summary columns before each group's questions (Score and placeholders for Task Completion, Overdue, Reopened).
    Responses are sorted by Response ID in descending order (latest first), matching current behavior.
    """
    # Sort responses by Response ID (keep current reverse sort behavior)
    responses = sorted(
        responses,
        key=lambda r: int(r.get('id', 0)) if str(r.get('id', '')).isdigit() else 0,
        reverse=True
    )

    # Helper function to replace option ID with text
    def replace_option_id_with_text(answers_data, options):
        if answers_data and 'answer' in answers_data and options:
            answer_value = answers_data['answer']
            raw_answer_value = answers_data.get('answer_id', answer_value)
            other_text = answers_data.get('other_text')

            def get_export_option_text(option_label):
                if option_label is None:
                    return option_label
                if other_text and str(option_label).strip().lower() == 'other':
                    return other_text
                return option_label

            def find_option_by_id(option_id):
                option_id = str(option_id).strip()
                for opt in options:
                    if opt and str(opt.get('id', '')).strip() == option_id:
                        return opt
                return None

            def replace_text_if_other(text_value):
                if other_text and str(text_value).strip().lower() == 'other':
                    return other_text
                return text_value

            if answer_value:
                answer_str = str(answer_value).strip()
                raw_answer_str = str(raw_answer_value).strip()

                # If serializer already converted the option to "Other", still prefer the custom text.
                if other_text and answer_str.lower() == 'other':
                    answers_data['answer'] = other_text
                elif ',' in answer_str:
                    # Handle multiple options like "1,2,3"
                    ids = [id.strip() for id in answer_str.split(',') if id.strip()]
                    replaced = []
                    for id_val in ids:
                        if id_val.isdigit():
                            opt = find_option_by_id(id_val)
                            if opt:
                                replaced.append(get_export_option_text(opt.get('option')))
                            else:
                                replaced.append(id_val)  # keep original if not found
                        else:
                            replaced.append(replace_text_if_other(id_val))
                    answers_data['answer'] = ', '.join(replaced)
                elif answer_str.isdigit():
                    # Single option
                    opt = find_option_by_id(answer_str)
                    if opt:
                        answers_data['answer'] = get_export_option_text(opt.get('option'))
                else:
                    # Handle malformed stored values such as "222986 [object Object]"
                    raw_option_ids = re.findall(r'\d+', raw_answer_str)
                    if raw_option_ids:
                        matched_options = []
                        for option_id in raw_option_ids:
                            opt = find_option_by_id(option_id)
                            if opt:
                                matched_options.append(get_export_option_text(opt.get('option')))
                        if matched_options:
                            answers_data['answer'] = ', '.join(matched_options)
                            return

                    answers_data['answer'] = replace_text_if_other(answer_value)

    # Create headers for the single sheet
    headers = [
        'Response ID',
        'Submission Date',
        'Initiated By',
        'Designation',
        'Department',
        'Location',
        'Current Owner',
        'Status',
        'Form Title',
        'Form Type',
        'Created By',
        'Form Created On',
        'Generated On'
    ]

    # Add audit-specific overall columns after Generated On for audit forms
    is_audit = form_info.get('form_type') == 'Audit'
    if is_audit:
        headers.append('Overall Status')
        headers.append('Total Score (%)')
        headers.append('Task Completion (%)')
        headers.append('Overdue Tasks (%)')
        headers.append('Reopened Tasks (%)')

    # When audit: build per-group column definitions + question columns in stable order
    # Otherwise: only question columns like current behavior
    column_definitions = []  # Each item: {type: 'group_score'|'group_task_completion'|'group_overdue_tasks'|'group_reopened_tasks'|'question', stage, text?, question_id?}

    # Optionally collect group scores up front for audit forms: map (submission_id, group_id) -> score/percentage
    group_scores = {}
    if is_audit:
        try:
            submission_ids = [r.get('id') for r in responses if r.get('id')]
            if submission_ids:
                histories = AuditFormSubmissionHistory.objects.filter(
                    form_submission_id__in=submission_ids,
                    organization=organization
                ).select_related('group_id')
                for h in histories:
                    # Prefer percentage if available, else use numeric score
                    key = (h.form_submission_id, getattr(h.group_id, 'id', None))
                    value = None
                    if h.group_percentage:
                        value = str(h.group_percentage)
                    elif h.group_score is not None:
                        value = f"{h.group_score}"
                    group_scores[key] = value
        except Exception:
            # Do not fail headers generation if score prefetch fails
            group_scores = {}

    # Build columns from the first response structure (all responses share same structure)
    for response in responses:
        if response.get('stages'):
            stages_list = response.get('stages', [])
            sorted_stages = sorted(stages_list, key=lambda stage: int(stage.get('order', 999)))

            for stage in sorted_stages:
                stage_name = stage.get('name', f"Stage {stage.get('order', '')}")
                if is_audit and not stage.get('is_audit_info') and stage_name != 'Audit Info':
                    # Add group columns once per group (avoid duplicates)
                    # Score
                    if not any(cd.get('type') == 'group_score' and cd.get('stage', {}).get('id') == stage.get('id') for cd in column_definitions):
                        column_definitions.append({'type': 'group_score', 'stage': stage, 'column_name': f"{stage_name} Score"})
                    # Placeholders for future metrics (keep " - " values)
                    if not any(cd.get('type') == 'group_task_completion' and cd.get('stage', {}).get('id') == stage.get('id') for cd in column_definitions):
                        column_definitions.append({'type': 'group_task_completion', 'stage': stage, 'column_name': f"{stage_name} Task Completion (%)"})
                    if not any(cd.get('type') == 'group_overdue_tasks' and cd.get('stage', {}).get('id') == stage.get('id') for cd in column_definitions):
                        column_definitions.append({'type': 'group_overdue_tasks', 'stage': stage, 'column_name': f"{stage_name} Overdue Tasks (%)"})
                    if not any(cd.get('type') == 'group_reopened_tasks' and cd.get('stage', {}).get('id') == stage.get('id') for cd in column_definitions):
                        column_definitions.append({'type': 'group_reopened_tasks', 'stage': stage, 'column_name': f"{stage_name} Reopened Tasks (%)"})

                # Questions for this stage
                if stage.get('questions'):
                    for question in stage['questions']:
                        q_text = question.get('question', '')
                        if q_text:
                            if not any(cd.get('type') == 'question' and cd.get('question_id') == question.get('id') for cd in column_definitions):
                                column_definitions.append({'type': 'question', 'question_id': question.get('id'), 'text': q_text, 'stage': stage})

                        # Logic questions
                        for logic in question.get('logics', []):
                            for logic_question in logic.get('logic_questions', []):
                                lq_text = logic_question.get('question', '')
                                if lq_text:
                                    if not any(cd.get('type') == 'question' and cd.get('question_id') == logic_question.get('id') for cd in column_definitions):
                                        column_definitions.append({'type': 'question', 'question_id': logic_question.get('id'), 'text': lq_text, 'stage': stage})
            break  # Only need structure from first response

    # Build headers from column definitions (in order), making duplicate question texts unique
    question_text_counts = {}  # Track duplicate question texts for unique headers
    for col_def in column_definitions:
        if col_def['type'] == 'group_score':
            headers.append(col_def['column_name'])
        elif col_def['type'] == 'group_task_completion':
            headers.append(col_def['column_name'])
        elif col_def['type'] == 'group_overdue_tasks':
            headers.append(col_def['column_name'])
        elif col_def['type'] == 'group_reopened_tasks':
            headers.append(col_def['column_name'])
        elif col_def['type'] == 'question':
            q_text = col_def['text']
            # Make duplicate question headers unique by appending a counter
            if q_text in question_text_counts:
                question_text_counts[q_text] += 1
                headers.append(f"{q_text} ({question_text_counts[q_text]})")
            else:
                question_text_counts[q_text] = 1
                headers.append(q_text)

    # Add metadata columns
    headers.extend([
        'Form ID',
        'Total Questions',
        'Completed Questions'
    ])

    # Create rows for each response
    rows = []

    for response in responses:
        # Base response data
        base_data = [
            response.get('id', 'N/A'),
            response.get('submission_initiated_on', 'N/A'),
            response.get('submission_initiated_by', 'N/A'),
            response.get('initiator_designation', 'N/A'),
            response.get('initiator_department', 'N/A'),
            response.get('initiator_location', 'N/A'),
            response.get('current_owner', 'N/A'),
            'Completed' if response.get('is_completed', False) else 'Pending',
            form_info.get('title', f'Form {form_id}'),
            form_info.get('form_type', 'N/A'),
            form_info.get('created_by', 'N/A'),
            form_info.get('created_at', 'N/A'),
            datetime.now().strftime('%d-%b-%Y %I:%M %p')
        ]

        # Insert audit overall columns after Generated On for audit forms
        if is_audit:
            audit_history = AuditFormSubmissionHistory.objects.filter(
                form_submission_id=response.get('id'),
                organization=organization
            ).first()
            base_data.append(audit_history.form_overall_status if audit_history else "N/A")
            base_data.append(str(audit_history.form_overall_score) if audit_history and audit_history.form_overall_score is not None else "N/A")
            # Calculate task metrics for this submission
            _resp_id = response.get('id')
            _t_total = 0
            _t_completed = 0
            _t_overdue = 0
            _t_reopened = 0
            try:
                from task.models import Task as _TM
                from django.utils import timezone as _tz2
                _sub = FormSubmision.objects.filter(id=_resp_id).first()
                if _sub:
                    _anchor = _sub.completed_on or _sub.submission_initiated_on
                    _tqs = _TM.objects.filter(followup_task_form_id=form_id, organization=organization)
                    _tqs_sub = _tqs.filter(form_submission_id=_resp_id)
                    if _tqs_sub.exists():
                        _bt = _tqs_sub
                    elif _anchor:
                        _bt = _tqs.filter(start_date=_anchor)
                    else:
                        _bt = _tqs_sub
                    _t_total = _bt.count()
                    _t_completed = _bt.filter(status='completed').count()
                    _t_overdue = _bt.filter(end_date__lt=_tz2.now(), status__in=['not_started', 'not_assigned', 'in_progress']).count()
                    _rf = (models.Q(reopened_remarks__isnull=False) & ~models.Q(reopened_remarks='')) | models.Q(audit_logs__task_action__in=['Reopened', 'Followup_Reopened'])
                    _t_reopened = _bt.filter(_rf).distinct().count()
            except Exception:
                pass
            base_data.append(f"{round(_t_completed / _t_total * 100)}%" if _t_total > 0 else "0%")  # Task Completion (%)
            base_data.append(f"{round(_t_overdue / _t_total * 100)}%" if _t_total > 0 else "0%")  # Overdue Tasks (%)
            base_data.append(f"{round(_t_reopened / _t_total * 100)}%" if _t_total > 0 else "0%")  # Reopened Tasks (%)

        # Create a dictionary of question answers for this response
        question_answers = {}

        if response.get('stages'):
            for stage in response['stages']:
                if stage.get('questions'):
                    for question in stage['questions']:
                        question_text = question.get('question', '')

                        # Handle main question answer
                        if question.get('answers') and question['answers'].get('answer'):
                            # Replace option IDs with text before processing
                            replace_option_id_with_text(question['answers'], question.get('options'))

                            answer = question['answers']['answer']
                            if isinstance(answer, str):
                                image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
                                potential_urls = answer.split("|") if "|" in answer else [answer]

                                text_parts = []
                                image_urls = []

                                for url in potential_urls:
                                    url = url.strip()
                                    if url and any(url.lower().endswith(ext) for ext in image_extensions):
                                        image_urls.append(url)
                                    elif url and 'http' in url:
                                        text_parts.append(url)
                                    elif url:
                                        text_parts.append(url)

                                if image_urls:
                                    answer_text = '; '.join(image_urls)
                                elif text_parts:
                                    answer_text = ' '.join(text_parts)
                                else:
                                    answer_text = answer[:500] + '...' if len(answer) > 500 else answer
                            else:
                                # Handle list/array answers (like checkboxes)
                                if isinstance(answer, list):
                                    answer_text = ', '.join(str(item) for item in answer)
                                    answer_text = answer_text[:500] + '...' if len(answer_text) > 500 else answer_text
                                else:
                                    answer_text = str(answer)[:500] + '...' if len(str(answer)) > 500 else str(answer)

                            q_id = question.get('id')
                            if q_id:
                                question_answers[q_id] = answer_text

                        # Handle sub-questions separately
                        for sub_question in question.get('sub_questions', []):
                            sub_question_text = sub_question.get('question', '')
                            if sub_question.get('answers') and sub_question['answers'].get('answer'):
                                # Replace option IDs with text for sub-questions
                                replace_option_id_with_text(sub_question['answers'], sub_question.get('options'))

                                sub_answer = sub_question['answers']['answer']
                                if isinstance(sub_answer, str):
                                    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
                                    potential_urls = sub_answer.split("|") if "|" in sub_answer else [sub_answer]

                                    text_parts = []
                                    image_urls = []

                                    for url in potential_urls:
                                        url = url.strip()
                                        if url and any(url.lower().endswith(ext) for ext in image_extensions):
                                            image_urls.append(url)
                                        elif url and 'http' in url:
                                            text_parts.append(url)
                                        elif url:
                                            text_parts.append(url)

                                    if image_urls:
                                        sub_answer_text = '; '.join(image_urls)
                                    elif text_parts:
                                        sub_answer_text = ' '.join(text_parts)
                                    else:
                                        sub_answer_text = sub_answer[:500] + '...' if len(sub_answer) > 500 else sub_answer
                                else:
                                    sub_answer_text = str(sub_answer)[:500] + '...' if len(str(sub_answer)) > 500 else str(sub_answer)

                                sq_id = sub_question.get('id')
                                if sq_id:
                                    question_answers[sq_id] = sub_answer_text

                        # Handle logic questions separately
                        for logic in question.get('logics', []):
                            for logic_question in logic.get('logic_questions', []):
                                logic_question_text = logic_question.get('question', '')
                                if logic_question.get('answers') and logic_question['answers'].get('answer'):
                                    # Replace option IDs with text for logic questions
                                    replace_option_id_with_text(logic_question['answers'], logic_question.get('options'))

                                    logic_answer = logic_question['answers']['answer']
                                    if isinstance(logic_answer, str):
                                        image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
                                        potential_urls = logic_answer.split("|") if "|" in logic_answer else [logic_answer]

                                        text_parts = []
                                        image_urls = []

                                        for url in potential_urls:
                                            url = url.strip()
                                            if url and any(url.lower().endswith(ext) for ext in image_extensions):
                                                image_urls.append(url)
                                            elif url and 'http' in url:
                                                text_parts.append(url)
                                            elif url:
                                                text_parts.append(url)

                                        if image_urls:
                                            logic_answer_text = '; '.join(image_urls)
                                        elif text_parts:
                                            logic_answer_text = ' '.join(text_parts)
                                        else:
                                            logic_answer_text = logic_answer[:500] + '...' if len(logic_answer) > 500 else logic_answer
                                    else:
                                        logic_answer_text = str(logic_answer)[:500] + '...' if len(str(logic_answer)) > 500 else str(logic_answer)

                                    lq_id = logic_question.get('id')
                                    if lq_id:
                                        question_answers[lq_id] = logic_answer_text

        # Add data per column definition
        for col_def in column_definitions:
            if col_def['type'] == 'group_score':
                stage = col_def['stage']
                group_id = stage.get('id')
                key = (response.get('id'), group_id)
                score_val = group_scores.get(key)
                # Append percentage if not already contains it and value exists
                if score_val is None or score_val == '':
                    base_data.append('')
                else:
                    # Avoid duplicating % if present
                    s = str(score_val)
                    base_data.append(s if s.endswith('%') else f"{s}%")
            elif col_def['type'] in ['group_task_completion', 'group_overdue_tasks', 'group_reopened_tasks']:
                # Calculate per-group task metrics
                stage = col_def['stage']
                group_id = stage.get('id')
                _g_val = "0%"
                try:
                    from task.models import Task as _TM2
                    from django.utils import timezone as _tz3
                    # Collect question IDs for this stage
                    _g_qids = set()
                    for _q in stage.get('questions', []):
                        _g_qids.add(_q.get('id'))
                        for _sq in _q.get('sub_questions', []):
                            _g_qids.add(_sq.get('id'))
                        for _lg in _q.get('logics', []):
                            for _lq in _lg.get('logic_questions', []):
                                _g_qids.add(_lq.get('id'))
                    _g_qids.discard(None)
                    if _g_qids:
                        _resp_id2 = response.get('id')
                        _g_tqs = _TM2.objects.filter(
                            followup_task_form_id=form_id,
                            organization=organization,
                            follow_task_sub_question_id__in=_g_qids
                        )
                        _g_tqs_sub = _g_tqs.filter(form_submission_id=_resp_id2)
                        if _g_tqs_sub.exists():
                            _g_bt = _g_tqs_sub
                        else:
                            _g_bt = _g_tqs
                        _g_total2 = _g_bt.count()
                        if _g_total2 > 0:
                            if col_def['type'] == 'group_task_completion':
                                _g_comp = _g_bt.filter(status='completed').count()
                                _g_val = f"{round(_g_comp / _g_total2 * 100)}%"
                            elif col_def['type'] == 'group_overdue_tasks':
                                _g_od = _g_bt.filter(end_date__lt=_tz3.now(), status__in=['not_started', 'not_assigned', 'in_progress']).count()
                                _g_val = f"{round(_g_od / _g_total2 * 100)}%"
                            elif col_def['type'] == 'group_reopened_tasks':
                                _g_rf2 = (models.Q(reopened_remarks__isnull=False) & ~models.Q(reopened_remarks='')) | models.Q(audit_logs__task_action__in=['Reopened', 'Followup_Reopened'])
                                _g_ro = _g_bt.filter(_g_rf2).distinct().count()
                                _g_val = f"{round(_g_ro / _g_total2 * 100)}%"
                except Exception:
                    pass
                base_data.append(_g_val)
            elif col_def['type'] == 'question':
                q_id = col_def.get('question_id')
                base_data.append(question_answers.get(q_id, ''))

        # Add metadata
        total_questions = len([c for c in column_definitions if c['type'] == 'question'])
        completed_questions = len([c for c in column_definitions if c['type'] == 'question' and question_answers.get(c.get('question_id'), '')])
        base_data.extend([
            str(form_id),
            str(total_questions),
            str(completed_questions)
        ])

        rows.append(base_data)

    return {
        'Form Responses': [headers] + rows
    }


class FormResponseCSVView(APIView):
    """
    Endpoint to generate Excel/CSV and download form responses.
    Supports both single and multiple submission IDs via query parameters.
    All Excel generation and image processing happens asynchronously in background.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, form_id):
        try:
            # === STEP 1: VALIDATE INPUT (Main Thread) ===
            
            # Validate email parameter
            email = request.query_params.get('email')
            if not email:
                return Response(
                    {"error": "Email parameter is required to send the report."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate submission IDs parameter
            submission_ids_param = request.query_params.get('submission_ids', '')
            if submission_ids_param:
                try:
                    submission_ids = [int(sid.strip()) for sid in submission_ids_param.split(',') if sid.strip()]
                except ValueError:
                    return Response({"error": "Invalid submission_ids format. Use comma-separated integers."}, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Fallback to single submission_id from query parameters
                submission_id = request.query_params.get('submission_id')
                if submission_id:
                    try:
                        submission_ids = [int(submission_id)]
                    except ValueError:
                        return Response({"error": "Invalid submission_id format. Must be an integer."}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"error": "No submission IDs provided. Use submission_ids (comma-separated) or submission_id parameter."}, status=status.HTTP_400_BAD_REQUEST)

            # Check form exists (lightweight check)
            if not Form.objects.filter(id=form_id, organization=request.user.organization).exists():
                return Response({"error": "Form not found."}, status=status.HTTP_404_NOT_FOUND)

            # === STEP 2: RETURN IMMEDIATELY & START BACKGROUND TASK ===
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            filename = f"{timestamp}_form_{form_id}_responses.xlsx"
            user_org_id = request.user.organization.id

            # Create tracking id and queue job status in cache (24h TTL)
            tracking_id = str(uuid.uuid4())
            cache_key = f"excel_job:{tracking_id}"
            short_download_url = _build_public_download_url(
                request,
                reverse("report-download-redirect", kwargs={"tracking_id": tracking_id}),
            )
            cache.set(cache_key, {
                "status": "QUEUED",
                "message": "Report generation queued",
                "form_id": form_id,
                "submission_ids": submission_ids,
                "email": email,
                "filename": filename,
                "bucket_name": None,
                "s3_key": None,
                "expires_in_seconds": None,
                "created_at": datetime.now().isoformat(),
                "started_at": None,
                "completed_at": None,
                "error": None,
            }, timeout=86400)

            def background_excel_generation(form_id, submission_ids, email, filename, user_org_id, tracking_id, short_download_url):
                """
                Generate entire Excel file and send email asynchronously.
                All heavy lifting happens in this function.
                """
                import sys
                import traceback as tb_module
                
                print(f"[Background Excel Task] ========== BACKGROUND THREAD STARTED ==========")
                sys.stdout.flush()
                
                try:
                    from datetime import datetime as dt
                    print(f"[Background Excel Task] START: form_id={form_id}, submission_ids={submission_ids}, email={email} | Time: {dt.now()}")
                    # Update status to RUNNING
                    try:
                        cache_key = f"excel_job:{tracking_id}"
                        job = cache.get(cache_key) or {}
                        job.update({
                            "status": "RUNNING",
                            "message": "Report generation started",
                            "started_at": dt.now().isoformat(),
                        })
                        cache.set(cache_key, job, timeout=86400)
                    except Exception:
                        pass
                    # --- Fetch form ---
                    print(f"[Background Excel Task] Fetching form {form_id}... | Time: {dt.now()}")
                    try:
                        form = Form.objects.get(id=form_id, organization_id=user_org_id)
                        print(f"[Background Excel Task] Form object fetched successfully | Time: {dt.now()}")
                    except Exception as form_fetch_err:
                        print(f"[Background Excel Task] ERROR fetching form: {str(form_fetch_err)} | Time: {dt.now()}")
                        logger.error(f"[Background Excel Task] ERROR fetching form: {str(form_fetch_err)}")
                        raise
                    
                    print(f"[Background Excel Task] Serializing form... | Time: {dt.now()}")
                    try:
                        start_ser = dt.now()
                        # Optimized prefetch to avoid N+1 during serialization
                        base = (
                            Form.objects
                            .filter(is_deleted=False, pk=form.id)
                            .select_related('folder', 'form_admin', 'deletedBy', 'archivedBy')
                        )

                        stage_access_qs = StageAccess.objects.select_related('allow_user', 'allow_group', 'form', 'stage')

                        if form.form_type == FormType.AUDIT:
                            qs = base.prefetch_related(
                                'assignee__user', 'assignee__group', 'assignee__leader',
                                'audit_info__questions',
                                'audit_info__questions__options',
                                'audit_info__questions__child_questions',
                                'audit_info__questions__child_questions__options',
                                'audit_info__questions__logic_parent_question__logic_questions__options',
                                'audit_info__questions__logic_parent_question__follow_ups',
                                'audit_info__questions__logic_parent_question__follow_ups__task_close_questions__options',
                                'audit_group__questions',
                                'audit_group__questions__options',
                                'audit_group__questions__child_questions',
                                'audit_group__questions__child_questions__options',
                                'audit_group__questions__logic_parent_question__logic_questions__options',
                                'audit_group__questions__logic_parent_question__follow_ups',
                                'audit_group__questions__logic_parent_question__follow_ups__task_close_questions__options',
                            )
                        else:
                            question_qs = (
                                Question.objects
                                .select_related('form', 'stage', 'parent_question')
                                .prefetch_related(
                                    'options',
                                    'child_questions',
                                    'child_questions__options',
                                    'child_questions__child_questions',
                                    'child_questions__child_questions__options',
                                    'child_questions__logic_parent_question__logic_questions__options',
                                    'child_questions__logic_parent_question__follow_ups',
                                    'child_questions__logic_parent_question__follow_ups__task_close_questions__options',
                                    'logic_parent_question__logic_questions__options',
                                    'logic_parent_question__follow_ups',
                                    'logic_parent_question__follow_ups__task_close_questions__options',
                                )
                            )
                            stage_qs = (
                                Stage.objects
                                .select_related('form')
                                .prefetch_related(
                                    models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                                    models.Prefetch('questions', queryset=question_qs),
                                )
                            )
                            qs = base.prefetch_related(
                                'assignee__user', 'assignee__group', 'assignee__leader',
                                models.Prefetch('stages', queryset=stage_qs),
                            )

                        optimized_instance = qs.get(pk=form.id)
                        formSchema = FormSerializer(optimized_instance, many=False).data
                        end_ser = dt.now()
                        print(f"[Background Excel Task] Form serialized successfully | Duration: {(end_ser - start_ser).total_seconds():.2f}s | Time: {dt.now()}")
                    except Exception as serializer_err:
                        print(f"[Background Excel Task] ERROR serializing form: {str(serializer_err)} | Time: {dt.now()}")
                        logger.error(f"[Background Excel Task] ERROR serializing form: {str(serializer_err)}")
                        raise
                    
                    print(f"[Background Excel Task] Form type: {form.get_form_type_display()} | Time: {dt.now()}")

                    # --- Fetch submissions with optimized queries ---
                    print(f"[Background Excel Task] Fetching submissions {submission_ids}... | Time: {dt.now()}")
                    sys.stdout.flush()
                    try:
                        submissions = FormSubmision.objects.filter(
                            id__in=submission_ids,
                            form_id=form_id,
                            organization_id=user_org_id
                        ).select_related(
                            'submission_initiated_by',
                            'completed_by'
                        ).prefetch_related(
                            models.Prefetch(
                                'answers',
                                queryset=Answer.objects.select_related('question').prefetch_related('question__options'),
                                to_attr='prefetched_answers'
                            ),
                            'stage_submissions_history',
                            'group_submissions_history'
                        )
                        print(f"[Background Excel Task] Submissions query executed | Time: {dt.now()}")
                        sys.stdout.flush()
                    except Exception as submission_fetch_err:
                        print(f"[Background Excel Task] ERROR fetching submissions: {str(submission_fetch_err)} | Time: {dt.now()}")
                        logger.error(f"[Background Excel Task] ERROR fetching submissions: {str(submission_fetch_err)}")
                        sys.stdout.flush()
                        raise

                    if not submissions.exists():
                        logger.warning(f"[Background Excel Task] No submissions found for form {form_id}")
                        print(f"[Background Excel Task] No submissions found for form {form_id}")
                        sys.stdout.flush()
                        # Update status to FAILED due to no submissions
                        try:
                            cache_key = f"excel_job:{tracking_id}"
                            job = cache.get(cache_key) or {}
                            job.update({
                                "status": "FAILED",
                                "message": "No submissions found for the provided IDs",
                                "completed_at": dt.now().isoformat(),
                                "error": "NO_SUBMISSIONS"
                            })
                            cache.set(cache_key, job, timeout=86400)
                        except Exception:
                            pass
                        return

                    print(f"[Background Excel Task] Found {submissions.count()} submission(s). Starting to populate response data... | Time: {dt.now()}")
                    sys.stdout.flush()
                    # --- Prepare multiple response data ---
                    responses_data = []

                    for submission_instance in submissions:
                        submission_id = submission_instance.id
                        print(f"[Background Excel Task] Processing submission {submission_id}...")

                        # --- Prepare response data like frontend ---
                        print(f"[Background Excel Task] Deep copying form schema for submission {submission_id}...")
                        response_data = copy.deepcopy(formSchema)
                        # Override with submission-specific data
                        response_data.update({
                            'id': submission_instance.id,
                            'submission_initiated_on': submission_instance.submission_initiated_on.isoformat() if submission_instance.submission_initiated_on else None,
                            'submission_initiated_by': getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                            'initiator_designation': getattr(submission_instance.submission_initiated_by, 'designation', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                            'initiator_department': getattr(submission_instance.submission_initiated_by, 'department', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                            'initiator_location': getattr(submission_instance.submission_initiated_by, 'location', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                            'current_owner': getattr(submission_instance.completed_by, 'first_name', '') + ' ' + getattr(submission_instance.completed_by, 'last_name', '') if submission_instance.completed_by else getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                            'is_completed': submission_instance.is_completed,
                        })

                        # --- Populate answers into stages using prefetched data ---
                        # Create a lookup dict for answers by question_id
                        print(f"[Background Excel Task] Creating answer lookup for submission {submission_id}. Total answers: {len(submission_instance.prefetched_answers)}")
                        answer_lookup = {answer.question_id: answer for answer in submission_instance.prefetched_answers}
                        print(f"[Background Excel Task] Answer lookup created. Processing form type: {response_data.get('form_type')}")

                        if response_data.get("form_type") == FormType.AUDIT:
                            stages = []

                            # Handle audit_info as the first stage
                            audit_info = response_data.get("audit_info")
                            if audit_info:
                                audit_info_stage = {
                                    'id': audit_info.get('id', 'audit_info'),
                                    'name': audit_info.get('name', 'Audit Info'),
                                    'is_audit_info': True,
                                    'order': 0,
                                    'questions': audit_info.get("questions", [])
                                }

                                # Process answers for audit_info questions
                                for question in audit_info_stage['questions']:
                                    answer = answer_lookup.get(question['id'])
                                    question['answers'] = AnswerSerializer(answer).data if answer else {}

                                    # Handle sub-questions for audit_info
                                    for subQuestion in question.get('sub_questions', []):
                                        sub_answer = answer_lookup.get(subQuestion['id'])
                                        subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}

                                    # Handle logic questions for audit_info
                                    for logics in question.get("logics", []):
                                        for logicQuestion in logics.get("logic_questions", []):
                                            logic_answer = answer_lookup.get(logicQuestion['id'])
                                            logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}

                                stages.append(audit_info_stage)

                            # For audit forms, create a deep copy of audit_group to avoid modifying shared formSchema
                            audit_groups_copy = copy.deepcopy(response_data.get("audit_group", []))
                            # For audit forms, we'll keep the original audit logic
                            for audit_group in audit_groups_copy:
                                for audit_group_question in audit_group.get("questions", []):
                                    answer = answer_lookup.get(audit_group_question['id'])
                                    audit_group_question['answers'] = AnswerSerializer(answer).data if answer else {}

                                    for logics in audit_group_question.get("logics", []):
                                        for logicQuestion in logics.get("logic_questions", []):
                                            logic_answer = answer_lookup.get(logicQuestion['id'])
                                            logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}

                                    for subQuestion in audit_group_question.get('sub_questions', []):
                                        sub_answer = answer_lookup.get(subQuestion['id'])
                                        subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}

                                stages.append(audit_group)

                            response_data['stages'] = stages
                        else:
                            # For standard/location forms, populate stages with questions and answers
                            for stage in formSchema.get('stages', []):
                                # Get stage completion history
                                stage_histories = StageSubmissionHistory.objects.filter(
                                    stage__id=stage['id'],
                                    form_submission__id=submission_id,
                                    organization_id=user_org_id
                                )

                                historyData = {}
                                if stage_histories.exists():
                                    stageHistory = stage_histories.first()
                                    historyData = StageSubmissionHistorySerializer(stageHistory, many=False).data

                                stage_data = {
                                    'id': stage['id'],
                                    'name': stage.get('name', f'Stage {stage.get("order", "")}'),
                                    'order': stage.get('order', 0),
                                    'is_completed': stage_histories.exists(),
                                    'completed_by': historyData.get('completed_by', None),
                                    'completed_on': historyData.get('completed_on', None),
                                    'questions': []
                                }

                                # Populate questions with answers - create copies to avoid modifying shared formSchema
                                for question in stage.get('questions', []):
                                    question_data = copy.deepcopy(question)  # Deep copy to avoid modifying original

                                    # Get answer for this question using prefetched data
                                    answer = answer_lookup.get(question['id'])
                                    question_data['answers'] = AnswerSerializer(answer).data if answer else {}

                                    # Handle sub-questions
                                    for subQuestion in question_data.get('sub_questions', []):
                                        sub_answer = answer_lookup.get(subQuestion['id'])
                                        subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}

                                    # Handle logic questions
                                    for logic in question_data.get("logics", []):
                                        for logicQuestion in logic.get("logic_questions", []):
                                            logic_answer = answer_lookup.get(logicQuestion['id'])
                                            logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}

                                    stage_data['questions'].append(question_data)

                                response_data['stages'].append(stage_data)

                        print(f"[Background Excel Task] Submission {submission_id} response data prepared successfully.")
                        responses_data.append(response_data)

                    print(f"[Background Excel Task] All {len(responses_data)} submission(s) processed. Preparing form info...")
                    # --- Prepare form info ---
                    form_info = {
                        'title': form.title or f'Form {form_id}',
                        'form_type': form.get_form_type_display() or 'standard',
                        'created_by': getattr(form.form_admin, 'first_name', '') + ' ' + getattr(form.form_admin, 'last_name', '') if form.form_admin else 'N/A',
                        'created_at': form.created_at.isoformat() if form.created_at else None
                    }

                    # --- Generate Excel data with batch processing ---
                    print(f"[Background Excel Task] Calling generate_excel_data for form {form_id}...")
                    excel_data = generate_excel_data(responses_data, form_info, form_id, user_org_id)
                    print(f"[Background Excel Task] Excel data prepared. Sheets: {list(excel_data.keys())}")
                    print(f"[Background Excel Task] Excel data prepared for form {form_id}; creating workbook...")

                    # --- Create Excel file with optimized writing ---
                    output = BytesIO()

                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # Write each section to different sheets
                        for sheet_name, data in excel_data.items():
                            if data:  # Only create sheet if there's data
                                print(f"[Background Excel Task] Writing sheet '{sheet_name}' with {len(data)-1} rows...")
                                # Use more efficient DataFrame creation
                                headers = data[0]
                                rows = data[1:]
                                df = pd.DataFrame(rows, columns=headers)

                                # Write to Excel more efficiently
                                print(f"[Background Excel Task] Converting DataFrame to Excel for '{sheet_name}'...")
                                df.to_excel(writer, sheet_name=sheet_name, index=False, engine_kwargs={'options': {'strings_to_urls': False}})

                                # Set column widths for better readability
                                worksheet = writer.sheets[sheet_name]

                                # Set specific widths for different column types
                                column_widths = {}
                                for idx, col_name in enumerate(headers):
                                    if col_name in ['Response ID', 'Form ID', 'Status']:
                                        column_widths[idx] = 15
                                    elif col_name in ['Submission Date', 'Form Created On', 'Generated On']:
                                        column_widths[idx] = 20
                                    elif col_name in ['Initiated By', 'Created By', 'Current Owner']:
                                        column_widths[idx] = 25
                                    elif col_name in ['Designation', 'Department', 'Location']:
                                        column_widths[idx] = 20
                                    elif col_name in ['Form Title', 'Form Type']:
                                        column_widths[idx] = 30
                                    else:
                                        column_widths[idx] = 35

                                # Apply the column widths more efficiently
                                for col_idx, width in column_widths.items():
                                    try:
                                        column_letter = worksheet.cell(row=1, column=col_idx + 1).column_letter
                                        worksheet.column_dimensions[column_letter].width = width
                                    except Exception:
                                        continue

                                # Optimized image processing with caching and parallel processing
                                try:
                                    print(f"[Background Excel Task] Starting image processing for sheet '{sheet_name}'...")
                                    import concurrent.futures

                                    # Image processing settings
                                    image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg')
                                    cloudinary_keywords = ('cloudinary', 'res.cloudinary', '/image/upload/')
                                    max_img_width_px = 150
                                    max_img_height_px = 120
                                    max_images_per_sheet = 200
                                    processed_images = 0

                                    # Global image cache across the entire request
                                    image_cache = {}
                                    processed_urls = set()
                                    adjusted_rows = set()

                                    def is_likely_image_url(url):
                                        """Check if URL is likely to be an image"""
                                        url_lower = url.lower()
                                        if any(ext in url_lower for ext in image_extensions):
                                            return True
                                        if any(keyword in url_lower for keyword in cloudinary_keywords):
                                            return True
                                        if 'image' in url_lower or 'img' in url_lower or 'photo' in url_lower:
                                            return True
                                        return False

                                    def process_single_image(url, cell, row_idx, col_idx):
                                        """Process a single image with caching"""
                                        try:
                                            if url in image_cache:
                                                pil_img = image_cache[url]
                                            else:
                                                resp = requests.get(url, timeout=5)
                                                if resp.status_code != 200 or len(resp.content) == 0:
                                                    return False

                                                if len(resp.content) > 2 * 1024 * 1024:
                                                    return False

                                                try:
                                                    pil_img = PILImage.open(BytesIO(resp.content)).convert("RGB")
                                                except Exception:
                                                    return False

                                                w, h = pil_img.size
                                                if w > 0 and h > 0:
                                                    scale = min(max_img_width_px / w, max_img_height_px / h, 1.0)
                                                    if scale < 1.0:
                                                        new_w = max(1, int(w * scale))
                                                        new_h = max(1, int(h * scale))
                                                        pil_img = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)

                                                image_cache[url] = pil_img

                                            img_stream = BytesIO()
                                            pil_img.save(img_stream, format='JPEG', quality=60)
                                            img_stream.seek(0)

                                            xl_img = XLImage(img_stream)
                                            xl_img.width = pil_img.width
                                            xl_img.height = pil_img.height

                                            anchor = cell.coordinate
                                            worksheet.add_image(xl_img, anchor)

                                            cell.value = None
                                            nonlocal adjusted_rows
                                            if row_idx not in adjusted_rows:
                                                worksheet.row_dimensions[row_idx].height = max(
                                                    worksheet.row_dimensions[row_idx].height or 15,
                                                    pil_img.height * 0.75
                                                )
                                                adjusted_rows.add(row_idx)

                                            return True

                                        except Exception as e:
                                            return False

                                    # Collect all potential image URLs to process
                                    image_tasks = []
                                    print(f"[Background Excel Task] Scanning for image URLs in sheet '{sheet_name}'...")
                                    for row_idx in range(2, min(worksheet.max_row + 1, 200)):
                                        for col_idx in range(1, worksheet.max_column + 1):
                                            if processed_images >= max_images_per_sheet:
                                                break

                                            try:
                                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                                value = cell.value

                                                if isinstance(value, str) and 'http' in value and len(value.strip()) > 10:
                                                    if value in processed_urls:
                                                        continue

                                                    if is_likely_image_url(value):
                                                        image_tasks.append((value, cell, row_idx, col_idx))
                                                        processed_urls.add(value)
                                            except Exception:
                                                continue

                                    # Process images in parallel (up to 10 concurrent)
                                    print(f"[Background Excel Task] Found {len(image_tasks)} potential image URLs to process...")
                                    if image_tasks:
                                        print(f"[Background Excel Task] Starting parallel image processing (max 10 workers)...")
                                        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                                            future_to_task = {
                                                executor.submit(process_single_image, url, cell, row_idx, col_idx): (url, cell, row_idx, col_idx)
                                                for url, cell, row_idx, col_idx in image_tasks[:max_images_per_sheet]
                                            }

                                            for future in concurrent.futures.as_completed(future_to_task):
                                                task = future_to_task[future]
                                                url, cell, row_idx, col_idx = task
                                                try:
                                                    result = future.result(timeout=8)
                                                    if result:
                                                        processed_images += 1
                                                except Exception:
                                                    continue
                                        print(f"[Background Excel Task] Image processing completed. Processed {processed_images} images for sheet '{sheet_name}'.")

                                except Exception as e:
                                    logger.warning(f"[Background Excel Task] Image processing error: {str(e)}")
                                    print(f"[Background Excel Task] Image processing error: {str(e)}")

                    output.seek(0)
                    print(f"[Background Excel Task] Excel file created. Size: {len(output.getvalue()) / (1024*1024):.2f} MB")

                    # --- Upload to S3 and send link email ---
                    print(f"[Background Excel Task] Uploading Excel to S3 for {email} with filename {filename}")
                    uf = UtilsFunctions()
                    upload_timestamp = dt.now().strftime("%d%m%y_%H%M%S")
                    local_file_path = f"reports/excel/{user_org_id}/{upload_timestamp}_{filename}"
                    bucket_name = getattr(settings, "S3_BUCKET_NAME_ANNOUNCEMENT", None) or settings.S3_BUCKET_NAME
                    output.seek(0)
                    
                    # Save file locally for development
                    import os
                    local_dir = os.path.join(settings.BASE_DIR, "media", "reports", "excel", str(user_org_id))
                    os.makedirs(local_dir, exist_ok=True)
                    local_file_full_path = os.path.join(local_dir, f"{upload_timestamp}_{filename}")
                    with open(local_file_full_path, 'wb') as f:
                        f.write(output.getvalue())
                    
                    # Use local file URL instead of S3
                    s3_url = _media_url(local_file_path)
                    expires_in_seconds = 86400
                    send_excel_link_email(email, short_download_url, filename, expires_in_seconds)
                    logger.info(f"[Background Excel Task] Excel report link sent successfully to {email} for form {form_id}")
                    print(f"[Background Excel Task] Email sent successfully to {email} for form {form_id}")
                    print(f"[Background Excel Task] COMPLETED SUCCESSFULLY for form {form_id}")

                    # Update status to SUCCESS
                    try:
                        cache_key = f"excel_job:{tracking_id}"
                        job = cache.get(cache_key) or {}
                        job.update({
                            "status": "SUCCESS",
                            "message": "Email sent",
                            "completed_at": dt.now().isoformat(),
                            "filename": filename,
                            "bucket_name": bucket_name,
                            "s3_key": local_file_path,
                            "expires_in_seconds": expires_in_seconds,
                        })
                        cache.set(cache_key, job, timeout=86400)
                    except Exception:
                        pass

                except Exception as e:
                    import traceback
                    logger.error(f"[Background Excel Task] Error generating Excel for form {form_id}: {str(e)}")
                    print(f"[Background Excel Task] ERROR for form {form_id}: {str(e)}")
                    print(f"[Background Excel Task] Traceback: {traceback.format_exc()}")
                    logger.error(f"[Background Excel Task] Traceback: {traceback.format_exc()}")
                    sys.stderr.write(f"[Background Excel Task] FATAL ERROR: {str(e)}\n")
                    sys.stderr.write(f"{traceback.format_exc()}\n")
                    sys.stderr.flush()
                    # Update status to FAILED
                    try:
                        cache_key = f"excel_job:{tracking_id}"
                        job = cache.get(cache_key) or {}
                        job.update({
                            "status": "FAILED",
                            "message": "Report generation failed",
                            "completed_at": dt.now().isoformat(),
                            "error": str(e)
                        })
                        cache.set(cache_key, job, timeout=86400)
                    except Exception:
                        pass
                    
                print(f"[Background Excel Task] ========== BACKGROUND THREAD ENDED ==========")
                sys.stdout.flush()

            # Start background thread (not daemon to ensure email is sent even if main process exits)
            threading.Thread(
                target=background_excel_generation,
                args=(form_id, submission_ids, email, filename, user_org_id, tracking_id, short_download_url),
                daemon=False
            ).start()

            return Response(
                {
                    "message": "Report is being generated and will be emailed to you shortly.",
                    "tracking_id": tracking_id
                },
                status=status.HTTP_202_ACCEPTED
            )

        except Exception as e:
            logger.error(f"[FormResponseCSVView] Error in main thread for form {form_id}: {str(e)}")
            return Response(
                {"error": "An error occurred processing your request."},
                status=status.HTTP_400_BAD_REQUEST
            )



class ExcelReportStatusView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, tracking_id):
        try:
            cache_key = f"excel_job:{tracking_id}"
            job = cache.get(cache_key)
            if not job:
                return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
            return Response(job, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class PDFReportStatusView(APIView):
    """Status endpoint for PDF-by-email jobs (same pattern as ExcelReportStatusView)."""
    permission_classes = [IsAuthenticated]

    def get(self, request, tracking_id):
        try:
            cache_key = f"pdf_job:{tracking_id}"
            job = cache.get(cache_key)
            if not job:
                return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
            return Response(job, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ReportDownloadRedirectView(APIView):
    """
    Redirect short report links to the actual file URL.
    For development, redirects to local file URLs. For production, redirects to S3 presigned URLs.
    Looks up both excel and pdf jobs by tracking_id.
    """
    permission_classes = [AllowAny]

    def get(self, request, tracking_id):
        try:
            job = cache.get(f"excel_job:{tracking_id}") or cache.get(f"pdf_job:{tracking_id}")
            if not job:
                return Response({"error": "Invalid or expired download link."}, status=status.HTTP_404_NOT_FOUND)

            if job.get("status") != "SUCCESS":
                return Response({"error": "Report is not ready yet."}, status=status.HTTP_409_CONFLICT)

            # For local development, redirect directly to the local file URL
            if job.get("s3_key", "").startswith("reports/"):
                # Local file path format: reports/pdfs/org_id/timestamp_filename
                local_file_url = _media_url(job.get('s3_key'))
                return HttpResponseRedirect(local_file_url)
            
            # For production with S3 (if bucket_name is not None) - now using local storage
            bucket_name = job.get("bucket_name")
            s3_key = job.get("s3_key")
            if bucket_name and s3_key and not s3_key.startswith("reports/"):
                # Use local storage instead of S3 presigned URL
                local_file_url = _media_url(s3_key)
                return HttpResponseRedirect(local_file_url)
            
            return Response({"error": "Download metadata is unavailable."}, status=status.HTTP_410_GONE)
        except Exception as e:
            logger.error(f"[ReportDownloadRedirectView] Error for tracking_id {tracking_id}: {str(e)}")
            return Response({"error": "Unable to process download link."}, status=status.HTTP_400_BAD_REQUEST)


class FormResponsePDFDownloadView(APIView):
    """
    Endpoint to generate PDF and download locally (no email involved).
    Supports both single and multiple submission IDs via query parameters or request body.
    When query param 'email' is provided, runs PDF generation in background and sends email (returns 202 + tracking_id).
    """
    permission_classes = [AllowAny]

    def get(self, request, form_id):
        import sys, traceback
        from datetime import datetime as _dt
        print(f"[PDF View] GET entered: form_id={form_id} | query_params={dict(request.query_params)} | Time: {_dt.now()}")
        sys.stdout.flush()
        try:
            # Handle anonymous users - remove organization filter
            organization = None if (hasattr(request.user, 'is_anonymous') and request.user.is_anonymous) else getattr(request.user, 'organization', None)
            
            # --- Fetch form ---
            if organization:
                form = get_object_or_404(Form, id=form_id, organization=organization)
            else:
                form = get_object_or_404(Form, id=form_id)
            print(f"[PDF View] Form found: {form.title or form_id} | Time: {_dt.now()}")
            sys.stdout.flush()

            # --- Get submission IDs from query parameters (before expensive serialization) ---
            submission_ids_param = request.query_params.get('submission_ids', '')
            if submission_ids_param:
                try:
                    submission_ids = [int(sid.strip()) for sid in submission_ids_param.split(',') if sid.strip()]
                except ValueError:
                    return Response({"error": "Invalid submission_ids format. Use comma-separated integers."}, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Fallback to single submission_id from query parameters
                submission_id = request.query_params.get('submission_id')
                if submission_id:
                    try:
                        submission_ids = [int(submission_id)]
                    except ValueError:
                        return Response({"error": "Invalid submission_id format. Must be an integer."}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"error": "No submission IDs provided. Use submission_ids (comma-separated) or submission_id parameter."}, status=status.HTTP_400_BAD_REQUEST)

            # --- Fetch submissions ---
            if organization:
                submissions = FormSubmision.objects.filter(
                    id__in=submission_ids,
                    form_id=form_id,
                    organization=organization
                )
            else:
                submissions = FormSubmision.objects.filter(
                    id__in=submission_ids,
                    form_id=form_id
                )

            if not submissions.exists():
                return Response({"error": "No submissions found for the provided IDs."}, status=status.HTTP_404_NOT_FOUND)

            if len(submissions) != len(submission_ids):
                found_ids = list(submissions.values_list('id', flat=True))
                missing_ids = [sid for sid in submission_ids if sid not in found_ids]
                return Response({
                    "error": f"Some submissions not found: {missing_ids}. Found: {found_ids}"
                }, status=status.HTTP_404_NOT_FOUND)

            # --- Optional: send PDF by email (background task, same flow as Excel) ---
            email = request.query_params.get('email', '').strip()
            print(f"[PDF View] submission_ids={submission_ids}, email param={repr(email)} (len={len(email) if email else 0}) | Time: {_dt.now()}")
            sys.stdout.flush()
            if email:
                from datetime import datetime as dt
                print(f"[PDF Email] Request received: form_id={form_id}, submission_ids={submission_ids}, email={email} | Time: {dt.now()}")
                sys.stdout.flush()
                timestamp = dt.now().strftime("%d_%m_%y_%H_%M")
                filename = f"{timestamp}_form_{form_id}_responses.pdf"
                user_org_id = organization.id if organization else 1
                tracking_id = str(uuid.uuid4())
                cache_key = f"pdf_job:{tracking_id}"
                short_download_url = _build_public_download_url(
                    request,
                    reverse("report-download-redirect", kwargs={"tracking_id": tracking_id}),
                )
                cache.set(cache_key, {
                    "status": "QUEUED",
                    "message": "PDF generation queued",
                    "form_id": form_id,
                    "submission_ids": submission_ids,
                    "email": email,
                    "filename": filename,
                    "bucket_name": None,
                    "s3_key": None,
                    "expires_in_seconds": None,
                    "created_at": dt.now().isoformat(),
                    "started_at": None,
                    "completed_at": None,
                    "error": None,
                }, timeout=86400)
                print(f"[PDF Email] Cache set QUEUED, tracking_id={tracking_id}. Starting background thread...")
                sys.stdout.flush()

                def background_pdf_generation(form_id, submission_ids, email, filename, user_org_id, tracking_id, short_download_url):
                    import sys
                    from datetime import datetime as dt
                    print(f"[Background PDF Task] ========== BACKGROUND THREAD STARTED ==========")
                    sys.stdout.flush()
                    try:
                        print(f"[Background PDF Task] START: form_id={form_id}, submission_ids={submission_ids}, email={email} | Time: {dt.now()}")
                        sys.stdout.flush()
                        cache_key = f"pdf_job:{tracking_id}"
                        job = cache.get(cache_key) or {}
                        job.update({"status": "RUNNING", "message": "PDF generation started", "started_at": dt.now().isoformat()})
                        cache.set(cache_key, job, timeout=86400)

                        print(f"[Background PDF Task] Fetching form {form_id}... | Time: {dt.now()}")
                        sys.stdout.flush()
                        form = Form.objects.get(id=form_id, organization_id=user_org_id)
                        organization = form.organization
                        print(f"[Background PDF Task] Form fetched. Serializing form schema (optimized prefetch)... | Time: {dt.now()}")
                        sys.stdout.flush()
                        # Optional form-schema cache (10 min TTL) to avoid ~2 min serialization on repeated PDF requests for same form
                        _PDF_FORM_SCHEMA_CACHE_TTL = 600
                        schema_cache_key = f"pdf_form_schema:{form_id}"
                        formSchema = cache.get(schema_cache_key)
                        if formSchema is not None:
                            print(f"[Background PDF Task] Form schema from cache (form_id={form_id}). Form type: {form.get_form_type_display()} | Time: {dt.now()}")
                            sys.stdout.flush()
                        else:
                            # Same optimized prefetch as Excel to avoid N+1 during serialization
                            base = (
                                Form.objects
                                .filter(is_deleted=False, pk=form.id)
                                .select_related('folder', 'form_admin', 'deletedBy', 'archivedBy')
                            )
                            stage_access_qs = StageAccess.objects.select_related('allow_user', 'allow_group', 'form', 'stage')
                            if form.form_type == FormType.AUDIT:
                                qs = base.prefetch_related(
                                    'assignee__user', 'assignee__group', 'assignee__leader',
                                    'audit_info__questions',
                                    'audit_info__questions__options',
                                    'audit_info__questions__child_questions',
                                    'audit_info__questions__child_questions__options',
                                    'audit_info__questions__logic_parent_question__logic_questions__options',
                                    'audit_info__questions__logic_parent_question__follow_ups',
                                    'audit_info__questions__logic_parent_question__follow_ups__task_close_questions__options',
                                    'audit_group__questions',
                                    'audit_group__questions__options',
                                    'audit_group__questions__child_questions',
                                    'audit_group__questions__child_questions__options',
                                    'audit_group__questions__logic_parent_question__logic_questions__options',
                                    'audit_group__questions__logic_parent_question__follow_ups',
                                    'audit_group__questions__logic_parent_question__follow_ups__task_close_questions__options',
                                )
                            else:
                                question_qs = (
                                    Question.objects
                                    .select_related('form', 'stage', 'parent_question')
                                    .prefetch_related(
                                        'options',
                                        'child_questions',
                                        'child_questions__options',
                                        'child_questions__child_questions',
                                        'child_questions__child_questions__options',
                                        'child_questions__logic_parent_question__logic_questions__options',
                                        'child_questions__logic_parent_question__follow_ups',
                                        'child_questions__logic_parent_question__follow_ups__task_close_questions__options',
                                        'logic_parent_question__logic_questions__options',
                                        'logic_parent_question__follow_ups',
                                        'logic_parent_question__follow_ups__task_close_questions__options',
                                    )
                                )
                                stage_qs = (
                                    Stage.objects
                                    .select_related('form')
                                    .prefetch_related(
                                        models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                                        models.Prefetch('questions', queryset=question_qs),
                                    )
                                )
                                qs = base.prefetch_related(
                                    'assignee__user', 'assignee__group', 'assignee__leader',
                                    models.Prefetch('stages', queryset=stage_qs),
                                )
                            optimized_instance = qs.get(pk=form.id)
                            formSchema = FormSerializer(optimized_instance, many=False).data
                            cache.set(schema_cache_key, formSchema, timeout=_PDF_FORM_SCHEMA_CACHE_TTL)
                            print(f"[Background PDF Task] Form schema serialized and cached. Form type: {form.get_form_type_display()} | Time: {dt.now()}")
                        sys.stdout.flush()

                        print(f"[Background PDF Task] Fetching submissions {submission_ids}... | Time: {dt.now()}")
                        sys.stdout.flush()
                        submissions = FormSubmision.objects.filter(
                            id__in=submission_ids,
                            form_id=form_id,
                            organization_id=user_org_id
                        ).select_related(
                            'submission_initiated_by',
                            'completed_by'
                        ).prefetch_related(
                            models.Prefetch(
                                'answers',
                                queryset=Answer.objects.select_related('question').prefetch_related('question__options'),
                                to_attr='prefetched_answers'
                            ),
                            'stage_submissions_history',
                            'group_submissions_history'
                        )
                        print(f"[Background PDF Task] Submissions query executed. exists={submissions.exists()} | Time: {dt.now()}")
                        sys.stdout.flush()
                        if not submissions.exists():
                            job = cache.get(cache_key) or {}
                            job.update({"status": "FAILED", "message": "No submissions found", "completed_at": dt.now().isoformat(), "error": "NO_SUBMISSIONS"})
                            cache.set(cache_key, job, timeout=86400)
                            return

                        def replace_option_id_with_text(answers_data, options):
                            if answers_data and 'answer' in answers_data and options:
                                answer_value = answers_data['answer']
                                raw_answer_value = answers_data.get('answer_id', answer_value)
                                other_text = answers_data.get('other_text')

                                def get_export_option_text(option_label):
                                    if option_label is None:
                                        return option_label
                                    if other_text and str(option_label).strip().lower() == 'other':
                                        return other_text
                                    return option_label

                                def find_option_by_id(option_id):
                                    option_id = str(option_id).strip()
                                    for opt in options:
                                        if opt and str(opt.get('id', '')).strip() == option_id:
                                            return opt
                                    return None

                                def replace_text_if_other(text_value):
                                    if other_text and str(text_value).strip().lower() == 'other':
                                        return other_text
                                    return text_value

                                if answer_value:
                                    answer_str = str(answer_value).strip()
                                    raw_answer_str = str(raw_answer_value).strip()

                                    if other_text and answer_str.lower() == 'other':
                                        answers_data['answer'] = other_text
                                    elif ',' in answer_str:
                                        ids = [id.strip() for id in answer_str.split(',') if id.strip()]
                                        replaced = []
                                        for id_val in ids:
                                            if id_val.isdigit():
                                                opt = find_option_by_id(id_val)
                                                if opt:
                                                    replaced.append(get_export_option_text(opt.get('option')))
                                                else:
                                                    replaced.append(id_val)
                                            else:
                                                replaced.append(replace_text_if_other(id_val))
                                        answers_data['answer'] = ', '.join(replaced)
                                    elif answer_str.isdigit():
                                        opt = find_option_by_id(answer_str)
                                        if opt:
                                            answers_data['answer'] = get_export_option_text(opt.get('option'))
                                    else:
                                        raw_option_ids = re.findall(r'\d+', raw_answer_str)
                                        if raw_option_ids:
                                            matched_options = []
                                            for option_id in raw_option_ids:
                                                opt = find_option_by_id(option_id)
                                                if opt:
                                                    matched_options.append(get_export_option_text(opt.get('option')))
                                            if matched_options:
                                                answers_data['answer'] = ', '.join(matched_options)
                                                return

                                        answers_data['answer'] = replace_text_if_other(answer_value)

                        responses_data = []
                        submission_count = 0
                        for submission_instance in submissions:
                            submission_id = submission_instance.id
                            submission_count += 1
                            print(f"[Background PDF Task] Processing submission {submission_count}/{submissions.count()}: id={submission_id} | Time: {dt.now()}")
                            sys.stdout.flush()
                            # Use prefetched answers (same as Excel) to avoid N+1
                            answer_lookup = {a.question_id: a for a in getattr(submission_instance, 'prefetched_answers', [])}
                            response_data = {
                                'id': submission_instance.id,
                                'submission_initiated_on': submission_instance.submission_initiated_on.isoformat() if submission_instance.submission_initiated_on else None,
                                'submission_initiated_by': getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                                'initiator_designation': str(getattr(submission_instance.submission_initiated_by, 'designation', 'N/A')) if submission_instance.submission_initiated_by else 'N/A',
                                'initiator_department': str(getattr(submission_instance.submission_initiated_by, 'department', 'N/A')) if submission_instance.submission_initiated_by else 'N/A',
                                'initiator_location': str(getattr(submission_instance.submission_initiated_by, 'location', 'N/A')) if submission_instance.submission_initiated_by else 'N/A',
                                'current_owner': getattr(submission_instance.completed_by, 'first_name', '') + ' ' + getattr(submission_instance.completed_by, 'last_name', '') if submission_instance.completed_by else getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                                'is_completed': submission_instance.is_completed,
                                'stages': []
                            }
                            if formSchema.get("form_type") == FormType.AUDIT:
                                stages = []
                                audit_info = formSchema.get("audit_info")
                                if audit_info:
                                    audit_info_copy = copy.deepcopy(audit_info)
                                    audit_info_stage = {'id': audit_info_copy.get('id', 'audit_info'), 'name': audit_info_copy.get('name', 'Audit Info'), 'is_audit_info': True, 'order': 0, 'questions': audit_info_copy.get("questions", [])}
                                    for question in audit_info_stage['questions']:
                                        answer = answer_lookup.get(question['id'])
                                        question['answers'] = AnswerSerializer(answer).data if answer else {}
                                        replace_option_id_with_text(question['answers'], question.get('options'))
                                        for subQuestion in question.get('sub_questions', []):
                                            sub_answer = answer_lookup.get(subQuestion['id'])
                                            subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}
                                            replace_option_id_with_text(subQuestion['answers'], subQuestion.get('options'))
                                        for logics in question.get("logics", []):
                                            for logicQuestion in logics.get("logic_questions", []):
                                                logic_answer = answer_lookup.get(logicQuestion['id'])
                                                logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}
                                                replace_option_id_with_text(logicQuestion['answers'], logicQuestion.get('options'))
                                    stages.append(audit_info_stage)
                                for audit_group in formSchema.get("audit_group", []):
                                    audit_group_copy = copy.deepcopy(audit_group)
                                    for audit_group_question in audit_group_copy.get("questions", []):
                                        answer = answer_lookup.get(audit_group_question['id'])
                                        audit_group_question['answers'] = AnswerSerializer(answer).data if answer else {}
                                        replace_option_id_with_text(audit_group_question['answers'], audit_group_question.get('options'))
                                        for logics in audit_group_question.get("logics", []):
                                            for logicQuestion in logics.get("logic_questions", []):
                                                logic_answer = answer_lookup.get(logicQuestion['id'])
                                                logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}
                                                replace_option_id_with_text(logicQuestion['answers'], logicQuestion.get('options'))
                                        for subQuestion in audit_group_question.get('sub_questions', []):
                                            sub_answer = answer_lookup.get(subQuestion['id'])
                                            subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}
                                            replace_option_id_with_text(subQuestion['answers'], subQuestion.get('options'))
                                    stages.append(audit_group_copy)
                                response_data['stages'] = stages
                            else:
                                stage_history_by_id = {h.stage_id: h for h in submission_instance.stage_submissions_history.all()}
                                for stage in formSchema.get('stages', []):
                                    stage_copy = copy.deepcopy(stage)
                                    stage_history = stage_history_by_id.get(stage_copy['id'])
                                    historyData = StageSubmissionHistorySerializer(stage_history, many=False).data if stage_history else {}
                                    stage_data = {'id': stage_copy['id'], 'name': stage_copy.get('name', f'Stage {stage_copy.get("order", "")}'), 'order': stage_copy.get('order', 0), 'is_completed': stage_history is not None, 'completed_by': historyData.get('completed_by'), 'completed_on': historyData.get('completed_on'), 'questions': []}
                                    for question in stage_copy.get('questions', []):
                                        question_data = {'id': question['id'], 'question': question.get('question', ''), 'question_type': question.get('question_type', ''), 'order': question.get('order', 0)}
                                        answer = answer_lookup.get(question['id'])
                                        if answer:
                                            question_data['answers'] = AnswerSerializer(answer).data
                                            replace_option_id_with_text(question_data['answers'], question.get('options'))
                                        sub_questions = []
                                        for subQuestion in question.get('sub_questions', []):
                                            sub_answer = answer_lookup.get(subQuestion['id'])
                                            sub_question_data = {'id': subQuestion['id'], 'question': subQuestion.get('question', ''), 'question_type': subQuestion.get('question_type', ''), 'answers': AnswerSerializer(sub_answer).data if sub_answer else {}}
                                            replace_option_id_with_text(sub_question_data['answers'], subQuestion.get('options'))
                                            sub_questions.append(sub_question_data)
                                        if sub_questions:
                                            question_data['sub_questions'] = sub_questions
                                        logics = []
                                        for logic in question.get("logics", []):
                                            logic_questions = []
                                            for logicQuestion in logic.get("logic_questions", []):
                                                logic_answer = answer_lookup.get(logicQuestion['id'])
                                                logic_question_data = {'id': logicQuestion['id'], 'question': logicQuestion.get('question', ''), 'question_type': logicQuestion.get('question_type', ''), 'answers': AnswerSerializer(logic_answer).data if logic_answer else {}}
                                                replace_option_id_with_text(logic_question_data['answers'], logicQuestion.get('options'))
                                                logic_questions.append(logic_question_data)
                                            if logic_questions:
                                                logics.append({'logic_questions': logic_questions})
                                        if logics:
                                            question_data['logics'] = logics
                                        stage_data['questions'].append(question_data)
                                    response_data['stages'].append(stage_data)
                            responses_data.append(response_data)
                            print(f"[Background PDF Task] Submission {submission_id} response_data built | Time: {dt.now()}")
                            sys.stdout.flush()

                        print(f"[Background PDF Task] All {len(responses_data)} submission(s) built. Building form_info... | Time: {dt.now()}")
                        sys.stdout.flush()
                        form_info = {
                            'title': form.title or f'Form {form_id}',
                            'form_type': form.get_form_type_display() or 'standard',
                            'created_by': getattr(form.form_admin, 'first_name', '') + ' ' + getattr(form.form_admin, 'last_name', '') if form.form_admin else 'N/A',
                            'created_at': form.created_at.isoformat() if form.created_at else None,
                            'prefix': getattr(form, 'prefix', None) or ''
                        }
                        if form.form_type == FormType.AUDIT:
                            print(f"[Background PDF Task] Calling generate_audit_pdf (audit mode)... | Time: {dt.now()}")
                            sys.stdout.flush()
                            pdf = generate_audit_pdf(responses_data, form_info, form_id, organization, is_audit_form=True)
                        else:
                            print(f"[Background PDF Task] Calling generate_audit_pdf (standard mode)... | Time: {dt.now()}")
                            sys.stdout.flush()
                            pdf = generate_audit_pdf(responses_data, form_info, form_id, organization, is_audit_form=False)
                        print(f"[Background PDF Task] PDF generated. Size: {len(pdf) / 1024:.1f} KB | Time: {dt.now()}")
                        sys.stdout.flush()
                        print(f"[Background PDF Task] Uploading PDF to S3... | Time: {dt.now()}")
                        sys.stdout.flush()
                        uf = UtilsFunctions()
                        upload_timestamp = dt.now().strftime("%d%m%y_%H%M%S")
                        local_file_path = f"reports/pdfs/{user_org_id}/{upload_timestamp}_{filename}"
                        bucket_name = getattr(settings, "S3_BUCKET_NAME_ANNOUNCEMENT", None) or settings.S3_BUCKET_NAME
                        pdf_buffer = io.BytesIO(pdf if isinstance(pdf, (bytes, bytearray)) else pdf.getvalue())
                        
                        # Save file locally for development
                        import os
                        local_dir = os.path.join(settings.BASE_DIR, "media", "reports", "pdfs", str(user_org_id))
                        os.makedirs(local_dir, exist_ok=True)
                        local_file_full_path = os.path.join(local_dir, f"{upload_timestamp}_{filename}")
                        with open(local_file_full_path, 'wb') as f:
                            f.write(pdf_buffer.getvalue())
                        
                        # Use local file URL instead of S3
                        s3_url = _media_url(local_file_path)
                        expires_in_seconds = 86400
                        print(f"[Background PDF Task] Sending PDF link email to {email}... | Time: {dt.now()}")
                        sys.stdout.flush()
                        send_pdf_link_email(email, short_download_url, filename, expires_in_seconds)
                        print(f"[Background PDF Task] Email sent successfully | Time: {dt.now()}")
                        sys.stdout.flush()
                        job = cache.get(cache_key) or {}
                        job.update({
                            "status": "SUCCESS",
                            "message": "Email with PDF link sent",
                            "completed_at": dt.now().isoformat(),
                            "filename": filename,
                            "bucket_name": bucket_name,
                            "s3_key": local_file_path,
                            "expires_in_seconds": expires_in_seconds
                        })
                        cache.set(cache_key, job, timeout=86400)
                        print(f"[Background PDF Task] COMPLETED SUCCESSFULLY for form_id={form_id}")
                        sys.stdout.flush()
                    except Exception as e:
                        import traceback
                        print(f"[Background PDF Task] EXCEPTION: {str(e)} | Time: {dt.now()}")
                        sys.stdout.flush()
                        print(f"[Background PDF Task] Traceback: {traceback.format_exc()}")
                        sys.stdout.flush()
                        logger.error(f"[Background PDF Task] Error: {str(e)}\n{traceback.format_exc()}")
                        try:
                            cache_key = f"pdf_job:{tracking_id}"
                            job = cache.get(cache_key) or {}
                            job.update({"status": "FAILED", "message": "PDF generation failed", "completed_at": dt.now().isoformat(), "error": str(e)})
                            cache.set(cache_key, job, timeout=86400)
                        except Exception:
                            pass
                    print(f"[Background PDF Task] ========== BACKGROUND THREAD ENDED ==========")
                    sys.stdout.flush()

                threading.Thread(
                    target=background_pdf_generation,
                    args=(form_id, submission_ids, email, filename, user_org_id, tracking_id, short_download_url),
                    daemon=False
                ).start()
                return Response(
                    {"message": "PDF is being generated and will be emailed to you shortly.", "tracking_id": tracking_id},
                    status=status.HTTP_202_ACCEPTED
                )

            # No email param: synchronous download path (build formSchema only here to avoid main-thread hang on email path)
            print(f"[PDF View] No email param - using download path | Time: {_dt.now()}")
            sys.stdout.flush()
            formSchema = FormSerializer(form, many=False, context={'request': request}).data
            # --- Helper function to replace option ID with text ---
            def replace_option_id_with_text(answers_data, options):
                if answers_data and 'answer' in answers_data and options:
                    answer_value = answers_data['answer']
                    raw_answer_value = answers_data.get('answer_id', answer_value)
                    other_text = answers_data.get('other_text')

                    def get_export_option_text(option_label):
                        if option_label is None:
                            return option_label
                        if other_text and str(option_label).strip().lower() == 'other':
                            return other_text
                        return option_label

                    def find_option_by_id(option_id):
                        option_id = str(option_id).strip()
                        for opt in options:
                            if opt and str(opt.get('id', '')).strip() == option_id:
                                return opt
                        return None

                    def replace_text_if_other(text_value):
                        if other_text and str(text_value).strip().lower() == 'other':
                            return other_text
                        return text_value

                    if answer_value:
                        answer_str = str(answer_value).strip()
                        raw_answer_str = str(raw_answer_value).strip()

                        if other_text and answer_str.lower() == 'other':
                            answers_data['answer'] = other_text
                        elif ',' in answer_str:
                            # Handle multiple options like "1,2,3"
                            ids = [id.strip() for id in answer_str.split(',') if id.strip()]
                            replaced = []
                            for id_val in ids:
                                if id_val.isdigit():
                                    opt = find_option_by_id(id_val)
                                    if opt:
                                        replaced.append(get_export_option_text(opt.get('option')))
                                    else:
                                        replaced.append(id_val)  # keep original if not found
                                else:
                                    replaced.append(replace_text_if_other(id_val))
                            answers_data['answer'] = ', '.join(replaced)
                        elif answer_str.isdigit():
                            # Single option
                            opt = find_option_by_id(answer_str)
                            if opt:
                                answers_data['answer'] = get_export_option_text(opt.get('option'))
                        else:
                            raw_option_ids = re.findall(r'\d+', raw_answer_str)
                            if raw_option_ids:
                                matched_options = []
                                for option_id in raw_option_ids:
                                    opt = find_option_by_id(option_id)
                                    if opt:
                                        matched_options.append(get_export_option_text(opt.get('option')))
                                if matched_options:
                                    answers_data['answer'] = ', '.join(matched_options)
                                    return

                            answers_data['answer'] = replace_text_if_other(answer_value)

            # --- Prepare multiple response data ---
            responses_data = []

            for submission_instance in submissions:
                submission_id = submission_instance.id
                submissions_data = FormSubmissionSerializer(submission_instance, context={'request': request}).data

                # --- Prepare response data like frontend ---
                response_data = {
                    'id': submission_instance.id,
                    'submission_initiated_on': submission_instance.submission_initiated_on.isoformat() if submission_instance.submission_initiated_on else None,
                    'submission_initiated_by': getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                    'initiator_designation': str(getattr(submission_instance.submission_initiated_by, 'designation', 'N/A')) if submission_instance.submission_initiated_by else 'N/A', # <--- Add str() cast here
                    'initiator_department': str(getattr(submission_instance.submission_initiated_by, 'department', 'N/A')) if submission_instance.submission_initiated_by else 'N/A',
                    'initiator_location': str(getattr(submission_instance.submission_initiated_by, 'location', 'N/A')) if submission_instance.submission_initiated_by else 'N/A',
                    'current_owner': getattr(submission_instance.completed_by, 'first_name', '') + ' ' + getattr(submission_instance.completed_by, 'last_name', '') if submission_instance.completed_by else getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                    'is_completed': submission_instance.is_completed,
                    'stages': []
                }

                # --- Populate answers into stages ---
                if formSchema.get("form_type") == FormType.AUDIT:
                    stages = []

                    # Handle audit_info as the first stage
                    audit_info = formSchema.get("audit_info")
                    if audit_info:
                        # Create a deep copy to avoid mutating shared objects across submissions
                        audit_info_copy = copy.deepcopy(audit_info)
                        audit_info_stage = {
                            'id': audit_info_copy.get('id', 'audit_info'),
                            'name': audit_info_copy.get('name', 'Audit Info'),
                            'is_audit_info': True,
                            'order': 0,
                            'questions': audit_info_copy.get("questions", [])
                        }

                        # Process answers for audit_info questions
                        for question in audit_info_stage['questions']:
                            answers = Answer.objects.filter(
                                question=question['id'],
                                submission=submission_id,
                                organization=form.organization
                            )
                            question['answers'] = AnswerSerializer(answers.first()).data if answers.exists() else {}
                            # Convert option IDs to text
                            replace_option_id_with_text(question['answers'], question.get('options'))

                            # Handle sub-questions for audit_info
                            for subQuestion in question.get('sub_questions', []):
                                sub_answers = Answer.objects.filter(
                                    question=subQuestion['id'],
                                    submission=submission_id,
                                    organization=form.organization
                                )
                                subQuestion['answers'] = AnswerSerializer(sub_answers.first()).data if sub_answers.exists() else {}
                                # Convert option IDs to text
                                replace_option_id_with_text(subQuestion['answers'], subQuestion.get('options'))

                            # Handle logic questions for audit_info
                            for logics in question.get("logics", []):
                                for logicQuestion in logics.get("logic_questions", []):
                                    logic_answer = Answer.objects.filter(
                                        question=logicQuestion['id'],
                                        submission=submission_id,
                                        organization=form.organization
                                    )
                                    logicQuestion['answers'] = AnswerSerializer(logic_answer.first()).data if logic_answer.exists() else {}
                                    # Convert option IDs to text
                                    replace_option_id_with_text(logicQuestion['answers'], logicQuestion.get('options'))

                        stages.append(audit_info_stage)

                    # For audit forms, we'll keep the original audit logic
                    for audit_group in formSchema.get("audit_group", []):
                        # Create a deep copy to avoid mutating shared objects across submissions
                        audit_group_copy = copy.deepcopy(audit_group)
                        for audit_group_question in audit_group_copy.get("questions", []):
                            answers = Answer.objects.filter(
                                question=audit_group_question['id'],
                                submission=submission_id,
                                organization=form.organization
                            )
                            audit_group_question['answers'] = AnswerSerializer(answers.first()).data if answers.exists() else {}
                            # Convert option IDs to text
                            replace_option_id_with_text(audit_group_question['answers'], audit_group_question.get('options'))

                            for logics in audit_group_question.get("logics", []):
                                for logicQuestion in logics.get("logic_questions", []):
                                    logic_answer = Answer.objects.filter(
                                        question=logicQuestion['id'],
                                        submission=submission_id,
                                        organization=form.organization
                                    )
                                    logicQuestion['answers'] = AnswerSerializer(logic_answer.first()).data if logic_answer.exists() else {}
                                    # Convert option IDs to text
                                    replace_option_id_with_text(logicQuestion['answers'], logicQuestion.get('options'))

                            for subQuestion in audit_group_question.get('sub_questions', []):
                                sub_answers = Answer.objects.filter(
                                    question=subQuestion['id'],
                                    submission=submission_id,
                                    organization=form.organization
                                )
                                subQuestion['answers'] = AnswerSerializer(sub_answers.first()).data if sub_answers.exists() else {}
                                # Convert option IDs to text
                                replace_option_id_with_text(subQuestion['answers'], subQuestion.get('options'))

                        stages.append(audit_group_copy)

                    response_data['stages'] = stages
                else:
                    # For standard/location forms, populate stages with questions and answers
                    for stage in formSchema.get('stages', []):
                        # Create a deep copy to avoid mutating shared objects across submissions
                        stage_copy = copy.deepcopy(stage)
                        
                        # Get stage completion history
                        stage_histories = StageSubmissionHistory.objects.filter(
                            stage__id=stage_copy['id'],
                            form_submission__id=submission_id,
                            organization=form.organization
                        )

                        historyData = {}
                        if stage_histories.exists():
                            stageHistory = stage_histories.first()
                            historyData = StageSubmissionHistorySerializer(stageHistory, many=False, context={'request': request}).data

                        stage_data = {
                            'id': stage_copy['id'],
                            'name': stage_copy.get('name', f'Stage {stage_copy.get("order", "")}'),
                            'order': stage_copy.get('order', 0),
                            'is_completed': stage_histories.exists(),
                            'completed_by': historyData.get('completed_by', None),
                            'completed_on': historyData.get('completed_on', None),
                            'questions': []
                        }

                        # Populate questions with answers
                        for question in stage_copy.get('questions', []):
                            question_data = {
                                'id': question['id'],
                                'question': question.get('question', ''),
                                'question_type': question.get('question_type', ''),
                                'order': question.get('order', 0),
                            }

                            # Get answer for this question
                            answers = Answer.objects.filter(
                                question=question['id'],
                                stage=stage_copy['id'],
                                submission=submission_id,
                                organization=form.organization
                            )
                            if answers.exists():
                                question_data['answers'] = AnswerSerializer(answers.first()).data
                                # Convert option IDs to text
                                replace_option_id_with_text(question_data['answers'], question.get('options'))

                            # Handle sub-questions
                            sub_questions = []
                            for subQuestion in question.get('sub_questions', []):
                                sub_answers = Answer.objects.filter(
                                    question=subQuestion['id'],
                                    stage=stage_copy['id'],
                                    submission=submission_id,
                                    organization=form.organization
                                )

                                sub_question_data = {
                                    'id': subQuestion['id'],
                                    'question': subQuestion.get('question', ''),
                                    'question_type': subQuestion.get('question_type', ''),
                                    'answers': AnswerSerializer(sub_answers.first()).data if sub_answers.exists() else {}
                                }
                                # Convert option IDs to text
                                replace_option_id_with_text(sub_question_data['answers'], subQuestion.get('options'))
                                sub_questions.append(sub_question_data)

                            if sub_questions:
                                question_data['sub_questions'] = sub_questions

                            # Handle logic questions
                            logics = []
                            for logic in question.get("logics", []):
                                logic_questions = []
                                for logicQuestion in logic.get("logic_questions", []):
                                    logic_answer = Answer.objects.filter(
                                        question=logicQuestion['id'],
                                        stage=stage_copy['id'],
                                        submission=submission_id,
                                        organization=form.organization
                                    )

                                    logic_question_data = {
                                        'id': logicQuestion['id'],
                                        'question': logicQuestion.get('question', ''),
                                        'question_type': logicQuestion.get('question_type', ''),
                                        'answers': AnswerSerializer(logic_answer.first()).data if logic_answer.exists() else {}
                                    }
                                    # Convert option IDs to text
                                    replace_option_id_with_text(logic_question_data['answers'], logicQuestion.get('options'))
                                    logic_questions.append(logic_question_data)

                                if logic_questions:
                                    logics.append({'logic_questions': logic_questions})

                            if logics:
                                question_data['logics'] = logics

                            stage_data['questions'].append(question_data)

                        response_data['stages'].append(stage_data)

                responses_data.append(response_data)

            # --- Prepare form info ---
            form_info = {
                'title': form.title or f'Form {form_id}',
                'form_type': form.get_form_type_display() or 'standard',
                'created_by': getattr(form.form_admin, 'first_name', '') + ' ' + getattr(form.form_admin, 'last_name', '') if form.form_admin else 'N/A',
                'created_at': form.created_at.isoformat() if form.created_at else None,
                'prefix': getattr(form, 'prefix', None) or ''
            }

            # --- Generate PDF with multiple responses ---
            # Use audit-style layout for both audit and standard forms.
            pdf_org = organization or form.organization
            if form.form_type == FormType.AUDIT:
                pdf = generate_audit_pdf(
                    responses_data,
                    form_info,
                    form_id,
                    pdf_org,
                    is_audit_form=True
                )
            else:
                pdf = generate_audit_pdf(
                    responses_data,
                    form_info,
                    form_id,
                    pdf_org,
                    is_audit_form=False
                )

            # --- Return as download ---
            submission_ids_str = '_'.join(map(str, sorted(submission_ids)))
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename=form_{form_id}_responses_{submission_ids_str}.pdf'
            response['Content-Length'] = len(pdf)
            return response

        except Exception as e:
            logger.error(f"Error generating PDF for form {form_id}: {str(e)}\n{traceback.format_exc()}")
            print(f"[PDF View] ERROR: {str(e)}")
            print(f"[PDF View] TRACEBACK:\n{traceback.format_exc()}")
            sys.stdout.flush()
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class AuditFormScoreDataView(APIView):
    """
    API endpoint to retrieve audit form score calculation data for Excel export.
    Returns all audit form submission history records with score calculations for a specific form.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, form_id):
        try:
            # Verify form exists and is an audit form
            form = get_object_or_404(
                Form,
                id=form_id,
                organization=request.user.organization,
                is_deleted=False,
                is_archived=False
            )
            
            if form.form_type != 'audit':
                return Response(
                    {"error": "This endpoint is only available for audit forms."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get submission IDs from query parameters (optional filter)
            submission_ids_param = request.query_params.get('submission_ids', '')
            submission_ids = None
            if submission_ids_param:
                try:
                    submission_ids = [int(sid.strip()) for sid in submission_ids_param.split(',') if sid.strip()]
                except ValueError:
                    return Response(
                        {"error": "Invalid submission_ids format. Use comma-separated integers."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Fetch audit form submission histories
            query = AuditFormSubmissionHistory.objects.filter(
                form_id=form_id,
                organization=request.user.organization
            ).select_related(
                'form_submission',
                'form_submission__form',
                'form_submission__submission_initiated_by',
                'form_submission__completed_by',
                'completed_by',
                'form_id',
                'group_id'
            )
            
            # Filter by submission IDs if provided
            if submission_ids:
                query = query.filter(form_submission_id__in=submission_ids)
            
            # Order by completion date (newest first)
            audit_histories = query.order_by('-completed_on')
            
            # Serialize the data
            from .serializers import AuditFormSubmissionHistorySerializer
            serializer = AuditFormSubmissionHistorySerializer(audit_histories, many=True)
            
            # Enhance response with additional details for Excel export
            enhanced_data = []
            for item in serializer.data:
                history_obj = audit_histories.filter(id=item['id']).first()
                if history_obj:
                    enhanced_item = {
                        **item,
                        'form_title': form.title,
                        'form_type': form.form_type,
                        'submission_initiated_by': history_obj.form_submission.submission_initiated_by.username if history_obj.form_submission.submission_initiated_by else None,
                        'submission_initiated_on': history_obj.form_submission.submission_initiated_on.isoformat() if history_obj.form_submission.submission_initiated_on else None,
                        'submission_completed_by': history_obj.form_submission.completed_by.username if history_obj.form_submission.completed_by else None,
                        'submission_completed_on': history_obj.form_submission.completed_on.isoformat() if history_obj.form_submission.completed_on else None,
                        'group_name': history_obj.group_id.name if history_obj.group_id else None,
                        'completed_by_username': history_obj.completed_by.username if history_obj.completed_by else None,
                        'completed_by_email': history_obj.completed_by.email if history_obj.completed_by else None,
                    }
                    enhanced_data.append(enhanced_item)
            
            return Response({
                'form_id': form_id,
                'form_title': form.title,
                'total_records': len(enhanced_data),
                'summary': enhanced_data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error retrieving audit form score data for form {form_id}: {str(e)}")
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


# Retrieves submissions for specified form IDs, grouped by form, with details like initiation and completion status.
class FormSubmissionGroupedList(APIView):
    permission_classes = [IsAdmin, IsEndUser]

    def post(self, request):
        try:
            data = request.data
            form_ids = data.get('forms', [])
            # Get forms and submissions filtered by organization and form IDs
            forms = Form.objects.filter(id__in=form_ids, organization=request.user.organization)
            
            response_data = []
            for form in forms:
                # Keep only true main-form submissions in grouped responses as well.
                # This prevents task-close generated followup submissions from appearing
                # as extra main responses (especially for audit forms).
                submissions = FormSubmision.objects.filter(
                    form=form,
                    organization=request.user.organization,
                ).filter(
                    Q(submission_initiated_stage__isnull=False) |
                    Q(
                        form__form_type=FormType.AUDIT,
                        group_submissions_history__isnull=False
                    )
                ).distinct()
                # Serialize individual submissions
                submission_serializer = FormSubmissionSerializer(submissions, many=True, context={'request': request})
                
                # Structure submission data according to the interface
                submission_list = [
                    {
                        'id': sub['id'],
                        'submission_initiated_on': sub.get('created_at', 0),  
                        'submission_initiated_stage': sub.get('stage', 0),    
                        'submission_initiated_by': sub.get('created_by', ''), 
                        'is_completed': 1 if sub.get('is_completed', False) else 0,
                        'completed_by': sub.get('completed_by', None) is not None,
                        'completed_on': sub.get('completed_at', None) is not None
                    } for sub in submission_serializer.data
                ]
                
                # Structure form data
                form_data = {
                    'id': str(form.id),
                    'title': form.title,  
                    'form_type': form.form_type if hasattr(form, 'form_type') else 'default',  
                    'submissions': submission_list
                }
                response_data.append(form_data)
            
            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error retrieving grouped form submissions: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)  

# Mobile view for user assigned forms
# Retrieves a list of form IDs and titles assigned to the authenticated user. 
class AssignedFormsList(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            if not request.user.is_authenticated:
                return Response({"error": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
            
            user = request.user
            # Fetch forms assigned to the user via StageAssignment
            form_assignments = FormAssignment.objects.filter(
                user=user,
                form__is_deleted=False,
                form__is_archived=False,
                form__organization=user.organization
            ).select_related('form').distinct()

            # Extract only id and title for each assigned form
            assigned_forms = [
                {
                    'id': str(assignment.form.id),
                    'title': assignment.form.title,
                    'type':assignment.form.form_type 
                }
                for assignment in form_assignments
            ]

            return Response(assigned_forms, status=status.HTTP_200_OK)
        except Exception as e:
            user_id = getattr(request.user, 'id', None) if request.user.is_authenticated else None
            logger.error(f"Error retrieving assigned forms for user {user_id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)  
        
# Mobile view for user sent forms
# Retrieves forms sent by a specified user, grouped by form with stage assignment details, for mobile app display.

class UserSentFormsMobileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, user_id):
        try:
            user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)
            forms_dict = {}
            
            # Get forms where user sent them (assigned_by=user) - existing logic
            stage_assignments = StageAssignment.objects.filter(
                assigned_by=user, organization=request.user.organization,
                form__is_archived=False
            ).select_related('form', 'stage', 'form_submission')

            for assignment in stage_assignments:
                form_id = assignment.form.id
                if form_id not in forms_dict:
                    form_data = CompactFormSerializer(
                        assignment.form, many=False, context={'request': request}
                    ).data
                    forms_dict[form_id] = {
                        'id': form_data['id'],
                        'title': form_data.get('title', ''),
                        'form_type': form_data.get('form_type', 'standard'),
                        'sent': []
                    }
                
                forms_dict[form_id]['sent'].append({
                    'stage_id': assignment.stage.id,
                    'stage_order': assignment.stage.order,
                    'stage_assignment_id': assignment.id,
                    'assignment_uuid': assignment.assignment_uuid,  # CharField, no str() needed
                    'form_submission_id': assignment.form_submission.id if assignment.form_submission else None,
                    'is_stage_submission_pending': not assignment.is_assignment_fullfilled,
                    'is_form_submission_pending': not assignment.form_submission.is_completed if assignment.form_submission else True
                })

            # Include audit/group assignments sent by this user (audit forms use GroupAssignment)
            group_assignments = GroupAssignment.objects.filter(
                assigned_by=user, organization=request.user.organization,
                form__is_archived=False
            ).select_related('form', 'group', 'form_submission')

            for g_assignment in group_assignments:
                form_id = g_assignment.form.id
                if form_id not in forms_dict:
                    form_data = CompactFormSerializer(
                        g_assignment.form, many=False, context={'request': request}
                    ).data
                    forms_dict[form_id] = {
                        'id': form_data['id'],
                        'title': form_data.get('title', ''),
                        'form_type': form_data.get('form_type', 'standard'),
                        'sent': []
                    }

                forms_dict[form_id]['sent'].append({
                    'stage_id': None,
                    'stage_order': g_assignment.group_order,
                    'stage_assignment_id': None,
                    'assignment_uuid': g_assignment.assignment_uuid,
                    'form_submission_id': g_assignment.form_submission.id if g_assignment.form_submission else None,
                    'is_stage_submission_pending': not g_assignment.is_assignment_fullfilled,
                    'is_form_submission_pending': not g_assignment.form_submission.is_completed if g_assignment.form_submission else True
                })

            # Helper: detect source of a form submission
            def get_submission_source(submission):
                """Returns (source, source_ref) for a submission.
                source: 'planner' | 'task' | 'form'
                source_ref: planner order_id, task id, or None
                """
                from planner.models import PlannerSubmission
                planner_sub = PlannerSubmission.objects.filter(
                    form_submission=submission
                ).select_related('planner_assignment').first()
                if planner_sub and planner_sub.planner_assignment:
                    return 'planner', planner_sub.planner_assignment.order_id
                # Check if this submission was triggered by a followup task
                followup_source = submission.followup_source_task.first() if hasattr(submission, 'followup_source_task') else None
                if not followup_source:
                    from task.models import Task
                    followup_source = Task.objects.filter(
                        followup_task_form_id=submission.form,
                        organization=submission.organization
                    ).first()
                if followup_source:
                    return 'task', followup_source.id
                return 'form', None

            # Get forms where user completed them (EXCLUDE task submissions)
            # First, get form IDs that are part of tasks assigned to this user (completed)
            from task.models import TaskTracking, Task
            completed_task_trackings = TaskTracking.objects.filter(
                assignee_user=user,
                status='completed',
                task__organization=user.organization
            ).select_related('task', 'task__form').values_list('task__form_id', flat=True).distinct()
            
            task_form_ids = set(completed_task_trackings)
            
            # Get stage submission histories where user completed them
            # Exclude submissions for forms that are part of tasks
            stage_submission_histories = StageSubmissionHistory.objects.filter(
                completed_by=user,
                organization=user.organization
            ).exclude(
                form_submission__form_id__in=task_form_ids
            ).select_related('form_submission__form', 'stage', 'form_submission').distinct()

            for history in stage_submission_histories:
                form = history.form_submission.form
                form_id = form.id
                submission = history.form_submission
                
                # Skip if already in forms_dict from StageAssignment
                if form_id not in forms_dict:
                    form_data = CompactFormSerializer(
                        form, many=False, context={'request': request}
                    ).data
                    forms_dict[form_id] = {
                        'id': form_data['id'],
                        'title': form_data.get('title', ''),
                        'form_type': form_data.get('form_type', 'standard'),
                        'sent': []
                    }
                
                # Check if this submission is already in the list
                submission_exists = any(
                    s.get('form_submission_id') == submission.id 
                    for s in forms_dict[form_id]['sent']
                )
                
                if not submission_exists:
                    sub_source, sub_source_ref = get_submission_source(submission)
                    forms_dict[form_id]['sent'].append({
                        'stage_id': history.stage.id if history.stage else None,
                        'stage_order': history.stage_order,
                        'stage_assignment_id': None,
                        'assignment_uuid': history.stage_assignment_uuid,
                        'form_submission_id': submission.id,
                        'is_stage_submission_pending': False,
                        'is_form_submission_pending': not submission.is_completed,
                        'submission_initiated_on': submission.submission_initiated_on.isoformat() if submission.submission_initiated_on else None,
                        'is_completed': submission.is_completed,
                        'completed_by': submission.completed_by.id if submission.completed_by else None,
                        'completed_on': submission.completed_on.isoformat() if submission.completed_on else None,
                        'id': submission.id,
                        'source': sub_source,
                        'source_ref': sub_source_ref,
                    })

            # Get audit form submission histories where user completed them
            # Exclude submissions for forms that are part of tasks
            audit_submission_histories = AuditFormSubmissionHistory.objects.filter(
                completed_by=user,
                organization=user.organization
            ).exclude(
                form_submission__form_id__in=task_form_ids
            ).select_related('form_submission__form', 'form_submission').distinct()

            for history in audit_submission_histories:
                form = history.form_submission.form
                form_id = form.id
                submission = history.form_submission
                
                if form_id not in forms_dict:
                    form_data = CompactFormSerializer(
                        form, many=False, context={'request': request}
                    ).data
                    forms_dict[form_id] = {
                        'id': form_data['id'],
                        'title': form_data.get('title', ''),
                        'form_type': form_data.get('form_type', 'standard'),
                        'sent': []
                    }
                
                # Check if this submission is already in the list
                submission_exists = any(
                    s.get('form_submission_id') == submission.id 
                    for s in forms_dict[form_id]['sent']
                )
                
                if not submission_exists:
                    sub_source, sub_source_ref = get_submission_source(submission)
                    forms_dict[form_id]['sent'].append({
                        'stage_id': None,
                        'stage_order': None,
                        'stage_assignment_id': None,
                        'assignment_uuid': history.group_assignment_uuid,
                        'form_submission_id': submission.id,
                        'is_stage_submission_pending': False,
                        'is_form_submission_pending': not submission.is_completed,
                        'submission_initiated_on': submission.submission_initiated_on.isoformat() if submission.submission_initiated_on else None,
                        'is_completed': submission.is_completed,
                        'completed_by': submission.completed_by.id if submission.completed_by else None,
                        'completed_on': submission.completed_on.isoformat() if submission.completed_on else None,
                        'source': sub_source,
                        'source_ref': sub_source_ref,
                    })

            response_data = list(forms_dict.values())
            return Response(response_data, status=status.HTTP_200_OK)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error retrieving sent forms for user {user_id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class UserReceivedFormsMobileView(APIView):  # Still using APIView to isolate userContextAPIView
    permission_classes = [IsAdmin, IsEndUser]

    def get(self, request, user_id=None):
        logger.debug(f"[UserReceivedFormsMobileView] Starting for user_id={user_id}, request.user={request.user.id}")
        try:
            if not user_id:
                logger.error("[UserReceivedFormsMobileView] User ID not provided")
                raise ValidationError("User Id Required.")
                            
            logger.debug(f"[UserReceivedFormsMobileView] Fetching user with id={user_id}, org={request.user.organization_id}")
            user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)
            logger.debug(f"[UserReceivedFormsMobileView] User fetched: id={user.id}")

            logger.debug("[UserReceivedFormsMobileView] Building StageAssignment query")
            stage_assignments_query = StageAssignment.objects.filter(
                user=user,
                form__is_deleted=False,
                form__is_archived=False,
                form__organization=user.organization
            ).select_related('form', 'stage', 'form_submission').distinct().order_by(
                '-form_submission__completed_on',
                '-form_submission__submission_initiated_on',
                '-id'
            )
            logger.debug(f"[UserReceivedFormsMobileView] Query built, count={stage_assignments_query.count()}")

            # Pre-check all assignments
            for assignment in stage_assignments_query:
                logger.debug(f"[UserReceivedFormsMobileView] Pre-check StageAssignment ID={assignment.id}, form_submission={assignment.form_submission}, form_submission_id={getattr(assignment, 'form_submission_id', None)}")
                if assignment.form_submission:
                    try:
                        logger.debug(f"[UserReceivedFormsMobileView] Pre-check is_completed for ID={assignment.id}: {assignment.form_submission.is_completed}")
                    except AttributeError as e:
                        logger.error(f"[UserReceivedFormsMobileView] Pre-check error for StageAssignment ID={assignment.id}: {str(e)}")
            
            # Group assignments by form
            forms_data = {}
            for assignment in stage_assignments_query:
                logger.debug(f"[UserReceivedFormsMobileView] Processing StageAssignment ID={assignment.id}, form_id={assignment.form.id}, form_submission={assignment.form_submission}")
                form_id = assignment.form.id
                if form_id not in forms_data:
                    forms_data[form_id] = {
                        'form': {
                            'id': form_id,
                            'title': assignment.form.title,
                            'form_type': assignment.form.form_type or 'standard'
                        },
                        'submissions': []
                    }
                    logger.debug(f"[UserReceivedFormsMobileView] Created form entry for form_id={form_id}")

                # Safely check is_completed
                is_form_submission_pending = True
                if assignment.form_submission:
                    try:
                        is_form_submission_pending = not assignment.form_submission.is_completed
                        logger.debug(f"[UserReceivedFormsMobileView] StageAssignment ID={assignment.id}: is_completed={assignment.form_submission.is_completed}")
                    except AttributeError as e:
                        logger.error(f"[UserReceivedFormsMobileView] Error for StageAssignment ID={assignment.id}: {str(e)}")
                        is_form_submission_pending = True
                
                forms_data[form_id]['submissions'].append({
                    'stage_id': assignment.stage.id,
                    'stage_order': assignment.stage.order,
                    'stage_assignment_id': assignment.id,
                    'assignment_uuid': assignment.assignment_uuid,
                    'form_submission_id': assignment.form_submission.id if assignment.form_submission else None,
                    'is_stage_submission_pending': not assignment.is_assignment_fullfilled,
                    'is_form_submission_pending': is_form_submission_pending
                })
                logger.debug(f"[UserReceivedFormsMobileView] Added submission for StageAssignment ID={assignment.id}")
            
            logger.debug(f"[UserReceivedFormsMobileView] Returning {len(forms_data)} forms")
            return Response(list(forms_data.values()), status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"[UserReceivedFormsMobileView] Error retrieving received forms for user {user_id or request.user.id}: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            
                
# Retrieves all folders belonging to the logged-in user's organization - Admin Side.        
class OrganizationFoldersView(userContextAPIView, APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        try:
            user = request.user
            # Fetch folders for the user's organization
            folders = Folder.objects.filter(organization=user.organization).select_related('created_by').distinct()
            
            # Serialize folder data using FolderSerializer
            serializer = FolderSerializer(folders, many=True, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Error retrieving folders for organization {user.organization.id}: {str(e)}")
            return Response({"error": f"Failed to retrieve folders: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)        

# Retrieves forms explicitly stored in a specified folder for the logged-in user's organization - Admin Side
class FormsInFolderView(userContextAPIView, APIView):
    permission_classes = [IsAdmin] 

    def get(self, request, folder_id=None):
        try:
            if not folder_id:
                raise ValidationError("Folder ID is required.")
            
            user = request.user
            # Verify folder belongs to user's organization
            folder = get_object_or_404(Folder, id=folder_id, organization=user.organization)
            latest_payload_status = FormPayloadFiles.objects.filter(
                form_id=OuterRef('pk')
            ).order_by('-created_at').values('status')[:1]
            
            # Fetch direct child folders for nested folder navigation (mobile/web parity)
            subfolders = Folder.objects.filter(
                parent=folder,
                organization=user.organization
            ).select_related('created_by').distinct()

            # Fetch forms explicitly stored in the specified folder with annotated response_count
            forms = Form.objects.filter(
                folder=folder,
                organization=user.organization,
                is_deleted=False
            ).annotate(
                response_count=Subquery(
                    FormSubmision.objects.filter(
                        form_id=OuterRef('pk')
                    ).filter(
                        Q(submission_initiated_stage__isnull=False) |
                        Q(group_submissions_history__isnull=False) |
                        Q(stage_submissions_history__isnull=False)
                    ).values('form_id').annotate(
                        c=Count('pk', distinct=True)
                    ).values('c')[:1],
                    output_field=IntegerField()
                ),
                status=Subquery(latest_payload_status)
            ).distinct()
            
            # Serialize both subfolders and forms from the same folder level
            folder_serializer = FolderSerializer(folder, context={'request': request})
            subfolder_serializer = FolderSerializer(subfolders, many=True, context={'request': request})
            form_serializer = CompactFormSerializer(forms, many=True, context={'request': request})

            return Response({
                "folder": folder_serializer.data,
                "subfolders": subfolder_serializer.data,
                "forms": form_serializer.data
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Error retrieving forms for folder {folder_id or 'unknown'} in organization {user.organization.id}: {str(e)}")
            return Response({"error": f"Failed to retrieve forms: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


class StageSubmissionHistoryView(APIView):
    # permission_classes = [IsEndUserOrAdmin]
    permission_classes = [AllowAny]

    def get(self, request, user_id=None):
        try:
            if not user_id:
                raise ValidationError("User Id Required.")

            user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)

            # Get stage submission histories - order by completed_on descending
            stage_submission_histories = StageSubmissionHistory.objects.filter(
                completed_by=user, 
                organization=user.organization
            ).select_related(
                'form_submission__form',
                'stage',
                'completed_by',
                'organization'
            ).order_by('-completed_on').distinct()

            # Get audit form submission histories - order by completed_on descending
            audit_submission_histories = AuditFormSubmissionHistory.objects.filter(
                completed_by=user,
                organization=user.organization
            ).select_related(
                'form_submission__form',
                'group_id',
                'completed_by',
                'organization'
            ).order_by('-completed_on').distinct()

            # Data structure: { form_id: { 'form': {...}, 'submissions': { submission_id: {...} } } }
            form_data = {}

            # Process StageSubmissionHistory records
            for history in stage_submission_histories:
                form = history.form_submission.form
                form_id = str(form.id)
                submission_id = history.form_submission.id

                if form_id not in form_data:
                    form_data[form_id] = {
                        "form": {
                            "id": form_id,
                            "title": form.title,
                            "form_type": form.form_type or "standard",
                            "prefix": getattr(form, 'prefix', None) or ''
                        },
                        "submissions": {}
                    }

                # Ensure this submission_id exists in submissions dict
                if submission_id not in form_data[form_id]["submissions"]:
                    form_data[form_id]["submissions"][submission_id] = {
                        "is_form_submission_pending": getattr(history.form_submission, "is_form_submission_pending", None),
                        "id": str(submission_id),
                        "submission_initiated_on": history.form_submission.submission_initiated_on.isoformat() if history.form_submission.submission_initiated_on else None,
                        "submission_initiated_stage": history.stage.id if history.stage else None,
                        "submission_initiated_by": history.form_submission.submission_initiated_by_id if hasattr(history.form_submission, "submission_initiated_by_id") else None,
                        "form_submission_id": submission_id,
                        "is_completed": history.form_submission.is_completed,
                        "completed_by": history.completed_by.id if history.completed_by else None,
                        "completed_on": history.form_submission.completed_on.isoformat() if history.form_submission.completed_on else None,
                        "history_type": "stage_submission",
                        "stage_assignment_uuid": history.stage_assignment_uuid,
                        "stage_order": history.stage_order,
                        "stage_id": history.stage.id if history.stage else None,
                        "stage_name": history.stage.name if history.stage else None,
                        # Add timestamp for sorting
                        "_sort_timestamp": history.completed_on.timestamp() if history.completed_on else history.form_submission.completed_on.timestamp() if history.form_submission.completed_on else 0
                    }

            # Process AuditFormSubmissionHistory records
            for history in audit_submission_histories:
                form = history.form_submission.form
                form_id = str(form.id)
                submission_id = history.form_submission.id

                if form_id not in form_data:
                    form_data[form_id] = {
                        "form": {
                            "id": form_id,
                            "title": form.title,
                            "form_type": form.form_type or "standard",
                            "prefix": getattr(form, 'prefix', None) or ''
                        },
                        "submissions": {}
                    }

                # If submission already exists from stage history, update with audit info
                if submission_id in form_data[form_id]["submissions"]:
                    # Merge the data and update sort timestamp if audit is newer
                    existing_timestamp = form_data[form_id]["submissions"][submission_id].get("_sort_timestamp", 0)
                    audit_timestamp = history.completed_on.timestamp() if history.completed_on else 0

                    form_data[form_id]["submissions"][submission_id].update({
                        "history_type": "both",  # Indicates both types exist
                        "group_assignment_uuid": history.group_assignment_uuid,
                        "audit_completed_on": history.completed_on.isoformat() if history.completed_on else None,
                        # Add overall audit form fields
                        "form_overall_status": history.form_overall_status,
                        "form_overall_score": str(history.form_overall_score) if history.form_overall_score is not None else None,
                        "form_critical_failed": history.form_critical_failed,
                        # Update timestamp if audit completion is newer
                        "_sort_timestamp": max(existing_timestamp, audit_timestamp)
                    })

                    # Initialize summarydata if not exists
                    if "summarydata" not in form_data[form_id]["submissions"][submission_id]:
                        form_data[form_id]["submissions"][submission_id]["summarydata"] = []
                else:
                    # Create new submission entry for audit history
                    form_data[form_id]["submissions"][submission_id] = {
                        "is_form_submission_pending": getattr(history.form_submission, "is_form_submission_pending", None),
                        "id": str(submission_id),
                        "submission_initiated_on": history.form_submission.submission_initiated_on.isoformat() if history.form_submission.submission_initiated_on else None,
                        "submission_initiated_by": history.form_submission.submission_initiated_by_id if hasattr(history.form_submission, "submission_initiated_by_id") else None,
                        "form_submission_id": submission_id,
                        "is_completed": history.form_submission.is_completed,
                        "completed_by": history.completed_by.id if history.completed_by else None,
                        "completed_on": history.form_submission.completed_on.isoformat() if history.form_submission.completed_on else None,
                        "history_type": "audit_submission",
                        "group_assignment_uuid": history.group_assignment_uuid,
                        "audit_completed_on": history.completed_on.isoformat() if history.completed_on else None,
                        "summarydata": [],  # Initialize summarydata array
                        # Add timestamp for sorting
                        "_sort_timestamp": history.completed_on.timestamp() if history.completed_on else history.form_submission.completed_on.timestamp() if history.form_submission.completed_on else 0
                    }

                # Add audit summary data to submission level (only group-specific data)
                form_data[form_id]["submissions"][submission_id]["summarydata"].append({
                    "group_assignment_uuid": history.group_assignment_uuid,
                    "groups_status": history.groups_status,
                    "group_score": str(history.group_score) if history.group_score is not None else None,
                    "group_percentage": history.group_percentage,
                    "group_name": history.group_id.name if history.group_id else None,
                    "group_order": history.group_id.order if history.group_id else None,
                    "completed_by_username": history.completed_by.username if history.completed_by else None,
                    "completed_by_email": history.completed_by.email if history.completed_by else None,
                    "completed_on": history.completed_on.isoformat() if history.completed_on else None,
                })

                # Sort summarydata by group_order after all entries are added
                form_data[form_id]["submissions"][submission_id]["summarydata"].sort(
                    key=lambda x: x.get("group_order", 999)
                )

            # Include planner-based completed submissions (collaborative audits + regular planner)
            from planner.models import PlannerSubmission as _PS2, CollaborativeSubmission as _CS2
            planner_subs = _PS2.objects.filter(
                submitted_by=user,
                form_submission__is_completed=True,
            ).select_related('form_submission__form', 'planner_assignment').order_by('-form_submission__completed_on')

            for ps in planner_subs:
                submission = ps.form_submission
                if not submission:
                    continue
                form = submission.form
                form_id = str(form.id)
                submission_id = submission.id

                # Skip if already present from stage/audit history
                if form_id in form_data and submission_id in form_data[form_id]["submissions"]:
                    continue

                if form_id not in form_data:
                    form_data[form_id] = {
                        "form": {
                            "id": form_id,
                            "title": form.title,
                            "form_type": form.form_type or "standard",
                            "prefix": getattr(form, 'prefix', None) or ''
                        },
                        "submissions": {}
                    }

                is_collab = _CS2.objects.filter(form_submission=submission).exists()
                form_data[form_id]["submissions"][submission_id] = {
                    "id": str(submission_id),
                    "submission_initiated_on": submission.submission_initiated_on.isoformat() if submission.submission_initiated_on else None,
                    "submission_initiated_by": submission.submission_initiated_by_id,
                    "form_submission_id": submission_id,
                    "is_completed": True,
                    "completed_by": submission.completed_by_id,
                    "completed_on": submission.completed_on.isoformat() if submission.completed_on else None,
                    "history_type": "collaborative" if is_collab else "planner",
                    "submission_type": "[Collaborative-Audit]" if is_collab else "[Planner]",
                    "source": "planner",
                    "source_ref": ps.planner_assignment.order_id if ps.planner_assignment else None,
                    "summarydata": [],
                    "_sort_timestamp": submission.completed_on.timestamp() if submission.completed_on else submission.submission_initiated_on.timestamp() if submission.submission_initiated_on else 0,
                }

            # Collect all submission IDs from the grouped data once
            all_sub_ids = {
                sub_id
                for form_entry in form_data.values()
                for sub_id in form_entry["submissions"].keys()
            }

            # Batch-resolve source (Planner / Task / Form) per submission
            # Use exact submission-level relationships so source_ref is dynamic
            # per submission instead of a static per-form/user mapping.
            from planner.models import PlannerAssignment as _PA, PlannerSubmission as _PS
            from form.models import FollowUpTask as _FUT

            planner_source_map = {}  # submission_id -> planner order_id
            for ps in _PS.objects.filter(
                form_submission_id__in=all_sub_ids
            ).select_related('planner_assignment'):
                order_id = ps.planner_assignment.order_id if ps.planner_assignment else None
                if order_id is not None:
                    planner_source_map[ps.form_submission_id] = order_id

            task_source_map = {}
            for fut in _FUT.objects.filter(
                followup_submission_id__in=all_sub_ids
            ).select_related('task_details'):
                task_id = fut.task_details_id if fut.task_details else None
                if task_id is not None:
                    task_source_map[fut.followup_submission_id] = task_id

            submissions_by_id = {
                submission.id: submission
                for submission in FormSubmision.objects.filter(
                    id__in=all_sub_ids,
                    organization=user.organization
                )
            }
            completion_logs = list(
                TaskAuditLog.objects.filter(
                    task_action__iexact='Followup_Completed',
                    task__form_id__in={submission.form_id for submission in submissions_by_id.values()},
                    task__organization=user.organization
                ).select_related('task').order_by('action_date_time')
            )
            creation_logs_by_task = {
                log.task_id: log
                for log in TaskAuditLog.objects.filter(
                    task_id__in={log.task_id for log in completion_logs},
                    task_action__iexact='Followup_Created'
                ).order_by('task_id', 'action_date_time')
            }
            root_planner_source_map = {}
            root_submissions = {}
            for sub_id, submission in submissions_by_id.items():
                completion_candidates = [
                    log for log in completion_logs
                    if log.task.form_id == submission.form_id
                    and log.action_by_id == submission.completed_by_id
                ]
                if not completion_candidates:
                    continue
                submission_time = submission.completed_on or submission.submission_initiated_on
                completion_log = min(
                    completion_candidates,
                    key=lambda log: abs((log.action_date_time - submission_time).total_seconds())
                    if submission_time else 0
                )
                creation_log = creation_logs_by_task.get(completion_log.task_id)
                task = completion_log.task
                if not creation_log or not task.followup_task_form_id or not creation_log.action_by_id:
                    continue
                root_candidates = FormSubmision.objects.filter(
                    form_id=task.followup_task_form_id_id,
                    submission_initiated_by_id=creation_log.action_by_id,
                    organization=user.organization,
                    submission_initiated_on__lte=creation_log.action_date_time
                ).order_by('-submission_initiated_on')
                root_submission = root_candidates.first()
                if root_submission:
                    root_submissions[sub_id] = root_submission.id

            root_submission_ids = set(root_submissions.values())
            planner_submission_ids = set(planner_source_map)
            if root_submission_ids:
                for ps in _PS.objects.filter(
                    form_submission_id__in=root_submission_ids
                ).select_related('planner_assignment'):
                    if ps.planner_assignment and ps.planner_assignment.order_id is not None:
                        for task_submission_id, root_submission_id in root_submissions.items():
                            if root_submission_id == ps.form_submission_id:
                                root_planner_source_map[task_submission_id] = ps.planner_assignment.order_id

            planner_candidates = list(
                _PA.objects.filter(
                    organization=user.organization,
                    is_completed=True,
                    completed_by_id__isnull=False,
                    form_id__in={submission.form_id for submission in submissions_by_id.values()}
                ).exclude(order_id__isnull=True).values(
                    'form_id', 'completed_by_id', 'completed_on', 'order_id'
                )
            )

            def get_completed_planner_order(submission):
                submission_time = submission.completed_on or submission.submission_initiated_on
                if not submission_time:
                    return None
                matches = [
                    planner for planner in planner_candidates
                    if planner['form_id'] == submission.form_id
                    and planner['completed_by_id'] == submission.submission_initiated_by_id
                    and planner['completed_on']
                ]
                if not matches:
                    return None
                planner = min(
                    matches,
                    key=lambda candidate: abs((candidate['completed_on'] - submission_time).total_seconds())
                )
                if abs((planner['completed_on'] - submission_time).total_seconds()) <= 60:
                    return planner['order_id']
                return None

            for sub_id, submission in submissions_by_id.items():
                if sub_id not in planner_submission_ids:
                    planner_order_id = get_completed_planner_order(submission)
                    if planner_order_id is not None:
                        planner_source_map[sub_id] = planner_order_id

            for task_submission_id, root_submission_id in root_submissions.items():
                if task_submission_id not in root_planner_source_map:
                    root_submission = FormSubmision.objects.filter(id=root_submission_id).first()
                    if root_submission:
                        planner_order_id = planner_source_map.get(root_submission_id) or get_completed_planner_order(root_submission)
                        if planner_order_id is not None:
                            root_planner_source_map[task_submission_id] = planner_order_id

            for form_entry in form_data.values():
                for submission in form_entry["submissions"].values():
                    sub_id = submission.get("form_submission_id") or submission.get("id")
                    if sub_id is not None:
                        try:
                            sub_id = int(sub_id)
                        except (ValueError, TypeError):
                            sub_id = None
                    if sub_id is not None and sub_id in planner_source_map:
                        submission["source"] = "planner"
                        submission["source_ref"] = planner_source_map[sub_id]
                    elif sub_id is not None and sub_id in root_planner_source_map:
                        submission["source"] = "planner"
                        submission["source_ref"] = root_planner_source_map[sub_id]
                    elif sub_id is not None and sub_id in task_source_map:
                        submission["source"] = "task"
                        submission["source_ref"] = task_source_map[sub_id]
                    else:
                        submission["source"] = "form"
                        submission["source_ref"] = None

            # ── Checkpoint summary (audit answers per submission) ──────────────
            from form.models import Answer as _Answer, Option as _Option
            # all_sub_ids already collected above before source resolution
            # Fetch all audit answers for these submissions in one query
            audit_answers = _Answer.objects.filter(
                submission_id__in=all_sub_ids,
                question_type='audit'
            ).values('submission_id', 'answer')

            # Batch fetch option texts for all audit answer values
            audit_option_ids = {int(a['answer']) for a in audit_answers if a['answer']}
            option_text_map = {}  # option_id -> option text
            for opt in _Option.objects.filter(id__in=audit_option_ids).values('id', 'option'):
                option_text_map[opt['id']] = opt['option']

            # Build checkpoint counts per submission
            checkpoint_by_sub = {}  # sub_id -> {ok, not_ok_corrected, not_ok_not_closed, total}
            for a in audit_answers:
                sid = a['submission_id']
                opt_text = option_text_map.get(int(a['answer']) if a['answer'] else 0, '') if a['answer'] else ''
                if sid not in checkpoint_by_sub:
                    checkpoint_by_sub[sid] = {'ok': 0, 'not_ok_corrected': 0, 'not_ok_not_closed': 0, 'total': 0}
                checkpoint_by_sub[sid]['total'] += 1
                lower = opt_text.lower()
                if 'not ok' in lower and 'corrected' in lower:
                    checkpoint_by_sub[sid]['not_ok_corrected'] += 1
                elif 'not ok' in lower or 'not going to close' in lower:
                    checkpoint_by_sub[sid]['not_ok_not_closed'] += 1
                elif 'ok' in lower:
                    checkpoint_by_sub[sid]['ok'] += 1

            # ── Followup task summary via Task + TaskAuditLog ────────────────────────────────
            # Tasks link to submissions via TaskAuditLog(action='Followup_Created').
            # For each submission, find tasks where the Followup_Created log's action_by
            # matches the submission's initiator and the form matches followup_task_form_id.
            from task.models import Task as _Task, TaskAuditLog as _TaskAuditLog

            followup_task_counts = {}  # sub_id -> {total, completed, in_progress, not_started, task_ids}

            # Build a map of all submissions: {sub_id -> submission_obj} for lookup
            all_sub_objs = FormSubmision.objects.filter(id__in=all_sub_ids).select_related('form')
            sub_map = {s.id: s for s in all_sub_objs}

            # Get all followup tasks (tasks with followup_task_form_id set) for these submissions' forms
            form_ids_in_view = list({s.form_id for s in all_sub_objs})
            followup_tasks = _Task.objects.filter(
                followup_task_form_id_id__in=form_ids_in_view
            ).values('id', 'status', 'followup_task_form_id_id')

            # Get TaskAuditLog entries for Followup_Created for these tasks
            task_ids_in_view = [t['id'] for t in followup_tasks]
            audit_logs = _TaskAuditLog.objects.filter(
                task_id__in=task_ids_in_view,
                task_action__iexact='Followup_Created'
            ).values('task_id', 'action_by_id', 'action_date_time').order_by('action_date_time')

            # Map task_id -> first Followup_Created log
            task_log_map = {}
            for log in audit_logs:
                if log['task_id'] not in task_log_map:
                    task_log_map[log['task_id']] = log

            # For each task, find the matching submission
            for task in followup_tasks:
                tid = task['id']
                st = task['status']
                form_id = task['followup_task_form_id_id']
                log = task_log_map.get(tid)
                if not log:
                    continue
                action_by_id = log['action_by_id']
                action_dt = log['action_date_time']
                # Find the submission: same form, initiated by same user, before task creation
                matched_sub = None
                for sub in all_sub_objs:
                    if sub.form_id != form_id:
                        continue
                    if sub.submission_initiated_by_id != action_by_id:
                        continue
                    if action_dt and sub.submission_initiated_on and sub.submission_initiated_on > action_dt:
                        continue
                    if matched_sub is None or sub.submission_initiated_on > matched_sub.submission_initiated_on:
                        matched_sub = sub
                if not matched_sub:
                    continue
                sid = matched_sub.id
                if sid not in followup_task_counts:
                    followup_task_counts[sid] = {
                        'total': 0, 'completed': 0, 'in_progress': 0, 'not_started': 0,
                        'task_ids': {'completed': [], 'in_progress': [], 'not_started': []}
                    }
                followup_task_counts[sid]['total'] += 1
                if st in ('completed', 'done'):
                    followup_task_counts[sid]['completed'] += 1
                    followup_task_counts[sid]['task_ids']['completed'].append(tid)
                elif st in ('in_progress', 'started', 'pending'):
                    followup_task_counts[sid]['in_progress'] += 1
                    followup_task_counts[sid]['task_ids']['in_progress'].append(tid)
                else:
                    followup_task_counts[sid]['not_started'] += 1
                    followup_task_counts[sid]['task_ids']['not_started'].append(tid)

            # Apply summaries to each submission
            for form_entry in form_data.values():
                for sub_id, submission in form_entry["submissions"].items():
                    cp = checkpoint_by_sub.get(sub_id, {
                        'ok': 0, 'not_ok_corrected': 0, 'not_ok_not_closed': 0, 'total': 0
                    })
                    submission["checkpoint_summary"] = cp
                    # Use real Task records matched via TaskAuditLog; empty if no tasks exist
                    submission["followup_tasks_summary"] = followup_task_counts.get(sub_id, {
                        'total': 0, 'completed': 0, 'in_progress': 0, 'not_started': 0,
                        'task_ids': {'completed': [], 'in_progress': [], 'not_started': []}
                    })

            # Fetch location answers for all submissions in one query
            from form.models import Answer
            submission_ids = {
                sub_id
                for form_entry in form_data.values()
                for sub_id in form_entry["submissions"].keys()
            }
            location_map = {}
            if submission_ids:
                for answer in Answer.objects.filter(
                    submission_id__in=submission_ids,
                    question_type='location',
                    location__isnull=False
                ).select_related('location').order_by('id'):
                    if answer.submission_id not in location_map:
                        location_map[answer.submission_id] = answer.location.name

            # Apply location names to submission entries
            for form_entry in form_data.values():
                for submission_id, submission in form_entry["submissions"].items():
                    if submission_id in location_map:
                        submission["location_name"] = location_map[submission_id]

            # Convert nested dict to final array structure and sort
            result = []
            for form_entry in form_data.values():
                # Sort submissions within each form by timestamp descending
                submissions_list = sorted(
                    form_entry["submissions"].values(),
                    key=lambda x: x.get("_sort_timestamp", 0),
                    reverse=True
                )
                
                # Remove the temporary sort key from final response
                for submission in submissions_list:
                    submission.pop("_sort_timestamp", None)
                
                result.append({
                    "form": form_entry["form"],
                    "submissions": submissions_list
                })

            # Sort forms by the most recent submission timestamp descending
            result.sort(
                key=lambda form_entry: max(
                    [sub.get("completed_on") or sub.get("audit_completed_on") or sub.get("submission_initiated_on") or "0" 
                     for sub in form_entry["submissions"]],
                    default="0"
                ),
                reverse=True
            )

            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error retrieving submission history for user {user_id or request.user.id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class FormRecipientListView(APIView):
    def get(self, request, form_id):
        """
        Retrieve the list of users assigned to a specific form.
        """
        try:
            form = get_object_or_404(Form, id=form_id, is_deleted=False)
            assignments = FormAssignment.objects.filter(form_id=form_id, assign_type='user', user__isnull=False)
            recipients = []
            for assignment in assignments:
                user = assignment.user
                if user:
                    user = CustomUser.objects.get(id=assignment.user_id)
                    recipients.append({
                        'user_id': user.id,
                        'username': user.first_name + ' ' + user.last_name if user.first_name and user.last_name else user.username,
                        'department': user.department.name if user.department else None,
                        'designation': user.designation.name if user.designation else None,
                        'form_shared_on': assignment.form_shared_on
                    })
            return Response(recipients, status=status.HTTP_200_OK)
        except Form.DoesNotExist:
            return Response({"detail": "Form not found or is deleted/archived."}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError as e:
            logger.error(f"Attribute error: {e}")
            return Response({"message": "Invalid data in assignment or user model."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return Response({"message": "Unable to retrieve recipients"}, status=status.HTTP_400_BAD_REQUEST)

class LatestFormSubmissionView(APIView):
    permission_classes = [IsEndUserOrAdmin]

    def get(self, request, form_id):
        try:
            latest_submission = FormSubmision.objects.filter(form_id=form_id).latest('submission_initiated_on')
            serializer = FormSubmissionSerializer(latest_submission)
            return Response(serializer.data)
        except FormSubmision.DoesNotExist:
            return Response({"message": "No submissions found for this form."}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error retrieving latest form submission for form {form_id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, form_id):
        """
        Unshare a form from all users.
        """
        try:
            form = get_object_or_404(Form, id=form_id, is_deleted=False, is_archived=False)
            assignments = FormAssignment.objects.filter(form_id=form_id, assign_type='user')
            if not assignments.exists():
                return Response({"detail": "No assignments found for this form."}, status=status.HTTP_404_NOT_FOUND)
            assignments.delete()
            return Response({"detail": "Form unshared from all users successfully."}, status=status.HTTP_200_OK)
        except Form.DoesNotExist:
            return Response({"detail": "Form not found or is deleted/archived."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"message": "Unable to unshare form"}, status=status.HTTP_400_BAD_REQUEST)

class FormRecipientUnshareView(APIView):
    def delete(self, request, form_id, user_id):
        """
        Unshare a form from a specific user.
        """
        try:
            form = get_object_or_404(Form, id=form_id, is_deleted=False )
            assignment = get_object_or_404(FormAssignment, form_id=form_id, assign_type='user', user_id=user_id)
            assignment.delete()
            return Response({"detail": "Form unshared from user successfully."}, status=status.HTTP_200_OK)
        except Form.DoesNotExist:
            return Response({"detail": "Form not found or is deleted/archived."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"message": "Unable to unshare form"}, status=status.HTTP_400_BAD_REQUEST)


class OrganizationFormsView(APIView):
    """
    API endpoint to get list of forms for a specific organization by organization ID.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, organization_id):
        try:
            # Verify the organization exists and user has access to it
            if request.user.organization_id != organization_id:
                raise PermissionDenied("You don't have access to this organization.")

            # Get all non-deleted forms for the specified organization, including archived revisions.
            forms = Form.objects.filter(
                organization_id=organization_id,
                is_deleted=False
            ).order_by('-created_at')

            # Serialize the forms data
            serializer = FormListSerializer(forms, many=True, context={'request': request})

            return Response({
                'organization_id': organization_id,
                'forms_count': forms.count(),
                'forms': serializer.data
            }, status=status.HTTP_200_OK)

        except PermissionDenied as e:
            return Response({"error": str(e)}, status=status.HTTP_403_FORBIDDEN)
        except Exception as e:
            logger.error(f"Error retrieving forms for organization {organization_id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class OrganizationFormsViewOptimized(APIView):
    """
    OPTIMIZED VERSION of OrganizationFormsView
    
    Key Optimizations:
    - select_related() for ForeignKey fields (form_admin, deletedBy, folder, archivedBy)
    - Annotated counts for stages, audit_groups, questions, and submissions
    - Conditional annotation for question_count based on form_type
    - Reduced queries from 100+ per 50 forms to ~3-5 queries total
    
    Performance Impact:
    - Before: ~100-500+ queries for 50 forms (2+ queries per form)
    - After: ~3-5 queries regardless of number of forms
    - Expected response time: <2 seconds for 1000+ forms
    
    Maintains exact same payload structure as OrganizationFormsView
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, organization_id):
        try:
            # Verify the organization exists and user has access to it
            if request.user.organization_id != organization_id:
                raise PermissionDenied("You don't have access to this organization.")

            latest_success = FormPayloadFiles.objects.filter(form_id=OuterRef('pk'),status=FormPayloadFiles.Status.SUCCESS).order_by('-created_at').values('status')[:1]
            
            # Build optimized queryset
            forms = Form.objects.filter(
                organization_id=organization_id,
                is_deleted=False
            ).select_related(
                # Optimize ForeignKey lookups - load in same query
                'form_admin',      # for form_admin display
                'deletedBy',       # for deleted_by display
                'folder',          # for folder reference
                'archivedBy',      # for archived info
                'organization'     # for organization reference
            ).annotate(
                # Annotate counts to avoid N+1 queries
                stage_count=Count('stages', distinct=True),
                audit_group_count=Count('audit_group', distinct=True),
                response_count=Subquery(
                    FormSubmision.objects.filter(
                        form_id=OuterRef('pk')
                    ).filter(
                        Q(submission_initiated_stage__isnull=False) |
                        Q(group_submissions_history__isnull=False) |
                        Q(stage_submissions_history__isnull=False)
                    ).values('form_id').annotate(
                        c=Count('pk', distinct=True)
                    ).values('c')[:1],
                    output_field=IntegerField()
                ),
                
                # Conditional question count based on form_type
                # For AUDIT forms: count questions in audit_groups
                # For other forms: count questions in stages
                question_count=Case(
                    When(
                        form_type=FormType.AUDIT,
                        then=Count('audit_group__questions', distinct=True)
                    ),
                    default=Count('stages__questions', distinct=True),
                    output_field=IntegerField()
                ),
                status=Subquery(latest_success)
            ).order_by('-created_at')

            # Get count before serialization
            forms_count = forms.count()

            # ✅ Serialize with optimized serializer
            from .serializers import FormListSerializerOptimized
            serializer = FormListSerializerOptimized(forms, many=True, context={'request': request})

            return Response({
                'organization_id': organization_id,
                'forms_count': forms_count,
                'forms': serializer.data
            }, status=status.HTTP_200_OK)

        except PermissionDenied as e:
            return Response({"error": str(e)}, status=status.HTTP_403_FORBIDDEN)
        except Exception as e:
            logger.error(f"Error retrieving forms for organization {organization_id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        
class FormListViewSet(userContextAPIView, ModelViewSet):
    permission_classes = [IsEndUserOrAdmin]
    queryset = Form.objects.filter(is_deleted=False, is_disabled=False)  # Base queryset without is_archived filter


    def get_serializer_class(self):
        if self.action == 'list':
            return FormListSerializer
        return FormSerializer
    
    def list(self, request, *args, **kwargs):
        # Optional: You can add pagination here
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class FormEditViewSet(userContextAPIView, GenericAPIView):
    permission_classes = [IsEndUserOrAdmin]
    queryset = Form.objects.all()
    serializer_class = FormSerializer

    @transaction.atomic
    def put(self, request, form_id):
        print(f"=== FORM EDIT REQUEST for form_id: {form_id} ===")
        print(f"Request data keys: {request.data.keys()}")
        print(f"Form type: {request.data.get('form_type')}")
        print(f"Title: {request.data.get('title')}")

        UF = UtilsFunctions()
        
        parent_form = get_object_or_404(
            Form,
            id=form_id,
            organization=request.user.organization,
            is_deleted=False
        )

        serializer = self.get_serializer(data=request.data, context={'request': request, 'skip_unique_checks': True})
        if not serializer.is_valid():
            clean = UF.extract_error_message(serializer.errors)
            print("clean ::", clean)
            print("error Message ::", serializer.errors)
            raise ValidationError(serializer.errors)
        new_form = serializer.save()

        parent_form.is_archived = True
        parent_form.last_archived_date = timezone.now()
        parent_form.archivedBy = request.user
        parent_form.save()

        # Copy assignments from parent to new form efficiently
        assignments = (
            FormAssignment.objects
            .filter(form=parent_form)
            .select_related('user', 'group', 'leader')
        )
        if assignments:
            to_create = [
                FormAssignment(
                    assign_type=a.assign_type,
                    form=new_form,
                    user=a.user,
                    group=a.group,
                    leader=a.leader,
                    form_shared_on=timezone.now(),
                )
                for a in assignments
            ]
            FormAssignment.bulk_manager = FormAssignment.objects
            FormAssignment.bulk_manager.bulk_create(to_create, ignore_conflicts=True)

        # Remove old assignments
        assignments.delete()

        # Update all PlannerAssignments and Tasks pointing to the old archived
        # form (and any older archived versions with the same title) to point
        # to the new form, so planners/tasks always use the latest version.
        from planner.models import PlannerAssignment
        archived_form_ids = list(
            Form.objects.filter(
                title=parent_form.title,
                organization=parent_form.organization,
                is_archived=True,
                is_deleted=False
            ).values_list('id', flat=True)
        )
        PlannerAssignment.objects.filter(form_id__in=archived_form_ids).update(form=new_form)
        Task.objects.filter(form_id__in=archived_form_ids).update(form=new_form)
        Task.objects.filter(followup_task_form_id_id__in=archived_form_ids).update(followup_task_form_id=new_form)

        # Re-fetch the newly created form with an optimized prefetch tree for fast serialization
        stage_access_qs = StageAccess.objects.select_related('allow_user', 'allow_group', 'form', 'stage')
        base = (
            Form.objects
            .filter(pk=new_form.id)
            .select_related('folder', 'form_admin', 'deletedBy', 'archivedBy')
        )
        if new_form.form_type == FormType.AUDIT:
            optimized_qs = base.prefetch_related(
                'assignee__user', 'assignee__group', 'assignee__leader',
                'audit_info__questions',
                'audit_info__questions__options',
                'audit_info__questions__child_questions',
                'audit_info__questions__child_questions__options',
                'audit_info__questions__logic_parent_question__logic_questions__options',
                'audit_info__questions__logic_parent_question__follow_ups',
                'audit_info__questions__logic_parent_question__follow_ups__task_close_questions__options',
                'audit_group__questions',
                'audit_group__questions__options',
                'audit_group__questions__child_questions',
                'audit_group__questions__child_questions__options',
                'audit_group__questions__logic_parent_question__logic_questions__options',
                'audit_group__questions__logic_parent_question__follow_ups',
                'audit_group__questions__logic_parent_question__follow_ups__task_close_questions__options',
            )
        else:
            question_qs = (
                Question.objects
                .select_related('form', 'stage', 'parent_question')
                .prefetch_related(
                    'options',
                    'child_questions',
                    'child_questions__options',
                    'child_questions__child_questions',
                    'child_questions__child_questions__options',
                    'child_questions__logic_parent_question__logic_questions__options',
                    'child_questions__logic_parent_question__follow_ups',
                    'child_questions__logic_parent_question__follow_ups__task_close_questions__options',
                    'logic_parent_question__logic_questions__options',
                    'logic_parent_question__follow_ups',
                    'logic_parent_question__follow_ups__task_close_questions__options',
                )
            )
            stage_qs = (
                Stage.objects
                .select_related('form')
                .prefetch_related(
                    models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                    models.Prefetch('questions', queryset=question_qs),
                )
            )
            optimized_qs = base.prefetch_related(
                'assignee__user', 'assignee__group', 'assignee__leader',
                models.Prefetch('stages', queryset=stage_qs),
            )

        optimized_instance = get_object_or_404(optimized_qs)
        serializer_out = self.get_serializer(optimized_instance)
        return Response(serializer_out.data, status=status.HTTP_201_CREATED)

    def get(self, request, form_id):
        form = get_object_or_404(
            Form,
            id=form_id,
            organization=request.user.organization,
            is_deleted=False
        )
        serializer = self.get_serializer(form, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)
 

class FormToggleView(APIView):
    permission_classes = [IsAdmin]

    def patch(self, request, pk):
        form = get_object_or_404(Form, pk=pk, organization=request.user.organization, is_deleted=False)
        serializer = FormToggleSerializer(form, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Logic for 1: Allow Sharing
class FormSubmissionShareView(APIView):
    permission_classes = [IsEndUserOrAdmin]

    def post(self, request, form_id, submission_id):
        form = get_object_or_404(Form, id=form_id, organization=request.user.organization, share_response=True)
        submission = get_object_or_404(FormSubmision, id=submission_id, form=form)

        # Highlight: Check if user is last stage completer
        last_stage = form.stages.order_by('-order').first()
        last_completion = StageSubmissionHistory.objects.filter(
            form_submission=submission, stage=last_stage
        ).order_by('-completed_on').first()
        if not last_completion or last_completion.completed_by != request.user:
            raise PermissionDenied("Only last stage submitter can share.")

        # Highlight: Validate and create shares
        serializer = FormSubmissionShareSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        users = serializer.validated_data.get('users', [])
        groups = serializer.validated_data.get('groups', [])
        leaders = serializer.validated_data.get('location_leaders', [])

        shared_count = 0
        skipped_count = 0

        # Share to users with duplicate check
        for user_id in users:
            # Check if already shared to this user
            existing_share = FormResponseShare.objects.filter(
                form_submission=submission,
                shared_to_user_id=user_id,
                share_type='user'
            ).first()

            if existing_share:
                skipped_count += 1
            else:
                FormResponseShare.objects.create(
                    form_submission=submission,
                    shared_to_user_id=user_id,
                    shared_by=request.user,
                    share_type='user',
                    organization=request.user.organization
                )
                shared_count += 1

        # Share to groups with duplicate check
        for group_id in groups:
            # Check if already shared to this group
            existing_share = FormResponseShare.objects.filter(
                form_submission=submission,
                shared_to_group_id=group_id,
                share_type='group'
            ).first()

            if existing_share:
                skipped_count += 1
            else:
                FormResponseShare.objects.create(
                    form_submission=submission,
                    shared_to_group_id=group_id,
                    shared_by=request.user,
                    share_type='group',
                    organization=request.user.organization
                )
                shared_count += 1

        # Share to location leaders with duplicate check
        for leader_id in leaders:
            # Check if already shared to this leader
            existing_share = FormResponseShare.objects.filter(
                form_submission=submission,
                shared_to_leader_id=leader_id,
                share_type='location_leader'
            ).first()

            if existing_share:
                skipped_count += 1
            else:
                FormResponseShare.objects.create(
                    form_submission=submission,
                    shared_to_leader_id=leader_id,
                    shared_by=request.user,
                    share_type='location_leader',
                    organization=request.user.organization
                )
                shared_count += 1

        message = f"Response shared successfully to {shared_count} recipient(s)"
        if skipped_count > 0:
            message += f", {skipped_count} duplicate(s) skipped"

        return Response({"message": message}, status=status.HTTP_200_OK)


# views.py - Full FormResponseView with detailed comments and fix
class FormResponseView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, form_id, submission_id):
        # Handle anonymous users - remove organization filter
        organization = None if request.user.is_anonymous else request.user.organization

        # Progressive loading: ?groups=1,2 returns only those stage/group orders
        # ?group_id=5 filters by AuditGroup.id (for collaborative mode)
        groups_param = request.query_params.get('groups', None)
        group_id_param = request.query_params.get('group_id', None)
        requested_orders = None
        if groups_param:
            try:
                requested_orders = [int(g.strip()) for g in groups_param.split(',') if g.strip()]
            except (ValueError, TypeError):
                requested_orders = None
        elif group_id_param:
            try:
                requested_group_id = int(group_id_param.strip())
                ag = AuditGroup.objects.filter(id=requested_group_id).first()
                if ag:
                    requested_orders = [ag.order]
            except (ValueError, TypeError):
                pass

        if request.user.is_anonymous:
            form = get_object_or_404(Form, id=form_id)
            submissions = FormSubmision.objects.filter(id=submission_id)
        else:
            form = get_object_or_404(Form, id=form_id, organization=organization)
            submissions = FormSubmision.objects.filter(id=submission_id, organization=organization)
        
        # Initialize response data with form details
        response_data = FormSerializer(form, many=False, context={'request': request}).data

        # Progressive loading: filter stages/audit_groups by order if requested
        if requested_orders:
            if response_data.get('form_type') == 'audit':
                response_data['audit_group'] = [
                    g for g in response_data.get('audit_group', [])
                    if g.get('order') in requested_orders
                ]
            else:
                response_data['stages'] = [
                    s for s in response_data.get('stages', [])
                    if s.get('order') in requested_orders
                ]
        if submissions.exists():
            current_submission = submissions.first()
            submissions_data = FormSubmissionSerializer(current_submission, context={'request': request}).data
            completed_by_user = current_submission.completed_by
            if completed_by_user:
                full_name = f"{completed_by_user.first_name or ''} {completed_by_user.last_name or ''}".strip()
                submissions_data['completed_by_details'] = {
                    "id": completed_by_user.id,
                    "name": full_name or completed_by_user.username or completed_by_user.email or "Unknown"
                }
            else:
                submissions_data['completed_by_details'] = None
            response_data['submissionsDetail'] = submissions_data
        
        # Helper function to replace option ID with text
        def replace_option_id_with_text(answers_data, options):
            if not answers_data or not options:
                return
            if isinstance(answers_data, list):
                for item in answers_data:
                    replace_option_id_with_text(item, options)
                return
            if 'answer' in answers_data:
                answer_value = answers_data['answer']
                if answer_value:
                    answer_str = str(answer_value).strip()
                    if ',' in answer_str:
                        # Handle multiple options like "1,2,3"
                        ids = [id.strip() for id in answer_str.split(',') if id.strip()]
                        replaced = []
                        for id_val in ids:
                            if id_val.isdigit():
                                option_id = id_val
                                found = False
                                for opt in options:
                                    if opt and str(opt.get('id', '')).strip() == option_id:
                                        replaced.append(opt['option'])
                                        found = True
                                        break
                                if not found:
                                    replaced.append(id_val)  # keep original if not found
                            else:
                                replaced.append(id_val)  # already text
                        answers_data['answer'] = ', '.join(replaced)
                    elif answer_str.isdigit():
                        # Single option
                        option_id = answer_str
                        for opt in options:
                            if opt and str(opt.get('id', '')).strip() == option_id:
                                answers_data['answer'] = opt['option']
                                break
                    # else: already text, leave as is

        def normalize_answers(qs):
            if not qs.exists():
                return {}
            data = AnswerSerializer(qs, many=True).data
            return data[0] if len(data) == 1 else data

        # Highlight: Handle audit form type - UPDATED to include audit_info questions
        if response_data.get("form_type") == FormType.AUDIT:
            # Handle audit_info questions - NEW CODE
            if response_data.get("audit_info"):
                for audit_info_question in response_data["audit_info"].get("questions", []):
                    answers = Answer.objects.filter(
                        question=audit_info_question['id'],
                        submission=submission_id,
                        organization=organization
                    )
                    audit_info_question['answers'] = normalize_answers(answers)
                    replace_option_id_with_text(audit_info_question['answers'], audit_info_question.get('options'))

                    # Handle sub-questions for audit_info
                    for subQuestion in audit_info_question.get('sub_questions', []):
                        sub_answers = Answer.objects.filter(
                            question=subQuestion['id'],
                            submission=submission_id,
                            organization=organization
                        )
                        subQuestion['answers'] = normalize_answers(sub_answers)
                        replace_option_id_with_text(subQuestion['answers'], subQuestion.get('options'))

                    # Handle logic questions for audit_info
                    for logics in audit_info_question.get("logics", []):
                        for logicQuestion in logics.get("logic_questions", []):
                            logic_answer = Answer.objects.filter(
                                question=logicQuestion['id'],
                                submission=submission_id,
                                organization=organization
                            )
                            logicQuestion['answers'] = normalize_answers(logic_answer)
                            replace_option_id_with_text(logicQuestion['answers'], logicQuestion.get('options'))

            # Handle audit_group questions - existing logic, no change
            for audit_group in response_data.get("audit_group", []):
                for audit_group_question in audit_group.get("questions", []):
                    answers = Answer.objects.filter(
                        question=audit_group_question['id'],
                        submission=submission_id,
                        organization=organization
                    )
                    audit_group_question['answers'] = normalize_answers(answers)
                    replace_option_id_with_text(audit_group_question['answers'], audit_group_question.get('options'))

                    for logics in audit_group_question.get("logics", []):
                        for logicQuestion in logics.get("logic_questions", []):
                            logic_answer = Answer.objects.filter(
                                question=logicQuestion['id'],
                                submission=submission_id,
                                organization=organization
                            )
                            logicQuestion['answers'] = normalize_answers(logic_answer)
                            replace_option_id_with_text(logicQuestion['answers'], logicQuestion.get('options'))

                    for subQuestion in audit_group_question.get('sub_questions', []):
                        sub_answers = Answer.objects.filter(
                            question=subQuestion['id'],
                            submission=submission_id,
                            organization=organization
                        )
                        subQuestion['answers'] = normalize_answers(sub_answers)
                        replace_option_id_with_text(subQuestion['answers'], subQuestion.get('options'))
        else:
            # Highlight: Handle non-audit form stages - existing logic, no change
            for stage in response_data.get('stages', []):
                stage_histories = StageSubmissionHistory.objects.filter(
                    stage__id=stage['id'],
                    form_submission__id=submission_id,
                    organization=organization
                )
                
                historyData={}
                if stage_histories.exists():
                    stageHistory = stage_histories.first()
                    historyData = StageSubmissionHistorySerializer(stageHistory, many=False, context={'request': request}).data
                
                stage['is_completed']=stage_histories.exists()
                stage['completed_by']= historyData.get('completed_by', None) if historyData else None
                stage['completed_on']= historyData.get('completed_on') if historyData else None
                
                for question in stage.get('questions', []):
                    answers = Answer.objects.filter(
                        question=question['id'],
                        stage=stage['id'],
                        submission=submission_id,
                        organization=organization
                    )
                    question['answers'] = normalize_answers(answers)
                    replace_option_id_with_text(question['answers'], question.get('options'))

                    for subQuestion in question.get('sub_questions', []):
                        sub_answers = Answer.objects.filter(
                            question=subQuestion['id'],
                            stage=stage['id'],
                            submission=submission_id,
                            organization=organization
                        )
                        subQuestion['answers'] = normalize_answers(sub_answers)
                        replace_option_id_with_text(subQuestion['answers'], subQuestion.get('options'))

                    for logics in question.get("logics", []):
                        for logicQuestion in logics.get("logic_questions", []):
                            logic_answer = Answer.objects.filter(
                                question=logicQuestion['id'],
                                stage=stage['id'],
                                submission=submission_id,
                                organization=organization
                            )
                            logicQuestion['answers'] = normalize_answers(logic_answer)
                            replace_option_id_with_text(logicQuestion['answers'], logicQuestion.get('options'))
        
        # Highlight: Fixed logic for toggle 1 - Check if user has share access or is owner
        # Purpose: Restricts response view to only last completer or users/groups/leaders with share access
        # Modification: Corrected syntax by combining all conditions in Q objects to avoid positional/keyword mix
        # Original issue: Positional Q objects after keyword 'form_submission' caused the error
        if organization:
            submission = get_object_or_404(FormSubmision, id=submission_id, organization=organization)
        else:
            submission = get_object_or_404(FormSubmision, id=submission_id)
        # if not FormResponseShare.objects.filter(
        #     models.Q(form_submission=submission) &
        #     (models.Q(shared_to_user=request.user) | models.Q(shared_to_group__members=request.user) | models.Q(shared_to_leader=request.user))
        # ).exists() and submission.completed_by != request.user:
        #     raise PermissionDenied("No access to this response.")
        
        # Highlight: New logic for toggle 1 - Include history for display
        # Purpose: Shows SR name/date/time for all stages (stable history), and only last completer name at form level
        # Modification: Added custom fields to stage and form-level data
        for stage in response_data.get('stages', []):
            history = StageSubmissionHistory.objects.filter(stage_id=stage['id'], form_submission=submission).first()
            if history:
                stage['completed_by_sr'] = f"{history.completed_by.first_name} {history.completed_by.last_name}"  # Full name for stages
                stage['completed_date'] = history.completed_on.date()
                stage['completed_time'] = history.completed_on.time()
                # Highlight: Added edit history display
                # Purpose: Show edited_by and edited_on for toggle 3 re-editing tracking
                stage['edited_by_sr'] = f"{history.edited_by.first_name} {history.edited_by.last_name}" if history.edited_by else None
                stage['edited_date'] = history.edited_on.date() if history.edited_on else None
                stage['edited_time'] = history.edited_on.time() if history.edited_on else None
        
        # Highlight: Form-level last completer name
        # Purpose: Displays only last completer's name (not full history) as per requirement
        last_completer = submission.completed_by
        response_data['last_completer_name'] = f"{last_completer.first_name} {last_completer.last_name}" if last_completer else 'N/A'

        # Add audit summary (scores/percentages) for audit forms
        if form.form_type == 'audit':
            if organization:
                histories = AuditFormSubmissionHistory.objects.filter(
                    form_submission=submission,
                    organization=organization
                ).select_related('group_id', 'form_id', 'completed_by').order_by('group_id__order', 'id')
            else:
                histories = AuditFormSubmissionHistory.objects.filter(
                    form_submission=submission
                ).select_related('group_id', 'form_id', 'completed_by').order_by('group_id__order', 'id')

            summary = []
            for history in histories:
                summary.append({
                    "id": history.id,
                    "form_submission": history.form_submission_id,
                    "group_assignment_uuid": history.group_assignment_uuid,
                    "form_overall_status": history.form_overall_status,
                    "form_overall_score": str(history.form_overall_score) if history.form_overall_score is not None else None,
                    "form_critical_failed": history.form_critical_failed,
                    "groups_status": history.groups_status,
                    "group_score": str(history.group_score) if history.group_score is not None else None,
                    "form_id": history.form_id_id if history.form_id_id else form.id,
                    "form_title": history.form_id.title if history.form_id else form.title,
                    "group_id": history.group_id_id,
                    "group_name": history.group_id.name if history.group_id else None,
                    "group_uuid": history.group_id.group_uuid if history.group_id else None,
                    "group_order": history.group_id.order if history.group_id else None,
                    "completed_by_username": history.completed_by.username if history.completed_by else None,
                    "completed_by_email": history.completed_by.email if history.completed_by else None,
                    "completed_on": history.completed_on.isoformat() if history.completed_on else None,
                })

            response_data['summary'] = summary

        return Response(response_data, status=status.HTTP_200_OK)

def manage_followup_tasks_for_submission(form_id, main_form, main_submission, user, mobile_edited_logic_ids=None):
    """Re-evaluate follow-up logic for a submission and synchronize follow-up tasks.

    Deletes existing tasks whose LogicFollowUp condition is no longer met and
    creates tasks for newly triggered conditions. This is used by both the
    form answer edit endpoint and the follow-up trigger endpoint.
    """
    created_tasks = []

    # Build answer dictionary from current submission answers
    answer_dict = {
        answer.question_id: answer.answer
        for answer in Answer.objects.filter(
            submission=main_submission,
            organization=user.organization
        ).select_related('question')
    }

    logic_followups_query = LogicFollowUp.objects.filter(
        models.Q(form_id=form_id) | models.Q(audit_group__form_id=form_id),
        followup_toggle=True,
    ).filter(
        models.Q(assign_form__isnull=False) | models.Q(task_close_questions__isnull=False)
    ).distinct()

    if mobile_edited_logic_ids:
        logic_followups_query = logic_followups_query.exclude(id__in=mobile_edited_logic_ids)

    logic_followups_with_form = logic_followups_query.select_related(
        'logic', 'question', 'assign_form', 'audit_group'
    )

    # Use the same logic condition evaluator as TriggerFollowupTasksView
    trigger_view = TriggerFollowupTasksView()
    filtered_logic_followups = [
        logic_followup
        for logic_followup in logic_followups_with_form
        if trigger_view._check_logic_condition(logic_followup, answer_dict)
    ]

    # Determine which (question, assigned_form) pairs are currently triggered
    triggered_keys = {
        (logic_followup.question_id, logic_followup.assign_form_id)
        for logic_followup in filtered_logic_followups
    }

    # Delete stale follow-up tasks for this submission
    all_logic_question_ids = list(logic_followups_with_form.values_list('question_id', flat=True))
    submission_date = (
        main_submission.completed_on
        or main_submission.submission_initiated_on
        or timezone.now()
    )

    if all_logic_question_ids and submission_date:
        existing_tasks = Task.objects.filter(
            followup_task_form_id=main_form,
            organization=main_submission.organization,
            start_date=submission_date,
            follow_task_sub_question_id__in=all_logic_question_ids,
        )
        for task in existing_tasks:
            if (task.follow_task_sub_question_id, task.form_id) not in triggered_keys:
                task.delete()

    # Create tasks for currently triggered follow-up logic
    for logic_followup in filtered_logic_followups:
        start_date = submission_date
        deadline_days = getattr(logic_followup, 'deadline', 7) or 7
        end_date = submission_date + timezone.timedelta(days=deadline_days)
        assigned_form = logic_followup.assign_form

        # Prevent duplicate tasks for the same submission/question
        existing_task = Task.objects.filter(
            followup_task_form_id=main_form,
            organization=main_submission.organization,
            follow_task_sub_question=logic_followup.question,
            form=assigned_form,
            start_date=start_date,
        ).first()

        if existing_task:
            created_tasks.append({
                'task_id': existing_task.id,
                'logic_id': logic_followup.logic.id if logic_followup.logic else None,
                'title': existing_task.task_name,
                'assignee_count': existing_task.assignees.count(),
                'has_assigned_form': assigned_form is not None,
                'has_task_close_questions': TaskCloseQuestion.objects.filter(task=existing_task).exists(),
                'source': 'web_configured'
            })
            continue

        task = Task.objects.create(
            task_name=logic_followup.title or 'Follow-up Task',
            description=logic_followup.description or '',
            form=assigned_form,
            followup_task_form_id=main_form,
            follow_task_sub_question=logic_followup.question,
            organization=main_submission.organization,
            status='not_started',
            start_date=start_date,
            end_date=end_date,
            created_by=user,
        )

        all_assignee_users = []

        for user_id in (logic_followup.assign_user_ids or []):
            try:
                from user.models import CustomUser
                assignee_user = CustomUser.objects.get(id=user_id, organization=user.organization)
                TaskAssignee.objects.create(
                    task=task,
                    assigned_user=assignee_user,
                    assigned_date_time=timezone.now()
                )
                all_assignee_users.append(assignee_user)
            except CustomUser.DoesNotExist:
                continue

        for group_id in (logic_followup.assign_group_ids or []):
            try:
                from user.models import Groups
                group = Groups.objects.get(id=group_id, organization=user.organization)
                TaskAssignee.objects.create(
                    task=task,
                    assigned_group=group,
                    assigned_date_time=timezone.now()
                )
                all_assignee_users.extend(group.members.filter(organization=user.organization))
            except Groups.DoesNotExist:
                continue

        for leader_id in (logic_followup.assign_leader_ids or []):
            try:
                from user.models import CustomUser
                leader = CustomUser.objects.get(id=leader_id, organization=user.organization)
                TaskAssignee.objects.create(
                    task=task,
                    assigned_leader=leader,
                    assigned_date_time=timezone.now()
                )
                all_assignee_users.append(leader)
            except CustomUser.DoesNotExist:
                continue

        TaskAuditLog.objects.create(
            task=task,
            task_action='Followup_Created',
            action_by=user,
            action_to=None
        )

        has_task_close_questions = logic_followup.task_close_questions.exists()
        if has_task_close_questions:
            for question in logic_followup.task_close_questions.all():
                TaskCloseQuestion.objects.create(
                    task=task,
                    question=question,
                    created_by=user,
                    organization=user.organization
                )

        created_tasks.append({
            'task_id': task.id,
            'logic_id': logic_followup.logic.id if logic_followup.logic else None,
            'title': task.task_name,
            'assignee_count': len(all_assignee_users),
            'has_assigned_form': assigned_form is not None,
            'has_task_close_questions': has_task_close_questions,
            'source': 'web_configured'
        })

    return created_tasks


class FormAnswerEditView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def put(self, request):
        try:
            serializer = FormAnswerEditSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)
            form = serializer.validated_data['form_instance']
            submission = serializer.validated_data['submission_instance']
            stage = serializer.validated_data.get('stage_instance')
            answers_data = serializer.validated_data['answers']

            responses = []
            for answer_data in answers_data:
                question_id = answer_data.get('question')
                answer_qs = Answer.objects.filter(
                    question_id=question_id,
                    submission=submission,
                    organization=request.user.organization
                )
                if answer_qs.exists():
                    answer_instance = answer_qs.first()
                    # Update only answer content, preserve submitted_by and submitted_on
                    answer_instance.answer = answer_data.get('answer', answer_instance.answer)
                    answer_instance.division_id = answer_data.get('division', answer_instance.division_id)
                    answer_instance.sub_division_id = answer_data.get('sub_division', answer_instance.sub_division_id)
                    answer_instance.location_id = answer_data.get('location', answer_instance.location_id)
                    answer_instance.user_id = answer_data.get('user', answer_instance.user_id)
                    if 'other_text' in answer_data:
                        answer_instance.other_text = answer_data.get('other_text')
                    if 'remarks' in answer_data:
                        answer_instance.remarks = answer_data.get('remarks', answer_instance.remarks)
                    if 'approved_stages' in answer_data:
                        answer_instance.approved_stages = answer_data.get('approved_stages', answer_instance.approved_stages)
                    if 'signature' in answer_data:
                        answer_instance.signature = answer_data.get('signature', answer_instance.signature)
                    answer_instance.save()  # submitted_by and submitted_on unchanged
                    responses.append(AnswerSerializer(answer_instance).data)
                else:
                    raise ValidationError(f"No answer found for question {question_id}.")

            # Log edit in StageSubmissionHistory if stage is specified
            if stage:
                stage_history_qs = StageSubmissionHistory.objects.filter(
                    stage=stage,
                    form_submission=submission,
                    organization=request.user.organization
                )
                if stage_history_qs.exists():
                    stage_history = stage_history_qs.first()
                    stage_history.edited_by = request.user
                    stage_history.edited_on = timezone.now()
                    stage_history.save()

            # Preserve submission's completed_by and completed_on
            # (No update to submission fields)

            # Remove answers for logic questions that are no longer visible
            # after this edit so stale data does not reappear on later re-edits.
            _cleanup_stale_logic_answers(
                form, submission, answers_data, request.user.organization
            )

            # Sync planner location to reflect edited location answer
            _sync_planner_location_from_submission(
                form, submission, request.user, request.user.organization
            )

            # Re-evaluate follow-up tasks against the updated answers.
            # This removes stale tasks whose logic is no longer met and creates
            # any newly triggered follow-up tasks.
            manage_followup_tasks_for_submission(
                form_id=form.id,
                main_form=form,
                main_submission=submission,
                user=request.user
            )

            edited_by_name = f"{request.user.first_name or ''} {request.user.last_name or ''}".strip() or request.user.username

            return Response(
                {
                    'message': 'Answers updated successfully',
                    'answers': responses,
                    'form_submission_id': submission.id,
                    'edited_by': edited_by_name
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            logger.error(f"Error updating answers: {str(e)}")
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            ) 


class GetFormResponseView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Fetch all submissions for the user's organization with related shares prefetched
            submissions = FormSubmision.objects.filter(
                organization=request.user.organization
            ).prefetch_related('shares')  # Corrected to 'shares' instead of 'formresponse_share_set'

            # Filter submissions the user has access to
            accessible_submissions = []
            for submission in submissions:
                has_access = False
                # Only treat a submission as accessible when it was explicitly shared
                # to the logged-in user (or one of their groups/leader shares).
                # Do NOT include submissions merely because the user completed them;
                # sharing must be triggered via FormSubmissionShareView.
                for share in submission.shares.all():
                    if share.shared_to_user == request.user or \
                       (share.shared_to_group and request.user.groups.filter(id=share.shared_to_group.id).exists()) or \
                       (share.shared_to_leader and share.shared_to_leader == request.user):
                        has_access = True
                        break
                if has_access and (submission.form.share_response or submission.form.auto_share_response):
                    accessible_submissions.append(submission)

            if not accessible_submissions:
                return Response({"message": "No accessible submissions found."}, status=status.HTTP_200_OK)

            # Serialize the list of accessible submissions
            serializer = FormResponseSummarySerializer(accessible_submissions, many=True)
            response_data = serializer.data

            logger.info(f"Retrieved {len(response_data)} responses for user {request.user.id}")
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error retrieving responses for user {request.user.id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
 

class FormFolderUpdateView(APIView):
    permission_classes = [IsAdmin]

    def patch(self, request, pk):
        form = get_object_or_404(Form, pk=pk, organization=request.user.organization, is_deleted=False)
        
        folder_id = request.data.get('folder_id')
        
        if folder_id is None:
            form.folder = None
            form.save(update_fields=['folder'])
            return Response({"message": "Form folder removed successfully."}, status=status.HTTP_200_OK)

        folder = get_object_or_404(Folder, id=folder_id, organization=request.user.organization)
        
        form.folder = folder
        form.save(update_fields=['folder'])
        
        return Response({"message": "Form folder updated successfully."}, status=status.HTTP_200_OK)

class TaskCloseQuestionsView(APIView):
    """
    Handle task close questions for followup tasks.
    When followup has no assigned form, user answers task close questions.
    """
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _normalize_value(value):
        return str(value or "").strip().lower().replace(" ", "_")

    @staticmethod
    def _get_user_field_value(user, field):
        if field == "department":
            return getattr(getattr(user, "department", None), "name", "")
        if field == "division":
            return getattr(getattr(user, "division", None), "name", "")
        if field == "subdivision":
            return getattr(getattr(user, "subdivision", None), "name", "")
        if field == "location":
            return getattr(getattr(user, "location", None), "name", "")
        if field == "designation":
            return getattr(getattr(user, "designation", None), "name", "")
        return ""

    @staticmethod
    def _user_matches_group(group, user):
        if group.members.filter(id=user.id).exists():
            return True

        if getattr(group, "type", "") != "rulebased":
            return False

        conditions = group.conditions.all()
        if not conditions.exists():
            return False

        def eval_condition(cond):
            user_value = TaskCloseQuestionsView._normalize_value(
                TaskCloseQuestionsView._get_user_field_value(user, cond.field)
            )
            cond_value = TaskCloseQuestionsView._normalize_value(cond.value)

            if cond.operator == "equals":
                return user_value == cond_value
            if cond.operator == "not_equal":
                return user_value != cond_value
            if cond.operator == "contains":
                return cond_value in user_value
            if cond.operator == "starts_with":
                return user_value.startswith(cond_value)
            if cond.operator == "ends_with":
                return user_value.endswith(cond_value)
            if cond.operator == "is_one_of":
                options = [TaskCloseQuestionsView._normalize_value(v) for v in cond.value.split(",")]
                return user_value in options
            return False

        if group.match_type == "or":
            return any(eval_condition(c) for c in conditions)
        return all(eval_condition(c) for c in conditions)

    @staticmethod
    def _is_user_assigned_to_task(task, user):
        if TaskAssignee.objects.filter(task=task, assigned_user=user).exists():
            return True
        if TaskAssignee.objects.filter(task=task, assigned_leader=user).exists():
            return True
        if task.created_by == user:
            return True

        group_assignees = TaskAssignee.objects.filter(
            task=task, assigned_group__isnull=False
        ).select_related("assigned_group")
        for assignee in group_assignees:
            group = assignee.assigned_group
            if group and TaskCloseQuestionsView._user_matches_group(group, user):
                return True

        # Fallback for followup tasks created from logic where TaskAssignee rows may not exist
        if not task.followup_task_form_id_id or not task.follow_task_sub_question_id:
            return False

        logic_followup = LogicFollowUp.objects.filter(
            followup_toggle=True,
            question_id=task.follow_task_sub_question_id
        ).filter(
            models.Q(form_id=task.followup_task_form_id_id) |
            models.Q(audit_group__form_id=task.followup_task_form_id_id)
        ).first()

        if not logic_followup:
            return False

        user_ids = set(logic_followup.assign_user_ids or [])
        group_ids = set(logic_followup.assign_group_ids or [])
        leader_ids = set(logic_followup.assign_leader_ids or [])

        if logic_followup.user_id:
            user_ids.add(logic_followup.user_id)
        if logic_followup.group_id:
            group_ids.add(logic_followup.group_id)
        if logic_followup.leader_id:
            leader_ids.add(logic_followup.leader_id)

        if user.id in user_ids or user.id in leader_ids:
            return True

        if group_ids:
            groups = Groups.objects.filter(id__in=group_ids, organization=task.organization)
            for group in groups:
                if TaskCloseQuestionsView._user_matches_group(group, user):
                    return True

        return False

    @staticmethod
    def _user_display(user):
        if not user:
            return None
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
        return {
            "id": user.id,
            "name": full_name or user.username or user.email or "Unknown"
        }

    @staticmethod
    def _get_task_close_questions(task, organization):
        """Return task-close question objects for the given task.

        We explicitly prefetch options to ensure the serializer can include them
        (and to avoid N+1 queries)."""
        question_ids = list(
            TaskCloseQuestion.objects.filter(
                task=task,
                organization=organization
            ).values_list('question_id', flat=True)
        )

        # If this is a followup task and no direct TaskCloseQuestion rows exist,
        # fall back to the configured LogicFollowUp task_close_questions.
        if not question_ids and task.followup_task_form_id:
            logic_followups_with_questions = LogicFollowUp.objects.filter(
                models.Q(form=task.followup_task_form_id) |
                models.Q(audit_group__form=task.followup_task_form_id),
                followup_toggle=True,
                assign_form__isnull=True,
                task_close_questions__isnull=False
            ).select_related('logic', 'question', 'audit_group', 'form')

            if task.follow_task_sub_question:
                logic_followups_with_questions = logic_followups_with_questions.filter(
                    question=task.follow_task_sub_question
                )

            question_ids = list(
                logic_followups_with_questions
                .values_list('task_close_questions__id', flat=True)
                .distinct()
            )

        if not question_ids:
            return []

        # Prefetch options to ensure API output includes them
        questions = list(
            Question.objects.filter(id__in=question_ids)
            .prefetch_related('options')
        )

        return questions

    @staticmethod
    def _has_answer_value(value):
        if value is None:
            return False
        if isinstance(value, str):
            return value.strip() != ""
        if isinstance(value, (list, dict)):
            return len(value) > 0
        return True

    def get(self, request, task_id=None):
        """Get task close questions for a followup task"""
        try:
            task = get_object_or_404(Task, id=task_id, organization=request.user.organization)

            # Check if user is assigned to this task or is the task creator
            is_assigned = self._is_user_assigned_to_task(task, request.user)

            if not is_assigned:
                return Response({"error": "Not assigned to this task"}, status=status.HTTP_403_FORBIDDEN)

            # First, try to get task close questions directly from Task via TaskCloseQuestion table
            task_close_question_entries = TaskCloseQuestion.objects.filter(
                task=task,
                organization=request.user.organization
            ).select_related('question')

            questions = [entry.question for entry in task_close_question_entries]

            # If no direct TaskCloseQuestion records found, check if this is a followup task
            # that should have task close questions from LogicFollowUp configuration
            # If no direct TaskCloseQuestion records found, check fallback LogicFollowUp configs
            if not questions and task.followup_task_form_id:

                logic_followups_with_questions = LogicFollowUp.objects.filter(
                    models.Q(form=task.followup_task_form_id) |
                    models.Q(audit_group__form=task.followup_task_form_id),
                    followup_toggle=True,
                    assign_form__isnull=True,
                    task_close_questions__isnull=False
                ).select_related('logic', 'question', 'audit_group', 'form')

                # 🔐 IMPORTANT FIX:
                # If task has a parent question, restrict to ONLY that LogicFollowUp
                if task.follow_task_sub_question:
                    logic_followups_with_questions = logic_followups_with_questions.filter(
                        question=task.follow_task_sub_question
                    )

                for logic_followup in logic_followups_with_questions:
                    for question in logic_followup.task_close_questions.all():
                        if question not in questions:
                            questions.append(question)


            questions = self._get_task_close_questions(task, request.user.organization)
            serializer = QuestionSerializer(questions, many=True, context={'request': request})

            return Response({
                'task_id': task_id,
                'task_name': task.task_name,
                'questions': serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request, task_id=None):
        """Submit answers for task close questions"""
        try:
            task = get_object_or_404(Task, id=task_id, organization=request.user.organization)

            # Check if user is assigned to this task or is the task creator
            is_assigned = self._is_user_assigned_to_task(task, request.user)

            if not is_assigned:
                return Response({"error": "Not assigned to this task"}, status=status.HTTP_403_FORBIDDEN)

            answers_data = request.data.get('answers', [])
            questions = self._get_task_close_questions(task, request.user.organization)
            allowed_question_ids = {question.id for question in questions}
            provided_answers = {}

            for answer_data in answers_data:
                question_id = answer_data.get('question_id')
                if question_id in allowed_question_ids:
                    provided_answers[question_id] = answer_data.get('answer')

            invalid_question_ids = [
                answer_data.get('question_id')
                for answer_data in answers_data
                if answer_data.get('question_id') not in allowed_question_ids
            ]
            if invalid_question_ids:
                return Response({
                    "error": "Answers contain invalid task close question IDs.",
                    "invalid_question_ids": invalid_question_ids
                }, status=status.HTTP_400_BAD_REQUEST)

            missing_required_questions = [
                {
                    "id": question.id,
                    "question": question.question
                }
                for question in questions
                if question.is_required and not self._has_answer_value(provided_answers.get(question.id))
            ]
            if missing_required_questions:
                return Response({
                    "error": "Required task close questions are missing answers.",
                    "missing_required_questions": missing_required_questions
                }, status=status.HTTP_400_BAD_REQUEST)

            # For task close questions, prefer the assigned followup form (task.form)
            # so submissions show under the followup form in completed-tasks.
            form_for_submission = task.form

            if not form_for_submission:
                # Fallback: if no assigned form, use the triggering main form
                form_for_submission = task.followup_task_form_id

            if not form_for_submission:
                return Response({
                    "error": "Cannot create submission: task has no associated form"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Ensure form_for_submission is a Form object
            if isinstance(form_for_submission, int):
                try:
                    form_for_submission = Form.objects.get(id=form_for_submission, organization=request.user.organization)
                except Form.DoesNotExist:
                    return Response({
                        "error": f"Form with ID {form_for_submission} not found"
                    }, status=status.HTTP_400_BAD_REQUEST)

            # Create form submission for task close answers
            try:
                submission = FormSubmision.objects.create(
                    form=form_for_submission,
                    submission_initiated_by=request.user,
                    organization=request.user.organization,
                    is_completed=True,
                    completed_by=request.user,
                    completed_on=timezone.now()
                )
                # Mark associated planner as completed when form is submitted
                from planner.models import PlannerAssignment
                planner_assignment_id = request.data.get('planner_assignment_id')
                if planner_assignment_id:
                    PlannerAssignment.objects.filter(
                        id=planner_assignment_id,
                        organization=request.user.organization,
                        is_completed=False
                    ).update(
                        is_completed=True,
                        completed_on=timezone.now(),
                        completed_by=request.user
                    )
                else:
                    # Do NOT bulk-complete planners by form match — only the
                    # specific planner_assignment_id should be marked completed.
                    # Bulk-matching caused all planners for the same form to
                    # auto-complete when any one was submitted.
                    pass
            except Exception as create_error:
                return Response({
                    "error": f"Failed to create form submission: {str(create_error)}"
                }, status=status.HTTP_400_BAD_REQUEST)

            question_type_map = {question.id: question.question_type for question in questions}

            # Save answers
            for answer_data in answers_data:
                Answer.objects.create(
                    Form=form_for_submission,
                    question_id=answer_data['question_id'],
                    question_type=answer_data.get('question_type') or question_type_map.get(answer_data['question_id']),
                    answer=answer_data['answer'],
                    submitted_by=request.user,
                    submission=submission,
                    organization=request.user.organization
                )

            # Update task completion status. Task model does not have a followup_submission field,
            # so only update the status here.
            task.status = 'completed'
            task.updated_by = request.user
            task.updated_on = timezone.now()
            task.save()

            # Create audit log
            TaskAuditLog.objects.create(
                task=task,
                task_action='Followup_Completed',
                action_by=request.user,
                action_to=None
            )

            # Close selected related tasks (same location + question combination)
            close_related_task_ids = request.data.get('close_related_task_ids', None)
            from task.utils import close_related_tasks
            related_tasks_result = close_related_tasks(
                task,
                request.user,
                request.user.organization,
                selected_task_ids=close_related_task_ids
            )

            return Response({
                'message': 'Task close questions submitted successfully' + (f" and {related_tasks_result['count']} related task(s) closed" if related_tasks_result['count'] > 0 else ""),
                'submission_id': submission.id,
                'task_id': task.id,
                'completed_by': self._user_display(submission.completed_by),
                'completed_on': submission.completed_on,
                'related_tasks_closed': related_tasks_result,
                'auto_close_info': {
                    'enabled': close_related_task_ids is not None and len(close_related_task_ids) > 0,
                    'description': 'Tasks with same Location and Question can be closed together'
                }
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class TaskCloseQuestionsAnswersView(APIView):
    """Get latest answers for task close questions"""
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _user_display(user):
        if not user:
            return None
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
        return {
            "id": user.id,
            "name": full_name or user.username or user.email or "Unknown"
        }

    def get(self, request, task_id=None):
        try:
            task = get_object_or_404(Task, id=task_id, organization=request.user.organization)

            # Check if user is assigned to this task or is the task creator
            is_assigned = TaskCloseQuestionsView._is_user_assigned_to_task(task, request.user)

            if not is_assigned:
                return Response({"error": "Not assigned to this task"}, status=status.HTTP_403_FORBIDDEN)

            # Build question list (same logic as TaskCloseQuestionsView.get)
            task_close_question_entries = TaskCloseQuestion.objects.filter(
                task=task,
                organization=request.user.organization
            ).select_related('question')
            questions = [entry.question for entry in task_close_question_entries]

            if not questions and task.followup_task_form_id:
                logic_followups_with_questions = LogicFollowUp.objects.filter(
                    models.Q(form=task.followup_task_form_id) |
                    models.Q(audit_group__form=task.followup_task_form_id),
                    followup_toggle=True,
                    assign_form__isnull=True,
                    task_close_questions__isnull=False
                ).select_related('logic', 'question', 'audit_group', 'form')

                if task.follow_task_sub_question:
                    logic_followups_with_questions = logic_followups_with_questions.filter(
                        question=task.follow_task_sub_question
                    )

                for logic_followup in logic_followups_with_questions:
                    for question in logic_followup.task_close_questions.all():
                        if question not in questions:
                            questions.append(question)

            if not questions:
                return Response({
                    "submission_id": None,
                    "completed_by": None,
                    "completed_on": None,
                    "answers": []
                }, status=status.HTTP_200_OK)

            question_ids = [q.id for q in questions]

            # Answers are saved under assigned followup form when available
            form_for_submission = task.form or task.followup_task_form_id

            answers_qs = Answer.objects.filter(
                question_id__in=question_ids,
                submission__form=form_for_submission,
                submission__organization=request.user.organization
            ).select_related('submission').order_by('-submission__completed_on', '-submission__id')

            if not answers_qs.exists():
                return Response({
                    "submission_id": None,
                    "completed_by": None,
                    "completed_on": None,
                    "answers": []
                }, status=status.HTTP_200_OK)

            latest_submission = answers_qs.first().submission
            latest_submission_id = latest_submission.id
            answers = AnswerSerializer(
                answers_qs.filter(submission_id=latest_submission_id),
                many=True
            ).data

            # Defensive: ignore any malformed/empty answer objects coming from serializer
            answers = [a for a in answers if a and a.get('question')]

            return Response({
                "submission_id": latest_submission_id,
                "completed_by": self._user_display(latest_submission.completed_by),
                "completed_on": latest_submission.completed_on,
                "answers": [
                    {
                        "question_id": a.get("question"),
                        "question_uuid": a.get("question_uuid"),
                        "question_type": a.get("question_type"),
                        "answer": a.get("answer"),
                        "answer_id": a.get("answer_id"),
                    }
                    for a in answers
                ]
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)





class TriggerFollowupTasksView(APIView):
    """
    Enhanced endpoint for mobile app to trigger followup tasks.
    Supports both web-configured and mobile-created followup tasks.

    Supports three scenarios:
    1. Web-configured: LogicFollowUp with followup_toggle=true and assign_form
    2. Mobile-created: User creates followup task via modal, links to web task_close_questions
    3. Mobile-editing: User edits existing web-configured task (isEditingWebTask=true)
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Create followup tasks supporting both web-configured and mobile-created tasks.

        Payload options:
        1. Minimal payload (existing web-configured tasks):
        {
            "form_id": 123,
            "main_form_submission_id": 456,
            "followup_task_form_id": 123
        }

        2. Enhanced payload (mobile-created tasks):
        {
            "form_id": 123,
            "main_form_submission_id": 456,
            "followup_task_form_id": 123,
            "mobile_created_tasks": [
                {
                    "title": "Custom Task Title",
                    "description": "Custom Description",
                    "assign_form_id": 789,
                    "assign_user_ids": [1,2,3],
                    "assign_group_ids": [4,5],
                    "assign_leader_ids": [6,7]
                }
            ]
        }

        3. Mobile editing payload (web-configured task editing):
        {
            "form_id": 123,
            "main_form_submission_id": 456,
            "followup_task_form_id": 123,
            "isEditingWebTask": true,
            "logicId": 789,
            "title": "Updated Task Title",
            "description": "Updated Description",
            "assign_user_ids": [1,2,3],
            "assign_group_ids": [4,5],
            "assign_leader_ids": [6,7]
        }

        Backend automatically handles all scenarios.
        """
        try:
            form_id = request.data.get('form_id')
            main_form_submission_id = request.data.get('main_form_submission_id')
            followup_task_form_id = request.data.get('followup_task_form_id')
            mobile_created_tasks = request.data.get('mobile_created_tasks', [])
            is_editing_web_task = request.data.get('isEditingWebTask', False)
            logic_id = request.data.get('logicId')
            
            # 🔧 CRITICAL FIX: Track mobile-edited LogicFollowUp IDs to exclude from web task creation
            mobile_edited_logic_ids = []

            # 🔧 CRITICAL FIX: Handle mobile editing BEFORE task creation
            if is_editing_web_task:
                logger.info(f"🎯 MOBILE EDITING REQUEST DETECTED: logicId={logic_id}")
                editing_response = self._handle_mobile_editing(request, logic_id)

                # If editing was successful, track the logic_id to exclude from web task creation
                # The LogicFollowUp has been updated, so new tasks will use the edited data
                # BUT we should NOT create new tasks - the existing tasks were already updated
                if editing_response.status_code == 200:
                    logger.info("✅ Mobile editing completed, proceeding with task creation")
                    # Add to exclude list - mobile already handled task creation/update
                    mobile_edited_logic_ids.append(logic_id)
                else:
                    logger.error("❌ Mobile editing failed, returning error")
                    return editing_response

            # Validate required fields for task creation (skip if mobile editing was handled)
            if not is_editing_web_task and not all([form_id, main_form_submission_id, followup_task_form_id]):
                return Response({
                    "error": "form_id, main_form_submission_id, and followup_task_form_id are all required"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get the main form submission
            try:
                main_submission = FormSubmision.objects.get(
                    id=main_form_submission_id,
                    form_id=form_id,
                    organization=request.user.organization
                )
            except FormSubmision.DoesNotExist:
                return Response({"error": "Main form submission not found"}, status=status.HTTP_404_NOT_FOUND)

            # Get the main form object for followup_task_form_id
            # Allow archived forms (planners may reference archived forms)
            try:
                main_form = Form.objects.get(
                    id=followup_task_form_id,
                    organization=request.user.organization,
                    is_deleted=False,
                )
            except Form.DoesNotExist:
                return Response({"error": "Main form not found"}, status=status.HTTP_404_NOT_FOUND)

            # Create followup tasks immediately after form submission
            created_tasks = []

            # FORCE CREATE MOBILE TASKS (Scenario-2)
            mobile_created_tasks = request.data.get("mobile_created_tasks", [])
            mobile_created_logic_ids = []
            for mobile_task in mobile_created_tasks:
                logic_followup_id = mobile_task.get('logic_followup_id')
                if logic_followup_id:
                    mobile_created_logic_ids.append(logic_followup_id)
                created_task = self._create_mobile_task(
                    mobile_task,
                    main_form,
                    main_submission,
                    request.user
                )
                if created_task:
                    created_tasks.append(created_task)

            # Handle web-configured tasks (existing logic)
            # Pass mobile-edited + mobile-created logic IDs to exclude duplicate web task creation
            excluded_logic_ids = mobile_edited_logic_ids + mobile_created_logic_ids
            web_configured_tasks = self._create_web_configured_tasks(
                form_id, main_form, main_submission, request.user, excluded_logic_ids
            )
            created_tasks.extend(web_configured_tasks)

            return Response({
                'message': 'Followup task processing completed',
                'tasks': created_tasks
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error creating followup tasks: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def _handle_mobile_editing(self, request, logic_id):
        """
        Handle mobile editing of web-configured followup tasks.
        Updates the LogicFollowUp record with new mobile data while preserving web assignees.
        """
        try:
            # Validate required fields for editing
            if not logic_id:
                return Response({
                    "error": "logicId is required for editing web tasks"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get the LogicFollowUp record with enhanced error handling
            try:
                logic_followup = LogicFollowUp.objects.select_related(
                    'form', 'stage', 'audit_info', 'audit_group', 'question', 'logic'
                ).get(id=logic_id)

                logger.info(f"🔍 Found LogicFollowUp {logic_id}: form={logic_followup.form}, stage={logic_followup.stage}, question={logic_followup.question}")

                # Verify the LogicFollowUp belongs to the user's organization
                # Check through related objects since LogicFollowUp doesn't have organization field
                organization_id = None
                
                # Try to get organization from multiple possible sources
                sources_checked = []
                
                if logic_followup.form:
                    organization_id = logic_followup.form.organization_id
                    sources_checked.append(f"form (ID: {logic_followup.form.id})")
                    logger.info(f"🔍 LogicFollowUp organization via form: {organization_id}")
                elif logic_followup.stage and logic_followup.stage.form:
                    organization_id = logic_followup.stage.form.organization_id
                    sources_checked.append(f"stage.form (ID: {logic_followup.stage.form.id})")
                    logger.info(f"🔍 LogicFollowUp organization via stage: {organization_id}")
                elif logic_followup.audit_info and logic_followup.audit_info.form:
                    organization_id = logic_followup.audit_info.form.organization_id
                    sources_checked.append(f"audit_info.form (ID: {logic_followup.audit_info.form.id})")
                    logger.info(f"🔍 LogicFollowUp organization via audit_info: {organization_id}")
                elif logic_followup.audit_group and logic_followup.audit_group.form:
                    organization_id = logic_followup.audit_group.form.organization_id
                    sources_checked.append(f"audit_group.form (ID: {logic_followup.audit_group.form.id})")
                    logger.info(f"🔍 LogicFollowUp organization via audit_group: {organization_id}")
                elif logic_followup.question and logic_followup.question.form:
                    organization_id = logic_followup.question.form.organization_id
                    sources_checked.append(f"question.form (ID: {logic_followup.question.form.id})")
                    logger.info(f"🔍 LogicFollowUp organization via question: {organization_id}")
                elif logic_followup.logic and logic_followup.logic.form:
                    organization_id = logic_followup.logic.form.organization_id
                    sources_checked.append(f"logic.form (ID: {logic_followup.logic.form.id})")
                    logger.info(f"🔍 LogicFollowUp organization via logic: {organization_id}")
                else:
                    logger.warning(f"🔍 LogicFollowUp {logic_id} has no related objects to determine organization")
                    logger.warning(f"🔍 Sources checked: {', '.join(sources_checked) if sources_checked else 'none'}")
                    return Response({
                        "error": f"Cannot determine organization for LogicFollowUp {logic_id}. No related form found.",
                        "logic_followup_id": logic_id,
                        "sources_checked": sources_checked
                    }, status=status.HTTP_400_BAD_REQUEST)

                logger.info(f"🔍 User's organization: {request.user.organization_id}, LogicFollowUp organization: {organization_id}")

                if organization_id != request.user.organization_id:
                    logger.warning(f"🔍 Access denied: organization mismatch")
                    return Response({
                        "error": "Access denied: LogicFollowUp belongs to different organization",
                        "user_organization": request.user.organization_id,
                        "logic_followup_organization": organization_id,
                        "logic_followup_id": logic_id
                    }, status=status.HTTP_403_FORBIDDEN)

            except LogicFollowUp.DoesNotExist:
                logger.error(f"🔍 LogicFollowUp {logic_id} not found")
                return Response({
                    "error": "LogicFollowUp record not found",
                    "logic_followup_id": logic_id,
                    "user_organization": request.user.organization_id
                }, status=status.HTTP_404_NOT_FOUND)

            # Get mobile data from request
            title = request.data.get('title')
            description = request.data.get('description')
            deadline = request.data.get('deadline')  # Add deadline handling
            assign_user_ids = request.data.get('assign_user_ids', [])
            assign_group_ids = request.data.get('assign_group_ids', [])
            assign_leader_ids = request.data.get('assign_leader_ids', [])

            # Preserve original web-assigned users
            original_user_ids = logic_followup.assign_user_ids or []
            original_group_ids = logic_followup.assign_group_ids or []
            original_leader_ids = logic_followup.assign_leader_ids or []

            # Combine original web assignees with new mobile assignees
            updated_user_ids = list(set(original_user_ids + assign_user_ids))
            updated_group_ids = list(set(original_group_ids + assign_group_ids))
            updated_leader_ids = list(set(original_leader_ids + assign_leader_ids))

            # Update the LogicFollowUp record
            logic_followup.title = title or logic_followup.title
            logic_followup.description = description or logic_followup.description
            logic_followup.deadline = deadline or logic_followup.deadline  # Add deadline update
            logic_followup.assign_user_ids = updated_user_ids
            logic_followup.assign_group_ids = updated_group_ids
            logic_followup.assign_leader_ids = updated_leader_ids
            logic_followup.save()

            # Update existing tasks created from this LogicFollowUp configuration
            from task.models import Task, TaskAssignee

            # Use enhanced task finding logic that handles both Audit and Standard/Location forms
            existing_tasks = self._find_tasks_for_logic_followup(logic_followup, request.user.organization)

            logger.info(f"🔍 Found {existing_tasks.count()} tasks to update for LogicFollowUp {logic_id}")

            # Debug: Log what tasks were found
            for task in existing_tasks:
                logger.info(f"🔍 Found task {task.id}: {task.task_name}, form={task.form}, followup_form={task.followup_task_form_id}, sub_question={task.follow_task_sub_question}")

            # If main_form_submission_id is provided but not found, avoid touching previous submissions
            main_form_submission_id = request.data.get('main_form_submission_id')
            if main_form_submission_id:
                try:
                    FormSubmision.objects.get(
                        id=main_form_submission_id,
                        organization=request.user.organization
                    )
                except FormSubmision.DoesNotExist:
                    logger.warning(
                        f"main_form_submission_id {main_form_submission_id} not found; "
                        f"skipping task_details updates to avoid touching previous submissions"
                    )
                    existing_tasks = Task.objects.none()

            for task in existing_tasks:
                # Keep task_details core fields in sync with edited LogicFollowUp
                updated_fields = []
                if title:
                    task.task_name = title
                    updated_fields.append('task_name')
                if description:
                    task.description = description
                    updated_fields.append('description')
                if deadline is not None:
                    try:
                        start_date = task.start_date or timezone.now()
                        task.end_date = start_date + timezone.timedelta(days=int(deadline))
                        updated_fields.append('end_date')
                    except Exception:
                        pass

                if updated_fields:
                    task.updated_by = request.user
                    task.updated_on = timezone.now()
                    updated_fields.extend(['updated_by', 'updated_on'])
                    task.save(update_fields=updated_fields)

                # Delete existing task assignees
                TaskAssignee.objects.filter(task=task).delete()

                # Create new task assignees based on updated LogicFollowUp
                assignee_count = 0

                # Add individual users
                for user_id in updated_user_ids:
                    try:
                        assignee_user = CustomUser.objects.get(id=user_id, organization=request.user.organization)
                        TaskAssignee.objects.create(
                            task=task,
                            assigned_user=assignee_user,
                            assigned_date_time=timezone.now()
                        )
                        assignee_count += 1
                    except CustomUser.DoesNotExist:
                        continue

                # Add group members
                for group_id in updated_group_ids:
                    try:
                        group = Groups.objects.get(id=group_id, organization=request.user.organization)
                        TaskAssignee.objects.create(
                            task=task,
                            assigned_group=group,
                            assigned_date_time=timezone.now()
                        )
                        assignee_count += 1
                    except Groups.DoesNotExist:
                        continue

                # Add location leaders
                for leader_id in updated_leader_ids:
                    try:
                        leader = CustomUser.objects.get(id=leader_id, organization=request.user.organization)
                        TaskAssignee.objects.create(
                            task=task,
                            assigned_leader=leader,
                            assigned_date_time=timezone.now()
                        )
                        assignee_count += 1
                    except CustomUser.DoesNotExist:
                        continue

                logger.info(f"Updated TaskAssignee for task {task.id}: {assignee_count} assignees")

            logger.info(f"Successfully updated LogicFollowUp {logic_id} with mobile data and {existing_tasks.count()} related tasks")

            return Response({
                'message': 'Successfully updated followup task configuration',
                'logic_id': logic_id,
                'mobile_edited_logic_id': logic_id,  # Track this for _create_web_configured_tasks to exclude
                'updated_fields': {
                    'title': logic_followup.title,
                    'description': logic_followup.description,
                    'total_assignees': len(updated_user_ids) + len(updated_group_ids) + len(updated_leader_ids),
                    'web_assignees_preserved': {
                        'users': len(original_user_ids),
                        'groups': len(original_group_ids),
                        'leaders': len(original_leader_ids)
                    },
                    'mobile_assignees_added': {
                        'users': len(assign_user_ids),
                        'groups': len(assign_group_ids),
                        'leaders': len(assign_leader_ids)
                    }
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error updating LogicFollowUp record: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def _create_mobile_task(self, mobile_task_data, main_form, main_submission, user):
        """Create a task from mobile-provided data and link web task_close_questions"""
        try:
            logger.info(f"🔍 Processing mobile task data: {mobile_task_data}")

            # Require logic_followup_id before creating any task record.
            # If missing, skip mobile task creation so web-configured flow can create only once.
            logic_followup_id = mobile_task_data.get('logic_followup_id')
            if not logic_followup_id:
                logger.error("❌ logic_followup_id missing for mobile-created task")
                return None

            # Get assigned form from mobile data - handle different possible keys
            assign_form_id = mobile_task_data.get('assign_form_id') or mobile_task_data.get('assigned_form_id') or mobile_task_data.get('form_id')

            assigned_form = None
            if assign_form_id:  # Only try to get form if assign_form_id is provided
                try:
                    assigned_form = Form.objects.get(
                        id=assign_form_id,
                        organization=user.organization,
                        is_deleted=False,
                    )
                except Form.DoesNotExist:
                    logger.warning(f"Mobile task assign_form_id {assign_form_id} not found, skipping")
                    return None
            # If no assign_form_id, assigned_form remains None (this is expected for Scenario 2)

            # Calculate deadline (use mobile-provided deadline or default 7 days)
            submission_date = main_submission.completed_on or main_submission.submission_initiated_on or timezone.now()
            start_date = submission_date
            deadline_days = mobile_task_data.get('deadline', 7) or 7
            end_date = submission_date + timezone.timedelta(days=deadline_days)

            # Prevent duplicate tasks for the same submission/question
            question_id = mobile_task_data.get("follow_task_sub_question_id")
            existing_task_qs = Task.objects.filter(
                followup_task_form_id=main_form,
                organization=main_submission.organization,
            )
            if question_id is not None:
                existing_task_qs = existing_task_qs.filter(
                    follow_task_sub_question_id=question_id
                )
            if assigned_form:
                existing_task_qs = existing_task_qs.filter(form=assigned_form)
            else:
                existing_task_qs = existing_task_qs.filter(form__isnull=True)
            # Prevent blocking new submissions by older tasks
            existing_task_qs = existing_task_qs.filter(start_date=start_date)

            existing_task = existing_task_qs.first()
            if existing_task:
                logger.warning(
                    f"Duplicate mobile task prevented for submission {main_submission.id} "
                    f"(existing task_id={existing_task.id})"
                )
                return {
                    'task_id': existing_task.id,
                    'title': existing_task.task_name,
                    'form': assigned_form.title if assigned_form else None,
                    'form_name': assigned_form.title if assigned_form else None,
                    'form_id': assigned_form.id if assigned_form else None,
                    'form_type': assigned_form.get_form_type_display() if assigned_form else None,
                    'assignee_count': existing_task.assignees.count(),
                    'has_assigned_form': True,
                    'has_task_close_questions': TaskCloseQuestion.objects.filter(task=existing_task).exists(),
                    'source': 'mobile_created'
                }

            # Create the task (allow multiple submissions)
            task = Task.objects.create(
                task_name=mobile_task_data.get('title', 'Follow-up Task'),
                description=mobile_task_data.get('description', ''),
                form=assigned_form,  # Mobile-selected form
                followup_task_form_id=main_form,  # Main form that triggered this
                organization=main_submission.organization,
                status='not_started',
                start_date=start_date,
                end_date=end_date,
                created_by=user,
            )
            logger.info(f"✅ Created task {task.id} for mobile task")

            # 🔑 SET follow_task_sub_question DIRECTLY FROM MOBILE (QUESTION ID)
            question_id = mobile_task_data.get('follow_task_sub_question_id')

            if question_id:
                try:
                    question = Question.objects.get(
                        id=question_id,
                        organization=user.organization
                    )
                    task.follow_task_sub_question = question
                    task.save(update_fields=['follow_task_sub_question'])
                    logger.info(
                        f"Set follow_task_sub_question={question.id} for mobile task {task.id}"
                    )
                except Question.DoesNotExist:
                    logger.warning(
                        f"Invalid follow_task_sub_question_id {question_id} received from mobile"
                    )
            else:
                logger.warning(
                    f"No follow_task_sub_question_id provided for mobile task {task.id}"
                )


            # Delete existing assignees for this task to avoid duplicates
            TaskAssignee.objects.filter(task=task).delete()

            # Create task assignees from mobile data - handle different possible key formats
            all_assignee_users = []
            assignee_count = 0

            # Handle individual users - try different possible key names
            assign_user_ids = (mobile_task_data.get('assign_user_ids') or
                              mobile_task_data.get('assigned_user_ids') or
                              mobile_task_data.get('user_ids') or
                              mobile_task_data.get('assignees') or
                              mobile_task_data.get('assign_users') or [])  # Add 'assign_users' from mobile app
            logger.info(f"🔍 Processing assign_user_ids: {assign_user_ids}")

            # Handle different data formats for assignees
            actual_user_ids = []
            if assign_user_ids:
                for item in assign_user_ids:
                    if isinstance(item, dict):
                        # Handle dict format: {"user_id": 1}
                        user_id = item.get('user_id')
                        if user_id:
                            actual_user_ids.append(user_id)
                    elif isinstance(item, int):
                        # Handle simple ID format: [1, 2, 3]
                        actual_user_ids.append(item)
                    # Skip other formats

            logger.info(f"🔍 Final user IDs to process: {actual_user_ids}")

            for user_id in actual_user_ids:
                try:
                    assignee_user = CustomUser.objects.get(id=user_id, organization=user.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_user=assignee_user,
                        assigned_date_time=timezone.now()
                    )
                    all_assignee_users.append(assignee_user)
                    assignee_count += 1
                    logger.info(f"✅ Created TaskAssignee for user {user_id}")
                except CustomUser.DoesNotExist:
                    logger.warning(f"❌ User {user_id} not found in organization {user.organization.id}")
                    continue
                except Exception as e:
                    logger.error(f"❌ Error creating TaskAssignee for user {user_id}: {str(e)}")
                    continue

            # Handle groups - try different possible key names
            assign_group_ids = (mobile_task_data.get('assign_group_ids') or
                               mobile_task_data.get('assigned_group_ids') or
                               mobile_task_data.get('group_ids') or [])
            logger.info(f"🔍 Processing assign_group_ids: {assign_group_ids}")
            for group_id in assign_group_ids:
                try:
                    group = Groups.objects.get(id=group_id, organization=user.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_group=group,
                        assigned_date_time=timezone.now()
                    )
                    # Add all group members
                    group_members = group.members.filter(organization=user.organization)
                    all_assignee_users.extend(group_members)
                    assignee_count += 1
                    logger.info(f"✅ Created TaskAssignee for group {group_id} with {group_members.count()} members")
                except Groups.DoesNotExist:
                    logger.warning(f"❌ Group {group_id} not found in organization {user.organization.id}")
                    continue
                except Exception as e:
                    logger.error(f"❌ Error creating TaskAssignee for group {group_id}: {str(e)}")
                    continue

            # Handle location leaders - try different possible key names
            assign_leader_ids = (mobile_task_data.get('assign_leader_ids') or
                                mobile_task_data.get('assigned_leader_ids') or
                                mobile_task_data.get('leader_ids') or
                                mobile_task_data.get('assignLocationLeaders') or [])  # Add mobile app key
            logger.info(f"🔍 Processing assign_leader_ids: {assign_leader_ids}")
            for leader_id in assign_leader_ids:
                try:
                    leader = CustomUser.objects.get(id=leader_id, organization=user.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_leader=leader,
                        assigned_date_time=timezone.now()
                    )
                    all_assignee_users.append(leader)
                    assignee_count += 1
                    logger.info(f"✅ Created TaskAssignee for leader {leader_id}")
                except CustomUser.DoesNotExist:
                    logger.warning(f"❌ Leader {leader_id} not found in organization {user.organization.id}")
                    continue
                except Exception as e:
                    logger.error(f"❌ Error creating TaskAssignee for leader {leader_id}: {str(e)}")
                    continue

            logger.info(f"✅ Task {task.id} has {assignee_count} assignees, {len(all_assignee_users)} total users")

            # Debug: Check if task assignees were actually created
            actual_assignee_count = TaskAssignee.objects.filter(task=task).count()
            logger.info(f"🔍 Verification: Task {task.id} has {actual_assignee_count} TaskAssignee records in database")

            # Create audit log
            TaskAuditLog.objects.create(
                task=task,
                task_action='Followup_Created',
                action_by=user,
                action_to=None
            )

            # 🔑 FIXED: Link ONLY the specific LogicFollowUp that triggered this mobile task

            logic_followups_to_link = []
            try:
                specific_logic_followup = LogicFollowUp.objects.get(
                    id=logic_followup_id,
                    followup_toggle=True,
                    # assign_form__isnull=True
                )

                # Safety check: ensure same form
                if (
                    (specific_logic_followup.form and specific_logic_followup.form == main_form) or
                    (specific_logic_followup.audit_group and specific_logic_followup.audit_group.form == main_form)
                ):
                    logic_followups_to_link = [specific_logic_followup]
                else:
                    logger.warning(
                        f"LogicFollowUp {logic_followup_id} does not belong to main form {main_form.id}"
                    )
            except LogicFollowUp.DoesNotExist:
                logger.warning(f"Invalid logic_followup_id {logic_followup_id} received from mobile")

            # ✅ Set correct parent question
            # Persist mobile-created followup details into Scenario-2 LogicFollowUp
            if assign_form_id is None and logic_followups_to_link:
                try:
                    mobile_title = mobile_task_data.get('title')
                    mobile_description = mobile_task_data.get('description')
                    mobile_deadline = mobile_task_data.get('deadline')
                    mobile_assign_user_ids = mobile_task_data.get('assign_user_ids') or []
                    mobile_assign_group_ids = mobile_task_data.get('assign_group_ids') or []
                    mobile_assign_leader_ids = mobile_task_data.get('assign_leader_ids') or []

                    for lf in logic_followups_to_link:
                        if mobile_title:
                            lf.title = mobile_title
                        if mobile_description:
                            lf.description = mobile_description
                        if mobile_deadline is not None:
                            lf.deadline = mobile_deadline
                        if mobile_assign_user_ids:
                            lf.assign_user_ids = mobile_assign_user_ids
                        if mobile_assign_group_ids:
                            lf.assign_group_ids = mobile_assign_group_ids
                        if mobile_assign_leader_ids:
                            lf.assign_leader_ids = mobile_assign_leader_ids
                        lf.save()
                except Exception:
                    pass

            if not task.follow_task_sub_question and logic_followups_to_link:
                primary_lf = logic_followups_to_link[0]
                if primary_lf.question:
                    task.follow_task_sub_question = primary_lf.question
                    task.save(update_fields=['follow_task_sub_question'])

            # ✅ Link ONLY relevant task close questions (no duplicates)
            seen_question_ids = set()
            has_task_close_questions = False

            for lf in logic_followups_to_link:
                for question in lf.task_close_questions.all():
                    if question.id not in seen_question_ids:
                        TaskCloseQuestion.objects.create(
                            task=task,
                            question=question,
                            created_by=user,
                            organization=user.organization
                        )
                        seen_question_ids.add(question.id)
                        has_task_close_questions = True

            logger.info(
                f"Linked {len(seen_question_ids)} task close questions "
                f"to mobile task {task.id} (logic_followup_id={logic_followup_id})"
            )


            return {
                'task_id': task.id,
                'title': task.task_name,
                'form': assigned_form.title if assigned_form else None,
                'form_name': assigned_form.title if assigned_form else None,  # Alternative field name
                'form_id': assigned_form.id if assigned_form else None,
                'form_type': assigned_form.get_form_type_display() if assigned_form else None,
                'assignee_count': assignee_count,
                'has_assigned_form': True,  # Mobile tasks always have assigned forms
                'has_task_close_questions': has_task_close_questions,
                'source': 'mobile_created'
            }

        except Exception as e:
            logger.error(f"Error creating mobile task: {str(e)}")
            return None

    def _create_web_configured_tasks(self, form_id, main_form, main_submission, user, mobile_edited_logic_ids=None):
        """Create tasks from web LogicFollowUp configurations (existing logic)

        Args:
            mobile_edited_logic_ids: List of LogicFollowUp IDs that were edited by mobile
                                    These will be excluded from web task creation to avoid duplicates
        """
        created_tasks = []

        # Get all answers for this form submission to check logic conditions
        from .models import Answer
        submission_answers = Answer.objects.filter(
            submission=main_submission,
            organization=user.organization
        ).select_related('question')

        # Create a dictionary of question_id -> answer_value for quick lookup
        answer_dict = {answer.question_id: answer.answer for answer in submission_answers}

        # Debug: Check what LogicFollowUp records exist for this form
        all_logic_followups = LogicFollowUp.objects.filter(
            models.Q(form_id=form_id) | models.Q(audit_group__form_id=form_id)
        ).select_related('logic', 'question', 'assign_form', 'audit_group')

        print(f"🔍 DEBUG: Found {all_logic_followups.count()} total LogicFollowUp records for form {form_id}")
        for lf in all_logic_followups:
            print(f"🔍 DEBUG: LogicFollowUp ID {lf.id}: followup_toggle={lf.followup_toggle}, assign_form={lf.assign_form}, form_id={lf.form_id}")

        # Get LogicFollowUp configurations with followup_toggle=true and actionable task data
        # These require logic conditions to determine when to trigger
        # 🔧 CRITICAL FIX: Exclude mobile-edited LogicFollowUp IDs to avoid duplicate task creation
        logic_followups_query = LogicFollowUp.objects.filter(
            models.Q(form_id=form_id) | models.Q(audit_group__form_id=form_id),
            followup_toggle=True,
        ).filter(
            models.Q(assign_form__isnull=False) |
            models.Q(task_close_questions__isnull=False)
        ).distinct()
        
        # Exclude mobile-edited LogicFollowUp IDs if provided
        if mobile_edited_logic_ids:
            logic_followups_query = logic_followups_query.exclude(id__in=mobile_edited_logic_ids)
            print(f"🔍 DEBUG: Excluding {len(mobile_edited_logic_ids)} mobile-edited LogicFollowUp IDs: {mobile_edited_logic_ids}")
        
        logic_followups_with_form = logic_followups_query.select_related('logic', 'question', 'assign_form', 'audit_group')

        print(f"🔍 DEBUG: Found {logic_followups_with_form.count()} web-configured LogicFollowUp records with followup_toggle=True and actionable task data")

        # Show which LogicFollowUp records we found
        for lf in logic_followups_with_form:
            print(f"🔍 DEBUG: Processing LogicFollowUp {lf.id}: title='{lf.title}', logic_value='{lf.logic.logic_value if lf.logic else None}'")

        # Debug: Show all answers found
        print(f"🔍 DEBUG: Found {len(answer_dict)} answers for submission {main_submission.id}")
        for q_id, answer in answer_dict.items():
            print(f"🔍 DEBUG: Question {q_id}: '{answer}'")

        # Filter logic followups based on whether their logic conditions are met
        filtered_logic_followups = []
        for logic_followup in logic_followups_with_form:
            # Debug: Show what we're checking
            logic = logic_followup.logic
            if logic:
                user_answer = answer_dict.get(logic.question_id)
                print(f"🔍 DEBUG: Checking LogicFollowUp {logic_followup.id}: question={logic.question_id}, logic_type='{logic.logic_type}', logic_value='{logic.logic_value}', user_answer='{user_answer}'")

            if self._check_logic_condition(logic_followup, answer_dict):
                filtered_logic_followups.append(logic_followup)
                print(f"✅ Logic condition met for LogicFollowUp {logic_followup.id}")
            else:
                print(f"❌ Logic condition NOT met for LogicFollowUp {logic_followup.id}")

        print(f"🔍 DEBUG: After logic filtering, {len(filtered_logic_followups)} LogicFollowUp records remain")

        # Remove stale follow-up tasks whose logic condition is no longer met
        triggered_keys = {
            (logic_followup.question_id, logic_followup.assign_form_id)
            for logic_followup in filtered_logic_followups
        }
        all_logic_question_ids = [lf.question_id for lf in logic_followups_with_form]
        submission_date = (
            main_submission.completed_on
            or main_submission.submission_initiated_on
            or timezone.now()
        )
        if all_logic_question_ids and submission_date:
            existing_tasks = Task.objects.filter(
                followup_task_form_id=main_form,
                organization=main_submission.organization,
                start_date=submission_date,
                follow_task_sub_question_id__in=all_logic_question_ids,
            )
            for task in existing_tasks:
                if (task.follow_task_sub_question_id, task.form_id) not in triggered_keys:
                    task.delete()

        # Process web-configured tasks (with assigned forms and logic filtering)
        for logic_followup in filtered_logic_followups:
            # Calculate deadline
            submission_date = main_submission.completed_on or main_submission.submission_initiated_on or timezone.now()
            start_date = submission_date
            deadline_days = getattr(logic_followup, 'deadline', 7) or 7
            end_date = submission_date + timezone.timedelta(days=deadline_days)

            # Get assigned form object
            assigned_form = logic_followup.assign_form

            # Prevent duplicate tasks for the same submission/question
            existing_task = Task.objects.filter(
                followup_task_form_id=main_form,
                organization=main_submission.organization,
                follow_task_sub_question=logic_followup.question,
                form=assigned_form,
                start_date=start_date,
            ).first()

            if existing_task:
                logger.warning(
                    f"Duplicate web task prevented for submission {main_submission.id} "
                    f"(logic_followup_id={logic_followup.id}, existing task_id={existing_task.id})"
                )
                created_tasks.append({
                    'task_id': existing_task.id,
                    'logic_id': logic_followup.logic.id if logic_followup.logic else None,
                    'title': existing_task.task_name,
                    'assignee_count': existing_task.assignees.count(),
                    'has_assigned_form': assigned_form is not None,
                    'has_task_close_questions': TaskCloseQuestion.objects.filter(task=existing_task).exists(),
                    'source': 'web_configured'
                })
                continue

            # Create task with parent question ID
            task = Task.objects.create(
                task_name=logic_followup.title or 'Follow-up Task',
                description=logic_followup.description or '',
                form=assigned_form,
                followup_task_form_id=main_form,
                follow_task_sub_question=logic_followup.question,
                organization=main_submission.organization,
                status='not_started',
                start_date=start_date,
                end_date=end_date,
                created_by=user,
            )

            # Create task assignees from stored configuration
            all_assignee_users = []

            # Handle assign_to='form_submitter' — assign to the user who submitted the form
            if logic_followup.assign_to == 'form_submitter':
                submitter = main_submission.submission_initiated_by or user
                TaskAssignee.objects.create(
                    task=task,
                    assigned_user=submitter,
                    assigned_date_time=timezone.now()
                )
                all_assignee_users.append(submitter)

            # Add individual users from stored configuration
            for user_id in (logic_followup.assign_user_ids or []):
                try:
                    from user.models import CustomUser
                    assignee_user = CustomUser.objects.get(id=user_id, organization=user.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_user=assignee_user,
                        assigned_date_time=timezone.now()
                    )
                    all_assignee_users.append(assignee_user)
                except CustomUser.DoesNotExist:
                    continue

            # Add group members from stored configuration
            for group_id in (logic_followup.assign_group_ids or []):
                try:
                    from user.models import Groups
                    group = Groups.objects.get(id=group_id, organization=user.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_group=group,
                        assigned_date_time=timezone.now()
                    )
                    all_assignee_users.extend(group.members.filter(organization=user.organization))
                except Groups.DoesNotExist:
                    continue

            # Add location leaders from stored configuration
            for leader_id in (logic_followup.assign_leader_ids or []):
                try:
                    from user.models import CustomUser
                    leader = CustomUser.objects.get(id=leader_id, organization=user.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_leader=leader,
                        assigned_date_time=timezone.now()
                    )
                    all_assignee_users.append(leader)
                except CustomUser.DoesNotExist:
                    continue

            # Create audit log
            TaskAuditLog.objects.create(
                task=task,
                task_action='Followup_Created',
                action_by=user,
                action_to=None
            )

            # Link task close questions
            has_task_close_questions = logic_followup.task_close_questions.exists()
            if has_task_close_questions:
                for question in logic_followup.task_close_questions.all():
                    TaskCloseQuestion.objects.create(
                        task=task,
                        question=question,
                        created_by=user,
                        organization=user.organization
                    )

            created_tasks.append({
                'task_id': task.id,
                'logic_id': logic_followup.logic.id if logic_followup.logic else None,
                'title': task.task_name,
                'assignee_count': len(all_assignee_users),
                'has_assigned_form': assigned_form is not None,
                'has_task_close_questions': has_task_close_questions,
                'source': 'web_configured'
            })

        # SCENARIO 2 code block removed - mobile-created tasks now handled by mobile app only

        return created_tasks

    def _check_logic_condition(self, logic_followup, answer_dict):
        """
        Check if the configured logic condition is met by the user's answer.

        Handles both option text and option ID comparisons:
        - If user_answer is an option ID, looks up the option text
        - Then compares with the configured logic_value

        This determines whether a followup task should be triggered:
        - If logic_type='is' and user_answer matches logic_value → trigger
        - If logic_type='is_not' and user_answer does NOT match logic_value → trigger
        - Otherwise → don't trigger
        """
        try:
            print(f"🔍 DEBUG: Starting logic check for LogicFollowUp {logic_followup.id}")

            # Get the logic associated with this followup
            logic = logic_followup.logic
            if not logic:
                print(f"🔍 DEBUG: ❌ LogicFollowUp {logic_followup.id} has no logic configuration")
                return False

            print(f"🔍 DEBUG: LogicFollowUp {logic_followup.id} logic config: question_id={logic.question_id}, logic_type='{logic.logic_type}', logic_value='{logic.logic_value}'")

            # Get the user's answer for this question
            user_answer = answer_dict.get(logic.question_id)
            print(f"🔍 DEBUG: LogicFollowUp {logic_followup.id}: user_answer for question {logic.question_id} = '{user_answer}'")

            if user_answer is None:
                print(f"🔍 DEBUG: ❌ No answer found for question {logic.question_id} in LogicFollowUp {logic_followup.id}")
                return False

            # Handle option ID to text conversion
            user_answer_str = str(user_answer).strip()
            print(f"🔍 DEBUG: LogicFollowUp {logic_followup.id}: user_answer_str = '{user_answer_str}'")

            # Check if user_answer is an option ID (numeric string)
            if user_answer_str.isdigit():
                print(f"🔍 DEBUG: LogicFollowUp {logic_followup.id}: user_answer_str '{user_answer_str}' is numeric, checking for option...")
                try:
                    # Try to find the option with this ID for this question
                    from .models import Option
                    option = Option.objects.filter(
                        id=int(user_answer_str),
                        question_id=logic.question_id
                    ).first()

                    if option:
                        # Use the option text instead of ID
                        user_answer_str = str(option.option).strip()
                        print(f"🔍 DEBUG: ✅ Converted option ID {user_answer} to text '{user_answer_str}' for LogicFollowUp {logic_followup.id}")
                    else:
                        print(f"🔍 DEBUG: ❌ No option found with ID {user_answer} for question {logic.question_id} in LogicFollowUp {logic_followup.id}")
                        # Debug: Show what options exist for this question
                        all_options = Option.objects.filter(question_id=logic.question_id)
                        print(f"🔍 DEBUG: Available options for question {logic.question_id}: {[(opt.id, opt.option) for opt in all_options]}")
                except Exception as e:
                    print(f"🔍 DEBUG: ❌ Failed to convert option ID {user_answer} to text: {str(e)} for LogicFollowUp {logic_followup.id}")

            logic_value_str = str(logic.logic_value).strip()
            print(f"🔍 DEBUG: LogicFollowUp {logic_followup.id}: comparing '{user_answer_str}' {logic.logic_type} '{logic_value_str}'")

            if logic.logic_type == 'is':
                # Trigger when user selects this specific option
                should_trigger = (user_answer_str == logic_value_str)
                print(f"🔍 DEBUG: LogicFollowUp {logic_followup.id}: 'is' comparison result = {should_trigger}")
            elif logic.logic_type == 'is_not':
                # Trigger when user does NOT select this specific option
                should_trigger = (user_answer_str != logic_value_str)
                print(f"🔍 DEBUG: LogicFollowUp {logic_followup.id}: 'is_not' comparison result = {should_trigger}")
            else:
                # Unknown logic type
                should_trigger = False
                print(f"🔍 DEBUG: ❌ Unknown logic_type '{logic.logic_type}' for LogicFollowUp {logic_followup.id}")

            print(f"🔍 DEBUG: Final result for LogicFollowUp {logic_followup.id}: should_trigger = {should_trigger}")
            return should_trigger

        except Exception as e:
            print(f"🔍 DEBUG: Error evaluating logic condition for LogicFollowUp {logic_followup.id}: {str(e)}")
            return False

    def _find_tasks_for_logic_followup(self, logic_followup, organization):
        """
        Find tasks that were created from this LogicFollowUp configuration.

        Tasks are linked to LogicFollowUp through:
        - followup_task_form_id = the main form (logic_followup.form or logic_followup.audit_group.form)
        - follow_task_sub_question = the question that triggered the followup (logic_followup.question)
        """
        from task.models import Task

        # Determine the main form for this LogicFollowUp
        main_form = None
        if logic_followup.audit_group:
            # For audit groups, the main form is the audit group's form
            main_form = logic_followup.audit_group.form
        elif logic_followup.form:
            # For regular forms, use the form directly
            main_form = logic_followup.form
        else:
            # Fallback: try to get form from stage if available
            if logic_followup.stage:
                main_form = logic_followup.stage.form

        if not main_form:
            logger.warning(f"Could not determine main form for LogicFollowUp {logic_followup.id}")
            return Task.objects.none()

        # Find tasks created from this LogicFollowUp configuration
        # Handle both web-configured tasks (with follow_task_sub_question) and mobile-created tasks
        tasks = Task.objects.filter(
            followup_task_form_id=main_form,
            organization=organization
        ).filter(
            models.Q(follow_task_sub_question=logic_followup.question)
        )

        logger.info(f"Found {tasks.count()} tasks for LogicFollowUp {logic_followup.id} (main_form={main_form.id if main_form else None}, question={logic_followup.question.id if logic_followup.question else None})")

        return tasks


class FormPayloadFilesViewSet(ReadOnlyModelViewSet):
    queryset = FormPayloadFiles.objects.all()
    serializer_class = FormPayloadFilesSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["method", "status", "form", "organization"]

    def retrieve(self, requets, *args, **kwargs):
        UF = UtilsFunctions()
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        payload = UF.get_s3_form_payload_with_url(instance.file_path)
        return Response(payload, status=status.HTTP_200_OK)


    @action(detail=True, methods=["put"], url_path="update-status")
    def update_status(self, request, pk=None):
        instance = self.get_object()
        form_id = request.data.get("form")
        status_value = request.data.get("status")
        if not form_id and not status_value:
            return Response(
                {"error": "form or status must be provided"},
                status=status.HTTP_400_BAD_REQUEST
            )
        if form_id:
            instance.form_id = form_id
        if status_value:
            instance.status = status_value
        if "error_message" in request.data:
            instance.error_message = request.data.get("error_message")
        instance.save()
        return Response(
            {"message": "Updated successfully", "data": FormPayloadFilesSerializer(instance).data},
            status=status.HTTP_200_OK
        )


def generate_csv_with_followup_data(responses, form_info, form_id, organization):
    """
    Generate CSV data from form responses with followup task information.
    Each row represents one question/answer within a response (long format).
    Response metadata is only on the first row of each response.
    Followup task data is shown on the row of the question that triggered the followup.
    Follow Q columns are dynamic based on the max number of close questions across all followup tasks.
    """
    responses = sorted(
        responses,
        key=lambda r: int(r.get('id', 0)) if str(r.get('id', '')).isdigit() else 0,
        reverse=True
    )

    is_audit = form_info.get('form_type') == 'Audit'
    form_prefix = form_info.get('prefix', 'NPX')

    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']

    def extract_image_urls(answer_text):
        if not isinstance(answer_text, str):
            return ''
        potential_urls = answer_text.split("|") if "|" in answer_text else [answer_text]
        image_urls = [u.strip() for u in potential_urls if u.strip() and any(u.lower().endswith(ext) for ext in image_extensions)]
        return '; '.join(image_urls) if image_urls else ''

    def extract_text_answer(answer_text):
        if not isinstance(answer_text, str):
            return str(answer_text) if answer_text else ''
        potential_urls = answer_text.split("|") if "|" in answer_text else [answer_text]
        text_parts = []
        for url in potential_urls:
            url = url.strip()
            if url and not any(url.lower().endswith(ext) for ext in image_extensions):
                text_parts.append(url)
        return ' '.join(text_parts) if text_parts else ''

    # ---- FIRST PASS: collect all followup tasks and find max close questions ----
    # Map: (response_id, question_id) -> [followup_task_data, ...]
    all_followup_map = {}
    max_close_questions = 0

    for response in responses:
        response_id = response.get('id')
        if not response_id:
            continue

        # Get the submission date to match Task.start_date
        # Fetch the actual FormSubmision to get the datetime object
        submission_date = None
        try:
            sub_obj = FormSubmision.objects.filter(id=response_id).first()
            if sub_obj:
                submission_date = sub_obj.completed_on or sub_obj.submission_initiated_on
        except Exception:
            pass

        # Query Task records (followup tasks are stored in task_details table)
        # Task.followup_task_form_id points to the main form
        # Task.follow_task_sub_question points to the triggering question
        # Task.form_submission links to the specific submission (new field)
        try:
            from task.models import Task as TaskModel
            task_qs = TaskModel.objects.filter(
                followup_task_form_id=form_id,
                organization=organization,
            ).select_related('follow_task_sub_question', 'form')

            # Match by form_submission_id (preferred) or start_date (fallback for old data)
            task_qs_sub = task_qs.filter(form_submission_id=response_id)
            if task_qs_sub.exists():
                followup_tasks = list(task_qs_sub.order_by('-created_on'))
            elif submission_date:
                # Fallback: match by start_date for tasks without form_submission link
                followup_tasks = list(task_qs.filter(start_date=submission_date, form_submission__isnull=True).order_by('-created_on'))
            else:
                followup_tasks = []
        except Exception:
            followup_tasks = []

        # Determine source type for this submission (Planner vs Form)
        source_type = 'Form'
        source_id = str(form_id)
        try:
            from planner.models import PlannerSubmission as _PS
            planner_sub = _PS.objects.filter(
                form_submission_id=response_id
            ).select_related('planner_assignment').first()
            if planner_sub and planner_sub.planner_assignment:
                source_type = 'Planner'
                source_id = planner_sub.planner_assignment.order_id or str(planner_sub.planner_assignment_id)
        except Exception:
            pass

        for fut in followup_tasks:
            task_title = fut.task_name or ''

            # Task status display mapping
            status_map = {
                'not_started': 'Not Started',
                'not_assigned': 'Not Assigned',
                'in_progress': 'In Progress',
                'completed': 'Completed',
                'cancelled': 'Cancelled',
            }
            task_status = status_map.get(fut.status, fut.status or 'Not Started')
            # If task has been reopened (has reopened_remarks) and is not yet re-completed, append " - Reopened"
            has_been_reopened = bool(fut.reopened_remarks)
            is_reopened = has_been_reopened and fut.status != 'completed'
            if has_been_reopened:
                task_status = task_status + ' - Reopened'

            # Reopen reason
            reopen_reason = fut.reopened_remarks or ''

            # Deadline
            deadline_str = fut.end_date.strftime('%d-%b-%Y %I:%M %p') if fut.end_date else ''

            # Get close questions from TaskCloseQuestion table
            close_qs = list(TaskCloseQuestion.objects.filter(
                task=fut,
                organization=organization
            ).select_related('question').order_by('id'))

            # Also try LogicFollowUp task_close_questions as fallback
            if not close_qs:
                try:
                    logic_followups_with_questions = LogicFollowUp.objects.filter(
                        models.Q(form_id=form_id) | models.Q(audit_group__form_id=form_id),
                        followup_toggle=True,
                    ).filter(
                        models.Q(assign_form__isnull=True) |
                        models.Q(task_close_questions__isnull=False)
                    ).distinct()

                    if fut.follow_task_sub_question:
                        logic_followups_with_questions = logic_followups_with_questions.filter(
                            question=fut.follow_task_sub_question
                        )

                    for lf in logic_followups_with_questions:
                        for question in lf.task_close_questions.all():
                            close_qs.append(type('obj', (object,), {'question': question, 'id': question.id})())
                except Exception:
                    pass

            # Get close question answers from Answer table
            # Answers are stored as Answer records with question_id matching close question IDs
            close_question_ids = [cq.question.id for cq in close_qs if hasattr(cq, 'question') and cq.question]
            close_answers_map = {}
            followup_submission_id = ''
            logger.info(f"[CSV+Followup] Task {fut.id}: close_question_ids={close_question_ids}, fut.form={fut.form_id}, fut.followup_task_form_id={fut.followup_task_form_id_id}")
            if close_question_ids:
                try:
                    form_for_submission = fut.form or fut.followup_task_form_id
                    answers_qs = Answer.objects.filter(
                        question_id__in=close_question_ids,
                        submission__form=form_for_submission,
                        submission__organization=organization,
                        question__is_task_close_question=True,
                    ).select_related('submission').order_by('-submission__completed_on', '-submission__id')

                    logger.info(f"[CSV+Followup] Task {fut.id}: answers_qs count={answers_qs.count()}")

                    if answers_qs.exists():
                        latest_submission = answers_qs.first().submission
                        followup_submission_id = f'{form_prefix}-SUB-{latest_submission.id}'
                        task_answers = list(answers_qs.filter(submission_id=latest_submission.id))
                        logger.info(f"[CSV+Followup] Task {fut.id}: latest_submission={latest_submission.id}, answer_count={len(task_answers)}")
                        for ans in task_answers:
                            close_answers_map[ans.question_id] = ans.answer or ''
                            logger.info(f"[CSV+Followup] Task {fut.id}: answer for question_id={ans.question_id} = '{ans.answer}'")
                except Exception as e:
                    logger.warning(f"[CSV+Followup] Task {fut.id}: Error fetching answers: {e}")

            # Build close question data list
            close_q_data = []
            for cq in close_qs:
                q = cq.question if hasattr(cq, 'question') else cq
                q_text = q.question if hasattr(q, 'question') else str(q)
                q_id = q.id if hasattr(q, 'id') else None
                q_answer = close_answers_map.get(q_id, '') if q_id else ''
                logger.info(f"[CSV+Followup] Task {fut.id}: close_q_data: q_id={q_id}, q_text='{q_text[:50]}', q_answer='{q_answer[:50]}'")
                close_q_data.append({
                    'question': q_text,
                    'answer': q_answer,
                })

            max_close_questions = max(max_close_questions, len(close_q_data))

            question_id = fut.follow_task_sub_question_id
            if question_id:
                key = (response_id, question_id)
                if key not in all_followup_map:
                    all_followup_map[key] = []
                # If reopened, keep questions but clear answers
                if is_reopened:
                    reopened_close_q_data = [{'question': cq['question'], 'answer': ''} for cq in close_q_data]
                else:
                    reopened_close_q_data = close_q_data

                all_followup_map[key].append({
                    'title': task_title,
                    'close_questions': reopened_close_q_data,
                    'source_type': source_type,
                    'source_id': source_id,
                    'task_id': str(fut.id),
                    'deadline': deadline_str,
                    'task_status': task_status,
                    'reopen_reason': reopen_reason,
                    'followup_submission_id': followup_submission_id if not is_reopened else '',
                    'is_bulk_imported': fut.is_bulk_imported,
                })

    # ---- BUILD DYNAMIC HEADERS ----
    base_headers = [
        'Response ID',
        'Submission Date',
        'Initiated By',
        'Designation',
        'Department',
        'Status',
        'Form Title',
        'Form Type',
        'Overall Status',
        'Total Score (%)',
        'Task Completion (%)',
        'Overdue Tasks (%)',
        'Reopened Tasks (%)',
        'Audited Location',
        'Ambient Temperature',
        'Total Production',
        'Source Type',
        'Source ID',
        'Group Title',
        'Item',
        'Response',
        'Image',
        'Remarks',
        'Consumed from / SAP Code or Product Name',
        'Quantity',
        'After Image',
    ]

    # Add Imported column after Source ID
    base_headers.insert(base_headers.index('Source ID') + 1, 'Imported')

    # Add new metadata columns before Follow Q columns
    metadata_headers = [
        'Followup Task ID',
        'Followup Deadline',
        'Followup Task Status',
        'Reopen Reason',
        'Followup Response Submission ID',
    ]

    # Add dynamic Follow Q columns
    followup_headers = ['Follow up action Title']
    for i in range(max_close_questions):
        n = i + 1
        followup_headers.append(f'Follow Q{n} Question')
        followup_headers.append(f'Follow Q{n} Answer')

    headers = base_headers + metadata_headers + followup_headers
    num_metadata_cols = len(metadata_headers)
    num_followup_cols = len(followup_headers)

    # ---- SECOND PASS: build rows ----
    rows = []

    for response in responses:
        response_id = response.get('id', 'N/A')
        display_response_id = f"{form_prefix}-{response_id}" if response_id != 'N/A' else 'N/A'
        is_first_row = True

        # Fetch audit history for this response
        audit_history = None
        if is_audit:
            try:
                audit_history = AuditFormSubmissionHistory.objects.filter(
                    form_submission_id=response_id,
                    organization=organization
                ).first()
            except Exception:
                pass

        # Calculate per-submission task metrics
        task_total = 0
        task_completed = 0
        task_overdue = 0
        task_reopened = 0
        try:
            from task.models import Task as TaskModel
            from django.utils import timezone as _tz
            sub_obj = FormSubmision.objects.filter(id=response_id).first()
            if sub_obj:
                sub_anchor = sub_obj.completed_on or sub_obj.submission_initiated_on
                task_qs = TaskModel.objects.filter(
                    followup_task_form_id=form_id,
                    organization=organization,
                )
                # Match by form_submission_id (preferred) or start_date (fallback)
                task_qs_sub = task_qs.filter(form_submission_id=response_id)
                if task_qs_sub.exists():
                    base_tasks = task_qs_sub
                elif sub_anchor:
                    base_tasks = task_qs.filter(start_date=sub_anchor)
                else:
                    base_tasks = task_qs_sub
                task_total = base_tasks.count()
                task_completed = base_tasks.filter(status='completed').count()
                task_overdue = base_tasks.filter(end_date__lt=_tz.now(), status__in=['not_started', 'not_assigned', 'in_progress']).count()
                reopened_filter = (
                    (models.Q(reopened_remarks__isnull=False) & ~models.Q(reopened_remarks='')) |
                    models.Q(audit_logs__task_action__in=['Reopened', 'Followup_Reopened'])
                )
                task_reopened = base_tasks.filter(reopened_filter).distinct().count()
        except Exception:
            pass

        task_completion_pct = f'{round(task_completed / task_total * 100)}%' if task_total > 0 else '0%'
        overdue_pct = f'{round(task_overdue / task_total * 100)}%' if task_total > 0 else '0%'
        reopened_pct = f'{round(task_reopened / task_total * 100)}%' if task_total > 0 else '0%'

        overall_status = audit_history.form_overall_status if audit_history else 'N/A'
        total_score = str(audit_history.form_overall_score) if audit_history and audit_history.form_overall_score is not None else 'N/A'

        # Extract audit info answers (Audited Location, Ambient Temperature, Total Production)
        audited_location = ''
        ambient_temperature = ''
        total_production = ''

        if response.get('stages'):
            for stage in response['stages']:
                if (stage.get('is_audit_info') or stage.get('name') == 'Audit Info') and stage.get('questions'):
                    for q in stage['questions']:
                        q_text = q.get('question', '').strip().lower()
                        q_answer = ''
                        if q.get('answers') and q['answers'].get('answer'):
                            q_answer = str(q['answers']['answer'])
                        if 'audited location' in q_text or q_text == 'audited location':
                            audited_location = q_answer
                        elif 'ambient temperature' in q_text or q_text == 'ambient temperature':
                            ambient_temperature = q_answer
                        elif 'total production' in q_text or q_text == 'total production':
                            total_production = q_answer

        # Process stages and questions
        if response.get('stages'):
            sorted_stages = sorted(response['stages'], key=lambda s: int(s.get('order', 999)))
            for stage in sorted_stages:
                stage_name = stage.get('name', f"Stage {stage.get('order', '')}")
                # Skip Audit Info stage for per-question rows (already captured above)
                if stage.get('is_audit_info') or stage_name == 'Audit Info':
                    continue

                if stage.get('questions'):
                    for question in stage['questions']:
                        q_id = question.get('id')
                        q_text = question.get('question', '')
                        answer_obj = question.get('answers', {})
                        answer_val = answer_obj.get('answer', '') if answer_obj else ''
                        remarks = answer_obj.get('remarks', '') if answer_obj else ''

                        # Extract image URLs and text answer from main question
                        image_urls = extract_image_urls(answer_val) if answer_val else ''
                        text_answer = extract_text_answer(answer_val) if answer_val else ''

                        # Collect Image, Remarks, Consumed from, Quantity, After Image
                        # from sub_questions AND logic questions (conditional fields)
                        consumed_from = ''
                        quantity = ''
                        after_image = ''
                        logic_remarks = ''

                        # Check sub-questions
                        for sub_q in question.get('sub_questions', []):
                            sub_q_text = sub_q.get('question', '').strip().lower()
                            sub_answer = sub_q.get('answers', {})
                            sub_answer_val = sub_answer.get('answer', '') if sub_answer else ''
                            if 'consumed' in sub_q_text or 'sap code' in sub_q_text or 'product name' in sub_q_text:
                                consumed_from = str(sub_answer_val) if sub_answer_val else ''
                            elif 'quantity' in sub_q_text:
                                quantity = str(sub_answer_val) if sub_answer_val else ''
                            elif 'after image' in sub_q_text or 'after_image' in sub_q_text:
                                after_image = extract_image_urls(sub_answer_val)
                            elif 'image' in sub_q_text and sub_answer_val:
                                if not image_urls:
                                    image_urls = extract_image_urls(sub_answer_val)
                            elif 'remark' in sub_q_text and sub_answer_val:
                                logic_remarks = logic_remarks + ('; ' if logic_remarks else '') + str(sub_answer_val)

                        # Check logic questions (conditional on answer option selected)
                        for logic in question.get('logics', []):
                            for logic_q in logic.get('logic_questions', []):
                                lq_text = logic_q.get('question', '').strip().lower()
                                lq_answer_obj = logic_q.get('answers', {})
                                lq_answer_val = lq_answer_obj.get('answer', '') if lq_answer_obj else ''

                                if not lq_answer_val:
                                    continue

                                if 'consumed' in lq_text or 'sap code' in lq_text or 'product name' in lq_text:
                                    consumed_from = str(lq_answer_val)
                                elif 'quantity' in lq_text:
                                    quantity = str(lq_answer_val)
                                elif 'after image' in lq_text or 'after_image' in lq_text:
                                    after_image = extract_image_urls(lq_answer_val)
                                elif 'image' in lq_text or 'upload' in lq_text:
                                    if not image_urls:
                                        image_urls = extract_image_urls(lq_answer_val)
                                elif 'remark' in lq_text:
                                    logic_remarks = logic_remarks + ('; ' if logic_remarks else '') + str(lq_answer_val)

                        # Use logic remarks if main question has no remarks
                        final_remarks = remarks or logic_remarks

                        # Check if this question has followup tasks
                        followup_data_list = all_followup_map.get((response_id, q_id), [])

                        # Build metadata + followup columns data
                        fut_metadata_cols = [''] * num_metadata_cols
                        fut_close_cols = [''] * num_followup_cols  # title + Q/A pairs

                        if followup_data_list:
                            # For multiple followups, concatenate metadata and titles
                            first_fut = followup_data_list[0]
                            fut_close_cols[0] = first_fut['title']
                            fut_metadata_cols[0] = first_fut['task_id']
                            fut_metadata_cols[1] = first_fut['deadline']
                            fut_metadata_cols[2] = first_fut['task_status']
                            fut_metadata_cols[3] = first_fut['reopen_reason']
                            fut_metadata_cols[4] = first_fut['followup_submission_id']

                            close_qs = first_fut['close_questions']
                            for idx, cq in enumerate(close_qs):
                                col_q = 1 + idx * 2  # Question column index
                                col_a = 2 + idx * 2  # Answer column index
                                if col_q < num_followup_cols:
                                    fut_close_cols[col_q] = cq['question']
                                if col_a < num_followup_cols:
                                    fut_close_cols[col_a] = cq['answer']

                            # If multiple followups, append additional data
                            if len(followup_data_list) > 1:
                                extra_titles = '; '.join([f['title'] for f in followup_data_list[1:] if f['title']])
                                if extra_titles:
                                    fut_close_cols[0] = fut_close_cols[0] + '; ' + extra_titles
                                extra_task_ids = '; '.join([f['task_id'] for f in followup_data_list[1:] if f['task_id']])
                                if extra_task_ids:
                                    fut_metadata_cols[0] = fut_metadata_cols[0] + '; ' + extra_task_ids
                                extra_statuses = '; '.join([f['task_status'] for f in followup_data_list[1:] if f['task_status']])
                                if extra_statuses:
                                    fut_metadata_cols[2] = fut_metadata_cols[2] + '; ' + extra_statuses
                                extra_reopen_reasons = '; '.join([f['reopen_reason'] for f in followup_data_list[1:] if f['reopen_reason']])
                                if extra_reopen_reasons:
                                    fut_metadata_cols[3] = fut_metadata_cols[3] + '; ' + extra_reopen_reasons
                                extra_sub_ids = '; '.join([f['followup_submission_id'] for f in followup_data_list[1:] if f['followup_submission_id']])
                                if extra_sub_ids:
                                    fut_metadata_cols[4] = fut_metadata_cols[4] + '; ' + extra_sub_ids

                        # Build the row - metadata columns repeat on every question row
                        # Source Type and Source ID come from followup data if available, else from planner lookup
                        row_source_type = ''
                        row_source_id = ''
                        if followup_data_list:
                            row_source_type = followup_data_list[0].get('source_type', '')
                            row_source_id = followup_data_list[0].get('source_id', '')

                        # Determine source type/id from planner if no followup data
                        if not row_source_type:
                            try:
                                from planner.models import PlannerSubmission as _PS
                                planner_sub = _PS.objects.filter(
                                    form_submission_id=response_id
                                ).select_related('planner_assignment').first()
                                if planner_sub and planner_sub.planner_assignment:
                                    row_source_type = 'Planner'
                                    row_source_id = planner_sub.planner_assignment.order_id or str(planner_sub.planner_assignment_id)
                                else:
                                    row_source_type = 'Form'
                                    row_source_id = str(form_id)
                            except Exception:
                                row_source_type = 'Form'
                                row_source_id = str(form_id)

                        # Determine if this submission has bulk-imported tasks
                        row_imported = 'Yes' if (followup_data_list and any(f.get('is_bulk_imported') for f in followup_data_list)) else 'No'

                        if is_first_row:
                            row = [
                                display_response_id,
                                response.get('submission_initiated_on', 'N/A'),
                                response.get('submission_initiated_by', 'N/A'),
                                response.get('initiator_designation', 'N/A'),
                                response.get('initiator_department', 'N/A'),
                                'Completed' if response.get('is_completed', False) else 'Pending',
                                form_info.get('title', f'Form {form_id}'),
                                form_info.get('form_type', 'N/A'),
                                overall_status if is_audit else '',
                                total_score if is_audit else '',
                                task_completion_pct if is_audit else '',
                                overdue_pct if is_audit else '',
                                reopened_pct if is_audit else '',
                                audited_location,
                                ambient_temperature,
                                total_production,
                                row_source_type,
                                row_source_id,
                                row_imported,
                                stage_name,
                                q_text,
                                text_answer,
                                image_urls,
                                final_remarks,
                                consumed_from,
                                quantity,
                                after_image,
                            ] + fut_metadata_cols + fut_close_cols
                            is_first_row = False
                        else:
                            row = [
                                display_response_id,
                                response.get('submission_initiated_on', 'N/A'),
                                response.get('submission_initiated_by', 'N/A'),
                                response.get('initiator_designation', 'N/A'),
                                response.get('initiator_department', 'N/A'),
                                'Completed' if response.get('is_completed', False) else 'Pending',
                                form_info.get('title', f'Form {form_id}'),
                                form_info.get('form_type', 'N/A'),
                                overall_status if is_audit else '',
                                total_score if is_audit else '',
                                task_completion_pct if is_audit else '',
                                overdue_pct if is_audit else '',
                                reopened_pct if is_audit else '',
                                audited_location,
                                ambient_temperature,
                                total_production,
                                row_source_type,
                                row_source_id,
                                row_imported,
                                stage_name,
                                q_text,
                                text_answer,
                                image_urls,
                                final_remarks,
                                consumed_from,
                                quantity,
                                after_image,
                            ] + fut_metadata_cols + fut_close_cols
                        rows.append(row)

        # If no rows were added at all (no questions), add a metadata-only row
        if is_first_row:
            # Determine source type/id for this response
            row_source_type = 'Form'
            row_source_id = str(form_id)
            try:
                from planner.models import PlannerSubmission as _PS
                planner_sub = _PS.objects.filter(
                    form_submission_id=response_id
                ).select_related('planner_assignment').first()
                if planner_sub and planner_sub.planner_assignment:
                    row_source_type = 'Planner'
                    row_source_id = planner_sub.planner_assignment.order_id or str(planner_sub.planner_assignment_id)
            except Exception:
                pass
            # Check if any task for this submission is bulk imported
            meta_imported = 'No'
            try:
                from task.models import Task as TaskModel
                if TaskModel.objects.filter(form_submission_id=response_id, is_bulk_imported=True).exists():
                    meta_imported = 'Yes'
            except Exception:
                pass
            row = [
                display_response_id,
                response.get('submission_initiated_on', 'N/A'),
                response.get('submission_initiated_by', 'N/A'),
                response.get('initiator_designation', 'N/A'),
                response.get('initiator_department', 'N/A'),
                'Completed' if response.get('is_completed', False) else 'Pending',
                form_info.get('title', f'Form {form_id}'),
                form_info.get('form_type', 'N/A'),
                overall_status if is_audit else '',
                total_score if is_audit else '',
                task_completion_pct if is_audit else '',
                overdue_pct if is_audit else '',
                reopened_pct if is_audit else '',
                audited_location,
                ambient_temperature,
                total_production,
                row_source_type,
                row_source_id,
                meta_imported,
            ] + [''] * (len(base_headers) - 19) + [''] * num_metadata_cols + [''] * num_followup_cols
            rows.append(row)

    return {
        'Form Responses with Followup': [headers] + rows
    }


class FormResponseCSVFollowupView(APIView):
    """
    Endpoint to generate CSV with followup data and email it to the user.
    Uses the same async background generation pattern as FormResponseCSVView.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, form_id):
        try:
            email = request.query_params.get('email')
            if not email:
                return Response(
                    {"error": "Email parameter is required to send the report."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            submission_ids_param = request.query_params.get('submission_ids', '')
            if submission_ids_param:
                try:
                    submission_ids = [int(sid.strip()) for sid in submission_ids_param.split(',') if sid.strip()]
                except ValueError:
                    return Response({"error": "Invalid submission_ids format. Use comma-separated integers."}, status=status.HTTP_400_BAD_REQUEST)
            else:
                submission_id = request.query_params.get('submission_id')
                if submission_id:
                    try:
                        submission_ids = [int(submission_id)]
                    except ValueError:
                        return Response({"error": "Invalid submission_id format. Must be an integer."}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"error": "No submission IDs provided. Use submission_ids (comma-separated) or submission_id parameter."}, status=status.HTTP_400_BAD_REQUEST)

            if not Form.objects.filter(id=form_id, organization=request.user.organization).exists():
                return Response({"error": "Form not found."}, status=status.HTTP_404_NOT_FOUND)

            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            filename = f"{timestamp}_form_{form_id}_responses_followup.xlsx"
            user_org_id = request.user.organization.id

            tracking_id = str(uuid.uuid4())
            cache_key = f"excel_job:{tracking_id}"
            short_download_url = _build_public_download_url(
                request,
                reverse("report-download-redirect", kwargs={"tracking_id": tracking_id}),
            )
            cache.set(cache_key, {
                "status": "QUEUED",
                "message": "Report generation queued",
                "form_id": form_id,
                "submission_ids": submission_ids,
                "email": email,
                "filename": filename,
                "bucket_name": None,
                "s3_key": None,
                "expires_in_seconds": None,
                "created_at": datetime.now().isoformat(),
                "started_at": None,
                "completed_at": None,
                "error": None,
            }, timeout=86400)

            def background_csv_followup_generation(form_id, submission_ids, email, filename, user_org_id, tracking_id, short_download_url):
                import sys
                import traceback as tb_module

                print(f"[Background CSV+Followup Task] ========== BACKGROUND THREAD STARTED ==========")
                sys.stdout.flush()

                try:
                    from datetime import datetime as dt
                    print(f"[Background CSV+Followup Task] START: form_id={form_id}, submission_ids={submission_ids}, email={email} | Time: {dt.now()}")

                    try:
                        cache_key = f"excel_job:{tracking_id}"
                        job = cache.get(cache_key) or {}
                        job.update({
                            "status": "RUNNING",
                            "message": "Report generation started",
                            "started_at": dt.now().isoformat(),
                        })
                        cache.set(cache_key, job, timeout=86400)
                    except Exception:
                        pass

                    print(f"[Background CSV+Followup Task] Fetching form {form_id}... | Time: {dt.now()}")
                    try:
                        form = Form.objects.get(id=form_id, organization_id=user_org_id)
                    except Exception as form_fetch_err:
                        print(f"[Background CSV+Followup Task] ERROR fetching form: {str(form_fetch_err)}")
                        logger.error(f"[Background CSV+Followup Task] ERROR fetching form: {str(form_fetch_err)}")
                        raise

                    print(f"[Background CSV+Followup Task] Serializing form... | Time: {dt.now()}")
                    try:
                        base = (
                            Form.objects
                            .filter(is_deleted=False, pk=form.id)
                            .select_related('folder', 'form_admin', 'deletedBy', 'archivedBy')
                        )

                        stage_access_qs = StageAccess.objects.select_related('allow_user', 'allow_group', 'form', 'stage')

                        if form.form_type == FormType.AUDIT:
                            qs = base.prefetch_related(
                                'assignee__user', 'assignee__group', 'assignee__leader',
                                'audit_info__questions',
                                'audit_info__questions__options',
                                'audit_info__questions__child_questions',
                                'audit_info__questions__child_questions__options',
                                'audit_info__questions__logic_parent_question__logic_questions__options',
                                'audit_info__questions__logic_parent_question__follow_ups',
                                'audit_info__questions__logic_parent_question__follow_ups__task_close_questions__options',
                                'audit_group__questions',
                                'audit_group__questions__options',
                                'audit_group__questions__child_questions',
                                'audit_group__questions__child_questions__options',
                                'audit_group__questions__logic_parent_question__logic_questions__options',
                                'audit_group__questions__logic_parent_question__follow_ups',
                                'audit_group__questions__logic_parent_question__follow_ups__task_close_questions__options',
                            )
                        else:
                            question_qs = (
                                Question.objects
                                .select_related('form', 'stage', 'parent_question')
                                .prefetch_related(
                                    'options',
                                    'child_questions',
                                    'child_questions__options',
                                    'child_questions__child_questions',
                                    'child_questions__child_questions__options',
                                    'child_questions__logic_parent_question__logic_questions__options',
                                    'child_questions__logic_parent_question__follow_ups',
                                    'child_questions__logic_parent_question__follow_ups__task_close_questions__options',
                                    'logic_parent_question__logic_questions__options',
                                    'logic_parent_question__follow_ups',
                                    'logic_parent_question__follow_ups__task_close_questions__options',
                                )
                            )
                            stage_qs = (
                                Stage.objects
                                .select_related('form')
                                .prefetch_related(
                                    models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                                    models.Prefetch('questions', queryset=question_qs),
                                )
                            )
                            qs = base.prefetch_related(
                                'assignee__user', 'assignee__group', 'assignee__leader',
                                models.Prefetch('stages', queryset=stage_qs),
                            )

                        optimized_instance = qs.get(pk=form.id)
                        formSchema = FormSerializer(optimized_instance, many=False).data
                    except Exception as serializer_err:
                        print(f"[Background CSV+Followup Task] ERROR serializing form: {str(serializer_err)}")
                        logger.error(f"[Background CSV+Followup Task] ERROR serializing form: {str(serializer_err)}")
                        raise

                    print(f"[Background CSV+Followup Task] Fetching submissions {submission_ids}... | Time: {dt.now()}")
                    sys.stdout.flush()
                    try:
                        submissions = FormSubmision.objects.filter(
                            id__in=submission_ids,
                            form_id=form_id,
                            organization_id=user_org_id
                        ).select_related(
                            'submission_initiated_by',
                            'completed_by'
                        ).prefetch_related(
                            models.Prefetch(
                                'answers',
                                queryset=Answer.objects.select_related('question').prefetch_related('question__options'),
                                to_attr='prefetched_answers'
                            ),
                            'stage_submissions_history',
                            'group_submissions_history'
                        )
                    except Exception as submission_fetch_err:
                        print(f"[Background CSV+Followup Task] ERROR fetching submissions: {str(submission_fetch_err)}")
                        logger.error(f"[Background CSV+Followup Task] ERROR fetching submissions: {str(submission_fetch_err)}")
                        raise

                    if not submissions.exists():
                        try:
                            cache_key = f"excel_job:{tracking_id}"
                            job = cache.get(cache_key) or {}
                            job.update({
                                "status": "FAILED",
                                "message": "No submissions found for the provided IDs",
                                "completed_at": dt.now().isoformat(),
                                "error": "NO_SUBMISSIONS"
                            })
                            cache.set(cache_key, job, timeout=86400)
                        except Exception:
                            pass
                        return

                    print(f"[Background CSV+Followup Task] Found {submissions.count()} submission(s). Processing... | Time: {dt.now()}")
                    responses_data = []

                    for submission_instance in submissions:
                        submission_id = submission_instance.id
                        response_data = copy.deepcopy(formSchema)
                        response_data.update({
                            'id': submission_instance.id,
                            'submission_initiated_on': submission_instance.submission_initiated_on.isoformat() if submission_instance.submission_initiated_on else None,
                            'submission_initiated_by': getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                            'initiator_designation': getattr(submission_instance.submission_initiated_by, 'designation', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                            'initiator_department': getattr(submission_instance.submission_initiated_by, 'department', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                            'initiator_location': getattr(submission_instance.submission_initiated_by, 'location', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                            'current_owner': getattr(submission_instance.completed_by, 'first_name', '') + ' ' + getattr(submission_instance.completed_by, 'last_name', '') if submission_instance.completed_by else getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                            'is_completed': submission_instance.is_completed,
                        })

                        answer_lookup = {answer.question_id: answer for answer in submission_instance.prefetched_answers}

                        if response_data.get("form_type") == FormType.AUDIT:
                            stages = []
                            audit_info = response_data.get("audit_info")
                            if audit_info:
                                audit_info_stage = {
                                    'id': audit_info.get('id', 'audit_info'),
                                    'name': audit_info.get('name', 'Audit Info'),
                                    'is_audit_info': True,
                                    'order': 0,
                                    'questions': audit_info.get("questions", [])
                                }
                                for question in audit_info_stage['questions']:
                                    answer = answer_lookup.get(question['id'])
                                    question['answers'] = AnswerSerializer(answer).data if answer else {}
                                    for subQuestion in question.get('sub_questions', []):
                                        sub_answer = answer_lookup.get(subQuestion['id'])
                                        subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}
                                    for logics in question.get("logics", []):
                                        for logicQuestion in logics.get("logic_questions", []):
                                            logic_answer = answer_lookup.get(logicQuestion['id'])
                                            logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}
                                stages.append(audit_info_stage)

                            audit_groups_copy = copy.deepcopy(response_data.get("audit_group", []))
                            for audit_group in audit_groups_copy:
                                for audit_group_question in audit_group.get("questions", []):
                                    answer = answer_lookup.get(audit_group_question['id'])
                                    audit_group_question['answers'] = AnswerSerializer(answer).data if answer else {}
                                    for logics in audit_group_question.get("logics", []):
                                        for logicQuestion in logics.get("logic_questions", []):
                                            logic_answer = answer_lookup.get(logicQuestion['id'])
                                            logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}
                                    for subQuestion in audit_group_question.get('sub_questions', []):
                                        sub_answer = answer_lookup.get(subQuestion['id'])
                                        subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}
                                stages.append(audit_group)
                            response_data['stages'] = stages
                        else:
                            for stage in formSchema.get('stages', []):
                                stage_histories = StageSubmissionHistory.objects.filter(
                                    stage__id=stage['id'],
                                    form_submission__id=submission_id,
                                    organization_id=user_org_id
                                )
                                historyData = {}
                                if stage_histories.exists():
                                    stageHistory = stage_histories.first()
                                    historyData = StageSubmissionHistorySerializer(stageHistory, many=False).data
                                stage_data = {
                                    'id': stage['id'],
                                    'name': stage.get('name', f'Stage {stage.get("order", "")}'),
                                    'order': stage.get('order', 0),
                                    'is_completed': stage_histories.exists(),
                                    'completed_by': historyData.get('completed_by', None),
                                    'completed_on': historyData.get('completed_on', None),
                                    'questions': []
                                }
                                for question in stage.get('questions', []):
                                    question_data = copy.deepcopy(question)
                                    answer = answer_lookup.get(question['id'])
                                    question_data['answers'] = AnswerSerializer(answer).data if answer else {}
                                    for subQuestion in question_data.get('sub_questions', []):
                                        sub_answer = answer_lookup.get(subQuestion['id'])
                                        subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}
                                    for logic in question_data.get("logics", []):
                                        for logicQuestion in logic.get("logic_questions", []):
                                            logic_answer = answer_lookup.get(logicQuestion['id'])
                                            logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}
                                    stage_data['questions'].append(question_data)
                                response_data['stages'].append(stage_data)

                        responses_data.append(response_data)

                    form_info = {
                        'title': form.title or f'Form {form_id}',
                        'form_type': form.get_form_type_display() or 'standard',
                        'created_by': getattr(form.form_admin, 'first_name', '') + ' ' + getattr(form.form_admin, 'last_name', '') if form.form_admin else 'N/A',
                        'created_at': form.created_at.isoformat() if form.created_at else None,
                        'prefix': form.prefix or 'NPX'
                    }

                    print(f"[Background CSV+Followup Task] Generating CSV with followup data... | Time: {dt.now()}")
                    excel_data = generate_csv_with_followup_data(responses_data, form_info, form_id, user_org_id)

                    print(f"[Background CSV+Followup Task] Creating Excel workbook... | Time: {dt.now()}")
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        for sheet_name, data in excel_data.items():
                            if data:
                                headers = data[0]
                                rows = data[1:]
                                df = pd.DataFrame(rows, columns=headers)
                                df.to_excel(writer, sheet_name=sheet_name, index=False, engine_kwargs={'options': {'strings_to_urls': False}})

                                worksheet = writer.sheets[sheet_name]
                                for col_idx in range(len(headers)):
                                    try:
                                        column_letter = worksheet.cell(row=1, column=col_idx + 1).column_letter
                                        worksheet.column_dimensions[column_letter].width = 30
                                    except Exception:
                                        continue

                    output.seek(0)
                    print(f"[Background CSV+Followup Task] Excel file created. Size: {len(output.getvalue()) / (1024*1024):.2f} MB")

                    import os
                    upload_timestamp = dt.now().strftime("%d%m%y_%H%M%S")
                    local_dir = os.path.join(settings.BASE_DIR, "media", "reports", "excel", str(user_org_id))
                    os.makedirs(local_dir, exist_ok=True)
                    local_file_full_path = os.path.join(local_dir, f"{upload_timestamp}_{filename}")
                    with open(local_file_full_path, 'wb') as f:
                        f.write(output.getvalue())

                    local_file_path = f"reports/excel/{user_org_id}/{upload_timestamp}_{filename}"
                    bucket_name = getattr(settings, "S3_BUCKET_NAME_ANNOUNCEMENT", None) or settings.S3_BUCKET_NAME
                    expires_in_seconds = 86400
                    send_excel_link_email(email, short_download_url, filename, expires_in_seconds)
                    logger.info(f"[Background CSV+Followup Task] Report link sent successfully to {email} for form {form_id}")
                    print(f"[Background CSV+Followup Task] Email sent successfully to {email} for form {form_id}")

                    try:
                        cache_key = f"excel_job:{tracking_id}"
                        job = cache.get(cache_key) or {}
                        job.update({
                            "status": "SUCCESS",
                            "message": "Email sent",
                            "completed_at": dt.now().isoformat(),
                            "filename": filename,
                            "bucket_name": bucket_name,
                            "s3_key": local_file_path,
                            "expires_in_seconds": expires_in_seconds,
                        })
                        cache.set(cache_key, job, timeout=86400)
                    except Exception:
                        pass

                except Exception as e:
                    import traceback
                    logger.error(f"[Background CSV+Followup Task] Error generating report for form {form_id}: {str(e)}")
                    print(f"[Background CSV+Followup Task] ERROR for form {form_id}: {str(e)}")
                    print(f"[Background CSV+Followup Task] Traceback: {traceback.format_exc()}")
                    try:
                        cache_key = f"excel_job:{tracking_id}"
                        job = cache.get(cache_key) or {}
                        job.update({
                            "status": "FAILED",
                            "message": f"Error: {str(e)}",
                            "completed_at": datetime.now().isoformat(),
                            "error": str(e),
                        })
                        cache.set(cache_key, job, timeout=86400)
                    except Exception:
                        pass

            thread = threading.Thread(
                target=background_csv_followup_generation,
                args=(form_id, submission_ids, email, filename, user_org_id, tracking_id, short_download_url),
                daemon=True
            )
            thread.start()

            return Response({
                "message": "CSV+Followup report generation started. You will receive an email with the download link.",
                "tracking_id": tracking_id,
                "status": "QUEUED"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"FormResponseCSVFollowupView error: {str(e)}")
            return Response({"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class FormResponseFollowupTableView(APIView):
    """
    Returns CSV+Followup data as JSON for display in the Analytics tab table.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, form_id):
        try:
            if not Form.objects.filter(id=form_id, organization=request.user.organization).exists():
                return Response({"error": "Form not found."}, status=status.HTTP_404_NOT_FOUND)

            submission_ids_param = request.query_params.get('submission_ids', '')
            if submission_ids_param:
                try:
                    submission_ids = [int(sid.strip()) for sid in submission_ids_param.split(',') if sid.strip()]
                except ValueError:
                    return Response({"error": "Invalid submission_ids format."}, status=status.HTTP_400_BAD_REQUEST)
            else:
                submissions_qs = FormSubmision.objects.filter(
                    form_id=form_id,
                    organization=request.user.organization
                )
                submission_ids = list(submissions_qs.values_list('id', flat=True))

            if not submission_ids:
                return Response({"headers": [], "rows": []}, status=status.HTTP_200_OK)

            form = Form.objects.get(id=form_id, organization=request.user.organization)
            user_org_id = request.user.organization.id

            # Build form schema (reuse same prefetch logic)
            base = (
                Form.objects
                .filter(is_deleted=False, pk=form.id)
                .select_related('folder', 'form_admin', 'deletedBy', 'archivedBy')
            )

            stage_access_qs = StageAccess.objects.select_related('allow_user', 'allow_group', 'form', 'stage')

            if form.form_type == FormType.AUDIT:
                qs = base.prefetch_related(
                    'assignee__user', 'assignee__group', 'assignee__leader',
                    'audit_info__questions',
                    'audit_info__questions__options',
                    'audit_info__questions__child_questions',
                    'audit_info__questions__child_questions__options',
                    'audit_info__questions__logic_parent_question__logic_questions__options',
                    'audit_info__questions__logic_parent_question__follow_ups',
                    'audit_info__questions__logic_parent_question__follow_ups__task_close_questions__options',
                    'audit_group__questions',
                    'audit_group__questions__options',
                    'audit_group__questions__child_questions',
                    'audit_group__questions__child_questions__options',
                    'audit_group__questions__logic_parent_question__logic_questions__options',
                    'audit_group__questions__logic_parent_question__follow_ups',
                    'audit_group__questions__logic_parent_question__follow_ups__task_close_questions__options',
                )
            else:
                question_qs = (
                    Question.objects
                    .select_related('form', 'stage', 'parent_question')
                    .prefetch_related(
                        'options',
                        'child_questions',
                        'child_questions__options',
                        'child_questions__child_questions',
                        'child_questions__child_questions__options',
                        'child_questions__logic_parent_question__logic_questions__options',
                        'child_questions__logic_parent_question__follow_ups',
                        'child_questions__logic_parent_question__follow_ups__task_close_questions__options',
                        'logic_parent_question__logic_questions__options',
                        'logic_parent_question__follow_ups',
                        'logic_parent_question__follow_ups__task_close_questions__options',
                    )
                )
                stage_qs = (
                    Stage.objects
                    .select_related('form')
                    .prefetch_related(
                        models.Prefetch('access_parent_stage', queryset=stage_access_qs),
                        models.Prefetch('questions', queryset=question_qs),
                    )
                )
                qs = base.prefetch_related(
                    'assignee__user', 'assignee__group', 'assignee__leader',
                    models.Prefetch('stages', queryset=stage_qs),
                )

            optimized_instance = qs.get(pk=form.id)
            formSchema = FormSerializer(optimized_instance, many=False).data

            submissions = FormSubmision.objects.filter(
                id__in=submission_ids,
                form_id=form_id,
                organization_id=user_org_id
            ).select_related(
                'submission_initiated_by',
                'completed_by'
            ).prefetch_related(
                models.Prefetch(
                    'answers',
                    queryset=Answer.objects.select_related('question').prefetch_related('question__options'),
                    to_attr='prefetched_answers'
                ),
                'stage_submissions_history',
                'group_submissions_history'
            )

            responses_data = []
            for submission_instance in submissions:
                submission_id = submission_instance.id
                response_data = copy.deepcopy(formSchema)
                response_data.update({
                    'id': submission_instance.id,
                    'submission_initiated_on': submission_instance.submission_initiated_on.isoformat() if submission_instance.submission_initiated_on else None,
                    'submission_initiated_by': getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                    'initiator_designation': getattr(submission_instance.submission_initiated_by, 'designation', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                    'initiator_department': getattr(submission_instance.submission_initiated_by, 'department', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                    'initiator_location': getattr(submission_instance.submission_initiated_by, 'location', 'N/A') if submission_instance.submission_initiated_by else 'N/A',
                    'current_owner': getattr(submission_instance.completed_by, 'first_name', '') + ' ' + getattr(submission_instance.completed_by, 'last_name', '') if submission_instance.completed_by else getattr(submission_instance.submission_initiated_by, 'first_name', '') + ' ' + getattr(submission_instance.submission_initiated_by, 'last_name', '') if submission_instance.submission_initiated_by else 'N/A',
                    'is_completed': submission_instance.is_completed,
                })

                answer_lookup = {answer.question_id: answer for answer in submission_instance.prefetched_answers}

                if response_data.get("form_type") == FormType.AUDIT:
                    stages = []
                    audit_info = response_data.get("audit_info")
                    if audit_info:
                        audit_info_stage = {
                            'id': audit_info.get('id', 'audit_info'),
                            'name': audit_info.get('name', 'Audit Info'),
                            'is_audit_info': True,
                            'order': 0,
                            'questions': audit_info.get("questions", [])
                        }
                        for question in audit_info_stage['questions']:
                            answer = answer_lookup.get(question['id'])
                            question['answers'] = AnswerSerializer(answer).data if answer else {}
                            for subQuestion in question.get('sub_questions', []):
                                sub_answer = answer_lookup.get(subQuestion['id'])
                                subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}
                            for logics in question.get("logics", []):
                                for logicQuestion in logics.get("logic_questions", []):
                                    logic_answer = answer_lookup.get(logicQuestion['id'])
                                    logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}
                        stages.append(audit_info_stage)

                    audit_groups_copy = copy.deepcopy(response_data.get("audit_group", []))
                    for audit_group in audit_groups_copy:
                        for audit_group_question in audit_group.get("questions", []):
                            answer = answer_lookup.get(audit_group_question['id'])
                            audit_group_question['answers'] = AnswerSerializer(answer).data if answer else {}
                            for logics in audit_group_question.get("logics", []):
                                for logicQuestion in logics.get("logic_questions", []):
                                    logic_answer = answer_lookup.get(logicQuestion['id'])
                                    logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}
                            for subQuestion in audit_group_question.get('sub_questions', []):
                                sub_answer = answer_lookup.get(subQuestion['id'])
                                subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}
                        stages.append(audit_group)
                    response_data['stages'] = stages
                else:
                    for stage in formSchema.get('stages', []):
                        stage_histories = StageSubmissionHistory.objects.filter(
                            stage__id=stage['id'],
                            form_submission__id=submission_id,
                            organization_id=user_org_id
                        )
                        historyData = {}
                        if stage_histories.exists():
                            stageHistory = stage_histories.first()
                            historyData = StageSubmissionHistorySerializer(stageHistory, many=False).data
                        stage_data = {
                            'id': stage['id'],
                            'name': stage.get('name', f'Stage {stage.get("order", "")}'),
                            'order': stage.get('order', 0),
                            'is_completed': stage_histories.exists(),
                            'completed_by': historyData.get('completed_by', None),
                            'completed_on': historyData.get('completed_on', None),
                            'questions': []
                        }
                        for question in stage.get('questions', []):
                            question_data = copy.deepcopy(question)
                            answer = answer_lookup.get(question['id'])
                            question_data['answers'] = AnswerSerializer(answer).data if answer else {}
                            for subQuestion in question_data.get('sub_questions', []):
                                sub_answer = answer_lookup.get(subQuestion['id'])
                                subQuestion['answers'] = AnswerSerializer(sub_answer).data if sub_answer else {}
                            for logic in question_data.get("logics", []):
                                for logicQuestion in logic.get("logic_questions", []):
                                    logic_answer = answer_lookup.get(logicQuestion['id'])
                                    logicQuestion['answers'] = AnswerSerializer(logic_answer).data if logic_answer else {}
                            stage_data['questions'].append(question_data)
                        response_data['stages'].append(stage_data)

                responses_data.append(response_data)

            form_info = {
                'title': form.title or f'Form {form_id}',
                'form_type': form.get_form_type_display() or 'standard',
                'created_by': getattr(form.form_admin, 'first_name', '') + ' ' + getattr(form.form_admin, 'last_name', '') if form.form_admin else 'N/A',
                'created_at': form.created_at.isoformat() if form.created_at else None,
                'prefix': form.prefix or 'NPX'
            }

            excel_data = generate_csv_with_followup_data(responses_data, form_info, form_id, user_org_id)

            sheet_name = list(excel_data.keys())[0]
            all_data = excel_data[sheet_name]
            headers = all_data[0]
            # Convert all cells to strings for JSON serialization
            rows = [[str(cell) if cell is not None else '' for cell in row] for row in all_data[1:]]

            # ---- Excel download ----
            if request.query_params.get('download') == 'excel':
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df = pd.DataFrame(rows, columns=headers)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, engine_kwargs={'options': {'strings_to_urls': False}})

                    worksheet = writer.sheets[sheet_name]
                    for col_idx in range(len(headers)):
                        try:
                            column_letter = worksheet.cell(row=1, column=col_idx + 1).column_letter
                            worksheet.column_dimensions[column_letter].width = 30
                        except Exception:
                            continue

                output.seek(0)
                timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
                filename = f"{timestamp}_form_{form_id}_responses_followup.xlsx"
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

            # ---- Compute chart data ----
            response_idx = headers.index('Response') if 'Response' in headers else -1
            fut_status_idx = headers.index('Followup Task Status') if 'Followup Task Status' in headers else -1

            # 1) Response status: OK / Not OK (Deviation Corrected) / Not OK (not going to Close) - Open
            response_status_counts = {}
            if response_idx >= 0:
                for row in rows:
                    val = row[response_idx].strip()
                    if val:
                        response_status_counts[val] = response_status_counts.get(val, 0) + 1

            total_responses = sum(response_status_counts.values()) if response_status_counts else 0
            response_status_data = [
                {
                    'name': status,
                    'count': count,
                    'percentage': round((count / total_responses * 100), 1) if total_responses > 0 else 0
                }
                for status, count in response_status_counts.items()
            ]

            # 2) Followup task status: triggered (total NC), completed, not_started, in_progress
            fut_status_counts = {}
            if fut_status_idx >= 0:
                for row in rows:
                    val = row[fut_status_idx].strip()
                    if val:
                        # Split by '; ' in case of multiple followups concatenated
                        for s in val.split('; '):
                            s = s.strip()
                            if s:
                                fut_status_counts[s] = fut_status_counts.get(s, 0) + 1

            total_followups = sum(fut_status_counts.values()) if fut_status_counts else 0
            followup_status_data = [
                {
                    'name': status,
                    'count': count,
                    'percentage': round((count / total_followups * 100), 1) if total_followups > 0 else 0
                }
                for status, count in fut_status_counts.items()
            ]

            return Response({
                'headers': headers,
                'rows': rows,
                'response_status_data': response_status_data,
                'followup_status_data': followup_status_data,
                'total_followups': total_followups,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"FormResponseFollowupTableView error: {str(e)}")
            return Response({"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PreviousSubmissionsView(APIView):
    """
    Returns the last 5 completed submissions' answers for a given form,
    optionally filtered by location_id. Answers are grouped by question_id.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        form_id = request.query_params.get('form_id')
        location_id = request.query_params.get('location_id')
        exclude_submission_id = request.query_params.get('exclude_submission_id')

        if not form_id:
            return Response(
                {"error": "form_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            form = Form.objects.get(id=form_id, is_deleted=False)
        except Form.DoesNotExist:
            return Response(
                {"error": "Form not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            # Get completed submissions for this form, ordered by most recent first
            submissions_qs = FormSubmision.objects.filter(
                form_id=form_id,
                is_completed=True,
            ).order_by('-completed_on')

            # Exclude a specific submission if requested (e.g. current submission being edited)
            if exclude_submission_id:
                try:
                    submissions_qs = submissions_qs.exclude(id=int(exclude_submission_id))
                except (ValueError, TypeError):
                    pass

            # If location_id is provided, filter submissions that have answers with that location
            if location_id:
                try:
                    loc_id = int(location_id)
                    submissions_qs = submissions_qs.filter(
                        answers__location_id=loc_id
                    ).distinct()
                except (ValueError, TypeError):
                    pass

            # Take last 5 submissions
            submissions = list(submissions_qs[:5])

            # Build response: for each question_id, list of answers from recent submissions
            question_answers = {}

            for sub in submissions:
                answers_qs = Answer.objects.filter(submission=sub).select_related(
                    'question', 'submitted_by'
                ).prefetch_related('question__options')

                for ans in answers_qs:
                    qid = str(ans.question_id)
                    if qid not in question_answers:
                        question_answers[qid] = []

                    submitted_by_name = ""
                    if ans.submitted_by:
                        submitted_by_name = f"{ans.submitted_by.first_name} {ans.submitted_by.last_name}".strip()
                        if not submitted_by_name:
                            submitted_by_name = ans.submitted_by.username

                    # Resolve option IDs to text for choice-based questions
                    display_answer = ans.answer
                    if ans.question_type in ('multiple_choice', 'checkboxes', 'audit', 'linear_scale', 'dropdown'):
                        raw = ans.answer or ''
                        parts = [p.strip() for p in raw.split('|') if p.strip()] if '|' in raw else [raw.strip()]
                        # Build option ID→text map from prefetched data
                        opt_map = {}
                        if ans.question:
                            for opt in ans.question.options.all():
                                opt_map[opt.id] = opt.option
                        option_texts = []
                        for part in parts:
                            try:
                                opt_id = int(part)
                                if opt_id in opt_map:
                                    option_texts.append(opt_map[opt_id])
                                    continue
                            except (ValueError, TypeError):
                                pass
                            option_texts.append(part)
                        display_answer = ' | '.join(option_texts) if option_texts else raw

                    question_answers[qid].append({
                        'answer': display_answer,
                        'other_text': ans.other_text,
                        'submitted_by': submitted_by_name,
                        'submitted_on': ans.submitted_on.isoformat() if ans.submitted_on else None,
                        'completed_on': sub.completed_on.isoformat() if sub.completed_on else None,
                        'submission_id': sub.id,
                        'question_type': ans.question_type,
                    })

            return Response(
                {"question_answers": question_answers},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
