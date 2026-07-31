from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Max, Q
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.utils import timezone
import re
import logging
import openpyxl
from io import BytesIO

from .models import PlannerAssignment, PlannerSubmission, PlannerAssignType, PlannerReasonHistory, PlannerFolder, CollaborativeSubmission, CollaborativeSubmissionStatus, GroupDelegation, GroupDelegationStatus
from .serializers import (
    PlannerAssignmentSerializer,
    PlannerBulkUploadSerializer,
    PlannerShareSerializer,
    CollaborativeSubmissionSerializer,
    GroupDelegationSerializer
)
from form.models import Form, CustomUser, Groups, AuditGroup, FormSubmision, Question
from user.models import Locations
from task.models import Task, TaskAssignee, TaskAuditLog

logger = logging.getLogger(__name__)


# Helper functions from task/views.py
def parse_date_value(date_val):
    """Parse date value from various formats"""
    from datetime import timezone, timedelta
    ist_tz = timezone(timedelta(hours=5, minutes=30))
    
    if isinstance(date_val, datetime):
        if date_val.tzinfo is None:
            return date_val.replace(tzinfo=ist_tz)
        else:
            return date_val.astimezone(ist_tz)
    
    s = str(date_val).strip()
    if not s:
        raise ValueError('Empty date')
    
    s = s.replace('Z', '+00:00')
    formats = [
        '%Y-%m-%dT%H:%M:%S%z',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            if fmt == '%Y-%m-%d':
                dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=ist_tz)
            return dt
        except ValueError:
            continue
    
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ist_tz)
        return dt
    except Exception:
        raise ValueError(f'Unable to parse date: {s}')


def extract_form_identifier(form_value):
    """Extract form identifier (ID or name) from form value"""
    # Check if it's a URL
    if form_value.startswith('http'):
        # Extract form ID from URL like /forms/form-123
        import re
        match = re.search(r'/forms/form-(\d+)', form_value)
        if match:
            return 'id', match.group(1)
    return 'name', form_value


def validate_form_identifier(identifier_type, identifier_value, organization):
    """Validate form identifier and return form object"""
    if identifier_type == 'id':
        try:
            form = Form.objects.get(id=int(identifier_value), organization=organization, is_deleted=False)
            return form
        except Form.DoesNotExist:
            return None
    else:
        # Search by title - prefer non-archived, then latest (highest ID)
        forms = Form.objects.filter(title__icontains=identifier_value, organization=organization, is_deleted=False)
        if forms.exists():
            non_archived = forms.filter(is_archived=False)
            if non_archived.exists():
                return non_archived.order_by('-id').first()
            return forms.order_by('-id').first()
        return None


def get_form_by_identifier(identifier_type, identifier_value, organization):
    """Get form by identifier"""
    return validate_form_identifier(identifier_type, identifier_value, organization)


def parse_id_list(value):
    """Parse a comma-separated string of IDs into a list of integers."""
    if not value:
        return []
    ids = []
    for part in str(value).split(','):
        part = part.strip()
        if part:
            try:
                ids.append(int(part))
            except ValueError:
                continue
    return ids


def resolve_location(location_value, organization):
    """Resolve a location by ID or name for a given organization."""
    if not location_value:
        return None
    location_value = str(location_value).strip()
    # Try by ID first
    try:
        location_id = int(location_value)
        location = Locations.objects.filter(id=location_id, organization=organization).first()
        if location:
            return location
    except ValueError:
        pass
    # Try by name (case-insensitive)
    location = Locations.objects.filter(name__iexact=location_value, organization=organization).first()
    return location


class PlannerBulkUploadView(APIView):
    """
    Accept JSON payload from frontend for planner bulk import.
    Creates PlannerAssignment objects instead of Task objects.
    Endpoint: POST /api/planner/bulk-import/
    Expected payload: { "data": [ {planner_name, form, start_date, end_date, description, rowIndex}, ... ] }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            payload = request.data
            
            # Support both {data: [...]} and direct array payload
            if isinstance(payload, dict) and 'data' in payload:
                rows = payload['data']
            elif isinstance(payload, list):
                rows = payload
            else:
                return Response({
                    'error': 'Invalid payload. Expecting JSON array under "data" or root array.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            errors = []
            success_count = 0
            created_assignments = []  # Track created assignments for sharing step
            
            with transaction.atomic():
                for row in rows:
                    row_index = row.get('rowIndex') or row.get('row')
                    order_type = str(row.get('order_type') or row.get('planner_name', '')).strip()
                    form_title = str(row.get('form', '')).strip()
                    location_value = row.get('location', '')
                    start_date_raw = row.get('start_date', '')
                    end_date_raw = row.get('end_date', '')
                    description = str(row.get('description', '')).strip()
                    share_users = parse_id_list(row.get('share_with_users', ''))
                    share_groups = parse_id_list(row.get('share_with_groups', ''))
                    share_locations = parse_id_list(row.get('share_with_locations', ''))
                    
                    # Repeat planner settings (can be set per-row or globally for all rows)
                    repeat_enabled = bool(row.get('repeat_enabled', payload.get('repeat_enabled', False)))
                    repeat_interval_days = int(row.get('repeat_interval_days', payload.get('repeat_interval_days', 0)) or 0)
                    early_notification_days = int(row.get('early_notification_days', payload.get('early_notification_days', 0)) or 0)
                    
                    # Folder assignment (optional)
                    folder_id = row.get('folder_id', payload.get('folder_id'))
                    folder = None
                    if folder_id:
                        folder = PlannerFolder.objects.filter(id=folder_id, organization=request.user.organization).first()
                    
                    # Collaborative audit settings (global for all rows)
                    collaborative_enabled = bool(row.get('collaborative_enabled', payload.get('collaborative_enabled', False)))
                    team_leader_id = row.get('team_leader', payload.get('team_leader'))
                    team_leader = None
                    if collaborative_enabled and team_leader_id:
                        team_leader = CustomUser.objects.filter(id=team_leader_id, organization=request.user.organization).first()
                    
                    # Collaborative participants (users + groups selected as delegate pool)
                    collaborative_participant_users = row.get('collaborative_participant_users', payload.get('collaborative_participant_users', []))
                    collaborative_participant_groups = row.get('collaborative_participant_groups', payload.get('collaborative_participant_groups', []))
                    
                    if not order_type or not form_title or not start_date_raw or not end_date_raw:
                        errors.append({
                            'row': row_index or None,
                            'error': 'Missing required fields: order_type (planner_name), form, start_date, end_date'
                        })
                        continue
                    
                    # Extract and validate form identifier
                    identifier_type, identifier_value = extract_form_identifier(form_title)
                    form = get_form_by_identifier(identifier_type, identifier_value, request.user.organization)
                    if not form:
                        errors.append({
                            'row': row_index or None,
                            'error': f'Form "{form_title}" not found or not accessible'
                        })
                        continue
                    
                    # Parse dates
                    try:
                        start_date = parse_date_value(start_date_raw)
                        end_date = parse_date_value(end_date_raw)
                    except ValueError as ve:
                        errors.append({
                            'row': row_index or None,
                            'error': f'Invalid date format: {ve}'
                        })
                        continue
                    
                    if start_date > end_date:
                        errors.append({
                            'row': row_index or None,
                            'error': 'End date must be on or after start date'
                        })
                        continue
                    
                    # Resolve location
                    location = resolve_location(location_value, request.user.organization)
                    
                    order_id = PlannerAssignment.generate_order_id()
                    common_assignment_kwargs = {
                        'order_id': order_id,
                        'planner_name': order_type,
                        'form': form,
                        'location': location,
                        'start_date': start_date,
                        'end_date': end_date,
                        'description': description if description else None,
                        'organization': request.user.organization,
                        'created_by': request.user,
                        'repeat_enabled': repeat_enabled,
                        'repeat_interval_days': repeat_interval_days,
                        'early_notification_days': early_notification_days,
                        'folder': folder,
                        'collaborative_enabled': collaborative_enabled,
                        'team_leader': team_leader,
                    }
                    
                    # Create planner assignment(s)
                    try:
                        created_in_row = []
                        
                        if collaborative_enabled and team_leader:
                            # Collaborative mode: create ONE assignment for Team Leader only
                            assignment = PlannerAssignment.objects.create(
                                **common_assignment_kwargs,
                                user=team_leader,
                                assign_type=PlannerAssignType.USER
                            )
                            created_in_row.append(assignment)
                            
                            # Create CollaborativeSubmission with participant pool
                            collab_submission = CollaborativeSubmission.objects.create(
                                planner_assignment=assignment,
                                team_leader=team_leader,
                                status=CollaborativeSubmissionStatus.DRAFT,
                                participant_users=collaborative_participant_users,
                                participant_groups=collaborative_participant_groups,
                                organization=request.user.organization
                            )

                            # Create GroupDelegation records for each audit group in the form
                            audit_groups = AuditGroup.objects.filter(form=assignment.form).order_by('order')
                            for group in audit_groups:
                                GroupDelegation.objects.create(
                                    collaborative_submission=collab_submission,
                                    audit_group=group,
                                    group_order=group.order,
                                    assigned_by=team_leader,
                                    status=GroupDelegationStatus.UNASSIGNED,
                                    organization=request.user.organization
                                )
                        elif share_users or share_groups or share_locations:
                            # Directly create assignments for each share target
                            for user_id in share_users:
                                user = CustomUser.objects.filter(id=user_id, organization=request.user.organization).first()
                                if not user:
                                    errors.append({
                                        'row': row_index or None,
                                        'error': f'User with ID {user_id} not found'
                                    })
                                    continue
                                assignment = PlannerAssignment.objects.create(
                                    **common_assignment_kwargs,
                                    user=user,
                                    assign_type=PlannerAssignType.USER
                                )
                                created_in_row.append(assignment)
                            
                            for group_id in share_groups:
                                group = Groups.objects.filter(id=group_id, organization=request.user.organization).first()
                                if not group:
                                    errors.append({
                                        'row': row_index or None,
                                        'error': f'Group with ID {group_id} not found'
                                    })
                                    continue
                                assignment = PlannerAssignment.objects.create(
                                    **common_assignment_kwargs,
                                    group=group,
                                    assign_type=PlannerAssignType.GROUP
                                )
                                created_in_row.append(assignment)
                            
                            for location_id in share_locations:
                                loc = Locations.objects.filter(id=location_id, organization=request.user.organization).first()
                                if not loc:
                                    errors.append({
                                        'row': row_index or None,
                                        'error': f'Location with ID {location_id} not found'
                                    })
                                    continue
                                location_users = CustomUser.objects.filter(location=loc, organization=request.user.organization)
                                if not location_users.exists():
                                    errors.append({
                                        'row': row_index or None,
                                        'error': f'No users found in location {loc.name}'
                                    })
                                    continue
                                for user in location_users:
                                    assignment = PlannerAssignment.objects.create(
                                        **common_assignment_kwargs,
                                        user=user,
                                        assign_type=PlannerAssignType.USER
                                    )
                                    created_in_row.append(assignment)
                        else:
                            # No share targets: create a placeholder assignment for later sharing
                            assignment = PlannerAssignment.objects.create(
                                **common_assignment_kwargs,
                                assign_type=PlannerAssignType.USER  # Default, will be updated during sharing
                            )
                            created_in_row.append(assignment)
                        
                        for assignment in created_in_row:
                            # Set repeat_generation_date for repeat-enabled planners
                            if assignment.repeat_enabled and assignment.repeat_interval_days > 0:
                                from datetime import timedelta
                                assignment.repeat_generation_date = assignment.start_date + timedelta(days=assignment.repeat_interval_days)
                                assignment.save(update_fields=['repeat_generation_date'])
                            
                            created_assignments.append({
                                'planner_assignment_id': assignment.id,
                                'order_id': assignment.order_id,
                                'planner_name': assignment.planner_name,
                                'location': assignment.location.id if assignment.location else None,
                                'location_name': assignment.location.name if assignment.location else None,
                                'form_id': form.id,
                                'form_title': form.title,
                                'row_index': row_index,
                                'repeat_enabled': assignment.repeat_enabled,
                                'repeat_interval_days': assignment.repeat_interval_days,
                                'early_notification_days': assignment.early_notification_days,
                            })
                        
                        success_count += len(created_in_row)
                    except Exception as e:
                        errors.append({
                            'row': row_index or None,
                            'error': f'Unexpected error: {str(e)}'
                        })
            
            return Response({
                'success_count': success_count,
                'errors': errors,
                'total_processed': len(rows),
                'created_assignments': created_assignments  # Return for sharing step
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            import traceback
            logger.error(f"Planner bulk upload error: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class PlannerShareView(APIView):
    """
    Share planner assignments to users/groups/leaders.
    Endpoint: POST /api/planner/share/
    Expected payload: {
        "planner_assignment_id": <id>,
        "users": [<user_id>, ...],
        "groups": [<group_id>, ...],
        "leaders": [<leader_id>, ...]
    }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            serializer = PlannerShareSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            planner_assignment_id = serializer.validated_data['planner_assignment_id']
            users = serializer.validated_data['users']
            groups = serializer.validated_data['groups']
            leaders = serializer.validated_data['leaders']
            locations = serializer.validated_data['locations']
            
            # Get the original planner assignment
            original_assignment = get_object_or_404(
                PlannerAssignment,
                id=planner_assignment_id,
                organization=request.user.organization
            )
            order_id = original_assignment.order_id or PlannerAssignment.generate_order_id()
            
            # Create assignments for each user
            user_assignments = []
            for user_id in users:
                user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)
                assignment = PlannerAssignment.objects.create(
                    order_id=order_id,
                    planner_name=original_assignment.planner_name,
                    form=original_assignment.form,
                    location=original_assignment.location,
                    user=user,
                    assign_type=PlannerAssignType.USER,
                    start_date=original_assignment.start_date,
                    end_date=original_assignment.end_date,
                    description=original_assignment.description,
                    organization=request.user.organization,
                    created_by=request.user,
                    repeat_enabled=original_assignment.repeat_enabled,
                    repeat_interval_days=original_assignment.repeat_interval_days,
                    early_notification_days=original_assignment.early_notification_days,
                    folder=original_assignment.folder,
                )
                if assignment.repeat_enabled and assignment.repeat_interval_days > 0:
                    from datetime import timedelta
                    assignment.repeat_generation_date = assignment.start_date + timedelta(days=assignment.repeat_interval_days)
                    assignment.save(update_fields=['repeat_generation_date'])
                user_assignments.append(assignment.id)
            
            # Create assignments for each group
            group_assignments = []
            for group_id in groups:
                group = get_object_or_404(Groups, id=group_id, organization=request.user.organization)
                assignment = PlannerAssignment.objects.create(
                    order_id=order_id,
                    planner_name=original_assignment.planner_name,
                    form=original_assignment.form,
                    location=original_assignment.location,
                    group=group,
                    assign_type=PlannerAssignType.GROUP,
                    start_date=original_assignment.start_date,
                    end_date=original_assignment.end_date,
                    description=original_assignment.description,
                    organization=request.user.organization,
                    created_by=request.user,
                    repeat_enabled=original_assignment.repeat_enabled,
                    repeat_interval_days=original_assignment.repeat_interval_days,
                    early_notification_days=original_assignment.early_notification_days,
                    folder=original_assignment.folder,
                )
                if assignment.repeat_enabled and assignment.repeat_interval_days > 0:
                    from datetime import timedelta
                    assignment.repeat_generation_date = assignment.start_date + timedelta(days=assignment.repeat_interval_days)
                    assignment.save(update_fields=['repeat_generation_date'])
                group_assignments.append(assignment.id)
            
            # Create assignments for each leader
            leader_assignments = []
            for leader_id in leaders:
                leader = get_object_or_404(CustomUser, id=leader_id, organization=request.user.organization)
                assignment = PlannerAssignment.objects.create(
                    order_id=order_id,
                    planner_name=original_assignment.planner_name,
                    form=original_assignment.form,
                    location=original_assignment.location,
                    leader=leader,
                    assign_type=PlannerAssignType.LOCATION_LEADER,
                    start_date=original_assignment.start_date,
                    end_date=original_assignment.end_date,
                    description=original_assignment.description,
                    organization=request.user.organization,
                    created_by=request.user,
                    repeat_enabled=original_assignment.repeat_enabled,
                    repeat_interval_days=original_assignment.repeat_interval_days,
                    early_notification_days=original_assignment.early_notification_days,
                    folder=original_assignment.folder,
                )
                if assignment.repeat_enabled and assignment.repeat_interval_days > 0:
                    from datetime import timedelta
                    assignment.repeat_generation_date = assignment.start_date + timedelta(days=assignment.repeat_interval_days)
                    assignment.save(update_fields=['repeat_generation_date'])
                leader_assignments.append(assignment.id)
            
            # Create assignments for each location (share with all users in that location)
            location_assignments = []
            for location_id in locations:
                loc = get_object_or_404(Locations, id=location_id, organization=request.user.organization)
                location_users = CustomUser.objects.filter(location=loc, organization=request.user.organization)
                for user in location_users:
                    assignment = PlannerAssignment.objects.create(
                        order_id=order_id,
                        planner_name=original_assignment.planner_name,
                        form=original_assignment.form,
                        location=original_assignment.location,
                        user=user,
                        assign_type=PlannerAssignType.USER,
                        start_date=original_assignment.start_date,
                        end_date=original_assignment.end_date,
                        description=original_assignment.description,
                        organization=request.user.organization,
                        created_by=request.user,
                        repeat_enabled=original_assignment.repeat_enabled,
                        repeat_interval_days=original_assignment.repeat_interval_days,
                        early_notification_days=original_assignment.early_notification_days,
                        folder=original_assignment.folder,
                    )
                    if assignment.repeat_enabled and assignment.repeat_interval_days > 0:
                        from datetime import timedelta
                        assignment.repeat_generation_date = assignment.start_date + timedelta(days=assignment.repeat_interval_days)
                        assignment.save(update_fields=['repeat_generation_date'])
                    location_assignments.append(assignment.id)
            
            # Delete the original assignment (it was a placeholder)
            original_assignment.delete()
            
            return Response({
                'message': 'Planner shared successfully',
                'user_assignments': user_assignments,
                'group_assignments': group_assignments,
                'leader_assignments': leader_assignments,
                'location_assignments': location_assignments,
                'total_assignments': len(user_assignments) + len(group_assignments) + len(leader_assignments) + len(location_assignments)
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Planner share error: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class PlannerListView(APIView):
    """
    Get planner assignments for the current user.
    Endpoint: GET /api/planner/my-planners/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user = request.user
            logger.info(f"Fetching planners for user: {user.id}, org: {user.organization}")

            # Admin/staff users see all planners for their organization
            if user.is_staff or user.is_superuser:
                all_planners = list(PlannerAssignment.objects.filter(
                    organization=user.organization,
                    form__is_archived=False
                ).select_related('form', 'location', 'folder'))
            else:
                # Get planners assigned directly to user
                user_planners = PlannerAssignment.objects.filter(
                    user=user,
                    organization=user.organization,
                    form__is_archived=False
                ).select_related('form', 'location', 'folder')

                # Get planners assigned to user's groups
                user_groups = user.user_groups.all() if hasattr(user, 'user_groups') else Groups.objects.none()
                group_planners = PlannerAssignment.objects.filter(
                    group__in=user_groups,
                    organization=user.organization,
                    form__is_archived=False
                ).select_related('form', 'location', 'folder')

                # Get planners assigned to user as location leader
                leader_planners = PlannerAssignment.objects.filter(
                    leader=user,
                    organization=user.organization,
                    form__is_archived=False
                ).select_related('form', 'location', 'folder')

                # Combine all planners
                all_planners = list(user_planners) + list(group_planners) + list(leader_planners)

                # Also include collaborative planners where user is a participant
                user_group_ids = list(user.user_groups.values_list('id', flat=True)) if hasattr(user, 'user_groups') else []
                all_collab_subs = CollaborativeSubmission.objects.filter(
                    organization=user.organization
                ).exclude(planner_assignment__user=user).select_related('planner_assignment')
                collab_planner_ids = []
                for cs in all_collab_subs:
                    p_users = cs.participant_users or []
                    p_groups = cs.participant_groups or []
                    if user.id in p_users or any(gid in p_groups for gid in user_group_ids):
                        collab_planner_ids.append(cs.planner_assignment_id)
                if collab_planner_ids:
                    collab_planners = PlannerAssignment.objects.filter(
                        id__in=collab_planner_ids,
                        organization=user.organization,
                        form__is_archived=False
                    ).select_related('form', 'location', 'folder')
                    all_planners.extend(list(collab_planners))

            # Remove duplicates - deduplicate by order_id so a user with both
            # a direct assignment and a group/leader assignment for the same
            # planner only sees one entry. Prefer direct user assignments.
            seen_order_ids = set()
            unique_planners = []
            # Sort so that user assignments come before group/leader assignments
            all_planners.sort(key=lambda p: 0 if p.user_id is not None else 1)
            for planner in all_planners:
                dedup_key = planner.order_id or f'no-order-{planner.id}'
                if dedup_key not in seen_order_ids:
                    seen_order_ids.add(dedup_key)
                    # If early_notification_days is set, only show the planner
                    # when current date is within the notification window before start_date
                    if planner.early_notification_days and planner.early_notification_days > 0:
                        from datetime import timedelta
                        notification_start = planner.start_date - timedelta(days=planner.early_notification_days)
                        if timezone.now() < notification_start:
                            continue  # Don't show yet — too early
                    unique_planners.append(planner)
            
            # Serialize
            planners_data = []
            for planner in unique_planners:
                planners_data.append({
                    'id': planner.id,
                    'order_id': planner.order_id,
                    'planner_name': planner.planner_name,
                    'order_type': planner.planner_name,
                    'location': planner.location.id if planner.location else None,
                    'location_name': planner.location.name if planner.location else None,
                    'form_id': planner.form.id,
                    'form_title': planner.form.title,
                    'form_type': planner.form.form_type,
                    'assign_type': planner.assign_type,
                    'start_date': planner.start_date.isoformat(),
                    'end_date': planner.end_date.isoformat(),
                    'description': planner.description,
                    'is_completed': planner.is_completed,
                    'non_completion_reason': planner.non_completion_reason,
                    'reason_status': planner.reason_status,
                    'rejection_reason': planner.rejection_reason,
                    'rejection_questions': planner.rejection_questions or [],
                    'rejection_answers': planner.rejection_answers or [],
                    'extended_due_date': planner.extended_due_date.isoformat() if planner.extended_due_date else None,
                    'extension_note': planner.extension_note,
                    'planner_shared_on': planner.planner_shared_on.isoformat(),
                    'started_by': getattr(planner.started_by, 'username', None) if planner.started_by else None,
                    'started_on': planner.started_on.isoformat() if planner.started_on else None,
                    'repeat_enabled': planner.repeat_enabled,
                    'repeat_interval_days': planner.repeat_interval_days,
                    'early_notification_days': planner.early_notification_days,
                    'parent_planner_id': planner.parent_planner_id,
                    'folder_id': planner.folder.id if planner.folder else None,
                    'folder_name': planner.folder.name if planner.folder else None,
                    'folder_color': planner.folder.color if planner.folder else None,
                    'collaborative_enabled': planner.collaborative_enabled,
                    'team_leader': planner.team_leader_id,
                })

            logger.info(f"Successfully fetched {len(planners_data)} planners for user {user.id}")
            return Response(planners_data, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Get planners error for user {request.user.id}: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to fetch planners: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerAllView(APIView):
    """
    Get all planners for the organization (for web admin view).
    Endpoint: GET /api/planner/all-planners/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user = request.user
            logger.info(f"Fetching all planners for organization: {user.organization}")
            
            # Get all planners for the organization
            all_planners = PlannerAssignment.objects.filter(
                organization=user.organization
            ).select_related('form', 'location', 'folder')
            
            # Serialize
            planners_data = []
            for planner in all_planners:
                # Get reason history
                history_qs = planner.reason_history.all().order_by('acted_on')
                history_data = []
                for h in history_qs:
                    history_data.append({
                        'id': h.id,
                        'action': h.action,
                        'non_completion_reason': h.non_completion_reason,
                        'rejection_reason': h.rejection_reason,
                        'rejection_questions': h.rejection_questions or [],
                        'rejection_answers': h.rejection_answers or [],
                        'extended_due_date': h.extended_due_date.isoformat() if h.extended_due_date else None,
                        'extension_note': h.extension_note,
                        'acted_by': h.acted_by.username if h.acted_by else None,
                        'acted_on': h.acted_on.isoformat() if h.acted_on else None,
                        'cycle_number': h.cycle_number,
                    })

                # Get group member names if assigned to a group
                group_members = []
                if planner.group:
                    group_members = list(planner.group.members.filter(
                        is_deleted=False, is_active=True
                    ).values_list('username', flat=True))

                # Get collaborative group delegation details if collaborative
                collab_groups = []
                if planner.collaborative_enabled:
                    collab = CollaborativeSubmission.objects.filter(
                        planner_assignment=planner,
                        organization=user.organization
                    ).first()
                    if collab:
                        for gd in collab.group_delegations.all().order_by('group_order'):
                            # Calculate answered/total for this group
                            from form.models import Question, Answer
                            main_q_ids = list(Question.objects.filter(
                                form=planner.form,
                                audit_group=gd.audit_group,
                                parent_question__isnull=True,
                                question_type='audit'
                            ).values_list('id', flat=True))
                            total_q = len(main_q_ids)
                            answered_q = 0
                            if collab.form_submission and main_q_ids:
                                answered_q = Answer.objects.filter(
                                    submission=collab.form_submission,
                                    question_id__in=main_q_ids,
                                    organization=user.organization
                                ).count()
                            progress_pct = round((answered_q / total_q) * 100) if total_q > 0 else 0
                            collab_groups.append({
                                'id': gd.id,
                                'audit_group_id': gd.audit_group_id,
                                'audit_group_name': gd.audit_group.name if gd.audit_group else None,
                                'group_order': gd.group_order,
                                'status': gd.status,
                                'assigned_user_names': list(gd.assigned_users.values_list('username', flat=True)),
                                'submitted_by_name': gd.submitted_by.username if gd.submitted_by else None,
                                'submitted_on': gd.submitted_on.isoformat() if gd.submitted_on else None,
                                'reviewed_on': gd.reviewed_on.isoformat() if gd.reviewed_on else None,
                                'rejection_comment': gd.rejection_comment,
                                'total_questions': total_q,
                                'answered_count': answered_q,
                                'progress_percentage': progress_pct,
                            })

                planners_data.append({
                    'id': planner.id,
                    'order_id': planner.order_id,
                    'planner_name': planner.planner_name,
                    'order_type': planner.planner_name,
                    'location': planner.location.id if planner.location else None,
                    'location_name': planner.location.name if planner.location else None,
                    'form_id': planner.form.id,
                    'form_title': planner.form.title,
                    'form_type': planner.form.form_type,
                    'assign_type': planner.assign_type,
                    'start_date': planner.start_date.isoformat(),
                    'end_date': planner.end_date.isoformat(),
                    'description': planner.description,
                    'is_completed': planner.is_completed,
                    'non_completion_reason': planner.non_completion_reason,
                    'reason_status': planner.reason_status,
                    'rejection_reason': planner.rejection_reason,
                    'rejection_questions': planner.rejection_questions or [],
                    'rejection_answers': planner.rejection_answers or [],
                    'extended_due_date': planner.extended_due_date.isoformat() if planner.extended_due_date else None,
                    'extension_note': planner.extension_note,
                    'extended_by': planner.extended_by.username if planner.extended_by else None,
                    'extended_on': planner.extended_on.isoformat() if planner.extended_on else None,
                    'planner_shared_on': planner.planner_shared_on.isoformat(),
                    'user': planner.user.username if planner.user else None,
                    'group': planner.group.name if planner.group else None,
                    'group_members': group_members,
                    'leader': planner.leader.username if planner.leader else None,
                    'completed_by': planner.completed_by.username if planner.completed_by else None,
                    'completed_on': planner.completed_on.isoformat() if planner.completed_on else None,
                    'started_by': planner.started_by.username if planner.started_by else None,
                    'started_on': planner.started_on.isoformat() if planner.started_on else None,
                    'reason_history': history_data,
                    'repeat_enabled': planner.repeat_enabled,
                    'repeat_interval_days': planner.repeat_interval_days,
                    'early_notification_days': planner.early_notification_days,
                    'parent_planner_id': planner.parent_planner_id,
                    'repeat_generation_date': planner.repeat_generation_date.isoformat() if planner.repeat_generation_date else None,
                    'folder_id': planner.folder.id if planner.folder else None,
                    'folder_name': planner.folder.name if planner.folder else None,
                    'folder_color': planner.folder.color if planner.folder else None,
                    'collaborative_enabled': planner.collaborative_enabled,
                    'team_leader': planner.team_leader.username if planner.team_leader else None,
                    'team_leader_id': planner.team_leader_id,
                    'collaborative_groups': collab_groups,
                })
            
            logger.info(f"Successfully fetched {len(planners_data)} planners for organization {user.organization}")
            return Response(planners_data, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Get all planners error for organization {request.user.organization}: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to fetch planners: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerStartView(APIView):
    """
    Mark a planner as started when user clicks on it.
    Endpoint: POST /api/planner/<planner_id>/start/
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, planner_id):
        try:
            user = request.user
            logger.info(f"Marking planner {planner_id} as started by user {user.id}")
            
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=user.organization,
                is_completed=False
            )
            
            # Check if user has access to this planner
            user_groups = user.user_groups.all() if hasattr(user, 'user_groups') else []
            if not (planner.user == user or planner.group in user_groups or planner.leader == user):
                return Response({'error': 'You do not have access to this planner'}, status=status.HTTP_403_FORBIDDEN)
            
            # Mark as started
            planner.started_by = user
            planner.started_on = timezone.now()
            planner.save()
            
            # Remove all other sibling assignments with the same order_id
            # so other users no longer see this planner on their mobile (first-come-first-served)
            if planner.order_id:
                siblings = PlannerAssignment.objects.filter(
                    order_id=planner.order_id,
                    organization=user.organization
                ).exclude(id=planner.id)
                deleted_count = siblings.count()
                siblings.delete()
                if deleted_count:
                    logger.info(f"Removed {deleted_count} sibling assignments for order_id {planner.order_id} after user {user.id} started")
            
            logger.info(f"Planner {planner_id} marked as started by user {user.id}")
            return Response({
                'message': 'Planner marked as started',
                'planner_id': planner.id,
                'order_id': planner.order_id,
                'form_id': planner.form.id,
                'form_title': planner.form.title,
                'location': planner.location.id if planner.location else None,
                'location_name': planner.location.name if planner.location else None,
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Mark planner as started error: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to mark planner as started: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerNonCompletionReasonView(APIView):
    """
    Record the reason for not completing an overdue planner.
    Endpoint: POST /api/planner/<planner_id>/non-completion-reason/
    Expected payload: { "reason": "...", "answers": [{"question_id": "...", "answer": "..."}, ...] }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, planner_id):
        try:
            user = request.user
            reason = request.data.get('reason', '').strip()
            answers = request.data.get('answers', [])

            if not reason:
                return Response({'error': 'Reason is required'}, status=status.HTTP_400_BAD_REQUEST)

            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=user.organization
            )

            user_groups = user.user_groups.all() if hasattr(user, 'user_groups') else []
            if not (planner.user == user or planner.group in user_groups or planner.leader == user):
                return Response({'error': 'You do not have access to this planner'}, status=status.HTTP_403_FORBIDDEN)

            if planner.non_completion_reason and planner.reason_status != 'rejected':
                return Response({'error': 'Reason has already been submitted and cannot be edited'}, status=status.HTTP_400_BAD_REQUEST)

            planner.non_completion_reason = reason
            planner.reason_status = 'pending'
            planner.rejection_reason = None
            if answers:
                planner.rejection_answers = answers
            planner.save()

            # Record history
            cycle_count = PlannerReasonHistory.objects.filter(planner=planner).count()
            PlannerReasonHistory.objects.create(
                planner=planner,
                action='submitted',
                non_completion_reason=reason,
                rejection_answers=answers if answers else [],
                acted_by=user,
                acted_on=timezone.now(),
                cycle_number=cycle_count + 1,
            )

            logger.info(f"Non-completion reason recorded for planner {planner_id} by user {user.id}")
            return Response({'message': 'Reason recorded successfully'}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Record non-completion reason error: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to record reason: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerExtendDueDateView(APIView):
    """
    Admin extends the due date of a planner with a note to the user.
    Endpoint: POST /api/planner/<planner_id>/extend-due-date/
    Expected payload: { "extended_due_date": "YYYY-MM-DD", "extension_note": "..." }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, planner_id):
        try:
            user = request.user
            extended_due_date = request.data.get('extended_due_date', '').strip()
            extension_note = request.data.get('extension_note', '').strip()

            if not extended_due_date:
                return Response({'error': 'Extended due date is required'}, status=status.HTTP_400_BAD_REQUEST)
            if not extension_note:
                return Response({'error': 'Extension note is required'}, status=status.HTTP_400_BAD_REQUEST)

            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=user.organization
            )

            from datetime import datetime
            try:
                parsed_date = datetime.strptime(extended_due_date, '%Y-%m-%d')
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)

            planner.extended_due_date = parsed_date
            planner.extension_note = extension_note
            planner.extended_by = user
            planner.extended_on = timezone.now()
            planner.end_date = parsed_date
            planner.reason_status = 'approved'
            planner.save()

            # Record history
            cycle_count = PlannerReasonHistory.objects.filter(planner=planner).count()
            PlannerReasonHistory.objects.create(
                planner=planner,
                action='approved',
                non_completion_reason=planner.non_completion_reason,
                rejection_answers=planner.rejection_answers or [],
                extended_due_date=parsed_date,
                extension_note=extension_note,
                acted_by=user,
                acted_on=timezone.now(),
                cycle_number=cycle_count + 1,
            )

            logger.info(f"Due date extended for planner {planner_id} by admin {user.id} to {extended_due_date}")
            return Response({
                'message': 'Due date extended successfully',
                'extended_due_date': extended_due_date,
                'extension_note': extension_note
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Extend due date error: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to extend due date: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerRejectReasonView(APIView):
    """
    Admin rejects the non-completion reason and sends feedback to the user for a new reason.
    Endpoint: POST /api/planner/<planner_id>/reject-reason/
    Expected payload: { "rejection_reason": "...", "rejection_questions": [{"type": "short_answer", "title": "...", "required": true}, ...] }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, planner_id):
        try:
            user = request.user
            rejection_reason = request.data.get('rejection_reason', '').strip()
            rejection_questions = request.data.get('rejection_questions', [])

            if not rejection_reason:
                return Response({'error': 'Rejection reason is required'}, status=status.HTTP_400_BAD_REQUEST)

            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=user.organization
            )

            if not planner.non_completion_reason:
                return Response({'error': 'No non-completion reason to reject'}, status=status.HTTP_400_BAD_REQUEST)

            planner.reason_status = 'rejected'
            planner.rejection_reason = rejection_reason
            planner.rejection_questions = rejection_questions
            planner.rejection_answers = []
            planner.non_completion_reason = None
            planner.save()

            # Record history
            cycle_count = PlannerReasonHistory.objects.filter(planner=planner).count()
            PlannerReasonHistory.objects.create(
                planner=planner,
                action='rejected',
                non_completion_reason=None,
                rejection_reason=rejection_reason,
                rejection_questions=rejection_questions,
                acted_by=user,
                acted_on=timezone.now(),
                cycle_number=cycle_count + 1,
            )

            logger.info(f"Non-completion reason rejected for planner {planner_id} by admin {user.id}")
            return Response({
                'message': 'Reason rejected successfully',
                'rejection_reason': rejection_reason,
                'rejection_questions': rejection_questions
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Reject reason error: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to reject reason: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerCompleteView(APIView):
    """
    Mark a planner as completed and extract follow-up tasks.
    Endpoint: POST /api/planner/complete/
    Expected payload: {
        "planner_assignment_id": <id>,
        "form_submission_id": <id>
    }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            planner_assignment_id = request.data.get('planner_assignment_id')
            form_submission_id = request.data.get('form_submission_id')
            
            if not planner_assignment_id or not form_submission_id:
                return Response({
                    'error': 'planner_assignment_id and form_submission_id are required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get planner assignment
            planner_assignment = get_object_or_404(
                PlannerAssignment,
                id=planner_assignment_id,
                organization=request.user.organization
            )
            
            # Mark planner as completed
            planner_assignment.is_completed = True
            planner_assignment.completed_on = timezone.now()
            planner_assignment.completed_by = request.user
            planner_assignment.save()
            
            # Create planner submission record
            from form.models import FormSubmision
            form_submission = get_object_or_404(
                FormSubmision,
                id=form_submission_id,
                form=planner_assignment.form,
                organization=request.user.organization
            )
            
            planner_submission = PlannerSubmission.objects.create(
                planner_assignment=planner_assignment,
                form_submission=form_submission,
                submitted_by=request.user
            )
            
            # Extract follow-up tasks from the form
            followup_tasks_created = self.extract_followup_tasks(
                planner_assignment,
                form_submission,
                request.user
            )
            
            planner_submission.followup_tasks_created = followup_tasks_created
            planner_submission.save()
            
            return Response({
                'message': 'Planner completed successfully',
                'followup_tasks_created': followup_tasks_created
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Planner complete error: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    def extract_followup_tasks(self, planner_assignment, form_submission, user):
        """
        Extract follow-up tasks from LogicFollowUp in the form and create Task objects.
        Mirrors the logic used by the form submission flow in form/views.py.
        """
        from form.models import LogicFollowUp, Logic
        from django.db.models import Q as QObj

        main_form = planner_assignment.form
        submission_date = form_submission.submission_initiated_on or timezone.now()

        # Get all LogicFollowUp for this form that have followup_toggle enabled
        # and an assigned form (same query pattern as form submission flow)
        logic_followups = LogicFollowUp.objects.filter(
            QObj(form_id=main_form.id) | QObj(audit_group__form_id=main_form.id),
            followup_toggle=True,
            assign_form__isnull=False
        ).select_related('logic', 'assign_form', 'question')

        tasks_created = 0

        with transaction.atomic():
            for followup in logic_followups:
                try:
                    assigned_form = followup.assign_form
                    deadline_days = getattr(followup, 'deadline', 7) or 7

                    # Check for duplicate task for this submission/question/form combo
                    existing_task = Task.objects.filter(
                        followup_task_form_id=main_form,
                        organization=planner_assignment.organization,
                        follow_task_sub_question=followup.question,
                        form=assigned_form,
                        start_date=submission_date
                    ).exists()

                    if existing_task:
                        continue

                    # Create task from follow-up
                    # form = the assigned follow-up form (what the assignee needs to fill)
                    # followup_task_form_id = the main/planner form that triggered it
                    # follow_task_sub_question = the question that triggered the follow-up
                    task = Task.objects.create(
                        task_name=followup.title or 'Follow-up Task',
                        description=followup.description or '',
                        form=assigned_form,
                        followup_task_form_id=main_form,
                        follow_task_sub_question=followup.question,
                        organization=planner_assignment.organization,
                        start_date=submission_date,
                        end_date=submission_date + timedelta(days=deadline_days),
                        created_by=user,
                        status='not_started'
                    )

                    # Create audit log with the same action name as form submission flow
                    # so the serializer's _get_main_form_submission can find it
                    TaskAuditLog.objects.create(
                        task=task,
                        task_action='Followup_Created',
                        action_by=user
                    )

                    # Assign task based on follow-up configuration
                    self.assign_followup_task(task, followup, planner_assignment, user)

                    tasks_created += 1

                except Exception as e:
                    logger.error(f"Error creating follow-up task: {str(e)}")
                    continue

        return tasks_created > 0
    
    def assign_followup_task(self, task, followup, planner_assignment, user):
        """
        Assign the follow-up task based on the LogicFollowUp configuration.
        """
        from form.models import CustomUser, Groups
        
        assign_to = followup.assign_to
        
        if assign_to == 'form_submitter':
            # Assign to the user who completed the planner
            TaskAssignee.objects.create(
                task=task,
                assigned_user=user,
                assigned_by=user
            )
            task.status = 'assigned'
            task.save()
        
        elif assign_to == 'user':
            # Assign to specific user
            if followup.user:
                TaskAssignee.objects.create(
                    task=task,
                    assigned_user=followup.user,
                    assigned_by=user
                )
                task.status = 'assigned'
                task.save()
            
            # Also assign to users in assign_user_ids
            for user_id in followup.assign_user_ids:
                try:
                    assign_user = CustomUser.objects.get(id=user_id, organization=task.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_user=assign_user,
                        assigned_by=user
                    )
                    task.status = 'assigned'
                except CustomUser.DoesNotExist:
                    continue
        
        elif assign_to == 'group':
            # Assign to group
            if followup.group:
                TaskAssignee.objects.create(
                    task=task,
                    assigned_group=followup.group,
                    assigned_by=user
                )
                task.status = 'assigned'
                task.save()
            
            # Also assign to groups in assign_group_ids
            for group_id in followup.assign_group_ids:
                try:
                    assign_group = Groups.objects.get(id=group_id, organization=task.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_group=assign_group,
                        assigned_by=user
                    )
                    task.status = 'assigned'
                except Groups.DoesNotExist:
                    continue
        
        elif assign_to == 'leader':
            # Assign to location leader
            if followup.leader:
                TaskAssignee.objects.create(
                    task=task,
                    assigned_user=followup.leader,
                    assigned_by=user
                )
                task.status = 'assigned'
                task.save()
            
            # Also assign to leaders in assign_leader_ids
            for leader_id in followup.assign_leader_ids:
                try:
                    assign_leader = CustomUser.objects.get(id=leader_id, organization=task.organization)
                    TaskAssignee.objects.create(
                        task=task,
                        assigned_user=assign_leader,
                        assigned_by=user
                    )
                    task.status = 'assigned'
                except CustomUser.DoesNotExist:
                    continue


class PlannerUpdateView(APIView):
    """
    Update a planner assignment (admin only).
    Endpoint: PUT /api/planner/<planner_id>/update/
    """
    permission_classes = [IsAuthenticated]

    def put(self, request, planner_id):
        try:
            user = request.user
            logger.info(f"Updating planner {planner_id} by user {user.id}")
            
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=user.organization
            )
            
            # Update fields
            planner_name = request.data.get('planner_name') or request.data.get('order_type')
            form_id = request.data.get('form_id')
            start_date = request.data.get('start_date')
            end_date = request.data.get('end_date')
            description = request.data.get('description')
            location_value = request.data.get('location')
            
            # Repeat settings — admin can edit these for any planner status
            repeat_enabled = request.data.get('repeat_enabled')
            repeat_interval_days = request.data.get('repeat_interval_days')
            early_notification_days = request.data.get('early_notification_days')
            folder_id = request.data.get('folder_id')
            
            if planner_name:
                planner.planner_name = planner_name
            if form_id:
                form = get_object_or_404(Form, id=form_id, organization=user.organization)
                planner.form = form
            if start_date:
                planner.start_date = parse_date_value(start_date)
            if end_date:
                planner.end_date = parse_date_value(end_date)
            if description is not None:
                planner.description = description
            if location_value is not None:
                if location_value:
                    location = resolve_location(location_value, user.organization)
                    if not location:
                        return Response({'error': f'Location "{location_value}" not found'}, status=status.HTTP_400_BAD_REQUEST)
                    planner.location = location
                else:
                    planner.location = None
            
            # Update repeat settings if provided
            if repeat_enabled is not None:
                planner.repeat_enabled = bool(repeat_enabled)
            if repeat_interval_days is not None:
                planner.repeat_interval_days = int(repeat_interval_days or 0)
            if early_notification_days is not None:
                planner.early_notification_days = int(early_notification_days or 0)
            
            # Recalculate repeat_generation_date if repeat settings changed
            if planner.repeat_enabled and planner.repeat_interval_days > 0:
                from datetime import timedelta
                planner.repeat_generation_date = planner.start_date + timedelta(days=planner.repeat_interval_days)
            else:
                planner.repeat_generation_date = None
            
            # Update folder if provided
            if folder_id is not None:
                if folder_id:
                    folder = PlannerFolder.objects.filter(id=folder_id, organization=user.organization).first()
                    if folder:
                        planner.folder = folder
                else:
                    planner.folder = None

            # Update team leader if provided (admin can reassign)
            team_leader_id = request.data.get('team_leader_id') or request.data.get('team_leader')
            if team_leader_id is not None:
                if team_leader_id:
                    new_leader = CustomUser.objects.filter(
                        id=int(team_leader_id),
                        organization=user.organization,
                        is_active=True,
                        is_deleted=False
                    ).first()
                    if not new_leader:
                        return Response({'error': 'Selected team leader not found in organization'}, status=status.HTTP_400_BAD_REQUEST)
                    planner.team_leader = new_leader
                    # Also update existing CollaborativeSubmission if it exists
                    collab = CollaborativeSubmission.objects.filter(
                        planner_assignment=planner,
                        organization=user.organization
                    ).exclude(status=CollaborativeSubmissionStatus.COMPLETED).first()
                    if collab:
                        collab.team_leader = new_leader
                        collab.save(update_fields=['team_leader'])
                else:
                    # Clearing team leader
                    planner.team_leader = None

            planner.save()
            
            logger.info(f"Planner {planner_id} updated successfully by user {user.id}")
            return Response({'message': 'Planner updated successfully'}, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Update planner error: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to update planner: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerDeleteView(APIView):
    """
    Delete a planner assignment (admin only).
    Endpoint: DELETE /api/planner/<planner_id>/delete/
    """
    permission_classes = [IsAuthenticated]

    def delete(self, request, planner_id):
        try:
            user = request.user
            logger.info(f"Deleting planner {planner_id} by user {user.id}")
            
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=user.organization
            )
            
            # Only block deletion of completed planners
            if planner.is_completed:
                return Response({
                    'error': 'Cannot delete completed planners'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            planner.delete()
            
            logger.info(f"Planner {planner_id} deleted successfully by user {user.id}")
            return Response({'message': 'Planner deleted successfully'}, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Delete planner error: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to delete planner: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerBulkDeleteView(APIView):
    """
    Bulk delete planner assignments (admin only).
    Endpoint: POST /api/planner/bulk-delete/
    Expected payload: { "planner_ids": [id1, id2, ...] }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            user = request.user
            planner_ids = request.data.get('planner_ids', [])
            
            if not planner_ids:
                return Response({'error': 'No planner IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"Bulk deleting planners {planner_ids} by user {user.id}")
            
            deleted_count = 0
            skipped_count = 0
            errors = []
            
            for planner_id in planner_ids:
                try:
                    planner = PlannerAssignment.objects.get(
                        id=planner_id,
                        organization=user.organization
                    )
                    
                    # Only block deletion of completed planners
                    if planner.is_completed:
                        skipped_count += 1
                        errors.append({
                            'planner_id': planner_id,
                            'error': 'Cannot delete completed planner'
                        })
                        continue
                    
                    planner.delete()
                    deleted_count += 1
                    
                except PlannerAssignment.DoesNotExist:
                    errors.append({
                        'planner_id': planner_id,
                        'error': 'Planner not found'
                    })
            
            logger.info(f"Bulk delete completed: {deleted_count} deleted, {skipped_count} skipped by user {user.id}")
            return Response({
                'message': f'Bulk delete completed: {deleted_count} deleted, {skipped_count} skipped',
                'deleted_count': deleted_count,
                'skipped_count': skipped_count,
                'errors': errors
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Bulk delete planner error: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to bulk delete planners: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerAssigneesView(APIView):
    """
    Get assignment details for a planner.
    Endpoint: GET /api/planner/<planner_id>/assignees/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, planner_id):
        try:
            user = request.user
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=user.organization
            )

            assignments = PlannerAssignment.objects.filter(
                order_id=planner.order_id,
                organization=user.organization
            ).select_related('user', 'group', 'leader', 'user__location', 'leader__location')

            data = []
            for assignment in assignments:
                if assignment.user:
                    assignee_name = f"{assignment.user.first_name} {assignment.user.last_name}".strip() or assignment.user.username
                    assignee_location = assignment.user.location.name if assignment.user.location else "N/A"
                elif assignment.group:
                    assignee_name = assignment.group.name
                    assignee_location = "N/A"
                elif assignment.leader:
                    assignee_name = f"{assignment.leader.first_name} {assignment.leader.last_name}".strip() or assignment.leader.username
                    assignee_location = assignment.leader.location.name if assignment.leader.location else "N/A"
                else:
                    assignee_name = "Unassigned"
                    assignee_location = "N/A"

                if assignment.is_completed:
                    status_label = "Completed"
                elif assignment.started_on:
                    status_label = "In Progress"
                else:
                    status_label = "Not Started"

                data.append({
                    'task_id': assignment.id,
                    'task_name': assignment.planner_name,
                    'assignee': assignee_name,
                    'assignee_location': assignee_location,
                    'status': status_label
                })

            logger.info(f"Fetched {len(data)} planner assignments for planner {planner_id}")
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Get planner assignees error: {str(e)}", exc_info=True)
            return Response({'error': f'Failed to fetch planner assignees: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class PlannerDownloadTemplateView(APIView):
    """
    Download an Excel template for planner bulk upload.
    Endpoint: GET /api/planner/download-template/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Planner Upload'

        headers = [
            'Planner Name',
            'Location',
            'Form',
            'Start Date',
            'End Date',
            'Description',
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='2196f3', end_color='2196f3', fill_type='solid')
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')

        example_data = [
            ['Preventive Maintenance - WTG 1', 'Deccan', 'Preventive Maintenance Form', '2024-12-01', '2024-12-31', 'Quarterly preventive maintenance for WTG 1'],
            ['Site Inspection - WTG 2', 'Tamil Nadu', 'Site Inspection Form', '2024-12-15', '2025-01-15', 'Monthly site inspection'],
        ]
        for row_idx, row_data in enumerate(example_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_num, value=value)

        col_widths = [32, 22, 28, 14, 14, 36]
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="planner_bulk_upload_template.xlsx"'
        return response


class PlannerFolderListView(APIView):
    """
    List and create planner folders.
    GET /api/planner/folders/  — list all folders for the organization
    POST /api/planner/folders/ — create a new folder
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        folders = PlannerFolder.objects.filter(organization=request.user.organization)
        data = []
        for f in folders:
            planner_count = PlannerAssignment.objects.filter(folder=f, organization=request.user.organization).count()
            data.append({
                'id': f.id,
                'name': f.name,
                'color': f.color,
                'created_on': f.created_on.isoformat() if f.created_on else None,
                'planner_count': planner_count,
                'order': f.order,
            })
        return Response(data, status=status.HTTP_200_OK)

    def post(self, request):
        name = (request.data.get('name') or '').strip()
        color = (request.data.get('color') or '#6366F1').strip()
        if not name:
            return Response({'error': 'Folder name is required'}, status=status.HTTP_400_BAD_REQUEST)
        # Set order to max+1 so new folders appear at the bottom
        max_order = PlannerFolder.objects.filter(organization=request.user.organization).aggregate(Max('order'))['order__max']
        next_order = (max_order or 0) + 1
        folder, created = PlannerFolder.objects.get_or_create(
            name=name,
            organization=request.user.organization,
            defaults={'color': color, 'created_by': request.user, 'order': next_order}
        )
        if not created:
            return Response({'error': 'A folder with this name already exists'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({
            'id': folder.id,
            'name': folder.name,
            'color': folder.color,
            'created_on': folder.created_on.isoformat(),
            'planner_count': 0,
            'order': folder.order,
        }, status=status.HTTP_201_CREATED)


class PlannerFolderDetailView(APIView):
    """
    Retrieve, update or delete a planner folder.
    GET /api/planner/folders/<folder_id>/   — fetch folder details
    PUT /api/planner/folders/<folder_id>/   — rename / change color
    DELETE /api/planner/folders/<folder_id>/ — delete folder (planners keep folder_id=null)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, folder_id):
        folder = get_object_or_404(PlannerFolder, id=folder_id, organization=request.user.organization)
        planner_count = PlannerAssignment.objects.filter(folder=folder).count()
        return Response({
            'id': folder.id,
            'name': folder.name,
            'color': folder.color,
            'created_on': folder.created_on.isoformat() if folder.created_on else None,
            'created_by': folder.created_by.get_full_name() if folder.created_by else None,
            'planner_count': planner_count,
        }, status=status.HTTP_200_OK)

    def put(self, request, folder_id):
        folder = get_object_or_404(PlannerFolder, id=folder_id, organization=request.user.organization)
        name = (request.data.get('name') or '').strip()
        color = (request.data.get('color') or folder.color).strip()
        if name:
            existing = PlannerFolder.objects.filter(name=name, organization=request.user.organization).exclude(id=folder_id).first()
            if existing:
                return Response({'error': 'A folder with this name already exists'}, status=status.HTTP_400_BAD_REQUEST)
            folder.name = name
        folder.color = color
        folder.save()
        return Response({
            'id': folder.id,
            'name': folder.name,
            'color': folder.color,
        }, status=status.HTTP_200_OK)

    def delete(self, request, folder_id):
        folder = get_object_or_404(PlannerFolder, id=folder_id, organization=request.user.organization)
        folder.delete()
        return Response({'message': 'Folder deleted'}, status=status.HTTP_200_OK)


class PlannerFolderReorderView(APIView):
    """
    Reorder planner folders by swapping order values.
    POST /api/planner/folders/reorder/
    Body: { "folder_ids": [3, 1, 5, 2] }  — ordered list of folder IDs
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        folder_ids = request.data.get('folder_ids', [])
        if not folder_ids or not isinstance(folder_ids, list):
            return Response({'error': 'folder_ids list is required'}, status=status.HTTP_400_BAD_REQUEST)

        folders = PlannerFolder.objects.filter(
            id__in=folder_ids,
            organization=request.user.organization
        )
        if len(folders) != len(folder_ids):
            return Response({'error': 'Some folders not found'}, status=status.HTTP_400_BAD_REQUEST)

        # Assign sequential order values based on the provided list
        for index, folder_id in enumerate(folder_ids):
            PlannerFolder.objects.filter(id=folder_id, organization=request.user.organization).update(order=index)

        return Response({'message': 'Folders reordered successfully'}, status=status.HTTP_200_OK)


class PlannerFolderStatsView(APIView):
    """
    Get completion stats per folder for the organization.
    GET /api/planner/folder-stats/
    Returns: [{ id, name, color, total, completed, percentage }]
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        folders = PlannerFolder.objects.filter(organization=request.user.organization)
        data = []
        for f in folders:
            total = PlannerAssignment.objects.filter(folder=f, organization=request.user.organization).count()
            completed = PlannerAssignment.objects.filter(folder=f, organization=request.user.organization, is_completed=True).count()
            percentage = round((completed / total) * 100) if total > 0 else 0
            data.append({
                'id': f.id,
                'name': f.name,
                'color': f.color,
                'total': total,
                'completed': completed,
                'percentage': percentage,
            })
        # Unassigned planners
        unassigned_total = PlannerAssignment.objects.filter(folder__isnull=True, organization=request.user.organization).count()
        unassigned_completed = PlannerAssignment.objects.filter(folder__isnull=True, organization=request.user.organization, is_completed=True).count()
        unassigned_percentage = round((unassigned_completed / unassigned_total) * 100) if unassigned_total > 0 else 0
        data.append({
            'id': None,
            'name': 'Unassigned',
            'color': '#9CA3AF',
            'total': unassigned_total,
            'completed': unassigned_completed,
            'percentage': unassigned_percentage,
        })
        return Response(data, status=status.HTTP_200_OK)


class MovePlannerToFolderView(APIView):
    """
    Move one or more planners to a folder.
    POST /api/planner/move-to-folder/
    Body: { "planner_ids": [1,2,3], "folder_id": 5 }  (folder_id=null to remove from folder)
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        planner_ids = request.data.get('planner_ids', [])
        folder_id = request.data.get('folder_id')

        if not planner_ids:
            return Response({'error': 'planner_ids is required'}, status=status.HTTP_400_BAD_REQUEST)

        folder = None
        if folder_id:
            folder = get_object_or_404(PlannerFolder, id=folder_id, organization=request.user.organization)

        updated = PlannerAssignment.objects.filter(
            id__in=planner_ids,
            organization=request.user.organization
        ).update(folder=folder)

        return Response({
            'message': f'Moved {updated} planner(s) to folder "{folder.name if folder else "No Folder"}"',
            'updated_count': updated,
        }, status=status.HTTP_200_OK)


# ============================================================================
# COLLABORATIVE AUDIT DELEGATION VIEWS
# ============================================================================

class CollaborativeStartView(APIView):
    """
    Team Leader starts a collaborative audit submission.
    Creates a CollaborativeSubmission + draft FormSubmision + GroupDelegation records for each audit group.
    Endpoint: POST /api/planner/<planner_id>/collaborative/start/
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization,
                is_completed=False
            )

            if not planner.collaborative_enabled:
                return Response({'error': 'Collaborative mode is not enabled for this planner'}, status=status.HTTP_400_BAD_REQUEST)

            if planner.form.form_type != 'audit':
                return Response({'error': 'Collaborative delegation is only available for audit forms'}, status=status.HTTP_400_BAD_REQUEST)

            team_leader = planner.team_leader or planner.user or request.user
            if request.user != team_leader and request.user != planner.created_by:
                return Response({'error': 'Only the Team Leader can start this collaborative audit'}, status=status.HTTP_403_FORBIDDEN)

            existing = CollaborativeSubmission.objects.filter(
                planner_assignment=planner,
                organization=request.user.organization
            ).exclude(status=CollaborativeSubmissionStatus.COMPLETED).first()
            if existing:
                # If DRAFT from bulk import, create FormSubmision and move to IN_PROGRESS
                if existing.status == CollaborativeSubmissionStatus.DRAFT and not existing.form_submission:
                    form_submission = FormSubmision.objects.create(
                        form=planner.form,
                        submission_initiated_by=team_leader,
                        is_completed=False,
                        organization=request.user.organization
                    )
                    existing.form_submission = form_submission
                    existing.status = CollaborativeSubmissionStatus.IN_PROGRESS
                    existing.started_on = timezone.now()
                    existing.save()

                    # Create GroupDelegation records if missing (safety net)
                    if existing.group_delegations.count() == 0:
                        audit_groups = AuditGroup.objects.filter(form=planner.form).order_by('order')
                        for group in audit_groups:
                            GroupDelegation.objects.create(
                                collaborative_submission=existing,
                                audit_group=group,
                                group_order=group.order,
                                assigned_by=team_leader,
                                status=GroupDelegationStatus.UNASSIGNED,
                                organization=request.user.organization
                            )

                    planner.started_by = team_leader
                    planner.started_on = timezone.now()
                    planner.save()
                    existing.refresh_from_db()

                serializer = CollaborativeSubmissionSerializer(existing)
                return Response({'message': 'Collaborative submission already exists', 'collaborative_submission': serializer.data}, status=status.HTTP_200_OK)

            form_submission = FormSubmision.objects.create(
                form=planner.form,
                submission_initiated_by=team_leader,
                is_completed=False,
                organization=request.user.organization
            )

            collab_submission = CollaborativeSubmission.objects.create(
                planner_assignment=planner,
                form_submission=form_submission,
                team_leader=team_leader,
                status=CollaborativeSubmissionStatus.IN_PROGRESS,
                started_on=timezone.now(),
                organization=request.user.organization
            )

            audit_groups = AuditGroup.objects.filter(form=planner.form).order_by('order')
            for group in audit_groups:
                GroupDelegation.objects.create(
                    collaborative_submission=collab_submission,
                    audit_group=group,
                    group_order=group.order,
                    assigned_by=team_leader,
                    status=GroupDelegationStatus.UNASSIGNED,
                    organization=request.user.organization
                )

            planner.started_by = team_leader
            planner.started_on = timezone.now()
            planner.save()

            serializer = CollaborativeSubmissionSerializer(collab_submission)
            return Response({'message': 'Collaborative audit started', 'collaborative_submission': serializer.data}, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"CollaborativeStartView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeDetailView(APIView):
    """
    Get collaborative submission details for a planner.
    Team Leader sees all groups + delegation status.
    Team Member sees only their assigned groups.
    Endpoint: GET /api/planner/<planner_id>/collaborative/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            collab = get_object_or_404(
                CollaborativeSubmission,
                planner_assignment=planner,
                organization=request.user.organization
            )

            is_team_leader = (request.user == collab.team_leader)
            is_admin = request.user.is_staff or request.user == planner.created_by

            if is_team_leader or is_admin:
                # Safety net: auto-create GroupDelegation records if missing
                if collab.group_delegations.count() == 0 and planner.form:
                    audit_groups = AuditGroup.objects.filter(form=planner.form).order_by('order')
                    for group in audit_groups:
                        GroupDelegation.objects.create(
                            collaborative_submission=collab,
                            audit_group=group,
                            group_order=group.order,
                            assigned_by=collab.team_leader,
                            status=GroupDelegationStatus.UNASSIGNED,
                            organization=request.user.organization
                        )
                    collab.refresh_from_db()

                serializer = CollaborativeSubmissionSerializer(collab)
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                my_delegations = collab.group_delegations.filter(
                    assigned_users=request.user,
                    organization=request.user.organization
                ).exclude(status=GroupDelegationStatus.UNASSIGNED)
                delegation_data = GroupDelegationSerializer(my_delegations, many=True).data
                total_groups = collab.group_delegations.count()
                completed_groups = collab.group_delegations.filter(
                    status__in=['submitted', 'reviewed']
                ).count()
                completion_pct = round((completed_groups / total_groups) * 100) if total_groups > 0 else 0
                return Response({
                    'id': collab.id,
                    'planner_assignment': collab.planner_assignment_id,
                    'form_submission': collab.form_submission_id,
                    'form_id': planner.form.id if planner.form else None,
                    'status': collab.status,
                    'planner_name': planner.planner_name,
                    'form_title': planner.form.title if planner.form else None,
                    'group_delegations': delegation_data,
                    'is_team_member': True,
                    'completed_groups': completed_groups,
                    'total_groups': total_groups,
                    'completion_percentage': completion_pct,
                }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeDetailView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeDelegateView(APIView):
    """
    Team Leader assigns audit groups to team members.
    Endpoint: POST /api/planner/<planner_id>/collaborative/delegate/
    Body: {
        "delegations": [
            {"audit_group_id": 1, "user_ids": [5, 6]},
            {"audit_group_id": 2, "user_ids": [7]},
            ...
        ]
    }
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            collab = get_object_or_404(
                CollaborativeSubmission,
                planner_assignment=planner,
                organization=request.user.organization
            )

            if request.user != collab.team_leader and not (request.user.is_staff or request.user == planner.created_by):
                return Response({'error': 'Only the Team Leader can delegate groups'}, status=status.HTTP_403_FORBIDDEN)

            delegations = request.data.get('delegations', [])
            if not delegations:
                return Response({'error': 'delegations list is required'}, status=status.HTTP_400_BAD_REQUEST)

            results = []
            for item in delegations:
                audit_group_id = item.get('audit_group_id')
                user_ids = item.get('user_ids', [])
                question_uuids = item.get('assigned_question_uuids', [])

                if not audit_group_id:
                    results.append({'audit_group_id': audit_group_id, 'error': 'audit_group_id is required'})
                    continue

                delegation = collab.group_delegations.filter(
                    audit_group_id=audit_group_id,
                    organization=request.user.organization
                ).first()

                if not delegation:
                    results.append({'audit_group_id': audit_group_id, 'error': 'Group delegation not found'})
                    continue

                if not user_ids:
                    # Unassign: clear users and reset status
                    delegation.assigned_users.clear()
                    delegation.status = GroupDelegationStatus.UNASSIGNED
                    delegation.save()
                    results.append({
                        'audit_group_id': audit_group_id,
                        'audit_group_name': delegation.audit_group.name if delegation.audit_group else None,
                        'assigned_user_ids': [],
                        'status': delegation.status
                    })
                    continue

                users = CustomUser.objects.filter(
                    id__in=user_ids,
                    organization=request.user.organization
                )
                if users.count() != len(user_ids):
                    results.append({'audit_group_id': audit_group_id, 'error': 'Some users not found in organization'})
                    continue

                # Only update status if the group hasn't been submitted/reviewed/rejected yet.
                # This prevents re-delegation from resetting a already-submitted group's status.
                if delegation.status not in [GroupDelegationStatus.SUBMITTED, GroupDelegationStatus.REVIEWED, GroupDelegationStatus.REJECTED]:
                    delegation.status = GroupDelegationStatus.ASSIGNED
                delegation.assigned_users.set(users)
                delegation.assigned_by = request.user
                if question_uuids:
                    delegation.assigned_question_uuids = question_uuids
                delegation.save()

                results.append({
                    'audit_group_id': audit_group_id,
                    'audit_group_name': delegation.audit_group.name if delegation.audit_group else None,
                    'assigned_user_ids': user_ids,
                    'status': delegation.status
                })

            all_assigned = not collab.group_delegations.filter(status=GroupDelegationStatus.UNASSIGNED).exists()
            if all_assigned and collab.status == CollaborativeSubmissionStatus.IN_PROGRESS:
                pass

            serializer = CollaborativeSubmissionSerializer(collab)
            return Response({
                'message': 'Delegation updated successfully',
                'results': results,
                'collaborative_submission': serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeDelegateView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeChangeLeaderView(APIView):
    """
    Admin changes the Team Leader of a collaborative audit.
    Also updates the PlannerAssignment.team_leader and any active CollaborativeSubmission.
    Optionally accepts 'delegations' to reassign groups at the same time.
    Endpoint: POST /api/planner/<planner_id>/collaborative/change-leader/
    Body: {
        "team_leader_id": 5,
        "delegations": [  // optional — same format as delegate endpoint
            {"audit_group_id": 1, "user_ids": [5, 6]},
            ...
        ]
    }
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            if not planner.collaborative_enabled:
                return Response({'error': 'Collaborative mode is not enabled for this planner'}, status=status.HTTP_400_BAD_REQUEST)

            # Only admin or planner creator can change team leader
            is_admin = request.user.is_staff or request.user == planner.created_by or \
                       (request.user.role and request.user.role.name in ('super_admin', 'admin'))
            if not is_admin:
                return Response({'error': 'Only admins can change the Team Leader'}, status=status.HTTP_403_FORBIDDEN)

            new_leader_id = request.data.get('team_leader_id')
            if not new_leader_id:
                return Response({'error': 'team_leader_id is required'}, status=status.HTTP_400_BAD_REQUEST)

            new_leader = CustomUser.objects.filter(
                id=int(new_leader_id),
                organization=request.user.organization,
                is_active=True,
                is_deleted=False
            ).first()
            if not new_leader:
                return Response({'error': 'Selected user not found in organization'}, status=status.HTTP_400_BAD_REQUEST)

            # Update planner
            old_leader = planner.team_leader
            planner.team_leader = new_leader
            planner.save(update_fields=['team_leader'])

            # Update active CollaborativeSubmission if exists
            collab = CollaborativeSubmission.objects.filter(
                planner_assignment=planner,
                organization=request.user.organization
            ).exclude(status=CollaborativeSubmissionStatus.COMPLETED).first()

            if collab:
                collab.team_leader = new_leader
                collab.save(update_fields=['team_leader'])

            # Optionally process delegations (same as CollaborativeDelegateView)
            delegations = request.data.get('delegations', [])
            delegation_results = []
            if delegations and collab:
                for item in delegations:
                    audit_group_id = item.get('audit_group_id')
                    user_ids = item.get('user_ids', [])

                    if not audit_group_id:
                        continue

                    delegation = collab.group_delegations.filter(
                        audit_group_id=audit_group_id,
                        organization=request.user.organization
                    ).first()
                    if not delegation:
                        continue

                    if not user_ids:
                        delegation.assigned_users.clear()
                        delegation.status = GroupDelegationStatus.UNASSIGNED
                        delegation.save()
                    else:
                        users = CustomUser.objects.filter(
                            id__in=user_ids,
                            organization=request.user.organization
                        )
                        delegation.assigned_users.set(users)
                        delegation.assigned_by = request.user
                        if delegation.status not in [GroupDelegationStatus.SUBMITTED, GroupDelegationStatus.REVIEWED, GroupDelegationStatus.REJECTED]:
                            delegation.status = GroupDelegationStatus.ASSIGNED
                        delegation.save()

                    delegation_results.append({
                        'audit_group_id': audit_group_id,
                        'audit_group_name': delegation.audit_group.name if delegation.audit_group else None,
                        'assigned_user_ids': user_ids,
                        'status': delegation.status
                    })

            response_data = {
                'message': 'Team Leader updated successfully',
                'old_team_leader': old_leader.username if old_leader else None,
                'new_team_leader': new_leader.username,
            }
            if delegation_results:
                response_data['delegation_results'] = delegation_results
            if collab:
                response_data['collaborative_submission'] = CollaborativeSubmissionSerializer(collab).data

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeChangeLeaderView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeSubmitGroupView(APIView):
    """
    Team member submits answers for their assigned audit group.
    Answers are saved against the shared FormSubmision.
    Endpoint: POST /api/planner/<planner_id>/collaborative/submit-group/
    Body: {
        "group_delegation_id": 1,
        "answers": [ {question, question_type, answer, ...}, ... ]
    }
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            collab = get_object_or_404(
                CollaborativeSubmission,
                planner_assignment=planner,
                organization=request.user.organization
            )

            delegation_id = request.data.get('group_delegation_id')
            answers = request.data.get('answers', [])

            if not delegation_id:
                return Response({'error': 'group_delegation_id is required'}, status=status.HTTP_400_BAD_REQUEST)

            delegation = get_object_or_404(
                GroupDelegation,
                id=delegation_id,
                collaborative_submission=collab,
                organization=request.user.organization
            )

            if request.user not in delegation.assigned_users.all() and request.user != collab.team_leader:
                return Response({'error': 'You are not assigned to this group'}, status=status.HTTP_403_FORBIDDEN)

            if delegation.status == GroupDelegationStatus.REVIEWED:
                return Response({'error': 'This group has already been reviewed and approved'}, status=status.HTTP_400_BAD_REQUEST)

            form_submission = collab.form_submission
            if not form_submission:
                return Response({'error': 'Form submission not found'}, status=status.HTTP_400_BAD_REQUEST)

            from form.models import Answer, AuditGroup, Question
            # Validate that all submitted answers belong to the assigned audit group
            # or are audit info questions (shared across all groups, audit_group=None)
            assigned_audit_group = delegation.audit_group
            if assigned_audit_group:
                valid_question_ids = set(
                    Question.objects.filter(
                        form=planner.form
                    ).filter(
                        Q(audit_group=assigned_audit_group) |
                        Q(audit_group__isnull=True)
                    ).values_list('id', flat=True)
                )
                invalid_answers = []
                for a in answers:
                    qid = a.get('question')
                    if not qid:
                        continue
                    try:
                        qid_int = int(qid)
                    except (ValueError, TypeError):
                        continue
                    if qid_int not in valid_question_ids:
                        invalid_answers.append(a)
                if invalid_answers:
                    return Response({
                        'error': f'Submitted answers contain questions not in the assigned audit group ({assigned_audit_group.name})',
                        'invalid_count': len(invalid_answers),
                    }, status=status.HTTP_400_BAD_REQUEST)

            conflicts = []
            # Skip conflict check for rejected groups — user is re-submitting after rejection
            is_rejected_group = delegation.status == GroupDelegationStatus.REJECTED
            if not is_rejected_group:
                for ans in answers:
                    question_id = ans.get('question')
                    if not question_id:
                        continue

                    try:
                        question = Question.objects.get(id=question_id, form=planner.form)
                    except Question.DoesNotExist:
                        continue

                    # Skip conflict check for shared audit_info questions (audit_group is null) —
                    # these are common across all groups and are expected to be filled by whoever answers first.
                    if question.audit_group_id is None:
                        continue

                    existing = Answer.objects.filter(
                        question_id=question_id,
                        submission=form_submission,
                        organization=request.user.organization
                    ).first()

                    if not existing or not existing.submitted_by_id or existing.submitted_by_id == request.user.id:
                        continue

                    # Only a real conflict if the incoming answer differs from what's already saved
                    incoming_answer = str(ans.get('answer', '') or '').strip()
                    existing_answer = str(existing.answer or '').strip()
                    if incoming_answer == existing_answer:
                        continue

                    conflicts.append({
                        'question_id': question_id,
                        'question_text': question.question,
                        'answered_by': existing.submitted_by.username if existing.submitted_by else 'Unknown',
                        'existing_answer': existing.answer,
                    })

            if conflicts:
                return Response({
                    'error': 'Conflict: Some questions were already answered by teammates',
                    'conflicts': conflicts,
                    'conflict_count': len(conflicts),
                }, status=status.HTTP_409_CONFLICT)

            for ans in answers:
                question_id = ans.get('question')
                if not question_id:
                    continue

                try:
                    question = Question.objects.get(id=question_id, form=planner.form)
                except Question.DoesNotExist:
                    continue

                existing = Answer.objects.filter(
                    question_id=question_id,
                    submission=form_submission,
                    organization=request.user.organization
                ).first()

                answer_data = {
                    'question_id': question_id,
                    'question_type': ans.get('question_type', question.question_type),
                    'answer': ans.get('answer', ''),
                    'Form_id': planner.form.id,
                    'submitted_by_id': request.user.id,
                    'submission_id': form_submission.id,
                    'organization_id': request.user.organization.id,
                    'other_text': ans.get('other_text'),
                    'remarks': ans.get('remarks'),
                }

                if ans.get('division'):
                    answer_data['division_id'] = ans.get('division')
                if ans.get('sub_division'):
                    answer_data['sub_division_id'] = ans.get('sub_division')
                if ans.get('location'):
                    answer_data['location_id'] = ans.get('location')
                if ans.get('user'):
                    answer_data['user_id'] = ans.get('user')

                if existing:
                    for key, value in answer_data.items():
                        if value is not None:
                            setattr(existing, key, value)
                    existing.save()
                else:
                    Answer.objects.create(**{k: v for k, v in answer_data.items() if v is not None})

            delegation.status = GroupDelegationStatus.SUBMITTED
            delegation.submitted_by = request.user
            delegation.submitted_on = timezone.now()
            delegation.rejection_comment = None
            delegation.save()

            all_submitted = not collab.group_delegations.filter(
                status__in=[GroupDelegationStatus.UNASSIGNED, GroupDelegationStatus.ASSIGNED, GroupDelegationStatus.IN_PROGRESS, GroupDelegationStatus.REJECTED]
            ).exists()
            if all_submitted and collab.status == CollaborativeSubmissionStatus.IN_PROGRESS:
                collab.status = CollaborativeSubmissionStatus.READY_FOR_REVIEW
                collab.save()

            return Response({
                'message': 'Group submitted successfully',
                'group_delegation_id': delegation.id,
                'status': delegation.status,
                'collaborative_status': collab.status,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeSubmitGroupView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativePollAnswersView(APIView):
    """
    Lightweight endpoint for silent polling of answers on a shared FormSubmision.
    Returns only question IDs that have answers and their values.
    Endpoint: GET /api/planner/<planner_id>/collaborative/poll-answers/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            collab = get_object_or_404(
                CollaborativeSubmission,
                planner_assignment=planner,
                organization=request.user.organization
            )

            form_submission = collab.form_submission
            if not form_submission:
                return Response({'answers': [], 'collaborative_status': collab.status}, status=status.HTTP_200_OK)

            from form.models import Answer
            answers = Answer.objects.filter(
                submission=form_submission,
                organization=request.user.organization
            ).exclude(answer__isnull=True).exclude(answer__exact='')

            answer_data = []
            for ans in answers:
                answer_data.append({
                    'question_id': ans.question_id,
                    'question_type': ans.question_type,
                    'answer': ans.answer,
                    'other_text': ans.other_text,
                    'division_id': ans.division_id,
                    'sub_division_id': ans.sub_division_id,
                    'location_id': ans.location_id,
                    'user_id': ans.user_id,
                    'submitted_by_id': ans.submitted_by_id,
                    'submitted_by_name': ans.submitted_by.username if ans.submitted_by else None,
                })

            return Response({
                'answers': answer_data,
                'collaborative_status': collab.status,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativePollAnswersView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeAutoSaveView(APIView):
    """
    Auto-save individual answers in real-time for collaborative mode.
    Does NOT change delegation status — just upserts answers to the shared FormSubmision.
    Endpoint: POST /api/planner/<planner_id>/collaborative/auto-save/
    Body: { group_delegation_id, answers: [{question, question_type, answer, ...}] }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            collab = get_object_or_404(
                CollaborativeSubmission,
                planner_assignment=planner,
                organization=request.user.organization
            )

            delegation_id = request.data.get('group_delegation_id')
            answers = request.data.get('answers', [])

            if not delegation_id:
                return Response({'error': 'group_delegation_id is required'}, status=status.HTTP_400_BAD_REQUEST)

            delegation = get_object_or_404(
                GroupDelegation,
                id=delegation_id,
                collaborative_submission=collab,
                organization=request.user.organization
            )

            if request.user not in delegation.assigned_users.all() and request.user != collab.team_leader:
                return Response({'error': 'You are not assigned to this group'}, status=status.HTTP_403_FORBIDDEN)

            # Mark delegation as IN_PROGRESS when user starts answering
            if delegation.status == GroupDelegationStatus.ASSIGNED:
                delegation.status = GroupDelegationStatus.IN_PROGRESS
                delegation.save(update_fields=['status'])

            form_submission = collab.form_submission
            if not form_submission:
                return Response({'error': 'Form submission not found'}, status=status.HTTP_400_BAD_REQUEST)

            from form.models import Answer, Question

            saved_count = 0
            conflicts = []

            for ans in answers:
                question_id = ans.get('question')
                if not question_id:
                    continue

                try:
                    question_id_int = int(question_id)
                except (ValueError, TypeError):
                    continue

                try:
                    question = Question.objects.get(id=question_id_int, form=planner.form)
                except Question.DoesNotExist:
                    continue

                # Check for conflict — existing answer by a different user
                existing = Answer.objects.filter(
                    question_id=question_id_int,
                    submission=form_submission,
                    organization=request.user.organization
                ).first()

                if existing and existing.submitted_by_id and existing.submitted_by_id != request.user.id:
                    conflicts.append({
                        'question_id': question_id_int,
                        'question_text': question.question,
                        'answered_by': existing.submitted_by.username if existing.submitted_by else 'Unknown',
                    })
                    continue

                # Upsert the answer
                answer_data = {
                    'question_id': question_id_int,
                    'question_type': ans.get('question_type', question.question_type),
                    'answer': ans.get('answer', ''),
                    'Form_id': planner.form.id,
                    'submitted_by_id': request.user.id,
                    'submission_id': form_submission.id,
                    'organization_id': request.user.organization.id,
                    'other_text': ans.get('other_text'),
                }

                if ans.get('division'):
                    answer_data['division_id'] = ans.get('division')
                if ans.get('sub_division'):
                    answer_data['sub_division_id'] = ans.get('sub_division')
                if ans.get('location'):
                    answer_data['location_id'] = ans.get('location')
                if ans.get('user'):
                    answer_data['user_id'] = ans.get('user')

                if existing:
                    for key, value in answer_data.items():
                        if value is not None:
                            setattr(existing, key, value)
                    existing.save()
                else:
                    Answer.objects.create(**{k: v for k, v in answer_data.items() if v is not None})

                saved_count += 1

            return Response({
                'message': 'Auto-saved',
                'saved_count': saved_count,
                'conflicts': conflicts,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeAutoSaveView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeReviewView(APIView):
    """
    Team Leader reviews submitted groups — approve, reject, or bulk approve.
    Endpoint: POST /api/planner/<planner_id>/collaborative/review/
    Body: {
        "action": "approve" | "reject" | "approve_all",
        "group_delegation_ids": [1, 2, 3],   (not needed for approve_all)
        "rejection_comment": "..."            (only for reject)
    }
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            collab = get_object_or_404(
                CollaborativeSubmission,
                planner_assignment=planner,
                organization=request.user.organization
            )

            if request.user != collab.team_leader and not (request.user.is_staff or request.user == planner.created_by):
                return Response({'error': 'Only the Team Leader can review groups'}, status=status.HTTP_403_FORBIDDEN)

            action = request.data.get('action')
            group_ids = request.data.get('group_delegation_ids', [])
            rejection_comment = request.data.get('rejection_comment', '')

            if action not in ['approve', 'reject', 'approve_all']:
                return Response({'error': 'Invalid action. Use: approve, reject, or approve_all'}, status=status.HTTP_400_BAD_REQUEST)

            if action == 'approve_all':
                groups = collab.group_delegations.filter(
                    status=GroupDelegationStatus.SUBMITTED,
                    organization=request.user.organization
                )
            else:
                if not group_ids:
                    return Response({'error': 'group_delegation_ids is required for approve/reject'}, status=status.HTTP_400_BAD_REQUEST)
                groups = collab.group_delegations.filter(
                    id__in=group_ids,
                    organization=request.user.organization
                )

            updated = []
            for group in groups:
                if action == 'approve':
                    if group.status != GroupDelegationStatus.SUBMITTED:
                        updated.append({'id': group.id, 'error': f'Group is in {group.status} status, cannot approve'})
                        continue
                    group.status = GroupDelegationStatus.REVIEWED
                    group.reviewed_on = timezone.now()
                    group.save()
                    updated.append({'id': group.id, 'status': 'reviewed'})

                elif action == 'reject':
                    if group.status not in [GroupDelegationStatus.SUBMITTED, GroupDelegationStatus.REVIEWED]:
                        updated.append({'id': group.id, 'error': f'Group is in {group.status} status, cannot reject'})
                        continue
                    group.status = GroupDelegationStatus.REJECTED
                    group.rejection_comment = rejection_comment
                    group.reviewed_on = timezone.now()
                    group.save()
                    updated.append({'id': group.id, 'status': 'rejected'})

            all_reviewed = not collab.group_delegations.exclude(status=GroupDelegationStatus.REVIEWED).exists()
            if all_reviewed and collab.status != CollaborativeSubmissionStatus.COMPLETED:
                collab.status = CollaborativeSubmissionStatus.READY_FOR_REVIEW
                collab.save()

            serializer = CollaborativeSubmissionSerializer(collab)
            return Response({
                'message': f'Review action "{action}" completed',
                'updated_groups': updated,
                'collaborative_submission': serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeReviewView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeCompleteView(APIView):
    """
    Team Leader final sign-off — completes the audit, locks scores, marks planner complete.
    Endpoint: POST /api/planner/<planner_id>/collaborative/complete/
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            collab = get_object_or_404(
                CollaborativeSubmission,
                planner_assignment=planner,
                organization=request.user.organization
            )

            if request.user != collab.team_leader and not (request.user.is_staff or request.user == planner.created_by):
                return Response({'error': 'Only the Team Leader can complete this audit'}, status=status.HTTP_403_FORBIDDEN)

            # Guard against double-completion — update timestamp but don't create duplicate records
            form_submission = collab.form_submission
            if collab.status == CollaborativeSubmissionStatus.COMPLETED:
                if form_submission:
                    form_submission.completed_on = timezone.now()
                    form_submission.completed_by = request.user
                    form_submission.save(update_fields=['completed_on', 'completed_by'])
                collab.completed_on = timezone.now()
                collab.save(update_fields=['completed_on'])
                # Create PlannerSubmission if missing (e.g. completed before fix)
                if not PlannerSubmission.objects.filter(planner_assignment=planner, form_submission=form_submission).exists():
                    PlannerSubmission.objects.create(
                        planner_assignment=planner,
                        form_submission=form_submission,
                        submitted_by=request.user,
                    )
                return Response({'message': 'Collaborative audit already completed', 'collaborative_submission': CollaborativeSubmissionSerializer(collab).data}, status=status.HTTP_200_OK)

            unreviewed = collab.group_delegations.exclude(status=GroupDelegationStatus.REVIEWED).count()
            if unreviewed > 0:
                return Response({'error': f'{unreviewed} group(s) are not yet reviewed. Approve all groups before completing.'}, status=status.HTTP_400_BAD_REQUEST)

            if form_submission:
                form_submission.is_completed = True
                form_submission.completed_by = request.user
                form_submission.completed_on = timezone.now()
                form_submission.save()

            collab.status = CollaborativeSubmissionStatus.COMPLETED
            collab.completed_on = timezone.now()
            collab.save()

            planner.is_completed = True
            planner.completed_on = timezone.now()
            planner.completed_by = request.user
            planner.save()

            PlannerSubmission.objects.create(
                planner_assignment=planner,
                form_submission=form_submission,
                submitted_by=request.user,
            )

            # Extract follow-up tasks (same logic as PlannerCompleteView)
            complete_view = PlannerCompleteView()
            followup_tasks_created = complete_view.extract_followup_tasks(
                planner, form_submission, request.user
            )

            serializer = CollaborativeSubmissionSerializer(collab)
            return Response({
                'message': 'Collaborative audit completed successfully',
                'collaborative_submission': serializer.data,
                'followup_tasks_created': followup_tasks_created
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeCompleteView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeMyGroupsView(APIView):
    """
    Get all group delegations assigned to the current user across all collaborative submissions.
    Endpoint: GET /api/planner/collaborative/my-groups/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            delegations = GroupDelegation.objects.filter(
                assigned_users=request.user,
                organization=request.user.organization
            ).exclude(status=GroupDelegationStatus.UNASSIGNED).select_related(
                'collaborative_submission',
                'audit_group',
                'collaborative_submission__planner_assignment',
                'collaborative_submission__planner_assignment__form'
            )

            result = []
            for d in delegations:
                collab = d.collaborative_submission
                planner = collab.planner_assignment
                result.append({
                    'delegation_id': d.id,
                    'audit_group_id': d.audit_group_id,
                    'audit_group_name': d.audit_group.name if d.audit_group else None,
                    'group_order': d.group_order,
                    'status': d.status,
                    'collaborative_submission_id': collab.id,
                    'collaborative_status': collab.status,
                    'planner_id': planner.id,
                    'planner_name': planner.planner_name,
                    'form_id': planner.form.id if planner.form else None,
                    'form_title': planner.form.title if planner.form else None,
                    'form_submission_id': collab.form_submission_id,
                    'rejection_comment': d.rejection_comment,
                    'assigned_question_uuids': d.assigned_question_uuids,
                })

            return Response({'my_groups': result}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeMyGroupsView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeOrgUsersView(APIView):
    """
    Get users in the organization that the Team Leader can assign groups to.
    Optionally filter by group.
    Endpoint: GET /api/planner/<planner_id>/collaborative/users/?group_id=<optional>
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            # Try to get the CollaborativeSubmission with participant pool
            collab = CollaborativeSubmission.objects.filter(
                planner_assignment=planner,
                organization=request.user.organization
            ).first()

            participant_user_ids = set()
            if collab and (collab.participant_users or collab.participant_groups):
                # Use the stored participant pool
                participant_user_ids.update(collab.participant_users or [])
                # Expand group members
                for group_id in (collab.participant_groups or []):
                    group = Groups.objects.filter(id=group_id, organization=request.user.organization).first()
                    if group:
                        participant_user_ids.update(group.members.filter(is_deleted=False, is_active=True).values_list('id', flat=True))
                # Include team leader so they can self-assign
                if collab.team_leader_id:
                    participant_user_ids.add(collab.team_leader_id)
            else:
                # Fallback: use all org users (for backward compatibility)
                group_id = request.query_params.get('group_id')
                if planner.group and not group_id:
                    group_id = planner.group.id

                if group_id:
                    group = get_object_or_404(Groups, id=group_id, organization=request.user.organization)
                    participant_user_ids = set(group.members.filter(is_active=True).values_list('id', flat=True))
                else:
                    participant_user_ids = set(CustomUser.objects.filter(
                        organization=request.user.organization,
                        is_active=True
                    ).values_list('id', flat=True))

            users = CustomUser.objects.filter(id__in=participant_user_ids, is_active=True)
            user_list = [{'id': u.id, 'username': u.username, 'email': u.email,
                          'first_name': u.first_name, 'last_name': u.last_name} for u in users]
            return Response({'users': user_list}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeOrgUsersView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CollaborativeOptionStatsView(APIView):
    """
    Returns option-wise statistics for audit questions in a collaborative submission.
    Dynamically counts how many times each audit option was selected across all answers.
    Endpoint: GET /api/planner/<planner_id>/collaborative/option-stats/
    Response: {
        "total_answered": 15,
        "total_questions": 20,
        "options": [
            {"id": 1, "option": "OK", "count": 10, "percentage": 66.7, "failed": false},
            {"id": 2, "option": "Not OK (deviation OK)", "count": 3, "percentage": 20.0, "failed": false},
            {"id": 3, "option": "Not OK (not going to close)", "count": 2, "percentage": 13.3, "failed": true},
        ],
        "per_group": [
            {"audit_group_id": 1, "audit_group_name": "Group 1", "options": [...]},
            ...
        ]
    }
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, planner_id):
        try:
            planner = get_object_or_404(
                PlannerAssignment,
                id=planner_id,
                organization=request.user.organization
            )

            collab = get_object_or_404(
                CollaborativeSubmission,
                planner_assignment=planner,
                organization=request.user.organization
            )

            form_submission = collab.form_submission
            if not form_submission:
                return Response({
                    'total_answered': 0,
                    'total_questions': 0,
                    'options': [],
                    'per_group': []
                }, status=status.HTTP_200_OK)

            from form.models import Answer, Question, Option, AuditGroup

            # Get all audit-type questions for this form
            audit_questions = Question.objects.filter(
                form=planner.form,
                question_type='audit',
                parent_question__isnull=True
            ).select_related('audit_group')

            total_questions = audit_questions.count()

            # Get all answers for these questions
            question_ids = list(audit_questions.values_list('id', flat=True))
            answers = Answer.objects.filter(
                submission=form_submission,
                question_id__in=question_ids,
                organization=request.user.organization
            ).select_related('question')

            total_answered = answers.count()

            # Collect all option IDs used in answers
            # Audit answers can be: option_id (string), "id1|id2", or [{id: x}, ...]
            answer_option_ids = []
            answer_to_group = {}  # option_id -> audit_group_id

            for ans in answers:
                raw = ans.answer
                if not raw:
                    continue
                # Parse answer value to extract option IDs
                ids = []
                if isinstance(raw, (list, tuple)):
                    for item in raw:
                        if isinstance(item, dict) and item.get('id') is not None:
                            ids.append(int(item['id']))
                        elif isinstance(item, (int, str)) and str(item).isdigit():
                            ids.append(int(item))
                elif isinstance(raw, str):
                    if '|' in raw or ',' in raw:
                        parts = re.split(r'[|,]', raw)
                        for p in parts:
                            p = p.strip()
                            if p.isdigit():
                                ids.append(int(p))
                    elif raw.isdigit():
                        ids.append(int(raw))
                elif isinstance(raw, dict) and raw.get('id') is not None:
                    ids.append(int(raw['id']))

                group_id = ans.question.audit_group_id if ans.question else None
                for oid in ids:
                    answer_option_ids.append(oid)
                    answer_to_group[oid] = group_id

            # Get all options for audit questions
            option_ids_in_form = set()
            for q in audit_questions:
                for opt in q.options.all():
                    option_ids_in_form.add(opt.id)

            options = Option.objects.filter(id__in=option_ids_in_form).order_by('order')

            # Build option stats
            from collections import Counter
            option_counter = Counter(answer_option_ids)

            option_stats = []
            for opt in options:
                count = option_counter.get(opt.id, 0)
                pct = round((count / total_answered * 100), 1) if total_answered > 0 else 0
                option_stats.append({
                    'id': opt.id,
                    'option': opt.option,
                    'count': count,
                    'percentage': pct,
                    'failed': opt.failed,
                })

            # Build per-group stats
            audit_groups = AuditGroup.objects.filter(form=planner.form).order_by('order')
            per_group = []
            for group in audit_groups:
                group_question_ids = list(
                    audit_questions.filter(audit_group=group).values_list('id', flat=True)
                )
                group_answers = [a for a in answers if a.question_id in group_question_ids]
                group_total = len(group_answers)

                group_option_ids = []
                for ans in group_answers:
                    raw = ans.answer
                    if not raw:
                        continue
                    ids = []
                    if isinstance(raw, (list, tuple)):
                        for item in raw:
                            if isinstance(item, dict) and item.get('id') is not None:
                                ids.append(int(item['id']))
                            elif isinstance(item, (int, str)) and str(item).isdigit():
                                ids.append(int(item))
                    elif isinstance(raw, str):
                        if '|' in raw or ',' in raw:
                            parts = re.split(r'[|,]', raw)
                            for p in parts:
                                p = p.strip()
                                if p.isdigit():
                                    ids.append(int(p))
                        elif raw.isdigit():
                            ids.append(int(raw))
                    elif isinstance(raw, dict) and raw.get('id') is not None:
                        ids.append(int(raw['id']))
                    group_option_ids.extend(ids)

                group_counter = Counter(group_option_ids)
                group_option_stats = []
                for opt in options:
                    count = group_counter.get(opt.id, 0)
                    pct = round((count / group_total * 100), 1) if group_total > 0 else 0
                    group_option_stats.append({
                        'id': opt.id,
                        'option': opt.option,
                        'count': count,
                        'percentage': pct,
                        'failed': opt.failed,
                    })
                per_group.append({
                    'audit_group_id': group.id,
                    'audit_group_name': group.name,
                    'total_answered': group_total,
                    'total_questions': len(group_question_ids),
                    'options': group_option_stats,
                })

            return Response({
                'total_answered': total_answered,
                'total_questions': total_questions,
                'options': option_stats,
                'per_group': per_group,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"CollaborativeOptionStatsView error: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
