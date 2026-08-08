from rest_framework.viewsets import ModelViewSet
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.db.models import Q, Prefetch
from django.db import models
from django.shortcuts import get_object_or_404
from django.core.mail import EmailMessage
from django.utils import timezone
import datetime
from collections import defaultdict
from rest_framework.parsers import MultiPartParser, FormParser
import openpyxl
import csv
import io
from rest_framework.permissions import IsAuthenticated



from .models import Task, TaskAssignee, TaskTracking, TaskStatus, TaskAuditLog
from .utils import find_related_tasks
from .serializers import (
    TaskSerializer, TaskCreateSerializer, TaskShareSerializer,
    TaskUpdateSerializer, TaskTrackingCreateSerializer, TaskListSerializer,
    TaskAssigneeSerializer, TaskTrackingSerializer, TaskAuditLogSerializer
)
from user.models import CustomUser, Groups, Locations
from form.models import Form, Stage, StageAssignment, FollowUpTask, LogicFollowUp
from vibro.views import userContextAPIView
from vibro.permissions import IsAdmin, IsEndUserOrAdmin
import logging
import re
from django.db.models import Count, Q
from django.http import HttpResponse
from io import BytesIO

logger = logging.getLogger(__name__)



def _is_user_assigned_via_logic_followup(task, user):
    if not task.followup_task_form_id_id or not task.follow_task_sub_question_id:
        return False, None

    logic_followup = LogicFollowUp.objects.filter(
        followup_toggle=True,
        question_id=task.follow_task_sub_question_id
    ).filter(
        Q(form_id=task.followup_task_form_id_id) |
        Q(audit_group__form_id=task.followup_task_form_id_id)
    ).first()

    if not logic_followup:
        return False, None

    user_ids = set(logic_followup.assign_user_ids or [])
    group_ids = set(logic_followup.assign_group_ids or [])
    leader_ids = set(logic_followup.assign_leader_ids or [])

    if logic_followup.user_id:
        user_ids.add(logic_followup.user_id)
    if logic_followup.group_id:
        group_ids.add(logic_followup.group_id)
    if logic_followup.leader_id:
        leader_ids.add(logic_followup.leader_id)

    if user.id in user_ids or user.id in leader_ids:
        return True, None

    if group_ids:
        matched_group = Groups.objects.filter(
            id__in=group_ids,
            members=user,
            organization=task.organization
        ).first()
        if matched_group:
            return True, matched_group

    return False, None

def extract_form_identifier(form_value):
    """
    Detect and extract form identifier from input
    Returns: ('type', 'value') where type is 'name' or 'id'
    """
    if not form_value:
        return ('name', '')

    form_str = str(form_value).strip()

    # Check if it's a URL (absolute or relative)
    if form_str.startswith(('http://', 'https://', '/')):
        # Extract ID from URL pattern: /forms/form-{id} or https://domain.com/forms/form-{id}
        # Handle both absolute URLs and relative paths
        path_part = form_str.split('?')[0]  # Remove query parameters
        if path_part.startswith(('http://', 'https://')):
            # For absolute URLs, extract the path after the domain
            from urllib.parse import urlparse
            parsed = urlparse(path_part)
            path_part = parsed.path

        # Match /forms/form-{id} pattern
        match = re.match(r'/forms/form-(\d+)$', path_part)
        if match:
            return ('id', match.group(1))

    # Default to name validation
    return ('name', form_str)


def validate_form_identifier(identifier_type, identifier_value, organization):
    """
    Validate form by name or ID within organization context
    """
    if identifier_type == 'id':
        try:
            form_id = int(identifier_value)
            return Form.objects.filter(
                id=form_id,
                organization=organization,
                is_deleted=False,
                is_archived=False
            ).exists()
        except (ValueError, TypeError):
            return False

    elif identifier_type == 'name':
        return Form.objects.filter(
            title__iexact=identifier_value,
            organization=organization,
            is_deleted=False,
            is_archived=False
        ).order_by('-created_at').exists()

    return False


def get_form_by_identifier(identifier_type, identifier_value, organization):
    """
    Get form object by name or ID within organization context
    """
    if identifier_type == 'id':
        try:
            form_id = int(identifier_value)
            return Form.objects.get(
                id=form_id,
                organization=organization,
                is_deleted=False,
                is_archived=False
            )
        except (ValueError, TypeError, Form.DoesNotExist):
            return None

    elif identifier_type == 'name':
        return Form.objects.filter(
            title__iexact=identifier_value,
            organization=organization,
            is_deleted=False,
            is_archived=False
        ).order_by('-created_at').first()

    return None


class TaskViewSet(userContextAPIView, ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Task.objects.all()
    serializer_class = TaskSerializer

    def get_queryset(self):
        queryset = Task.objects.filter(
            organization=self.request.user.organization
        ).select_related(
            'form', 'created_by', 'organization', 'followup_task_form_id'
        ).prefetch_related(
            Prefetch('assignees', queryset=TaskAssignee.objects.select_related('assigned_user', 'assigned_group')),
            Prefetch('tracking_records', queryset=TaskTracking.objects.select_related('assignee_user', 'assignee_group')),
            Prefetch('followup_task', queryset=FollowUpTask.objects.select_related('started_by'))
        ).order_by('-created_on')


        # Apply followup task visibility logic
        # Hide followup tasks that have been started by other users
        user = self.request.user

        # Get all followup tasks that are started by someone other than current user
        hidden_followup_tasks = FollowUpTask.objects.filter(
            started_by__isnull=False
        ).exclude(
            started_by=user
        ).values_list('task_details_id', flat=True)

        # Exclude these tasks from the queryset
        if hidden_followup_tasks:
            queryset = queryset.exclude(id__in=hidden_followup_tasks)

        return queryset

    def get_serializer_class(self):
        if self.action == 'create':
            return TaskCreateSerializer
        elif self.action == 'list':
            return TaskListSerializer
        return TaskSerializer

    def perform_create(self, serializer):
        serializer.save(
            organization=self.request.user.organization,
            created_by=self.request.user
        )

    def retrieve(self, request, *args, **kwargs):
        """Enhanced retrieve method that includes activity logs and parent question"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        data = serializer.data

        # Add parent form question for followup tasks
        if instance.follow_task_sub_question:
            data['parent_question'] = instance.follow_task_sub_question.question

        # Add activity logs from existing TaskAuditLog table (no changes to table)
        activity_logs = TaskAuditLog.objects.filter(
            task=instance
        ).select_related('action_by', 'action_to').order_by('-action_date_time')

        data['activity_logs'] = [{
            'id': log.id,
            'action': log.task_action,
            'action_by': {
                'id': log.action_by.id if log.action_by else None,
                'name': f"{log.action_by.first_name} {log.action_by.last_name}".strip() if log.action_by else None
            },
            'action_to': {
                'id': log.action_to.id if log.action_to else None,
                'name': f"{log.action_to.first_name} {log.action_to.last_name}".strip() if log.action_to else None
            },
            'created_at': log.action_date_time
        } for log in activity_logs]

        return Response(data)

    def perform_update(self, serializer):
        serializer.save(
            updated_by=self.request.user,
            updated_on=timezone.now()
        )

    @action(detail=True, methods=['post'])
    def share(self, request, pk=None):
        """Share task with users/groups and optionally send emails"""
        task = self.get_object()
        serializer = TaskShareSerializer(data=request.data, context={'request': request})

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        users = serializer.validated_data.get('users', [])
        groups = serializer.validated_data.get('groups', [])
        send_email = serializer.validated_data.get('send_email', False)

        created_assignees = []
        email_recipients = []

        # Share with individual users
        for user_id in users:
            user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)

            # Check if already assigned
            existing = TaskAssignee.objects.filter(
                task=task,
                assigned_user=user
            ).exists()

            if not existing:
                assignee = TaskAssignee.objects.create(
                    task=task,
                    assigned_user=user,
                    assigned_date_time=timezone.now()
                )
                created_assignees.append(assignee)
                email_recipients.append(user.email)

        # Share with groups
        for group_id in groups:
            group = get_object_or_404(Groups, id=group_id, organization=request.user.organization)

            # Check if group already assigned
            existing = TaskAssignee.objects.filter(
                task=task,
                assigned_group=group
            ).exists()

            if not existing:
                assignee = TaskAssignee.objects.create(
                    task=task,
                    assigned_group=group,
                    assigned_date_time=timezone.now()
                )
                created_assignees.append(assignee)

                # Add all group members to email recipients
                for member in group.members.filter(organization=request.user.organization):
                    email_recipients.append(member.email)

        # Send emails if requested
        if send_email and email_recipients:
            self._send_task_assignment_emails(task, email_recipients)

        # Create share audit log for all share attempts (regardless of if new assignees created)
        action_to = None
        if users:
            action_to = get_object_or_404(CustomUser, id=users[0], organization=request.user.organization)
        TaskAuditLog.objects.create(task=task, task_action='assigned', action_by=request.user, action_to=action_to)

        # Update task status if assignees were created
        if created_assignees:
            task.status = 'not_started'
            task.save()

        response_data = {
            'message': f'Task shared with {len(created_assignees)} assignee(s) successfully.',
            'created_assignees': TaskAssigneeSerializer(created_assignees, many=True).data,
            'emails_sent': send_email and len(email_recipients) > 0
        }

        return Response(response_data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
        """Update task status"""
        task = self.get_object()
        serializer = TaskUpdateSerializer(task, data=request.data, partial=True)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Get old status for audit log check
        old_status = task.status
        new_status = serializer.validated_data.get('status')
        is_status_changing_to_completed = new_status == 'completed' and old_status != 'completed'

        serializer.save(updated_by=request.user, updated_on=timezone.now())

        # Add audit log for task completion
        if is_status_changing_to_completed:
            TaskAuditLog.objects.create(
                task=task,
                task_action='Task Completed',
                action_by=request.user,
                action_to=None
            )

        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'])
    def mark_complete(self, request, pk=None):
        """Mark task as complete when complete: true is received"""
        task = self.get_object()

        # Check if complete is true
        if request.data.get('complete') != True:
            return Response({'error': 'complete must be true to mark task as completed'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if this is a followup task with task close questions - prevent manual completion
        is_followup_task = task.followup_task_form_id is not None
        if is_followup_task:
            from form.models import TaskCloseQuestion
            has_task_close_questions = TaskCloseQuestion.objects.filter(
                task=task,
                organization=request.user.organization
            ).exists()
            if has_task_close_questions:
                return Response({
                    'error': 'Cannot manually complete followup task with task close questions. Complete the task close questions instead.'
                }, status=status.HTTP_400_BAD_REQUEST)

        # Get old status for audit log check
        old_status = task.status
        is_status_changing_to_completed = old_status != 'completed'

        # Check if this is a followup task for appropriate audit message
        audit_action = 'Followup_Completed' if is_followup_task else 'Task Completed'

        with transaction.atomic():
            # Update task status
            task.status = 'completed'
            task.updated_by = request.user
            task.updated_on = timezone.now()
            task.save(update_fields=['status', 'updated_by', 'updated_on'])

            # Update all TaskTracking records for this task
            TaskTracking.objects.filter(task=task).update(status='completed')

            # Add audit log for task completion if changing to completed
            if is_status_changing_to_completed:
                TaskAuditLog.objects.create(
                    task=task,
                    task_action=audit_action,  # "Followup_Completed" or "Task Completed"
                    action_by=request.user,
                    action_to=None
                )

        return Response({'message': 'Task marked as completed successfully.', 'task': TaskSerializer(task, context={'request': request}).data}, status=status.HTTP_200_OK)



    @action(detail=True, methods=['patch', 'post'])
    def reopen(self, request, pk=None):
        """Reopen a completed task, clear assignees, and prepare for sharing"""
        task = self.get_object()

        # Get remarks from request (for followup tasks)
        remarks = request.data.get('remarks', '').strip()
        normalized_remarks = remarks or None

        # Robust followup detection (do not rely on FollowUpTask table)
        from form.models import TaskCloseQuestion
        has_task_close_questions = TaskCloseQuestion.objects.filter(
            task=task,
            organization=request.user.organization
        ).exists()
        is_followup_task = (
            task.followup_task_form_id is not None
            or task.follow_task_sub_question is not None
            or has_task_close_questions
        )

        # Followup reopen ownership: only main form submitter can reopen.
        # Ownership is inferred from the first Followup_Created audit action.
        if is_followup_task:
            owner_log = TaskAuditLog.objects.filter(
                task=task,
                task_action='Followup_Created'
            ).order_by('action_date_time').first()
            if owner_log and owner_log.action_by_id != request.user.id:
                return Response(
                    {"error": "Only the main form submitter can reopen this followup task."},
                    status=status.HTTP_403_FORBIDDEN
                )
        logger.info(
            "Reopen: task_id=%s, remarks='%s', followup_form_id=%s, follow_sub_q=%s, "
            "has_task_close_questions=%s, is_followup_task=%s",
            task.id,
            remarks,
            task.followup_task_form_id_id,
            task.follow_task_sub_question_id,
            has_task_close_questions,
            is_followup_task,
        )

        # Support both old format (new_users/new_groups/new_leaders) and new format (assigned_users)
        assigned_users = request.data.get('assigned_users', [])
        if assigned_users:
            # New format: assigned_users array
            new_users = assigned_users
            new_groups = []
            new_leaders = []
        else:
            # Old format: separate arrays
            new_users = request.data.get('new_users', [])
            new_groups = request.data.get('new_groups', [])
            new_leaders = request.data.get('new_leaders', [])
        restore_followup_assignees = False

        # If this is a followup task and no new assignees were provided, restore original assignees
        if is_followup_task and not (new_users or new_groups or new_leaders):
            restored_users: list = []
            restored_groups: list = []
            restored_leaders: list = []

            # Infer from LogicFollowUp config using main form + triggering question
            if task.followup_task_form_id_id and task.follow_task_sub_question_id:
                logic_followup = LogicFollowUp.objects.filter(
                    followup_toggle=True,
                    question_id=task.follow_task_sub_question_id,
                ).filter(
                    Q(form_id=task.followup_task_form_id_id) |
                    Q(audit_group__form_id=task.followup_task_form_id_id)
                ).order_by('-id').first()
                if logic_followup:
                    restored_users = list(logic_followup.assign_user_ids or [])
                    restored_groups = list(logic_followup.assign_group_ids or [])
                    restored_leaders = list(logic_followup.assign_leader_ids or [])

            if restored_users or restored_groups or restored_leaders:
                new_users = restored_users
                new_groups = restored_groups
                new_leaders = restored_leaders
                restore_followup_assignees = True

        with transaction.atomic():
            if not is_followup_task:
                # Clear all assignees for regular tasks
                task.assignees.all().delete()

                # Reset regular task status to not_assigned
                task.status = 'not_assigned'
                task.save(update_fields=['status'])
            else:
                # Keep assignees for followup tasks and reset status to not_started
                task.status = 'not_started'
                task.save(update_fields=['status'])
                # If we are restoring assignees, clear existing to avoid duplicates
                if restore_followup_assignees:
                    task.assignees.all().delete()

            # Add new assignees if provided
            for user_id in new_users:
                user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)
                TaskAssignee.objects.create(
                    task=task,
                    assigned_user=user,
                    assigned_date_time=timezone.now()
                )

            for group_id in new_groups:
                group = get_object_or_404(Groups, id=group_id, organization=request.user.organization)
                TaskAssignee.objects.create(
                    task=task,
                    assigned_group=group,
                    assigned_date_time=timezone.now()
                )

            for leader_id in new_leaders:
                leader = get_object_or_404(CustomUser, id=leader_id, organization=request.user.organization)
                TaskAssignee.objects.create(
                    task=task,
                    assigned_leader=leader,
                    assigned_date_time=timezone.now()
                )

            # Save reopen details + audit log
            task.reopened_remarks = normalized_remarks
            if is_followup_task:
                # Update task description with latest reopen reason (followup only)
                if remarks:
                    original_description = task.description or ""
                    # Remove any previous reopen tag so only the latest is kept
                    original_description = re.sub(r"\[REOPENED:\s*[^\]]*\]", "", original_description).strip()
                    reopen_text = f"[REOPENED: {remarks}]"
                    if original_description:
                        task.description = f"{original_description}\n\n{reopen_text}".strip()
                    else:
                        task.description = reopen_text
                    task.save(update_fields=['reopened_remarks', 'description'])
                else:
                    task.save(update_fields=['reopened_remarks'])

                # Create audit log for followup reopening
                TaskAuditLog.objects.create(
                    task=task,
                    task_action='Followup_Reopened',
                    action_by=request.user,
                    action_to=None
                )
            else:
                task.save(update_fields=['reopened_remarks'])
                # Regular task reopening
                TaskAuditLog.objects.create(
                    task=task,
                    task_action='Reopened',
                    action_by=request.user,
                    action_to=None
                )

        message = 'Task has been reopened and is ready for reassignment.'
        if remarks:
            message += ' Remarks have been recorded.'
        if new_users or new_groups or new_leaders:
            message += f' Reassigned to {len(new_users)} users, {len(new_groups)} groups, {len(new_leaders)} leaders.'

        # Reopen auto-closed related tasks (tasks that were auto-closed when this task was closed)
        reopened_related = []
        try:
            from .utils import get_location_from_task, get_location_from_submission, get_main_form_submission_for_task
            # Query directly for related tasks INCLUDING completed ones (find_related_tasks excludes completed)
            if task.followup_task_form_id and task.follow_task_sub_question:
                location_info = get_location_from_task(task)
                if location_info:
                    candidate_tasks = Task.objects.filter(
                        organization=request.user.organization,
                        followup_task_form_id=task.followup_task_form_id,
                        follow_task_sub_question=task.follow_task_sub_question,
                        status='completed',
                    ).exclude(id=task.id)

                    for candidate in candidate_tasks:
                        try:
                            has_auto_close_log = TaskAuditLog.objects.filter(
                                task=candidate,
                                task_action='Auto_Closed_Related_Task'
                            ).exists()
                            if has_auto_close_log:
                                # Verify same location
                                candidate_submission = get_main_form_submission_for_task(candidate)
                                if not candidate_submission:
                                    continue
                                candidate_location = get_location_from_submission(candidate_submission)
                                if candidate_location != location_info:
                                    continue
                                candidate.status = 'not_started'
                                candidate.save(update_fields=['status'])
                                TaskAuditLog.objects.create(
                                    task=candidate,
                                    task_action='Followup_Reopened',
                                    action_by=request.user,
                                    action_to=None
                                )
                                reopened_related.append({
                                    'id': candidate.id,
                                    'task_name': candidate.task_name,
                                })
                        except Exception:
                            continue
        except Exception as e:
            logger.warning("Reopen: failed to reopen auto-closed related tasks for task %s: %s", task.id, str(e))

        if reopened_related:
            message += f' {len(reopened_related)} auto-closed related task(s) also reopened.'

        return Response({
            'message': message,
            'task': TaskSerializer(task, context={'request': request}).data,
            'reopened_related_tasks': reopened_related,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'])
    def start_followup(self, request, pk=None):
        """Start a followup task (locks it for the current user and removes other assignments)"""
        task = self.get_object()

        # Check if user is assigned to this task
        is_assigned = TaskAssignee.objects.filter(
            task=task,
            assigned_user=request.user
        ).exists() or TaskAssignee.objects.filter(
            task=task,
            assigned_group__members=request.user
        ).exists()

        # For followup tasks, also allow users in the same organization to start them
        # This handles cases where followup task assignment failed during creation
        if not is_assigned and task.followup_task_form_id and task.organization == request.user.organization:
            # Allow any user in the organization to start followup tasks
            is_assigned = True
            logger.info(f"Allowing user {request.user.id} to start followup task {task.id} as organization member")

        if not is_assigned:
            return Response({"error": "You are not assigned to this followup task"}, status=status.HTTP_403_FORBIDDEN)

        # Check if this is a followup task by checking followup_task_form_id
        if not task.followup_task_form_id:
            return Response({"error": "This is not a followup task"}, status=status.HTTP_400_BAD_REQUEST)

        # Check if task is already in progress
        if task.status in ('followup_started', 'in_progress'):
            return Response({"error": "This followup task has already been started"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            # Ensure the current user has a direct assigned_user record
            # before removing other assignments. This prevents the task from
            # vanishing when the user was only assigned via a group.
            user_assignee = TaskAssignee.objects.filter(
                task=task,
                assigned_user=request.user
            ).first()

            if not user_assignee:
                TaskAssignee.objects.create(
                    task=task,
                    assigned_user=request.user,
                    assigned_date_time=timezone.now()
                )

            # Delete all assignments except for the current user's direct assignment
            # This ensures only the user who started it can continue working on it
            TaskAssignee.objects.filter(
                task=task
            ).exclude(
                assigned_user=request.user
            ).delete()

            # Update task status to indicate it's started
            task.status = 'in_progress'
            task.save()

            # Create audit log
            TaskAuditLog.objects.create(
                task=task,
                task_action='Followup_Started',
                action_by=request.user,
                action_to=None
            )

        return Response({
            'message': 'Followup task started successfully',
            'task': TaskSerializer(task, context={'request': request}).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'])
    def start(self, request, pk=None):
        """Start a regular (non-followup) task — sets status to in_progress."""
        task = self.get_object()

        # Check if user is assigned to this task
        is_assigned = TaskAssignee.objects.filter(
            task=task,
            assigned_user=request.user
        ).exists() or TaskAssignee.objects.filter(
            task=task,
            assigned_group__members=request.user
        ).exists()

        if not is_assigned:
            return Response({"error": "You are not assigned to this task"}, status=status.HTTP_403_FORBIDDEN)

        # This endpoint is for regular tasks only (no followup)
        if task.followup_task_form_id:
            return Response({"error": "Use start_followup for followup tasks"}, status=status.HTTP_400_BAD_REQUEST)

        if task.status in ('in_progress', 'completed'):
            return Response({"error": "Task already started or completed"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            task.status = 'in_progress'
            task.save()

            TaskAuditLog.objects.create(
                task=task,
                task_action='Started',
                action_by=request.user,
                action_to=None
            )

        return Response({
            'message': 'Task started successfully',
            'task': TaskSerializer(task, context={'request': request}).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'])
    def complete(self, request, pk=None):
        """Complete a regular (non-followup) task."""
        try:
            task = self.get_object()
        except Exception:
            from django.shortcuts import get_object_or_404
            task = get_object_or_404(Task, id=pk, organization=request.user.organization)

        if task.followup_task_form_id:
            return Response({"error": "Use complete_followup for followup tasks"}, status=status.HTTP_400_BAD_REQUEST)

        if task.status != 'in_progress':
            return Response({"error": "Task must be in progress to complete"}, status=status.HTTP_400_BAD_REQUEST)

        close_related_task_ids = request.data.get('close_related_task_ids', None)

        with transaction.atomic():
            task.status = 'completed'
            task.save()

            TaskAuditLog.objects.create(
                task=task,
                task_action='Completed',
                action_by=request.user,
                action_to=None
            )

        # Close selected related tasks with same location + question
        from .utils import close_related_tasks
        related_tasks_result = close_related_tasks(
            task,
            request.user,
            request.user.organization,
            selected_task_ids=close_related_task_ids
        )

        return Response({
            'message': 'Task completed successfully' + (f" and {related_tasks_result['count']} related task(s) closed" if related_tasks_result['count'] > 0 else ""),
            'task': TaskSerializer(task, context={'request': request}).data,
            'related_tasks_closed': related_tasks_result
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reassign_back(self, request, pk=None):
        """Reassign a started task back to group/user/leader, resetting status to not_started.
        The user who started the task gives up ownership and the task becomes available
        again for the specified assignees.

        Payload: {
            "users": [user_id, ...],   # optional
            "groups": [group_id, ...], # optional
            "leaders": [user_id, ...]  # optional
        }
        """
        task = self.get_object()

        # Verify the current user is the one who started it
        user_assignee = TaskAssignee.objects.filter(
            task=task,
            assigned_user=request.user
        ).exists()

        if not user_assignee and task.organization != request.user.organization:
            return Response(
                {"error": "You are not authorized to reassign this task"},
                status=status.HTTP_403_FORBIDDEN
            )

        users = request.data.get('users', [])
        groups = request.data.get('groups', [])
        leaders = request.data.get('leaders', [])

        if not users and not groups and not leaders:
            return Response(
                {"error": "At least one user, group, or leader must be specified"},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            # Remove the current user's direct assignment
            TaskAssignee.objects.filter(
                task=task,
                assigned_user=request.user
            ).delete()

            # Add new assignees
            for user_id in users:
                try:
                    u = CustomUser.objects.get(id=user_id, organization=task.organization)
                    if not TaskAssignee.objects.filter(task=task, assigned_user=u).exists():
                        TaskAssignee.objects.create(
                            task=task,
                            assigned_user=u,
                            assigned_date_time=timezone.now()
                        )
                except CustomUser.DoesNotExist:
                    continue

            for group_id in groups:
                try:
                    g = Groups.objects.get(id=group_id, organization=task.organization)
                    if not TaskAssignee.objects.filter(task=task, assigned_group=g).exists():
                        TaskAssignee.objects.create(
                            task=task,
                            assigned_group=g,
                            assigned_date_time=timezone.now()
                        )
                except Groups.DoesNotExist:
                    continue

            for leader_id in leaders:
                try:
                    l = CustomUser.objects.get(id=leader_id, organization=task.organization)
                    if not TaskAssignee.objects.filter(task=task, assigned_leader=l).exists():
                        TaskAssignee.objects.create(
                            task=task,
                            assigned_leader=l,
                            assigned_date_time=timezone.now()
                        )
                except CustomUser.DoesNotExist:
                    continue

            # Reset task status
            task.status = 'not_started'
            task.save()

            # Create audit log
            TaskAuditLog.objects.create(
                task=task,
                task_action='Reassigned_Back',
                action_by=request.user
            )

        return Response({
            'message': 'Task reassigned successfully',
            'task': TaskSerializer(task, context={'request': request}).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'])
    def complete_followup(self, request, pk=None):
        """Complete a followup task after task close questions are answered
        
        This will also automatically close all related tasks (same location + question)
        and return information about related tasks closed.
        """
        from .utils import close_related_tasks, get_related_tasks_info
        
        # For complete_followup called after task close questions, be more lenient
        # since assignment was already verified during task close questions submission
        try:
            task = self.get_object()
        except Exception:
            # If get_object fails, try direct lookup (bypass visibility filters)
            from django.shortcuts import get_object_or_404
            task = get_object_or_404(Task, id=pk, organization=request.user.organization)

        # Check if this is a followup task
        if not task.followup_task_form_id:
            return Response({"error": "This is not a followup task"}, status=status.HTTP_400_BAD_REQUEST)

        # Update task status to completed
        if task.status != 'completed':
            task.status = 'completed'
            task.updated_by = request.user
            task.updated_on = timezone.now()
            task.save()

        # Create audit log for main task
        TaskAuditLog.objects.create(
            task=task,
            task_action='Followup_Completed',
            action_by=request.user,
            action_to=None
        )

        close_related_task_ids = request.data.get('close_related_task_ids', None)

        # Close selected related tasks (same location + question combination)
        related_tasks_result = close_related_tasks(
            task,
            request.user,
            request.user.organization,
            selected_task_ids=close_related_task_ids
        )

        return Response({
            'message': 'Task completed successfully' + (f" and {related_tasks_result['count']} related task(s) closed" if related_tasks_result['count'] > 0 else ""),
            'task': TaskSerializer(task, context={'request': request}).data,
            'related_tasks_closed': related_tasks_result,
            'auto_close_info': {
                'enabled': close_related_task_ids is not None and len(close_related_task_ids) > 0,
                'description': 'Tasks with same Location and Question can be closed together'
            }
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def close_related(self, request, pk=None):
        """Close selected related tasks for an already-completed task.
        
        Payload: { 'related_task_ids': [id1, id2, ...] }
        Only tasks that are related by same location + question are actually closed.
        """
        from .utils import close_related_tasks, find_related_tasks
        
        try:
            task = self.get_object()
        except Exception:
            from django.shortcuts import get_object_or_404
            task = get_object_or_404(Task, id=pk, organization=request.user.organization)
        
        related_task_ids = request.data.get('related_task_ids', None)
        if related_task_ids is None:
            return Response({'error': 'related_task_ids is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not isinstance(related_task_ids, list):
            return Response({'error': 'related_task_ids must be a list'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate that provided IDs belong to related tasks
        related_tasks = find_related_tasks(task, request.user.organization)
        valid_related_ids = set(related_tasks.values_list('id', flat=True))
        selected_valid_ids = [int(x) for x in related_task_ids if int(x) in valid_related_ids]
        
        result = close_related_tasks(
            task,
            request.user,
            request.user.organization,
            selected_task_ids=selected_valid_ids
        )
        
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def related_tasks(self, request, pk=None):
        """Get preview of related tasks that will be closed together
        
        Returns information about tasks with same location + question
        that will be automatically closed when this task is completed.
        """
        from .utils import get_related_tasks_info
        
        try:
            task = self.get_object()
        except Exception:
            from django.shortcuts import get_object_or_404
            task = get_object_or_404(Task, id=pk, organization=request.user.organization)

        if not task.follow_task_sub_question:
            return Response({
                'has_related_tasks': False,
                'count': 0,
                'tasks': [],
                'message': 'No related tasks found (not a follow-up task)',
            }, status=status.HTTP_200_OK)

        related_info = get_related_tasks_info(task, request.user.organization)
        
        return Response({
            'has_related_tasks': related_info['has_related_tasks'],
            'count': related_info['count'],
            'tasks': related_info['tasks'],
            'message': f"Closing this task will also close {related_info['count']} related task(s) with the same Location and Question" if related_info['count'] > 0 else 'No related tasks',
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'])
    def extend_due_date(self, request, pk=None):
        """Extend the task due date - dedicated endpoint for frontend integration"""
        task = self.get_object()

        # Get the new end_date from request
        new_end_date = request.data.get('end_date')
        if not new_end_date:
            return Response({'error': 'end_date is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Validate date format
        try:
            # Convert string to datetime if needed
            if isinstance(new_end_date, str):
                new_end_date = datetime.datetime.fromisoformat(new_end_date.replace('Z', '+00:00'))
        except ValueError:
            return Response({'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

        # Get old end_date
        old_end_date = task.end_date

        # Check if due date was actually extended
        if new_end_date and old_end_date and new_end_date > old_end_date:
            # Update the task directly (bypass serializer to avoid duplicate logging)
            task.end_date = new_end_date
            task.save(update_fields=['end_date'])

            # Create only the "Due Extended" audit log
            TaskAuditLog.objects.create(
                task=task,
                task_action='Due Extended',
                action_by=request.user,
                action_to=None
            )
        else:
            return Response({'error': 'New end date must be after current end date'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'message': 'Task due date extended successfully.',
            'task': TaskSerializer(task, context={'request': request}).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def submit_mobile(self, request, pk=None):
        """Simplified mobile task submission - completes task with current user as assignee"""
        task = self.get_object()

        # Check if user is already assigned to this task
        user_assigned = TaskAssignee.objects.filter(
            task=task,
            assigned_user=request.user
        ).exists()

        group_assigned = TaskAssignee.objects.filter(
            task=task,
            assigned_group__members=request.user
        ).exists()

        logic_assigned, logic_group = _is_user_assigned_via_logic_followup(task, request.user)

        if not user_assigned and not group_assigned and not logic_assigned:
            return Response(
                {'error': 'You are not assigned to this task and cannot submit it.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get assignee_group if user is assigned through group
        assignee_group = None
        if not user_assigned and group_assigned:
            assignee = TaskAssignee.objects.filter(
                task=task,
                assigned_group__members=request.user
            ).first()
            if assignee:
                assignee_group = assignee.assigned_group

        if not assignee_group and logic_group:
            assignee_group = logic_group

        # Get comments if provided
        comments = request.data.get('comments', 'Submitted from mobile')

        # Find existing task tracking record or create new one
        # Look for existing tracking record created when START was pressed
        if user_assigned:
            tracking = TaskTracking.objects.filter(
                task=task,
                assignee_user=request.user
            ).first()
        elif group_assigned and assignee_group:
            tracking = TaskTracking.objects.filter(
                task=task,
                assignee_user=request.user,  # Even group-assigned users have assignee_user set
                assignee_group=assignee_group
            ).first()
        else:
            tracking = None

        if tracking:
            # Update existing tracking record
            tracking.actual_end_date = timezone.now()
            tracking.status = 'completed'
            if comments:
                tracking.comments = comments
            tracking.save()
        else:
            # Create new tracking record for completion
            tracking_data = {
                'task': task,
                'assignee_user': request.user,  # Always set assignee_user (required by model)
                'assignee_group': assignee_group,
                'comments': comments,
                'actual_end_date': timezone.now(),
                'status': 'completed'
            }

            # If task was not started, set start date too
            if task.status == 'not_started':
                tracking_data['actual_start_date'] = timezone.now()

            tracking = TaskTracking.objects.create(**tracking_data)

        # Check if this task has task close questions - don't complete immediately if it does
        from form.models import TaskCloseQuestion
        has_task_close_questions = TaskCloseQuestion.objects.filter(
            task=task,
            organization=request.user.organization
        ).exists()

        is_followup_task = task.followup_task_form_id is not None

        close_related_task_ids = request.data.get('close_related_task_ids', None)

        if has_task_close_questions:
            # For tasks with task close questions, don't mark as completed yet
            # The task close questions flow will handle completion
            related_tasks_result = {
                'related_task_ids': [],
                'count': 0,
                'details': [],
                'message': 'No related tasks found'
            }
        else:
            # For tasks without task close questions, mark as completed
            task.status = 'completed'
            task.save(update_fields=['status'])

            # Create audit log
            audit_action = 'Followup_Completed' if is_followup_task else 'Completed'
            TaskAuditLog.objects.create(
                task=task,
                task_action=audit_action,
                action_by=request.user,
                action_to=None
            )

            # Close selected related tasks with same location + question
            from .utils import close_related_tasks
            related_tasks_result = close_related_tasks(
                task,
                request.user,
                request.user.organization,
                selected_task_ids=close_related_task_ids
            )

        return Response({
            'message': 'Task submitted successfully from mobile.' + (f" and {related_tasks_result['count']} related task(s) closed" if related_tasks_result['count'] > 0 else ""),
            'task': TaskSerializer(task, context={'request': request}).data,
            'tracking_record': TaskTrackingSerializer(tracking).data,
            'has_task_close_questions': has_task_close_questions,
            'is_followup_task': is_followup_task,
            'related_tasks_closed': related_tasks_result
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def assignees(self, request, pk=None):
        """Get all assignees for a task"""
        task = self.get_object()
        assignees = task.assignees.all()
        serializer = TaskAssigneeSerializer(assignees, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def tracking(self, request, pk=None):
        """Get tracking records for a task"""
        task = self.get_object()
        tracking_records = task.tracking_records.all()
        serializer = TaskTrackingSerializer(tracking_records, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def audit_logs(self, request, pk=None):
        """Get audit log records for a task"""
        task = self.get_object()
        audit_logs = task.audit_logs.all()
        serializer = TaskAuditLogSerializer(audit_logs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def count(self, request):
        """Get total count of tasks and counts by status for the user's organization"""
        queryset = self.get_queryset()
        total_count = queryset.count()
        not_started_count = queryset.filter(status__in=['not_started', 'not_assigned']).count()
        in_progress_count = queryset.filter(status='in_progress').count()
        completed_count = queryset.filter(status='completed').count()
        return Response({
            'total_task_count': total_count,
            'not_started': not_started_count,
            'in_progress': in_progress_count,
            'completed': completed_count
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def planner_stats(self, request):
        """Get planner statistics for the user's organization"""
        org = request.user.organization
        queryset = self.get_queryset()
        today = timezone.now().date()

        # Overall counts
        total_tasks = queryset.count()
        tasks_not_started = queryset.filter(status__in=['not_started', 'not_assigned']).count()
        tasks_in_progress = queryset.filter(status='in_progress').count()
        tasks_completed = queryset.filter(status='completed').count()
        overdue_tasks = queryset.filter(end_date__lt=today).count()
        tasks_today = queryset.filter(created_on__date=today).count()

        # Forms with tasks
        total_forms = Form.objects.filter(
            organization=org,
            is_deleted=False,
            is_archived=False,
            task__organization=org
        ).distinct().count()

        forms_today = Form.objects.filter(
            organization=org,
            is_deleted=False,
            is_archived=False,
            task__organization=org,
            created_at__date=today
        ).distinct().count()

        # Forms breakdown
        forms_breakdown = Form.objects.filter(
            organization=org,
            is_deleted=False,
            is_archived=False,
            task__organization=org
        ).distinct().annotate(
            total_tasks=Count('task', filter=Q(task__organization=org)),
            not_started=Count('task', filter=Q(task__organization=org, task__status__in=['not_started', 'not_assigned'])),
            in_progress=Count('task', filter=Q(task__organization=org, task__status='in_progress')),
            completed=Count('task', filter=Q(task__organization=org, task__status='completed')),
            overdue=Count('task', filter=Q(task__organization=org, task__end_date__lt=today))
        ).values('id', 'title', 'total_tasks', 'not_started', 'in_progress', 'completed', 'overdue').order_by('-total_tasks')

        forms_breakdown_list = [{
            'form_id': f['id'],
            'form_name': f['title'],
            'total_tasks': f['total_tasks'],
            'not_started': f['not_started'],
            'in_progress': f['in_progress'],
            'completed': f['completed'],
            'overdue': f['overdue']
        } for f in forms_breakdown]

        return Response({
            'total_forms': total_forms,
            'total_tasks': total_tasks,
            'tasks_not_started': tasks_not_started,
            'tasks_in_progress': tasks_in_progress,
            'tasks_completed': tasks_completed,
            'forms_today': forms_today,
            'tasks_today': tasks_today,
            'overdue_tasks': overdue_tasks,
            'forms_breakdown': forms_breakdown_list
        }, status=status.HTTP_200_OK)

    def destroy(self, request, *args, pk=None, **kwargs):
        """Custom delete method to return a success message"""
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(
            {"message": f"Task '{instance.task_name}' has been successfully deleted."},
            status=status.HTTP_200_OK
        )


    def _send_task_assignment_emails(self, task, recipients):
        """Send task assignment emails to recipients"""
        try:
            subject = f"New Task Assigned: {task.task_name}"

            created_by_name = f"{task.created_by.first_name} {task.created_by.last_name}".strip() or task.created_by.username

            message = f"""
                        You have been assigned a new task:

                        Task Name: {task.task_name}
                        Description: {task.description or 'N/A'}
                        Form: {task.form.title}
                        Start Date: {task.start_date.strftime('%d-%b-%Y %I:%M %p')}
                        End Date: {task.end_date.strftime('%d-%b-%Y %I:%M %p')}
                        Assigned By: {created_by_name}

                        Please complete this task within the specified timeframe.
                        """

            email = EmailMessage(
                subject=subject,
                body=message,
                from_email='no-reply@vibro.com',
                to=recipients,
            )

            email.send(fail_silently=False)

        except Exception as e:
            # Log the error but don't fail the task creation
            print(f"Failed to send task assignment emails: {str(e)}")

class TaskAssigneeViewSet(userContextAPIView, ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = TaskAssignee.objects.all()
    serializer_class = TaskAssigneeSerializer

    def get_queryset(self):
        return TaskAssignee.objects.filter(
            task__organization=self.request.user.organization
        ).select_related('task', 'assigned_user', 'assigned_group')
    
    def get_object(self):
        """Allow using task ID as pk if TaskAssignee ID doesn't exist"""
        pk = self.kwargs.get('pk')
        queryset = self.get_queryset()

        try:
            # Try to get by TaskAssignee id
            obj = queryset.get(pk=pk)
            return obj
        except TaskAssignee.DoesNotExist:
            # Try to get first assignee for the task with this id
            try:
                obj = queryset.filter(task_id=pk).first()
                if obj:
                    return obj
                else:
                    raise TaskAssignee.DoesNotExist()
            except TaskAssignee.DoesNotExist:
                from django.http import Http404
                raise Http404("No TaskAssignee matches the given query.")

class TaskTrackingViewSet(userContextAPIView, ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = TaskTracking.objects.all()
    serializer_class = TaskTrackingCreateSerializer

    def get_queryset(self):
        return TaskTracking.objects.filter(
            task__organization=self.request.user.organization
        ).select_related('task', 'assignee_user', 'assignee_group')

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return TaskTrackingCreateSerializer
        return TaskTrackingSerializer

class TaskAuditLogViewSet(userContextAPIView, ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = TaskAuditLog.objects.all()
    serializer_class = TaskAuditLogSerializer

    def get_queryset(self):
        queryset = TaskAuditLog.objects.filter(
            task__organization=self.request.user.organization
        ).select_related('task', 'action_by', 'action_to')

        # Optional filtering by task
        task_id = self.request.query_params.get('task')
        if task_id:
            queryset = queryset.filter(task_id=task_id)

        return queryset

    def retrieve(self, request, pk=None):
        """Custom retrieve method: return audit logs for a task when pk is treated as task_id"""
        queryset = self.get_queryset()
        task_id = pk

        # Check if the task exists and belongs to user's organization
        try:
            from .models import Task
            task = Task.objects.get(id=task_id, organization=request.user.organization)
        except Task.DoesNotExist:
            from django.http import Http404
            raise Http404("No Task matches the given query.")

        # Get audit logs for this task
        audit_logs = queryset.filter(task_id=task_id)
        serializer = TaskAuditLogSerializer(audit_logs, many=True)
        return Response(serializer.data)




class UserAssignedTasksView(APIView):
    """Get tasks assigned to current user with enhanced status and assignee info"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        # Get tasks where user is directly assigned
        # Exclude tasks whose form is archived, but include tasks with no form (followup tasks)
        user_assigned_tasks = TaskAssignee.objects.filter(
            assigned_user=user,
            task__organization=user.organization
        ).exclude(
            task__followup_task__deadline_at__lt=timezone.now()
        ).exclude(
            task__form__is_archived=True
        ).select_related('task__form', 'task__created_by')

        # Get tasks where user's groups are assigned
        group_assigned_tasks = TaskAssignee.objects.filter(
            assigned_group__members=user,
            task__organization=user.organization
        ).exclude(
            task__followup_task__deadline_at__lt=timezone.now()
        ).exclude(
            task__form__is_archived=True
        ).select_related('task__form', 'task__created_by')

        # Combine and deduplicate
        all_assignees = user_assigned_tasks.union(group_assigned_tasks)
        tasks = [assignee.task for assignee in all_assignees]

        # Remove duplicates
        seen_ids = set()
        unique_tasks = []
        for task in tasks:
            if task.id not in seen_ids:
                seen_ids.add(task.id)
                unique_tasks.append(task)

        # Enhance each task with status and assigned users
        enhanced_tasks = []

        # Batch: get all form IDs referenced by tasks, then check which have location questions
        all_form_ids = set()
        for task in unique_tasks:
            if task.form_id:
                all_form_ids.add(task.form_id)
            if task.followup_task_form_id_id:
                all_form_ids.add(task.followup_task_form_id_id)
        from form.models import Question as _Q
        location_form_ids = set(
            _Q.objects.filter(form_id__in=all_form_ids, question_type='location')
            .values_list('form_id', flat=True)
        ) if all_form_ids else set()

        for task in unique_tasks:
            latest_activity = TaskAuditLog.objects.filter(
                task=task
            ).order_by('-action_date_time').first()

            effective_status = str(task.status or 'not_started').strip().lower()
            if effective_status == 'not_assigned':
                effective_status = 'not_started'

            is_followup_task = bool(task.followup_task_form_id_id or task.follow_task_sub_question_id)
            if is_followup_task and latest_activity:
                latest_action = str(latest_activity.task_action or '').strip().lower().replace(' ', '_')
                if latest_action == 'followup_started':
                    effective_status = 'in_progress'
                elif latest_action == 'followup_completed':
                    effective_status = 'completed'

            if effective_status == 'completed':
                continue

            # Get assigned user names (existing TaskAssignee table, no changes)
            assigned_users = TaskAssignee.objects.filter(
                task=task, assigned_user__isnull=False
            ).values_list('assigned_user__first_name', 'assigned_user__last_name')

            assigned_user_names = [
                f"{first} {last}".strip()
                for first, last in assigned_users
            ]

            # Serialize the task
            serializer = TaskListSerializer(task, context={'request': request, 'location_form_ids': location_form_ids})
            task_data = serializer.data

            # Add enhanced fields
            task_data.update({
                'derived_status': effective_status,
                'assigned_users': assigned_user_names
            })

            enhanced_tasks.append(task_data)

        return Response(enhanced_tasks)

class UserCompletedTasksView(APIView):
    """Get completed tasks with form submissions for TODO sent screen"""
    permission_classes = [IsAuthenticated]

    def get(self, request, user_id=None):
        try:
            if user_id:
                user = get_object_or_404(CustomUser, id=user_id, organization=request.user.organization)
            else:
                user = request.user

            from form.models import FormSubmision, Answer, StageSubmissionHistory, TaskCloseQuestion

            # Helper caches for assigned users and start time
            logic_followup_cache = {}
            task_started_cache = {}
            task_assignees_cache = {}

            def get_logic_followup(main_form_id: int, question_id: int):
                key = (main_form_id, question_id)
                if key in logic_followup_cache:
                    return logic_followup_cache[key]
                lf = (
                    LogicFollowUp.objects.filter(
                        followup_toggle=True,
                        question_id=question_id,
                    )
                    .filter(
                        Q(form_id=main_form_id) |
                        Q(audit_group__form_id=main_form_id)
                    )
                    .order_by('-id')
                    .first()
                )
                logic_followup_cache[key] = lf
                return lf

            def get_task_started_at(task: Task | None) -> str | None:
                if not task:
                    return None
                if task.id in task_started_cache:
                    return task_started_cache[task.id]
                start_log = TaskAuditLog.objects.filter(
                    task=task,
                    task_action='Followup_Started'
                ).order_by('-action_date_time').first()
                value = start_log.action_date_time.isoformat() if start_log else None
                task_started_cache[task.id] = value
                return value

            root_planner_cache = {}

            def get_root_planner_source(task: Task | None):
                if not task:
                    return None
                if task.id in root_planner_cache:
                    return root_planner_cache[task.id]
                creation_log = TaskAuditLog.objects.filter(
                    task=task,
                    task_action__iexact='Followup_Created'
                ).order_by('action_date_time').first()
                if not creation_log or not task.followup_task_form_id_id or not creation_log.action_by_id:
                    root_planner_cache[task.id] = None
                    return None
                from form.models import FormSubmision
                from planner.models import PlannerAssignment, PlannerSubmission
                root_submission = FormSubmision.objects.filter(
                    form_id=task.followup_task_form_id_id,
                    submission_initiated_by_id=creation_log.action_by_id,
                    organization=user.organization,
                    submission_initiated_on__lte=creation_log.action_date_time
                ).order_by('-submission_initiated_on').first()
                if not root_submission:
                    root_planner_cache[task.id] = None
                    return None
                planner_submission = PlannerSubmission.objects.filter(
                    form_submission=root_submission
                ).select_related('planner_assignment').first()
                if planner_submission and planner_submission.planner_assignment:
                    planner_order_id = planner_submission.planner_assignment.order_id
                    root_planner_cache[task.id] = planner_order_id
                    return planner_order_id
                root_time = root_submission.completed_on or root_submission.submission_initiated_on
                planner_candidates_qs = PlannerAssignment.objects.filter(
                    form_id=root_submission.form_id,
                    organization=user.organization,
                    is_completed=True,
                    completed_by_id=root_submission.submission_initiated_by_id,
                    completed_on__isnull=False
                ).exclude(order_id__isnull=True)
                if root_time and planner_candidates_qs.exists():
                    planner_assignment = min(
                        planner_candidates_qs,
                        key=lambda pa: abs((pa.completed_on - root_time).total_seconds())
                    )
                    if abs((planner_assignment.completed_on - root_time).total_seconds()) <= 60:
                        root_planner_cache[task.id] = planner_assignment.order_id
                        return planner_assignment.order_id
                root_planner_cache[task.id] = None
                return None

            def get_assigned_user_names(task: Task | None):
                if not task:
                    return []
                if task.id in task_assignees_cache:
                    return task_assignees_cache[task.id]

                names = []

                # Prefer LogicFollowUp config for original assignees
                if task.followup_task_form_id_id and task.follow_task_sub_question_id:
                    lf = get_logic_followup(task.followup_task_form_id_id, task.follow_task_sub_question_id)
                    if lf:
                        user_ids = list(lf.assign_user_ids or [])
                        if user_ids:
                            users = CustomUser.objects.filter(id__in=user_ids, organization=user.organization)
                            for u in users:
                                names.append(f"{u.first_name} {u.last_name}".strip() or u.username)
                        group_ids = list(lf.assign_group_ids or [])
                        if group_ids:
                            groups = Groups.objects.filter(id__in=group_ids, organization=user.organization)
                            for g in groups:
                                names.append(f"Group: {g.name}")
                        leader_ids = list(lf.assign_leader_ids or [])
                        if leader_ids:
                            leaders = CustomUser.objects.filter(id__in=leader_ids, organization=user.organization)
                            for l in leaders:
                                names.append(f"{l.first_name} {l.last_name}".strip() or l.username)

                # Fallback to current TaskAssignee if LogicFollowUp is missing
                if not names:
                    assigned_users = TaskAssignee.objects.filter(
                        task=task, assigned_user__isnull=False
                    ).select_related('assigned_user')
                    for assignee in assigned_users:
                        u = assignee.assigned_user
                        names.append(f"{u.first_name} {u.last_name}".strip() or u.username)

                    assigned_groups = TaskAssignee.objects.filter(
                        task=task, assigned_group__isnull=False
                    ).select_related('assigned_group')
                    for assignee in assigned_groups:
                        names.append(f"Group: {assignee.assigned_group.name}")

                    assigned_leaders = TaskAssignee.objects.filter(
                        task=task, assigned_leader__isnull=False
                    ).select_related('assigned_leader')
                    for assignee in assigned_leaders:
                        u = assignee.assigned_leader
                        names.append(f"{u.first_name} {u.last_name}".strip() or u.username)

                task_assignees_cache[task.id] = names
                return names

            # Get all task IDs assigned to this user (direct/group/leader)
            user_assigned_task_ids = list(TaskAssignee.objects.filter(
                assigned_user=user,
                task__organization=user.organization
            ).values_list('task_id', flat=True).distinct())
            
            group_assigned_task_ids = list(TaskAssignee.objects.filter(
                assigned_group__members=user,
                task__organization=user.organization
            ).values_list('task_id', flat=True).distinct())

            leader_assigned_task_ids = list(TaskAssignee.objects.filter(
                assigned_leader=user,
                task__organization=user.organization
            ).values_list('task_id', flat=True).distinct())
            
            # Combine both sets of task IDs
            task_ids = set(user_assigned_task_ids) | set(group_assigned_task_ids) | set(leader_assigned_task_ids)
            # Include followup tasks where this user is main-form owner (via Followup_Created log)
            owner_task_ids = set(
                TaskAuditLog.objects.filter(
                    task__organization=user.organization,
                    task__followup_task_form_id__isnull=False,
                    task_action='Followup_Created',
                    action_by=user
                ).values_list('task_id', flat=True).distinct()
            )
            # Include tasks where user has followup activity, even if assignment rows are missing.
            activity_task_ids = set(
                TaskAuditLog.objects.filter(
                    task__organization=user.organization,
                    task__followup_task_form_id__isnull=False,
                    action_by=user,
                    task_action__in=['Followup_started', 'Followup_Started', 'Followup_Completed', 'Completed', 'Task Completed']
                ).values_list('task_id', flat=True).distinct()
            )
            task_ids = task_ids | owner_task_ids | activity_task_ids
            logger.info(
                "completed-tasks user=%s org=%s direct_assigned=%s group_assigned=%s leader_assigned=%s owner_tasks=%s activity_tasks=%s combined_task_ids=%s",
                user.id,
                user.organization_id,
                len(user_assigned_task_ids),
                len(group_assigned_task_ids),
                len(leader_assigned_task_ids),
                len(owner_task_ids),
                len(activity_task_ids),
                len(task_ids),
            )

            # Only followup tasks (exclude archived forms)
            followup_tasks = Task.objects.filter(
                id__in=task_ids,
                followup_task_form_id__isnull=False,
                organization=user.organization
            ).exclude(
                followup_task_form_id__is_archived=True
            )
            logger.info(
                "completed-tasks user=%s followup_task_count=%s",
                user.id,
                followup_tasks.count(),
            )

            followup_form_ids = set(
                followup_tasks.exclude(form_id__isnull=True).values_list('form_id', flat=True)
            )
            logger.info(
                "completed-tasks user=%s followup_form_count=%s followup_form_ids_sample=%s",
                user.id,
                len(followup_form_ids),
                list(followup_form_ids)[:10],
            )

            # Completed submissions for related followup forms.
            # Restrict to submissions initiated by this user so that Sent only shows
            # what the user personally submitted. Other users' submissions for the same
            # form (e.g. original submitter after task was reassigned) are excluded from
            # this view but remain intact in the DB for reports and audit analysis.
            if followup_form_ids:
                form_submissions = FormSubmision.objects.filter(
                    form_id__in=followup_form_ids,
                    is_completed=True,
                    organization=user.organization,
                    submission_initiated_by=user
                ).select_related('form').order_by('-completed_on')
            else:
                form_submissions = FormSubmision.objects.none()
            logger.info(
                "completed-tasks user=%s followup_submissions_count=%s",
                user.id,
                form_submissions.count(),
            )

            # Group by form (similar to form sent screen structure)
            tasks_dict = {}
            skipped_only_task_close = 0
            followup_submission_rows = 0
            
            for submission in form_submissions:
                form = submission.form
                form_id = form.id
                
                if form_id not in tasks_dict:
                    tasks_dict[form_id] = {
                        'id': form.id,
                        'title': form.title,
                        'form_type': form.form_type or 'standard',
                        'prefix': getattr(form, 'prefix', None),
                        'group_type': 'form',
                        'sent': []
                    }
                
                # Add submission if not already in list (avoid duplicates)
                submission_exists = any(
                    s.get('form_submission_id') == submission.id 
                    for s in tasks_dict[form_id]['sent']
                )
                
                if not submission_exists:
                    # Skip submissions that only contain task-close answers
                    has_stage_answers = Answer.objects.filter(
                        submission=submission,
                        organization=user.organization,
                        question__is_task_close_question=False
                    ).exists()
                    has_stage_history = StageSubmissionHistory.objects.filter(
                        form_submission=submission,
                        organization=user.organization
                    ).exists()
                    if not (has_stage_answers or has_stage_history):
                        skipped_only_task_close += 1
                        continue

                    # Match followup submission only if there is a completion log near the submission time
                    related_task = None
                    ts = submission.completed_on or submission.submission_initiated_on
                    if ts:
                        window_start = ts - timezone.timedelta(minutes=5)
                        window_end = ts + timezone.timedelta(minutes=5)
                        followup_completion = TaskAuditLog.objects.filter(
                            task__form=form,
                            task__followup_task_form_id__isnull=False,
                            task__organization=user.organization,
                            task_action='Followup_Completed',
                            action_date_time__gte=window_start,
                            action_date_time__lte=window_end
                        ).select_related('task').order_by('-action_date_time').first()
                        related_task = followup_completion.task if followup_completion else None
                        # Keep only tasks relevant to current user context
                        if related_task and related_task.id not in task_ids:
                            related_task = None

                    # Check if this is a followup task submission
                    is_followup_task = related_task is not None
                    submission_type = "[Followup-Task]" if is_followup_task else None
                    can_reopen = bool(related_task and related_task.id in owner_task_ids)
                    task_status = related_task.status if related_task else 'completed'

                    # Skip submissions where the linked task is still open.
                    # In-progress / not-started tasks belong in the New tab only.
                    # The submission record is preserved in DB for reports/audits.
                    if related_task and task_status in ('in_progress', 'not_started', 'not_assigned'):
                        continue
                    assigned_users = get_assigned_user_names(related_task)
                    task_started_at = get_task_started_at(related_task)
                    main_form_title = related_task.followup_task_form_id.title if related_task and related_task.followup_task_form_id else None
                    root_planner_id = get_root_planner_source(related_task)

                    is_auto_closed = related_task and TaskAuditLog.objects.filter(
                        task=related_task,
                        task_action='Auto_Closed_Related_Task'
                    ).exists() if related_task else False

                    tasks_dict[form_id]['sent'].append({
                        'task_id': related_task.id if related_task else None,
                        'task_name': related_task.task_name if related_task else None,
                        'task_status': task_status,
                        'status': task_status,
                        'main_form_title': main_form_title,
                        'assigned_users': assigned_users,
                        'task_started_at': task_started_at,
                        'form_submission_id': submission.id,
                        'form_id': form.id,
                        'is_completed': True,
                        'is_auto_closed': is_auto_closed,
                        'completed_by': submission.completed_by_id,
                        'completed_on': submission.completed_on.isoformat() if submission.completed_on else None,
                        'submission_initiated_on': submission.submission_initiated_on.isoformat() if submission.submission_initiated_on else None,
                        'id': submission.id,
                        'submission_type': submission_type,  # "[Followup-Task]" for followup tasks
                        'can_reopen': can_reopen,  # true for followup tasks
                        'source': 'planner' if root_planner_id is not None else 'task',
                        'source_ref': root_planner_id if root_planner_id is not None else related_task.id if related_task else None,
                    })
                    followup_submission_rows += 1

            logger.info(
                "completed-tasks user=%s form_groups=%s followup_sent_rows=%s skipped_only_task_close=%s",
                user.id,
                len(tasks_dict),
                followup_submission_rows,
                skipped_only_task_close,
            )

            # Include planner-based completed form submissions (collaborative audits + regular planner)
            from planner.models import PlannerSubmission, CollaborativeSubmission
            planner_subs = PlannerSubmission.objects.filter(
                submitted_by=user,
                form_submission__is_completed=True,
            ).select_related('form_submission__form', 'planner_assignment').order_by('-form_submission__completed_on')

            planner_submission_rows = 0
            for ps in planner_subs:
                submission = ps.form_submission
                if not submission:
                    continue
                form = submission.form
                form_id = form.id

                # Skip if this submission is already in tasks_dict (from followup task section)
                if form_id in tasks_dict:
                    already_exists = any(
                        s.get('form_submission_id') == submission.id
                        for s in tasks_dict[form_id]['sent']
                    )
                    if already_exists:
                        continue

                if form_id not in tasks_dict:
                    tasks_dict[form_id] = {
                        'id': form.id,
                        'title': form.title,
                        'form_type': form.form_type or 'standard',
                        'prefix': getattr(form, 'prefix', None),
                        'group_type': 'form',
                        'sent': []
                    }

                # Check if this is a collaborative audit
                is_collab = CollaborativeSubmission.objects.filter(
                    form_submission=submission,
                    organization=user.organization
                ).exists()

                tasks_dict[form_id]['sent'].append({
                    'task_id': None,
                    'task_name': None,
                    'task_status': 'completed',
                    'status': 'completed',
                    'main_form_title': None,
                    'assigned_users': [],
                    'task_started_at': None,
                    'form_submission_id': submission.id,
                    'form_id': form.id,
                    'is_completed': True,
                    'is_auto_closed': False,
                    'completed_by': submission.completed_by_id,
                    'completed_on': submission.completed_on.isoformat() if submission.completed_on else None,
                    'submission_initiated_on': submission.submission_initiated_on.isoformat() if submission.submission_initiated_on else None,
                    'id': submission.id,
                    'submission_type': '[Collaborative-Audit]' if is_collab else '[Planner]',
                    'can_reopen': False,
                    'source': 'planner',
                    'source_ref': ps.planner_assignment.order_id if ps.planner_assignment else None,
                })
                planner_submission_rows += 1

            logger.info(
                "completed-tasks user=%s planner_submission_rows=%s",
                user.id,
                planner_submission_rows,
            )

            response_data = list(tasks_dict.values())
            # Task-close-only submissions
            task_close_dict = {}
            task_close_base_qs = Task.objects.filter(
                id__in=task_ids,
                followup_task_form_id__isnull=False,
                organization=user.organization
            ).distinct()
            scenario2_tasks = task_close_base_qs
            logger.info(
                "completed-tasks user=%s task_close_candidates=%s",
                user.id,
                scenario2_tasks.count(),
            )

            task_close_completion_matches = 0
            task_close_without_completion_log = 0
            task_close_submission_matches = 0
            task_close_no_question_matches = 0
            task_close_window_miss_fallback_hits = 0

            for task in scenario2_tasks:
                # Skip tasks that are not completed — only completed tasks belong in Sent.
                if task.status in ('in_progress', 'not_started', 'not_assigned'):
                    continue
                completion_log = TaskAuditLog.objects.filter(
                    task=task,
                    task_action='Followup_Completed',
                    task__organization=user.organization
                ).order_by('-action_date_time').first()
                window_start = None
                window_end = None
                if completion_log:
                    task_close_completion_matches += 1
                    window_start = completion_log.action_date_time - timezone.timedelta(minutes=10)
                    window_end = completion_log.action_date_time + timezone.timedelta(minutes=10)
                else:
                    task_close_without_completion_log += 1

                question_ids = set(
                    TaskCloseQuestion.objects.filter(
                        task=task,
                        organization=user.organization
                    ).values_list('question_id', flat=True)
                )
                # Fallback: derive task-close questions from LogicFollowUp config.
                if not question_ids and task.followup_task_form_id:
                    logic_followups_with_questions = LogicFollowUp.objects.filter(
                        Q(form=task.followup_task_form_id) |
                        Q(audit_group__form=task.followup_task_form_id),
                        followup_toggle=True,
                        assign_form__isnull=True,
                        task_close_questions__isnull=False
                    )
                    if task.follow_task_sub_question:
                        logic_followups_with_questions = logic_followups_with_questions.filter(
                            question=task.follow_task_sub_question
                        )
                    fallback_question_ids = logic_followups_with_questions.values_list(
                        'task_close_questions__id',
                        flat=True
                    )
                    question_ids.update([qid for qid in fallback_question_ids if qid is not None])

                task_submission_form = task.form or task.followup_task_form_id
                if not task_submission_form:
                    continue

                # Primary: match by explicit task-close question IDs.
                # Fallback: if question links are missing, match submissions with any task-close-flagged answer.
                if question_ids:
                    base_submissions_qs = FormSubmision.objects.filter(
                        form=task_submission_form,
                        is_completed=True,
                        organization=user.organization,
                        answers__question_id__in=question_ids
                    )
                else:
                    task_close_no_question_matches += 1
                    base_submissions_qs = FormSubmision.objects.filter(
                        form=task_submission_form,
                        is_completed=True,
                        organization=user.organization,
                        answers__question__is_task_close_question=True
                    )
                if window_start and window_end:
                    submissions_qs = base_submissions_qs.filter(
                        (Q(completed_on__gte=window_start) & Q(completed_on__lte=window_end)) |
                        (Q(submission_initiated_on__gte=window_start) & Q(submission_initiated_on__lte=window_end))
                    ).distinct().order_by('-completed_on', '-id')
                else:
                    submissions_qs = base_submissions_qs.distinct().order_by('-completed_on', '-id')

                latest_submission = submissions_qs.first()
                if not latest_submission:
                    fallback_submissions_qs = base_submissions_qs.distinct().order_by('-completed_on', '-id')
                    latest_submission = fallback_submissions_qs.first()
                    if latest_submission:
                        task_close_window_miss_fallback_hits += 1
                if not latest_submission:
                    continue
                task_close_submission_matches += 1

                main_form = latest_submission.form
                main_form_id = main_form.id
                if main_form_id not in task_close_dict:
                    task_close_dict[main_form_id] = {
                        'id': main_form.id,
                        'title': main_form.title,
                        'form_type': main_form.form_type or 'standard',
                        'prefix': getattr(main_form, 'prefix', None),
                        'group_type': 'task_close',
                        'sent': []
                    }

                root_planner_id = get_root_planner_source(task)
                is_auto_closed = TaskAuditLog.objects.filter(
                    task=task,
                    task_action='Auto_Closed_Related_Task'
                ).exists()

                task_close_dict[main_form_id]['sent'].append({
                    'task_id': task.id,
                    'task_name': task.task_name,
                    'task_status': task.status,
                    'status': task.status,
                    'main_form_title': task.followup_task_form_id.title if task.followup_task_form_id else None,
                    'assigned_users': get_assigned_user_names(task),
                    'task_started_at': get_task_started_at(task),
                    'form_submission_id': latest_submission.id,
                    'form_id': main_form.id,
                    'is_completed': True,
                    'is_auto_closed': is_auto_closed,
                    'completed_by': latest_submission.completed_by_id,
                    'completed_on': latest_submission.completed_on.isoformat() if latest_submission.completed_on else None,
                    'submission_initiated_on': latest_submission.submission_initiated_on.isoformat() if latest_submission.submission_initiated_on else None,
                    'id': latest_submission.id,
                    'submission_type': "[Task-Close]",
                    'can_reopen': task.id in owner_task_ids,
                    'source': 'planner' if root_planner_id is not None else 'task',
                    'source_ref': root_planner_id if root_planner_id is not None else task.id,
                })

            logger.info(
                "completed-tasks user=%s task_close_completion_matches=%s task_close_without_completion_log=%s task_close_no_question_matches=%s task_close_submission_matches=%s task_close_window_miss_fallback_hits=%s task_close_groups=%s",
                user.id,
                task_close_completion_matches,
                task_close_without_completion_log,
                task_close_no_question_matches,
                task_close_submission_matches,
                task_close_window_miss_fallback_hits,
                len(task_close_dict),
            )

            response_data.extend(task_close_dict.values())
            response_sent_ids = []
            for group in response_data:
                for sent_item in group.get('sent', []):
                    sent_id = sent_item.get('form_submission_id')
                    if sent_id is not None:
                        response_sent_ids.append(sent_id)
            logger.info(
                "completed-tasks user=%s final_response_groups=%s final_response_sent_rows=%s sent_id_sample=%s",
                user.id,
                len(response_data),
                len(response_sent_ids),
                response_sent_ids[:10],
            )
            return Response(response_data, status=status.HTTP_200_OK)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error retrieving completed tasks for user {user_id or request.user.id}: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class OrganizationFormsForTaskView(APIView):
    """Get forms available for task creation (organization-wise filtering)"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        forms = Form.objects.filter(
            organization=request.user.organization,
            is_deleted=False,
            is_archived=False
        ).select_related('form_admin').order_by('-created_at')

        form_data = [{
            'id': form.id,
            'title': form.title,
            'form_type': form.get_form_type_display(),
            'created_by': f"{form.form_admin.first_name} {form.form_admin.last_name}".strip() or form.form_admin.username if form.form_admin else 'N/A'
        } for form in forms]

        return Response(form_data)


class TaskDownloadTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tasks'

        # Write headers
        headers = ['task_name', 'description', 'form', 'start_date', 'end_date']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write example row
        example_data = [
            'Sample Task Name',
            'This is a task description',
            'Your Form Title Here',
            '2023-12-01T09:00:00Z',
            '2023-12-31T17:00:00Z'
        ]
        for col_num, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col_num, value=value)

        # Set column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        # Create HttpResponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="task_bulk_upload_template.xlsx"'
        return response


class TaskBulkValidateView(APIView):
    """Validate task bulk upload file for missing fields and form name mismatches"""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        """Validate task Excel file before upload"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        if not (file.name.endswith('.xlsx') or file.name.endswith('.xls')):
            return Response({'error': 'Only Excel files (.xlsx, .xls) are allowed'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # Read headers from first row
            headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]

            # Create header mapping
            required_headers = ['task_name', 'form', 'start_date', 'end_date']
            header_indices = {}
            for req_header in required_headers:
                if req_header not in headers:
                    return Response({'error': f'Missing required column: {req_header}'}, status=status.HTTP_400_BAD_REQUEST)
                header_indices[req_header] = headers.index(req_header)

            # Optional columns
            optional_headers = ['description']
            for opt_header in optional_headers:
                if opt_header in headers:
                    header_indices[opt_header] = headers.index(opt_header)
                else:
                    header_indices[opt_header] = None

            errors = []
            valid_records = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    task_name = str(row[header_indices['task_name']]).strip() if row[header_indices['task_name']] else ''
                    form_title = str(row[header_indices['form']]).strip() if row[header_indices['form']] else ''
                    start_date_str = str(row[header_indices['start_date']]).strip() if row[header_indices['start_date']] else ''
                    end_date_str = str(row[header_indices['end_date']]).strip() if row[header_indices['end_date']] else ''
                    description = str(row[header_indices['description']]).strip() if header_indices['description'] is not None and row[header_indices['description']] else ''

                    row_errors = []

                    # Check for missing required fields
                    missing_fields = []
                    if not task_name:
                        missing_fields.append('task_name')
                    if not form_title:
                        missing_fields.append('form')
                    if not start_date_str:
                        missing_fields.append('start_date')
                    if not end_date_str:
                        missing_fields.append('end_date')

                    if missing_fields:
                        row_errors.append({
                            'row': row_idx,
                            'field': ', '.join(missing_fields),
                            'message': f'Missing required fields: {", ".join(missing_fields)}'
                        })

                    # Validate form identifier (name or URL)
                    if form_title:
                        identifier_type, identifier_value = extract_form_identifier(form_title)
                        if not validate_form_identifier(identifier_type, identifier_value, request.user.organization):
                            row_errors.append({
                                'row': row_idx,
                                'field': 'form',
                                'message': f'Form "{form_title}" not found or not accessible in your organization'
                            })

                    # Validate date formats
                    def parse_date_value(date_val):
                        # Create IST timezone (UTC+5:30)
                        ist_tz = datetime.timezone(datetime.timedelta(hours=5, minutes=30))

                        # Accept datetime, ISO strings, and 'YYYY-MM-DD HH:MM' without seconds
                        if isinstance(date_val, datetime.datetime):
                            if date_val.tzinfo is None:
                                # Assume naive datetime is in IST
                                return date_val.replace(tzinfo=ist_tz)
                            else:
                                # Convert to IST if it has timezone info
                                return date_val.astimezone(ist_tz)

                        s = str(date_val).strip()
                        if not s:
                            raise ValueError('Empty date')

                        s = s.replace('Z', '+00:00')
                        formats = [
                            '%Y-%m-%dT%H:%M:%S%z',  # ISO with timezone
                            '%Y-%m-%dT%H:%M:%S',    # ISO without timezone
                            '%Y-%m-%d %H:%M:%S',    # Space separated with seconds
                            '%Y-%m-%d %H:%M',       # Space separated without seconds - set seconds to 0
                            '%Y-%m-%d',             # Date only - assume IST 00:00:00
                        ]

                        for fmt in formats:
                            try:
                                dt = datetime.datetime.strptime(s, fmt)
                                if fmt == '%Y-%m-%d':
                                    # Date-only format: assume start of day in IST
                                    dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
                                elif fmt == '%Y-%m-%d %H:%M':
                                    # Time format without seconds: ensure seconds are 0
                                    dt = dt.replace(second=0, microsecond=0)
                                if dt.tzinfo is None:
                                    dt = dt.replace(tzinfo=ist_tz)
                                return dt
                            except ValueError:
                                continue

                        try:
                            dt = datetime.datetime.fromisoformat(s)
                            if dt.tzinfo is None:
                                dt = dt.replace(tzinfo=ist_tz)
                            return dt
                        except Exception:
                            raise ValueError(f'Unable to parse date: {s}')

                    # Validate start_date format
                    if start_date_str:
                        try:
                            parse_date_value(start_date_str)
                        except ValueError as ve:
                            row_errors.append({
                                'row': row_idx,
                                'field': 'start_date',
                                'message': f'Invalid start date format: {ve}'
                            })

                    # Validate end_date format
                    if end_date_str:
                        try:
                            parse_date_value(end_date_str)
                        except ValueError as ve:
                            row_errors.append({
                                'row': row_idx,
                                'field': 'end_date',
                                'message': f'Invalid end date format: {ve}'
                            })

                    if row_errors:
                        errors.extend(row_errors)
                    else:
                        # This row passed validation
                        valid_records.append({
                            'row': row_idx,
                            'task_name': task_name,
                            'form': form_title,
                            'start_date': start_date_str,
                            'end_date': end_date_str,
                            'description': description
                        })

                except Exception as e:
                    errors.append({
                        'row': row_idx,
                        'field': 'general',
                        'message': f'Unexpected error: {str(e)}'
                    })

            if errors:
                return Response({
                    'error': 'Validation failed',
                    'errors': errors,
                    'valid_records_count': len(valid_records),
                    'total_rows': len(list(ws.rows)) - 1
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': 'Validation successful',
                'errors': [],
                'valid_records': valid_records,
                'valid_records_count': len(valid_records),
                'total_rows': len(list(ws.rows)) - 1
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': f'Failed to process Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class TaskBulkImportView(APIView):
    """Accept JSON payload from frontend (parsed rows) for bulk import.
    This supports frontends that parse the file client-side and send rows as JSON.
    Endpoint: POST /api/tasks/bulk-import/
    Expected payload: { "data": [ {task_name, form, start_date, end_date, description, rowIndex}, ... ] }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            payload = request.data

            # support both {data: [...]} and direct array payload
            if isinstance(payload, dict) and 'data' in payload:
                rows = payload['data']
            elif isinstance(payload, list):
                rows = payload
            else:
                return Response({'error': 'Invalid payload. Expecting JSON array under "data" or root array.'}, status=status.HTTP_400_BAD_REQUEST)

            errors = []
            success_count = 0

            def parse_date_value(date_val):
                # Create IST timezone (UTC+5:30)
                ist_tz = datetime.timezone(datetime.timedelta(hours=5, minutes=30))

                # Accept datetime, ISO strings, and 'YYYY-MM-DD HH:MM' without seconds
                if isinstance(date_val, datetime.datetime):
                    if date_val.tzinfo is None:
                        # Assume naive datetime is in IST
                        return date_val.replace(tzinfo=ist_tz)
                    else:
                        # Convert to IST if it has timezone info
                        return date_val.astimezone(ist_tz)

                s = str(date_val).strip()
                if not s:
                    raise ValueError('Empty date')

                s = s.replace('Z', '+00:00')
                formats = [
                    '%Y-%m-%dT%H:%M:%S%z',  # ISO with timezone
                    '%Y-%m-%dT%H:%M:%S',    # ISO without timezone
                    '%Y-%m-%d %H:%M:%S',    # Space separated with seconds
                    '%Y-%m-%d %H:%M',       # Space separated without seconds
                    '%Y-%m-%d',             # Date only - assume IST 00:00:00
                ]

                for fmt in formats:
                    try:
                        dt = datetime.datetime.strptime(s, fmt)
                        if fmt == '%Y-%m-%d':
                            # Date-only format: assume start of day in IST
                            dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=ist_tz)
                        return dt
                    except ValueError:
                        continue

                try:
                    dt = datetime.datetime.fromisoformat(s)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=ist_tz)
                    return dt
                except Exception:
                    raise ValueError(f'Unable to parse date: {s}')

            with transaction.atomic():
                for row in rows:
                    row_index = row.get('rowIndex') or row.get('row')
                    task_name = str(row.get('task_name', '')).strip()
                    form_title = str(row.get('form', '')).strip()
                    start_date_raw = row.get('start_date', '')
                    end_date_raw = row.get('end_date', '')
                    description = str(row.get('description', '')).strip()

                    if not task_name or not form_title or not start_date_raw or not end_date_raw:
                        errors.append({'row': row_index or None, 'error': 'Missing required fields: task_name, form, start_date, end_date'})
                        continue

                    # Extract and validate form identifier
                    identifier_type, identifier_value = extract_form_identifier(form_title)
                    if not validate_form_identifier(identifier_type, identifier_value, request.user.organization):
                        errors.append({'row': row_index or None, 'error': f'Form "{form_title}" not found or not accessible'})
                        continue

                    # Get the form object
                    form = get_form_by_identifier(identifier_type, identifier_value, request.user.organization)
                    if not form:
                        errors.append({'row': row_index or None, 'error': f'Form "{form_title}" not found or not accessible'})
                        continue

                    # Parse dates
                    try:
                        start_date = parse_date_value(start_date_raw)
                        end_date = parse_date_value(end_date_raw)
                    except ValueError as ve:
                        errors.append({'row': row_index or None, 'error': f'Invalid date format: {ve}'})
                        continue

                    if start_date > end_date:
                        errors.append({'row': row_index or None, 'error': 'End date must be on or after start date'})
                        continue

                    # Create task
                    try:
                        task = Task.objects.create(
                            task_name=task_name,
                            description=description if description else None,
                            form=form,
                            organization=request.user.organization,
                            start_date=start_date,
                            end_date=end_date,
                            created_by=request.user,
                            status='not_assigned'
                        )

                        TaskAuditLog.objects.create(
                            task=task,
                            task_action='created via bulk upload',
                            action_by=request.user
                        )

                        success_count += 1
                    except Exception as e:
                        errors.append({'row': row_index or None, 'error': f'Unexpected error: {str(e)}'})

            return Response({'success_count': success_count, 'errors': errors, 'total_processed': len(rows)}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class FormAssigneesView(APIView):
    """Get assignees associated with a specific form through tasks"""
    permission_classes = [IsAuthenticated]

    def get(self, request, form_id):
        assignees = TaskAssignee.objects.filter(
            task__form_id=form_id,
            task__organization=request.user.organization
        ).select_related('task', 'assigned_user__location', 'assigned_group')

        data = []
        status_map = {
            'not_started': 'Not Started',
            'not_assigned': 'Not Assigned',
            'in_progress': 'In Progress',
            'completed': 'Completed',
            'cancelled': 'Cancelled'
        }

        for assignee in assignees:
            task = assignee.task
            if assignee.assigned_user:
                assignee_name = f"{assignee.assigned_user.first_name} {assignee.assigned_user.last_name}".strip() or assignee.assigned_user.username
                assignee_location = assignee.assigned_user.location.name if assignee.assigned_user.location else "N/A"
            else:
                assignee_name = assignee.assigned_group.name
                assignee_location = "N/A"

            status = status_map.get(task.status, task.status)

            data.append({
                'task_id': task.id,
                'task_name': task.task_name,
                'assignee': assignee_name,
                'assignee_location': assignee_location,
                'status': status
            })

        return Response(data)

class OrganizationUsersForTaskView(APIView):
    """Get users available for task assignment (organization-wise filtering)"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        users = CustomUser.objects.filter(
            organization=request.user.organization,
            is_deleted=False,
            is_archived=False
        ).exclude(disable=True).order_by('username')

        user_data = [{
            'id': user.id,
            'username': user.username,
            'full_name': f"{user.first_name} {user.last_name}".strip() or user.username,
            'email': user.email,
            'phone_number': user.phone,
            'designation': user.designation.name if user.designation else None,
            'department': user.department.name if user.department else None
        } for user in users]

        return Response(user_data)

class OrganizationGroupsForTaskView(APIView):
    """Get groups available for task assignment (organization-wise filtering)"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        groups = Groups.objects.filter(
            organization=request.user.organization,
            is_deleted=False,
            is_archived=False
        ).order_by('name')

        group_data = [{
            'id': group.id,
            'name': group.name,
            'description': group.description,
            'member_count': group.members.count()
        } for group in groups]

        return Response(group_data)


class FormsAssociatedWithTasksView(APIView):
    """Get forms that are associated with tasks for planner module"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        forms = Form.objects.filter(
            organization=request.user.organization,
            is_deleted=False,
            is_archived=False,
            task__organization=request.user.organization  # Forms that have at least one task in the same organization
        ).select_related('form_admin').distinct().order_by('-created_at')

        # Get all task IDs for these forms to fetch assignees efficiently
        task_ids = Task.objects.filter(
            form__in=forms,
            organization=request.user.organization
        ).values_list('id', flat=True)

        # Fetch all assignees for these tasks
        assignees = TaskAssignee.objects.filter(
            task_id__in=task_ids
        ).select_related('task', 'assigned_user', 'assigned_group')

        # Build a set of unique assignees per form
        form_assignee_set = defaultdict(set)
        form_status_counts = defaultdict(lambda: defaultdict(int))
        for assignee in assignees:
            if assignee.assigned_user:
                form_assignee_set[assignee.task.form_id].add(('user', assignee.assigned_user.id))
            if assignee.assigned_group:
                form_assignee_set[assignee.task.form_id].add(('group', assignee.assigned_group.id))

            # Count by status for this assignee
            form_status_counts[assignee.task.form_id][assignee.task.status] += 1

        form_data = [{
            'id': form.id,
            'title': form.title,
            'form_type': form.get_form_type_display(),
            'created_by': f"{form.form_admin.first_name} {form.form_admin.last_name}".strip() or form.form_admin.username if form.form_admin else 'N/A',
            'created_at': form.created_at.isoformat() if form.created_at else None,
            'assignee_count': len(form_assignee_set.get(form.id, set())),
            'not_started_count': form_status_counts[form.id]['not_started'],
            'in_progress_count': form_status_counts[form.id]['in_progress'],
            'completed_count': form_status_counts[form.id]['completed'],
        } for form in forms]

        return Response(form_data)

